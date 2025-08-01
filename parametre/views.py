from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg, Min, Max, Sum
from django.contrib.auth.models import User, Group
from django.contrib import messages
from .models import Region, Ville, Operateur, HistoriqueMotDePasse
from article.models import Article
from commande.models import Commande, EtatCommande, EnumEtatCmd
from django.contrib.messages import success, error
from django.views.decorators.http import require_POST
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash # Required for password change
import csv
import json
from datetime import datetime, timedelta
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

@login_required
def dashboard(request):
    """Page d'accueil de l'interface administrateur - Redirige vers le dashboard KPIs"""
    # Rediriger directement vers le dashboard KPIs pour les administrateurs
    if request.user.is_staff or hasattr(request.user, 'profil_operateur') and request.user.profil_operateur.type_operateur == 'ADMIN':
        return redirect('/kpis/')
    
    # Fallback vers l'ancien dashboard pour les autres types d'utilisateurs
    # Statistiques pour le dashboard
    stats = {
        'total_articles': Article.objects.count(),
        'articles_disponibles': Article.objects.filter(qte_disponible__gt=0).count(),
        'total_commandes': Commande.objects.count(),
        'total_operateurs': Operateur.objects.count(),
        'operateurs_actifs': Operateur.objects.filter(actif=True).count(),
    }
    
    # Ajouter les statistiques des états de commandes
    try:
        from commande.models import EtatCommande, EnumEtatCmd
        
        # Compter les commandes par état en utilisant les libellés des états
        commandes_non_affectees = EtatCommande.objects.filter(
            date_fin__isnull=True,
            enum_etat__libelle__icontains='Non affectée'
        ).count()
        
        commandes_affectees = EtatCommande.objects.filter(
            date_fin__isnull=True,
            enum_etat__libelle__icontains='Affectée'
        ).count()
        
        commandes_erronnees = EtatCommande.objects.filter(
            date_fin__isnull=True,
            enum_etat__libelle__icontains='Erronée'
        ).count()
        
        commandes_doublons = EtatCommande.objects.filter(
            date_fin__isnull=True,
            enum_etat__libelle__icontains='Doublon'
        ).count()
        
        # Commandes nouvelles = commandes sans état actuel
        commandes_avec_etat = EtatCommande.objects.filter(date_fin__isnull=True).values_list('commande_id', flat=True)
        commandes_nouvelles = Commande.objects.exclude(id__in=commandes_avec_etat).count()
        
        stats.update({
            'commandes_non_affectees': commandes_non_affectees,
            'commandes_affectees': commandes_affectees,
            'commandes_erronnees': commandes_erronnees,
            'commandes_doublons': commandes_doublons,
            'commandes_nouvelles': commandes_nouvelles,
        })
    except ImportError:
        # Si les modèles ne sont pas disponibles, mettre des valeurs par défaut
        stats.update({
            'commandes_non_affectees': 0,
            'commandes_affectees': 0,
            'commandes_erronnees': 0,
            'commandes_doublons': 0,
            'commandes_nouvelles': 0,
        })

    # Récupérer quelques opérateurs actifs pour l'affichage des initiales
    active_operators_for_display = Operateur.objects.filter(actif=True, user__is_superuser=False).order_by('id')[:3]

    context = {
        'stats': stats,
        'active_operators_for_display': active_operators_for_display,
    }
    return render(request, 'composant_generale/admin/home.html', context)

@staff_member_required
@login_required
def liste_operateurs(request):
    """Liste des opérateurs"""
    operateurs = Operateur.objects.select_related('user').order_by('nom', 'prenom').exclude(type_operateur='ADMIN')
    
    # Statistiques pour les cartes
    total_operateurs = Operateur.objects.exclude(type_operateur='ADMIN').count()
    operateurs_actifs = Operateur.objects.filter(actif=True).exclude(type_operateur='ADMIN').count()
    operateurs_inactifs = Operateur.objects.filter(actif=False).exclude(type_operateur='ADMIN').count()
    administrateurs = Operateur.objects.filter(type_operateur='ADMIN').count()
    
    # Filtrage par type si spécifié
    type_filter = request.GET.get('type')
    if type_filter:
        operateurs = operateurs.filter(type_operateur=type_filter)
    
    # Recherche
    search = request.GET.get('search')
    if search:
        operateurs = operateurs.filter(
            Q(nom__icontains=search) | 
            Q(prenom__icontains=search) |
            Q(mail__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(operateurs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'type_filter': type_filter,
        'search': search,
        'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN'],
        'total_operateurs': total_operateurs,
        'operateurs_actifs': operateurs_actifs,
        'operateurs_inactifs': operateurs_inactifs,
        'administrateurs': administrateurs,
    }
    return render(request, 'parametre/liste_operateurs.html', context)

@staff_member_required
@login_required
def creer_operateur(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        mail = request.POST.get('mail')
        telephone = request.POST.get('telephone')
        adresse = request.POST.get('adresse')
        type_operateur = request.POST.get('type_operateur')

        # Validation de base
        if not all([username, password, nom, prenom, mail, type_operateur]):
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")
            return render(request, 'parametre/creer_operateur.html', {'form_data': request.POST, 'types_operateur': Operateur.TYPE_OPERATEUR_CHOICES})
        
        if type_operateur == 'ADMIN':
            messages.error(request, "Les opérateurs de type 'Administrateur' ne peuvent pas être créés via ce formulaire.")
            return render(request, 'parametre/creer_operateur.html', {'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})

        if User.objects.filter(username=username).exists():
            messages.error(request, "Ce nom d'utilisateur existe déjà.")
            return render(request, 'parametre/creer_operateur.html', {'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})
        
        if User.objects.filter(email=mail).exists():
            messages.error(request, "Cet email est déjà utilisé.")
            return render(request, 'parametre/creer_operateur.html', {'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})

        try:
            # Créer l'utilisateur Django
            user = User.objects.create_user(
                username=username,
                email=mail,
                password=password,
                first_name=prenom,
                last_name=nom
            )
            user.save()

            # Créer le profil Operateur
            operateur = Operateur.objects.create(
                user=user,
                nom=nom,
                prenom=prenom,
                mail=mail,
                type_operateur=type_operateur,
                telephone=telephone,
                adresse=adresse
            )
            operateur.save()

            # Ajouter l'utilisateur au groupe correspondant au type d'opérateur
            group_name_map = {
                'CONFIRMATION': 'operateur_confirme',
                'LOGISTIQUE': 'operateur_logistique',
                'ADMIN': 'administrateur', # Assurez-vous que ce groupe existe
            }
            group_name = group_name_map.get(type_operateur)
            if group_name:
                group, created = Group.objects.get_or_create(name=group_name)
                user.groups.add(group)

            messages.success(request, f"L'opérateur {prenom} {nom} ({type_operateur}) a été créé avec succès.")
            return redirect('app_admin:liste_operateurs')

        except Exception as e:
            messages.error(request, f"Une erreur est survenue lors de la création de l'opérateur : {e}")
            if 'user' in locals() and user.pk:
                user.delete()
            return render(request, 'parametre/creer_operateur.html', {'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})

    return render(request, 'parametre/creer_operateur.html', {'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})

@staff_member_required
@login_required
def liste_regions(request):
    """Liste des régions avec statistiques"""
    regions = Region.objects.annotate(
        nb_villes=Count('villes'),
        tarif_moyen=Avg('villes__frais_livraison')
    ).order_by('nom_region')
    
    # Recherche
    search = request.GET.get('search')
    if search:
        regions = regions.filter(nom_region__icontains=search)
    
    # Pagination
    paginator = Paginator(regions, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search': search,
        'total_regions': Region.objects.count(),
        'total_villes': Ville.objects.count(),
    }
    return render(request, 'parametre/liste_regions.html', context)

@staff_member_required
@login_required
def detail_region(request, region_id):
    """Détail d'une région avec ses villes"""
    region = get_object_or_404(Region, id=region_id)
    villes = region.villes.order_by('nom')
    
    # Statistiques de la région
    stats = {
        'nb_villes': villes.count(),
        'tarif_min': villes.aggregate(min_tarif=Min('frais_livraison'))['min_tarif'],
        'tarif_max': villes.aggregate(max_tarif=Max('frais_livraison'))['max_tarif'],
        'tarif_moyen': villes.aggregate(avg_tarif=Avg('frais_livraison'))['avg_tarif'],
    }
    
    # Recherche dans les villes
    search = request.GET.get('search')
    if search:
        villes = villes.filter(nom__icontains=search)
    
    # Pagination des villes
    paginator = Paginator(villes, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'region': region,
        'page_obj': page_obj,
        'stats': stats,
        'search': search,
    }
    return render(request, 'parametre/detail_region.html', context)

@staff_member_required
@login_required
def liste_villes(request):
    """Liste complète des villes"""
    villes = Ville.objects.select_related('region').order_by('region__nom_region', 'nom')
    
    # Filtrage par région
    region_filter = request.GET.get('region')
    if region_filter:
        villes = villes.filter(region_id=region_filter)
    
    # Filtrage par tarif
    tarif_min = request.GET.get('tarif_min')
    tarif_max = request.GET.get('tarif_max')
    if tarif_min:
        villes = villes.filter(frais_livraison__gte=tarif_min)
    if tarif_max:
        villes = villes.filter(frais_livraison__lte=tarif_max)
    
    # Recherche
    search = request.GET.get('search')
    if search:
        villes = villes.filter(
            Q(nom__icontains=search) | 
            Q(region__nom_region__icontains=search)
        )
    
    # Pagination
    paginator = Paginator(villes, 30)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Liste des régions pour le filtre
    regions = Region.objects.order_by('nom_region')
    
    context = {
        'page_obj': page_obj,
        'regions': regions,
        'region_filter': region_filter,
        'tarif_min': tarif_min,
        'tarif_max': tarif_max,
        'search': search,
    }
    return render(request, 'parametre/liste_villes.html', context)

@staff_member_required
@login_required
def synchronisation(request):
    """Page de synchronisation"""
    return render(request, 'parametre/synchronisation.html')

@staff_member_required
@login_required
def logs(request):
    """Page des logs système"""
    return render(request, 'parametre/logs.html')

@staff_member_required
@login_required
def detail_operateur(request, pk):
    """Afficher les détails d'un opérateur avec ses commandes affectées"""
    operateur = get_object_or_404(Operateur.objects.select_related('user'), pk=pk)
    
    # Récupérer les commandes affectées à cet opérateur
    from commande.models import EtatCommande, Commande
    from django.db.models import Sum, Count, Q
    from django.core.paginator import Paginator
    
    # Commandes actuellement affectées à cet opérateur
    commandes_affectees = Commande.objects.filter(
        etats__operateur=operateur,
        etats__date_fin__isnull=True
    ).distinct().order_by('-date_cmd')
    
    # Historique de toutes les commandes traitées par cet opérateur
    commandes_historique = Commande.objects.filter(
        etats__operateur=operateur
    ).distinct().order_by('-date_cmd')
    
    # Statistiques
    total_commandes_affectees = commandes_affectees.count()
    total_commandes_traitees = commandes_historique.count()
    
    # Montant total des commandes affectées
    montant_total_affectees = commandes_affectees.aggregate(
        total=Sum('total_cmd')
    )['total'] or 0
    
    # Montant total de toutes les commandes traitées
    montant_total_traitees = commandes_historique.aggregate(
        total=Sum('total_cmd')
    )['total'] or 0
    
    # Statistiques par état pour cet opérateur
    etats_stats = EtatCommande.objects.filter(
        operateur=operateur
    ).values(
        'enum_etat__libelle', 'enum_etat__couleur'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Historique des modifications de mots de passe
    historique_mots_de_passe = HistoriqueMotDePasse.objects.filter(
        operateur=operateur
    ).select_related('administrateur').order_by('-date_modification')[:10]  # Les 10 dernières modifications
    
    # Dernière modification de mot de passe
    derniere_modification_mdp = historique_mots_de_passe.first() if historique_mots_de_passe.exists() else None
    
    # Pagination pour les commandes affectées
    paginator = Paginator(commandes_affectees, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Recherche dans les commandes affectées
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_affectees = commandes_affectees.filter(
            Q(num_cmd__icontains=search_query) |
            Q(id_yz__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query)
        )
        paginator = Paginator(commandes_affectees, 10)
        page_obj = paginator.get_page(page_number)
    
    context = {
        'operateur': operateur,
        'page_obj': page_obj,
        'search_query': search_query,
        'total_commandes_affectees': total_commandes_affectees,
        'total_commandes_traitees': total_commandes_traitees,
        'montant_total_affectees': montant_total_affectees,
        'montant_total_traitees': montant_total_traitees,
        'etats_stats': etats_stats,
        'historique_mots_de_passe': historique_mots_de_passe,
        'derniere_modification_mdp': derniere_modification_mdp,
    }
    return render(request, 'parametre/detail_operateur.html', context)

@staff_member_required
@login_required
def modifier_operateur(request, pk):
    """Modifier un opérateur existant"""
    operateur = get_object_or_404(Operateur.objects.select_related('user'), pk=pk)
    user = operateur.user

    if request.method == 'POST':
        username = request.POST.get('username')
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        mail = request.POST.get('mail')
        telephone = request.POST.get('telephone')
        adresse = request.POST.get('adresse')
        type_operateur = request.POST.get('type_operateur')
        actif = request.POST.get('actif') == 'on'
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        # Validation de base
        if not all([username, nom, prenom, mail, type_operateur]):
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")
            return render(request, 'parametre/modifier_operateur.html', {'operateur': operateur, 'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})
        
        if type_operateur == 'ADMIN':
            messages.error(request, "Les opérateurs de type 'Administrateur' ne peuvent pas être gérés via ce formulaire.")
            return render(request, 'parametre/modifier_operateur.html', {'operateur': operateur, 'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})

        # Vérifier si le nom d'utilisateur ou l'email existe déjà pour un AUTRE utilisateur
        if User.objects.filter(username=username).exclude(pk=user.pk).exists():
            messages.error(request, "Ce nom d'utilisateur existe déjà pour un autre utilisateur.")
            return render(request, 'parametre/modifier_operateur.html', {'operateur': operateur, 'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})
        
        if User.objects.filter(email=mail).exclude(pk=user.pk).exists():
            messages.error(request, "Cet email est déjà utilisé par un autre utilisateur.")
            return render(request, 'parametre/modifier_operateur.html', {'operateur': operateur, 'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})

        # Validation du mot de passe
        if password or confirm_password:
            if password != confirm_password:
                messages.error(request, "Les mots de passe ne correspondent pas.")
                return render(request, 'parametre/modifier_operateur.html', {'operateur': operateur, 'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})
            if password and len(password) < 8:
                messages.error(request, "Le mot de passe doit contenir au moins 8 caractères.")
                return render(request, 'parametre/modifier_operateur.html', {'operateur': operateur, 'form_data': request.POST, 'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN']})

        try:
            # Mettre à jour l'utilisateur Django
            user.username = username
            user.email = mail
            user.first_name = prenom
            user.last_name = nom
            user.is_active = actif
            if password:
                user.set_password(password)
            user.save()

            # Mettre à jour le profil Operateur
            operateur.nom = nom
            operateur.prenom = prenom
            operateur.mail = mail
            operateur.type_operateur = type_operateur
            operateur.telephone = telephone
            operateur.adresse = adresse
            operateur.actif = actif
            operateur.save()
            
            # Gérer l'appartenance aux groupes
            group_name_map = {
                'CONFIRMATION': 'operateur_confirme',
                'LOGISTIQUE': 'operateur_logistique',
            }
            
            # Retirer l'utilisateur de tous les groupes d'opérateur précédents
            for group_name in group_name_map.values():
                group = Group.objects.filter(name=group_name).first()
                if group and user.groups.filter(name=group_name).exists():
                    user.groups.remove(group)

            # Ajouter l'utilisateur au nouveau groupe correspondant au type d'opérateur
            new_group_name = group_name_map.get(type_operateur)
            if new_group_name:
                group, created = Group.objects.get_or_create(name=new_group_name)
                user.groups.add(group)

            messages.success(request, f"L'opérateur {prenom} {nom} a été modifié avec succès.")
            return redirect('app_admin:liste_operateurs')

        except Exception as e:
            messages.error(request, f"Une erreur est survenue lors de la modification de l'opérateur : {e}")

    context = {
        'operateur': operateur,
        'types_operateur': [choice for choice in Operateur.TYPE_OPERATEUR_CHOICES if choice[0] != 'ADMIN'],
        'form_data': {
            'username': operateur.user.username,
            'nom': operateur.nom,
            'prenom': operateur.prenom,
            'mail': operateur.mail,
            'telephone': operateur.telephone,
            'adresse': operateur.adresse,
            'type_operateur': operateur.type_operateur,
            'actif': 'on' if operateur.actif else '',
        }
    }
    return render(request, 'parametre/modifier_operateur.html', context)

@staff_member_required
@login_required
def supprimer_operateur(request, pk):
    """Supprimer un opérateur"""
    if request.method == 'POST':
        operateur = get_object_or_404(Operateur, pk=pk)
        user = operateur.user

        try:
            # Supprimer l'opérateur et l'utilisateur associé
            user.delete()
            messages.success(request, f"L'opérateur {operateur.prenom} {operateur.nom} a été supprimé avec succès.")
        except Exception as e:
            messages.error(request, f"Une erreur est survenue lors de la suppression de l'opérateur : {e}")

    return redirect('app_admin:liste_operateurs')

@require_POST
@staff_member_required
@login_required
def supprimer_operateurs_masse(request):
    selected_ids = request.POST.getlist('ids[]')
    if not selected_ids:
        messages.error(request, "Aucun opérateur sélectionné pour la suppression.")
        return redirect('app_admin:liste_operateurs')

    try:
        operateurs_a_supprimer = Operateur.objects.filter(pk__in=selected_ids)
        count = operateurs_a_supprimer.count()
        for operateur in operateurs_a_supprimer:
            user_to_delete = operateur.user
            operateur.delete()
            user_to_delete.delete()
        messages.success(request, f"{count} opérateur(s) supprimé(s) avec succès.")
    except Exception as e:
        messages.error(request, f"Une erreur est survenue lors de la suppression en masse : {e}")
    
    return redirect('app_admin:liste_operateurs')

@staff_member_required
@login_required
def modifier_region(request, pk):
    region = get_object_or_404(Region, pk=pk)
    if request.method == 'POST':
        nom_region = request.POST.get('nom_region')
        if nom_region:
            region.nom_region = nom_region
            region.save()
            messages.success(request, "La région a été modifiée avec succès.")
            return redirect('app_admin:detail_region', region_id=region.pk)
        else:
            messages.error(request, "Le nom de la région ne peut pas être vide.")
    context = {
        'region': region
    }
    return render(request, 'parametre/modifier_region.html', context)

@staff_member_required
@login_required
def creer_region(request):
    if request.method == 'POST':
        nom_region = request.POST.get('nom_region')
        if nom_region:
            if Region.objects.filter(nom_region__iexact=nom_region).exists():
                messages.error(request, "Une région avec ce nom existe déjà.")
            else:
                Region.objects.create(nom_region=nom_region)
                messages.success(request, f"La région '{nom_region}' a été créée avec succès.")
                return redirect('app_admin:liste_regions')
        else:
            messages.error(request, "Le nom de la région ne peut pas être vide.")
    return render(request, 'parametre/creer_region.html')

@staff_member_required
@login_required
def modifier_ville(request, pk):
    ville = get_object_or_404(Ville, pk=pk)
    regions = Region.objects.all().order_by('nom_region')

    if request.method == 'POST':
        nom_ville = request.POST.get('nom')
        frais_livraison = request.POST.get('frais_livraison')
        frequence_livraison = request.POST.get('frequence_livraison')
        region_id = request.POST.get('region')

        if not all([nom_ville, frais_livraison, frequence_livraison, region_id]):
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")
        else:
            try:
                region = get_object_or_404(Region, pk=region_id)
                ville.nom = nom_ville
                ville.frais_livraison = float(frais_livraison)
                ville.frequence_livraison = frequence_livraison
                ville.region = region
                ville.save()
                messages.success(request, f"La ville '{nom_ville}' a été modifiée avec succès.")
                return redirect('app_admin:liste_villes')
            except ValueError:
                messages.error(request, "Les frais de livraison doivent être un nombre valide.")
            except Exception as e:
                messages.error(request, f"Une erreur est survenue lors de la modification de la ville : {e}")

    context = {
        'ville': ville,
        'regions': regions,
        'formatted_frais_livraison': f'{ville.frais_livraison:.2f}' if ville.frais_livraison is not None else ''
    }
    return render(request, 'parametre/modifier_ville.html', context)

@staff_member_required
@login_required
def creer_ville(request):
    regions = Region.objects.all().order_by('nom_region')
    if request.method == 'POST':
        nom_ville = request.POST.get('nom')
        frais_livraison = request.POST.get('frais_livraison')
        frequence_livraison = request.POST.get('frequence_livraison')
        region_id = request.POST.get('region')

        if not all([nom_ville, frais_livraison, frequence_livraison, region_id]):
            messages.error(request, "Tous les champs obligatoires doivent être remplis.")
        else:
            try:
                region = get_object_or_404(Region, pk=region_id)
                if Ville.objects.filter(nom__iexact=nom_ville, region=region).exists():
                    messages.error(request, "Une ville avec ce nom existe déjà dans cette région.")
                else:
                    Ville.objects.create(
                        nom=nom_ville,
                        frais_livraison=float(frais_livraison),
                        frequence_livraison=frequence_livraison,
                        region=region
                    )
                    messages.success(request, f"La ville '{nom_ville}' a été créée avec succès.")
                    return redirect('app_admin:liste_villes')
            except ValueError:
                messages.error(request, "Les frais de livraison doivent être un nombre valide.")
            except Exception as e:
                messages.error(request, f"Une erreur est survenue lors de la création de la ville : {e}")

    context = {
        'regions': regions,
    }
    return render(request, 'parametre/creer_ville.html', context)

@staff_member_required
@login_required
def supprimer_ville(request, pk):
    ville = get_object_or_404(Ville, pk=pk)
    if request.method == 'POST':
        try:
            ville.delete()
            messages.success(request, f"La ville '{ville.nom}' a été supprimée avec succès.")
        except Exception as e:
            messages.error(request, f"Une erreur est survenue lors de la suppression de la ville : {e}")
        return redirect('app_admin:liste_villes')
    return redirect('app_admin:liste_villes')

@require_POST
@staff_member_required
@login_required
def supprimer_villes_masse(request):
    selected_ids = request.POST.getlist('ids[]')
    if not selected_ids:
        messages.error(request, "Aucune ville sélectionnée pour la suppression.")
        return redirect('app_admin:liste_villes')

    try:
        count = Ville.objects.filter(pk__in=selected_ids).delete()[0]
        messages.success(request, f"{count} ville(s) supprimée(s) avec succès.")
    except Exception as e:
        messages.error(request, f"Une erreur est survenue lors de la suppression en masse : {e}")
    
    return redirect('app_admin:liste_villes')

@staff_member_required
@login_required
def supprimer_region(request, pk):
    region = get_object_or_404(Region, pk=pk)
    if request.method == 'POST':
        try:
            region.delete()
            messages.success(request, f"La région '{region.nom_region}' a été supprimée avec succès.")
        except Exception as e:
            messages.error(request, f"Une erreur est survenue lors de la suppression de la région : {e}")
        return redirect('app_admin:liste_regions')
    return redirect('app_admin:liste_regions')

@require_POST
@staff_member_required
@login_required
def supprimer_regions_masse(request):
    selected_ids = request.POST.getlist('ids[]')
    if not selected_ids:
        messages.error(request, "Aucune région sélectionnée pour la suppression.")
        return redirect('app_admin:liste_regions')

    try:
        # On récupère les régions, puis on les supprime
        count = Region.objects.filter(pk__in=selected_ids).delete()[0]
        messages.success(request, f"{count} région(s) supprimée(s) avec succès.")
    except Exception as e:
        messages.error(request, f"Une erreur est survenue lors de la suppression en masse : {e}")
    
    return redirect('app_admin:liste_regions')

@staff_member_required
@login_required
def detail_ville(request, pk):
    ville = get_object_or_404(Ville, pk=pk)
    context = {
        'ville': ville
    }
    return render(request, 'parametre/detail_ville.html', context)

@staff_member_required
@login_required
def admin_profile(request):
    try:
        operateur = Operateur.objects.get(user=request.user)
    except Operateur.DoesNotExist:
        operateur = None
    context = {
        'user': request.user,
        'operateur': operateur,
    }
    return render(request, 'parametre/admin_profile.html', context)

@staff_member_required
@login_required
def modifier_admin_profile(request):
    try:
        operateur = Operateur.objects.get(user=request.user)
    except Operateur.DoesNotExist:
        # If admin doesn't have an Operateur profile, create one.
        # This might happen if the admin was created directly via Django admin and not as an Operateur.
        operateur = Operateur.objects.create(
            user=request.user,
            nom=request.user.last_name if request.user.last_name else '',
            prenom=request.user.first_name if request.user.first_name else '',
            mail=request.user.email,
            type_operateur='ADMIN'
        )

    if request.method == 'POST':
        # Handle User fields
        request.user.first_name = request.POST.get('first_name', request.user.first_name)
        request.user.last_name = request.POST.get('last_name', request.user.last_name)
        request.user.email = request.POST.get('email', request.user.email)
        request.user.save()

        # Handle Operateur fields
        operateur.telephone = request.POST.get('telephone')
        operateur.adresse = request.POST.get('adresse')
        if 'photo' in request.FILES:
            operateur.photo = request.FILES['photo']
        elif request.POST.get('clear_photo') == 'on': # For clearing existing photo
            operateur.photo = None

        operateur.save()

        messages.success(request, "Votre profil a été mis à jour avec succès.")
        return redirect('app_admin:profile')

    context = {
        'user': request.user,
        'operateur': operateur,
    }
    return render(request, 'parametre/modifier_admin_profile.html', context)

@staff_member_required
@login_required
def changer_mot_de_passe_admin(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important! Keeps the user logged in
            messages.success(request, 'Votre mot de passe a été mis à jour avec succès !')
            return redirect('app_admin:profile')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'parametre/changer_mot_de_passe_admin.html', {'form': form})

# ======================== VUES SERVICE APRÈS-VENTE POUR ADMIN ========================

@staff_member_required
@login_required
def sav_commandes_retournees(request):
    """Vue admin pour afficher les commandes retournées"""
    from commande.models import Commande, EtatCommande
    
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Retournée',
        etats__date_fin__isnull=True
    ).select_related('client', 'ville', 'ville__region').prefetch_related(
        'etats__enum_etat', 'etats__operateur', 'paniers__article'
    ).order_by('-etats__date_debut').distinct()
    
    # Pagination
    paginator = Paginator(commandes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Commandes Retournées',
        'subtitle': 'Commandes retournées par les clients ou opérateurs logistiques',
        'icon': 'fa-undo',
        'color': 'red'
    }
    return render(request, 'parametre/sav/liste_commandes_sav.html', context)

@staff_member_required
@login_required
def sav_commandes_reportees(request):
    """Vue admin pour afficher les commandes reportées"""
    from commande.models import Commande, EtatCommande
    
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Reportée',
        etats__date_fin__isnull=True
    ).select_related('client', 'ville', 'ville__region').prefetch_related(
        'etats__enum_etat', 'etats__operateur', 'paniers__article'
    ).order_by('-etats__date_debut').distinct()
            
    # Pagination
    paginator = Paginator(commandes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Commandes Reportées',
        'subtitle': 'Commandes dont la livraison a été reportée',
        'icon': 'fa-clock',
        'color': 'orange'
    }
    return render(request, 'parametre/sav/liste_commandes_sav.html', context)

@staff_member_required
@login_required
def sav_livrees_partiellement(request):
    """Vue admin pour afficher les commandes livrées partiellement"""
    from commande.models import Commande, EtatCommande
    
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Livrée Partiellement',
        etats__date_fin__isnull=True
    ).select_related('client', 'ville', 'ville__region').prefetch_related(
        'etats__enum_etat', 'etats__operateur', 'paniers__article'
    ).order_by('-etats__date_debut').distinct()
    
    # Pagination
    paginator = Paginator(commandes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Commandes Livrées Partiellement',
        'subtitle': 'Commandes avec livraison partielle',
        'icon': 'fa-box-open',
        'color': 'yellow'
    }
    return render(request, 'parametre/sav/liste_commandes_sav.html', context)

@staff_member_required
@login_required
def sav_annulees_sav(request):
    """Vue admin pour afficher les commandes annulées au niveau SAV"""
    from commande.models import Commande, EtatCommande
    
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Annulée (SAV)',
        etats__date_fin__isnull=True
    ).select_related('client', 'ville', 'ville__region').prefetch_related(
        'etats__enum_etat', 'etats__operateur', 'paniers__article'
    ).order_by('-etats__date_debut').distinct()
    
    # Pagination
    paginator = Paginator(commandes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Commandes Annulées (SAV)',
        'subtitle': 'Commandes annulées lors de la livraison',
        'icon': 'fa-times-circle',
        'color': 'gray'
    }
    return render(request, 'parametre/sav/liste_commandes_sav.html', context)

@staff_member_required
@login_required
def sav_livrees_avec_changement(request):
    """Vue admin pour afficher les commandes livrées avec changement"""
    from commande.models import Commande, EtatCommande
    
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Livrée avec changement',
        etats__date_fin__isnull=True
    ).select_related('client', 'ville', 'ville__region').prefetch_related(
        'etats__enum_etat', 'etats__operateur', 'paniers__article'
    ).order_by('-etats__date_debut').distinct()
    
    # Pagination
    paginator = Paginator(commandes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Commandes Livrées avec Changement',
        'subtitle': 'Commandes livrées avec modifications',
        'icon': 'fa-exchange-alt',
        'color': 'blue'
    }
    return render(request, 'parametre/sav/liste_commandes_sav.html', context)

@staff_member_required
@login_required
def sav_livrees(request):
    """Vue admin pour afficher les commandes livrées avec succès"""
    from commande.models import Commande, EtatCommande
    
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Livrée',
        etats__date_fin__isnull=True
    ).select_related('client', 'ville', 'ville__region').prefetch_related(
        'etats__enum_etat', 'etats__operateur', 'paniers__article'
    ).order_by('-etats__date_debut').distinct()
    
    # Pagination
    paginator = Paginator(commandes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'title': 'Commandes Livrées',
        'subtitle': 'Commandes livrées avec succès',
        'icon': 'fa-check-circle',
        'color': 'green'
    }
    return render(request, 'parametre/sav/liste_commandes_sav.html', context)

@staff_member_required
@login_required
@require_POST
def sav_creer_nouvelle_commande(request, commande_id):
    """Créer une nouvelle commande pour les articles défectueux retournés"""
    from commande.models import Commande, Panier, EtatCommande, EnumEtatCmd
    from django.db import transaction
    from django.utils import timezone
    import json
    
    try:
        commande_originale = get_object_or_404(Commande, id=commande_id)
        
        # Récupérer les articles défectueux depuis la requête POST
        articles_defectueux = json.loads(request.POST.get('articles_defectueux', '[]'))
        commentaire = request.POST.get('commentaire', '')
        
        if not articles_defectueux:
            messages.error(request, "Aucun article défectueux spécifié.")
            return redirect('app_admin:sav_commandes_retournees')
        
        with transaction.atomic():
            # Créer une nouvelle commande
            nouvelle_commande = Commande.objects.create(
                client=commande_originale.client,
                ville=commande_originale.ville,
                total_cmd=0,  # Sera recalculé
                num_cmd=f"SAV-{commande_originale.num_cmd}",
                id_yz=f"SAV-{commande_originale.id_yz}",
                is_upsell=False
            )
            
            total = 0
            # Créer les paniers pour les articles défectueux
            for article_data in articles_defectueux:
                article_id = article_data['article_id']
                quantite = int(article_data['quantite'])
                
                # Récupérer l'article original
                panier_original = commande_originale.paniers.filter(
                    article_id=article_id
                ).first()
                
                if panier_original:
                    Panier.objects.create(
                        commande=nouvelle_commande,
                        article=panier_original.article,
                        quantite=quantite,
                        sous_total=panier_original.article.prix_unitaire * quantite
                    )
                    total += panier_original.article.prix_unitaire * quantite
            
            # Mettre à jour le total de la commande
            nouvelle_commande.total_cmd = total
            nouvelle_commande.save()
            
            # Créer l'état initial "Non affectée"
            enum_etat = EnumEtatCmd.objects.get(libelle='Non affectée')
            EtatCommande.objects.create(
                commande=nouvelle_commande,
                enum_etat=enum_etat,
                operateur=request.user.profil_operateur,
                date_debut=timezone.now(),
                commentaire=f"Commande SAV créée pour articles défectueux. {commentaire}"
            )
            
            messages.success(request, 
                f"Nouvelle commande SAV créée avec succès : {nouvelle_commande.num_cmd}")
            return redirect('commande:detail', commande_id=nouvelle_commande.id)
            
    except Exception as e:
        messages.error(request, f"Erreur lors de la création de la commande SAV : {str(e)}")
        return redirect('app_admin:sav_commandes_retournees')

@staff_member_required
@login_required
@require_POST
def sav_renvoyer_preparation(request, commande_id):
    """Renvoyer la commande aux opérateurs de préparation suite aux modifications du client"""
    from commande.models import Commande, EtatCommande, EnumEtatCmd
    from django.db import transaction
    from django.utils import timezone
    
    try:
        commande = get_object_or_404(Commande, id=commande_id)
        commentaire = request.POST.get('commentaire', '')
        modifications = request.POST.get('modifications', '')
        
        with transaction.atomic():
            # Fermer l'état actuel
            etat_actuel = commande.etat_actuel
            if etat_actuel:
                etat_actuel.date_fin = timezone.now()
                etat_actuel.save()
            
            # Créer un nouvel état "En préparation"
            enum_etat = EnumEtatCmd.objects.get(libelle='Préparation en cours')
            EtatCommande.objects.create(
                commande=commande,
                enum_etat=enum_etat,
                operateur=request.user.profil_operateur,
                date_debut=timezone.now(),
                commentaire=f"Renvoyé en préparation suite à modification client. {modifications}. {commentaire}"
            )
            
            messages.success(request, 
                f"Commande {commande.num_cmd} renvoyée en préparation avec succès.")
            return redirect('commande:detail', commande_id=commande.id)
            
    except Exception as e:
        messages.error(request, f"Erreur lors du renvoi en préparation : {str(e)}")
        return redirect('app_admin:sav_commandes_retournees')


@staff_member_required
@login_required
def repartition_automatique(request):
    """Page de répartition automatique des commandes"""
    from django.db.models import Count, Sum
    
    # Statistiques pour les KPI
    total_commandes = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).distinct().count()
    
    total_commandes_en_livraison = Commande.objects.filter(
        etats__enum_etat__libelle='En livraison'
    ).distinct().count()
    
    total_montant_confirmees = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).distinct().aggregate(total=Sum('total_cmd'))['total'] or 0
    
    operateurs_disponibles = Operateur.objects.filter(actif=True).exclude(type_operateur='ADMIN').count()
    
    # Statistiques par région
    stats_par_region = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).values('ville__region__nom_region').annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('-nb_commandes')
    
    # Données de prévisualisation (simulation)
    preview_data = {}
    operateurs_actifs = Operateur.objects.filter(actif=True).exclude(type_operateur='ADMIN')[:5]
    
    for operateur in operateurs_actifs:
        preview_data[operateur.id] = {
            'nom_operateur': f"{operateur.prenom} {operateur.nom}",
            'nb_commandes': 0,
            'regions': [],
            'commandes': []
        }
    
    # Historique des répartitions (simulation)
    historique_repartitions = []
    
    # Vérifier si des commandes préparées existent pour les exportations
    commandes_preparees_exist = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée'
    ).exists()
    
    # Vérifier si openpyxl est disponible
    try:
        import openpyxl
        openpyxl_available = True
    except ImportError:
        openpyxl_available = False
    
    context = {
        'total_commandes': total_commandes,
        'total_commandes_en_livraison': total_commandes_en_livraison,
        'total_montant_confirmees': total_montant_confirmees,
        'operateurs_disponibles': operateurs_disponibles,
        'stats_par_region': stats_par_region,
        'preview_data': preview_data,
        'historique_repartitions': historique_repartitions,
        'commandes_preparees_exist': commandes_preparees_exist,
        'openpyxl_available': openpyxl_available,
    }
    
    return render(request, 'parametre/repartition_automatique.html', context)


@staff_member_required
@login_required
def details_region_view(request):
    """Page de détails par région"""
    from django.db.models import Count, Sum
    
    # Gérer les exports
    export_type = request.GET.get('export')
    region_name = request.GET.get('region')
    
    if export_type == 'csv_region_detail' and region_name:
        return export_region_detail_csv(request, region_name)
    elif export_type == 'excel_region_detail' and region_name:
        return export_region_detail_excel(request, region_name)
    elif export_type == 'csv_region':
        return export_regions_csv(request)
    elif export_type == 'excel_region':
        return export_regions_excel(request)
    elif export_type == 'csv_ville':
        return export_villes_csv(request)
    elif export_type == 'excel_ville':
        return export_villes_excel(request)
    elif export_type == 'csv_combine':
        return export_combine_csv(request)
    elif export_type == 'excel_combine':
        return export_combine_excel(request)
    
    # Statistiques globales
    total_commandes = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).distinct().count()
    
    total_montant = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).distinct().aggregate(total=Sum('total_cmd'))['total'] or 0
    
    regions = Region.objects.count()
    
    # Statistiques par région
    stats_par_region = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).values('ville__region__nom_region').annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('-nb_commandes')
    
    # Statistiques par ville
    stats_par_ville = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).values('ville__nom', 'ville__region__nom_region').annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('-nb_commandes')
    
    # Vérifier si des commandes préparées existent pour les exportations
    commandes_preparees_exist = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée'
    ).exists()
    
    # Vérifier si openpyxl est disponible
    try:
        import openpyxl
        openpyxl_available = True
    except ImportError:
        openpyxl_available = False
    
    context = {
        'total_commandes': total_commandes,
        'total_montant': total_montant,
        'regions': Region.objects.all(),
        'stats_par_region': stats_par_region,
        'stats_par_ville': stats_par_ville,
        'commandes_preparees_exist': commandes_preparees_exist,
        'openpyxl_available': openpyxl_available,
    }
    
    return render(request, 'parametre/details_region.html', context)


@staff_member_required
@login_required
def export_region_detail_csv(request, region_name):
    """Export CSV détaillé pour une région spécifique"""
    from commande.models import Commande
    
    # Récupérer les commandes PRÉPARÉES de la région
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        ville__region__nom_region=region_name
    ).select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'paniers__article'
    ).distinct().order_by('-date_cmd')
    
    # Créer la réponse CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"region_{region_name.lower().replace(' ', '_')}_detail_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Écrire l'en-tête BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-têtes
    headers = [
        'N° Commande', 'Client', 'Téléphone', 'Ville', 'Région', 
        'Articles et Quantités', 'Prix Total (MAD)', 'Adresse', 'État', 'Date Commande',
        'Heure Préparation', 'Heure Exportation'
    ]
    writer.writerow(headers)
    
    # Traiter chaque commande
    for commande in commandes:
        # Construire la liste des articles avec quantités
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = ", ".join(articles_list) if articles_list else "Aucun article"
        
        # État actuel de la commande
        etat_actuel = commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else "Non défini"
        
        # Trouver l'heure de préparation (dernier état "Préparée")
        heure_preparation = None
        for etat in commande.etats.filter(enum_etat__libelle='Préparée').order_by('-date_debut'):
            if etat.date_debut:
                heure_preparation = etat.date_debut.strftime('%d/%m/%Y %H:%M')
                break
        
        # Heure d'exportation
        heure_exportation = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        
        # Écrire la ligne
        row = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}" if commande.client else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region if commande.ville and commande.ville.region else "N/A",
            articles_consolides,
            f"{commande.total_cmd:.2f}" if commande.total_cmd else "0.00",
            commande.adresse or "N/A",
            etat_actuel,
            commande.date_creation.strftime('%d/%m/%Y %H:%M') if commande.date_creation else "N/A",
            heure_preparation or "N/A",
            heure_exportation
        ]
        writer.writerow(row)
    
    return response


@staff_member_required
@login_required
def export_region_detail_excel(request, region_name):
    """Export Excel détaillé pour une région spécifique"""
    from commande.models import Commande
    
    # Récupérer les commandes PRÉPARÉES de la région
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        ville__region__nom_region=region_name
    ).select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'paniers__article'
    ).distinct().order_by('-date_cmd')
    
    # Créer le fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Région {region_name}"
    
    # En-têtes
    headers = [
        'N° Commande', 'Client', 'Téléphone', 'Ville', 'Région', 
        'Articles et Quantités', 'Prix Total (MAD)', 'Adresse', 'État', 'Date Commande',
        'Heure Préparation', 'Heure Exportation'
    ]
    
    # Ajouter les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Traiter chaque commande
    for row, commande in enumerate(commandes, 2):
        # Construire la liste des articles avec quantités
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = ", ".join(articles_list) if articles_list else "Aucun article"
        
        # État actuel de la commande
        etat_actuel = commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else "Non défini"
        
        # Trouver l'heure de préparation (dernier état "Préparée")
        heure_preparation = None
        for etat in commande.etats.filter(enum_etat__libelle='Préparée').order_by('-date_debut'):
            if etat.date_debut:
                heure_preparation = etat.date_debut.strftime('%d/%m/%Y %H:%M')
                break
        
        # Heure d'exportation
        heure_exportation = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        
        # Écrire la ligne
        row_data = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}" if commande.client else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region if commande.ville and commande.ville.region else "N/A",
            articles_consolides,
            commande.total_cmd or 0,
            commande.adresse or "N/A",
            etat_actuel,
            commande.date_creation.strftime('%d/%m/%Y %H:%M') if commande.date_creation else "N/A",
            heure_preparation or "N/A",
            heure_exportation
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"region_{region_name.lower().replace(' ', '_')}_detail_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@staff_member_required
@login_required
def export_villes_csv(request):
    """Export CSV pour toutes les villes"""
    from commande.models import Commande
    
    # Récupérer les commandes PRÉPARÉES
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée'
    ).select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'paniers__article'
    ).distinct().order_by('-date_cmd')
    
    # Créer la réponse CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"villes_consolidees_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Écrire l'en-tête BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-têtes
    headers = [
        'N° Commande', 'Client', 'Téléphone', 'Ville', 'Région', 
        'Articles et Quantités', 'Prix Total (MAD)', 'Adresse', 'État', 'Date Commande',
        'Heure Préparation', 'Heure Exportation'
    ]
    writer.writerow(headers)
    
    # Traiter chaque commande
    for commande in commandes:
        # Construire la liste des articles avec quantités
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = ", ".join(articles_list) if articles_list else "Aucun article"
        
        # État actuel de la commande
        etat_actuel = commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else "Non défini"
        
        # Trouver l'heure de préparation (dernier état "Préparée")
        heure_preparation = None
        for etat in commande.etats.filter(enum_etat__libelle='Préparée').order_by('-date_debut'):
            if etat.date_debut:
                heure_preparation = etat.date_debut.strftime('%d/%m/%Y %H:%M')
                break
        
        # Heure d'exportation
        heure_exportation = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        
        # Écrire la ligne
        row = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}" if commande.client else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region if commande.ville and commande.ville.region else "N/A",
            articles_consolides,
            f"{commande.total_cmd:.2f}" if commande.total_cmd else "0.00",
            commande.adresse or "N/A",
            etat_actuel,
            commande.date_creation.strftime('%d/%m/%Y %H:%M') if commande.date_creation else "N/A",
            heure_preparation or "N/A",
            heure_exportation
        ]
        writer.writerow(row)
    
    return response


@staff_member_required
@login_required
def export_villes_excel(request):
    """Export Excel pour toutes les villes"""
    from commande.models import Commande
    
    # Récupérer les commandes PRÉPARÉES
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée'
    ).select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'paniers__article'
    ).distinct().order_by('-date_cmd')
    
    # Créer le fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Villes Consolidées"
    
    # En-têtes
    headers = [
        'N° Commande', 'Client', 'Téléphone', 'Ville', 'Région', 
        'Articles et Quantités', 'Prix Total (MAD)', 'Adresse', 'État', 'Date Commande',
        'Heure Préparation', 'Heure Exportation'
    ]
    
    # Ajouter les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Traiter chaque commande
    for row, commande in enumerate(commandes, 2):
        # Construire la liste des articles avec quantités
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = ", ".join(articles_list) if articles_list else "Aucun article"
        
        # État actuel de la commande
        etat_actuel = commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else "Non défini"
        
        # Trouver l'heure de préparation (dernier état "Préparée")
        heure_preparation = None
        for etat in commande.etats.filter(enum_etat__libelle='Préparée').order_by('-date_debut'):
            if etat.date_debut:
                heure_preparation = etat.date_debut.strftime('%d/%m/%Y %H:%M')
                break
        
        # Heure d'exportation
        heure_exportation = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        
        # Écrire la ligne
        row_data = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}" if commande.client else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region if commande.ville and commande.ville.region else "N/A",
            articles_consolides,
            commande.total_cmd or 0,
            commande.adresse or "N/A",
            etat_actuel,
            commande.date_creation.strftime('%d/%m/%Y %H:%M') if commande.date_creation else "N/A",
            heure_preparation or "N/A",
            heure_exportation
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"villes_consolidees_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@staff_member_required
@login_required
def get_modal_data_ajax(request):
    """Vue AJAX pour récupérer les données des modales en temps réel"""
    from django.utils import timezone
    from datetime import datetime
    
    try:
        # Statistiques globales
        total_commandes = Commande.objects.filter(
            etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
        ).distinct().count()
        
        total_montant = Commande.objects.filter(
            etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
        ).distinct().aggregate(total=Sum('total_cmd'))['total'] or 0
        
        # Statistiques par région avec pourcentages
        stats_region_avec_pourcentage = []
        stats_par_region = Commande.objects.filter(
            etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
        ).values('ville__region__nom_region').annotate(
            nb_commandes=Count('id'),
            total_montant=Sum('total_cmd')
        ).order_by('-nb_commandes')
        
        for stat in stats_par_region:
            pourcentage = (stat['nb_commandes'] / total_commandes * 100) if total_commandes > 0 else 0
            stats_region_avec_pourcentage.append({
                'region': stat['ville__region__nom_region'],
                'nb_commandes': stat['nb_commandes'],
                'total_montant': stat['total_montant'] or 0,
                'pourcentage': round(pourcentage, 1)
            })
        
        # Top 10 des villes
        top_10_villes = Commande.objects.filter(
            etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
        ).values('ville__nom', 'ville__region__nom_region').annotate(
            nb_commandes=Count('id'),
            total_montant=Sum('total_cmd')
        ).order_by('-nb_commandes')[:10]
        
        # Statistiques globales détaillées
        nb_regions_actives = len(stats_region_avec_pourcentage)
        nb_villes_actives = Commande.objects.filter(
            etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
        ).values('ville').distinct().count()
        
        moyenne_commandes_par_region = round(total_commandes / nb_regions_actives, 1) if nb_regions_actives > 0 else 0
        moyenne_commandes_par_ville = round(total_commandes / nb_villes_actives, 1) if nb_villes_actives > 0 else 0
        
        stats_globales = {
            'total_commandes': total_commandes,
            'total_montant': total_montant,
            'nb_regions_actives': nb_regions_actives,
            'nb_villes_actives': nb_villes_actives,
            'moyenne_commandes_par_region': moyenne_commandes_par_region,
            'moyenne_commandes_par_ville': moyenne_commandes_par_ville,
            'derniere_mise_a_jour': timezone.now().strftime('%d/%m/%Y à %H:%M:%S')
        }
        
        return JsonResponse({
            'success': True,
            'stats_region_avec_pourcentage': stats_region_avec_pourcentage,
            'top_10_villes': list(top_10_villes),
            'stats_globales': stats_globales
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@staff_member_required
@login_required
def export_combine_csv(request):
    """Export CSV combiné des visualisations : régions et villes"""
    from django.db.models import Count, Sum
    from django.http import HttpResponse
    import csv
    from django.utils import timezone
    
    # Récupérer les statistiques par région
    stats_par_region = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).values('ville__region__nom_region').annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('-nb_commandes')
    
    # Récupérer les statistiques par ville (Top 10)
    stats_par_ville = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).values('ville__nom', 'ville__region__nom_region').annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('-nb_commandes')[:10]
    
    # Créer la réponse CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"visualisations_combinees_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Écrire l'en-tête BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # Section 1 : Répartition par Région
    writer.writerow(['RÉPARTITION PAR RÉGION'])
    writer.writerow(['Région', 'Nombre de Commandes', 'Total Montant (MAD)', 'Pourcentage'])
    
    total_commandes = sum(stat['nb_commandes'] for stat in stats_par_region)
    
    for stat in stats_par_region:
        pourcentage = (stat['nb_commandes'] / total_commandes * 100) if total_commandes > 0 else 0
        writer.writerow([
            stat['ville__region__nom_region'] or 'Non définie',
            stat['nb_commandes'],
            f"{stat['total_montant'] or 0:.2f}",
            f"{pourcentage:.1f}%"
        ])
    
    writer.writerow([])  # Ligne vide
    
    # Section 2 : Top 10 des Villes
    writer.writerow(['TOP 10 DES VILLES'])
    writer.writerow(['Ville', 'Région', 'Nombre de Commandes', 'Total Montant (MAD)'])
    
    for stat in stats_par_ville:
        writer.writerow([
            stat['ville__nom'] or 'Non définie',
            stat['ville__region__nom_region'] or 'Non définie',
            stat['nb_commandes'],
            f"{stat['total_montant'] or 0:.2f}"
        ])
    
    writer.writerow([])  # Ligne vide
    
    # Section 3 : Statistiques Globales
    writer.writerow(['STATISTIQUES GLOBALES'])
    writer.writerow(['Métrique', 'Valeur'])
    writer.writerow(['Total Commandes', total_commandes])
    writer.writerow(['Nombre de Régions', len(stats_par_region)])
    writer.writerow(['Nombre de Villes', len(stats_par_ville)])
    writer.writerow(['Date d\'Export', timezone.now().strftime('%d/%m/%Y à %H:%M:%S')])
    
    return response


@staff_member_required
@login_required
def export_combine_excel(request):
    """Export Excel combiné des visualisations : régions et villes"""
    from django.db.models import Count, Sum
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.utils import timezone
    
    # Récupérer les statistiques par région
    stats_par_region = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).values('ville__region__nom_region').annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('-nb_commandes')
    
    # Récupérer les statistiques par ville (Top 10)
    stats_par_ville = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée']
    ).values('ville__nom', 'ville__region__nom_region').annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('-nb_commandes')[:10]
    
    # Créer le fichier Excel
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_font = Font(bold=True, color='FFFFFF', size=14)
    title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Feuille 1 : Répartition par Région
    ws1 = wb.active
    ws1.title = "Répartition par Région"
    
    # Titre
    ws1.merge_cells('A1:D1')
    title_cell = ws1['A1']
    title_cell.value = "RÉPARTITION PAR RÉGION"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # En-têtes
    headers = ['Région', 'Nombre de Commandes', 'Total Montant (MAD)', 'Pourcentage']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Données
    total_commandes = sum(stat['nb_commandes'] for stat in stats_par_region)
    
    for row, stat in enumerate(stats_par_region, 4):
        pourcentage = (stat['nb_commandes'] / total_commandes * 100) if total_commandes > 0 else 0
        
        ws1.cell(row=row, column=1, value=stat['ville__region__nom_region'] or 'Non définie').border = border
        ws1.cell(row=row, column=2, value=stat['nb_commandes']).border = border
        ws1.cell(row=row, column=3, value=f"{stat['total_montant'] or 0:.2f}").border = border
        ws1.cell(row=row, column=4, value=f"{pourcentage:.1f}%").border = border
    
    # Ajuster la largeur des colonnes
    for column in ws1.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Feuille 2 : Top 10 des Villes
    ws2 = wb.create_sheet("Top 10 des Villes")
    
    # Titre
    ws2.merge_cells('A1:D1')
    title_cell = ws2['A1']
    title_cell.value = "TOP 10 DES VILLES"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # En-têtes
    headers = ['Ville', 'Région', 'Nombre de Commandes', 'Total Montant (MAD)']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Données
    for row, stat in enumerate(stats_par_ville, 4):
        ws2.cell(row=row, column=1, value=stat['ville__nom'] or 'Non définie').border = border
        ws2.cell(row=row, column=2, value=stat['ville__region__nom_region'] or 'Non définie').border = border
        ws2.cell(row=row, column=3, value=stat['nb_commandes']).border = border
        ws2.cell(row=row, column=4, value=f"{stat['total_montant'] or 0:.2f}").border = border
    
    # Ajuster la largeur des colonnes
    for column in ws2.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws2.column_dimensions[column_letter].width = adjusted_width
    
    # Feuille 3 : Statistiques Globales
    ws3 = wb.create_sheet("Statistiques Globales")
    
    # Titre
    ws3.merge_cells('A1:B1')
    title_cell = ws3['A1']
    title_cell.value = "STATISTIQUES GLOBALES"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # En-têtes
    headers = ['Métrique', 'Valeur']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Données
    stats_data = [
        ['Total Commandes', total_commandes],
        ['Nombre de Régions', len(stats_par_region)],
        ['Nombre de Villes', len(stats_par_ville)],
        ['Date d\'Export', timezone.now().strftime('%d/%m/%Y à %H:%M:%S')]
    ]
    
    for row, (metric, value) in enumerate(stats_data, 4):
        ws3.cell(row=row, column=1, value=metric).border = border
        ws3.cell(row=row, column=2, value=value).border = border
    
    # Ajuster la largeur des colonnes
    for column in ws3.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws3.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"visualisations_combinees_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

@staff_member_required
@login_required
def export_regions_csv(request):
    """Export CSV pour toutes les régions"""
    from commande.models import Commande
    
    # Récupérer les commandes PRÉPARÉES groupées par région
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée'
    ).select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'paniers__article'
    ).distinct().order_by('ville__region__nom_region', '-date_cmd')
    
    # Créer la réponse CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"regions_consolidees_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Écrire l'en-tête BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-têtes
    headers = [
        'N° Commande', 'Client', 'Téléphone', 'Ville', 'Région', 
        'Articles et Quantités', 'Prix Total (MAD)', 'Adresse', 'État', 'Date Commande',
        'Heure Préparation', 'Heure Exportation'
    ]
    writer.writerow(headers)
    
    # Traiter chaque commande
    for commande in commandes:
        # Construire la liste des articles avec quantités
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = ", ".join(articles_list) if articles_list else "Aucun article"
        
        # État actuel de la commande
        etat_actuel = commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else "Non défini"
        
        # Trouver l'heure de préparation (dernier état "Préparée")
        heure_preparation = None
        for etat in commande.etats.filter(enum_etat__libelle='Préparée').order_by('-date_debut'):
            if etat.date_debut:
                heure_preparation = etat.date_debut.strftime('%d/%m/%Y %H:%M')
                break
        
        # Heure d'exportation
        heure_exportation = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        
        # Écrire la ligne
        row = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}" if commande.client else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region if commande.ville and commande.ville.region else "N/A",
            articles_consolides,
            f"{commande.total_cmd:.2f}" if commande.total_cmd else "0.00",
            commande.adresse or "N/A",
            etat_actuel,
            commande.date_creation.strftime('%d/%m/%Y %H:%M') if commande.date_creation else "N/A",
            heure_preparation or "N/A",
            heure_exportation
        ]
        writer.writerow(row)
    
    return response


@staff_member_required
@login_required
def export_regions_excel(request):
    """Export Excel pour toutes les régions"""
    from commande.models import Commande
    
    # Récupérer les commandes PRÉPARÉES groupées par région
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée'
    ).select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'paniers__article'
    ).distinct().order_by('ville__region__nom_region', '-date_cmd')
    
    # Créer le fichier Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Régions Consolidées"
    
    # En-têtes
    headers = [
        'N° Commande', 'Client', 'Téléphone', 'Ville', 'Région', 
        'Articles et Quantités', 'Prix Total (MAD)', 'Adresse', 'État', 'Date Commande',
        'Heure Préparation', 'Heure Exportation'
    ]
    
    # Ajouter les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Traiter chaque commande
    for row, commande in enumerate(commandes, 2):
        # Construire la liste des articles avec quantités
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = ", ".join(articles_list) if articles_list else "Aucun article"
        
        # État actuel de la commande
        etat_actuel = commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else "Non défini"
        
        # Trouver l'heure de préparation (dernier état "Préparée")
        heure_preparation = None
        for etat in commande.etats.filter(enum_etat__libelle='Préparée').order_by('-date_debut'):
            if etat.date_debut:
                heure_preparation = etat.date_debut.strftime('%d/%m/%Y %H:%M')
                break
        
        # Heure d'exportation
        heure_exportation = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        
        # Écrire la ligne
        row_data = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}" if commande.client else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region if commande.ville and commande.ville.region else "N/A",
            articles_consolides,
            commande.total_cmd or 0,
            commande.adresse or "N/A",
            etat_actuel,
            commande.date_creation.strftime('%d/%m/%Y %H:%M') if commande.date_creation else "N/A",
            heure_preparation or "N/A",
            heure_exportation
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Créer la réponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"regions_consolidees_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response