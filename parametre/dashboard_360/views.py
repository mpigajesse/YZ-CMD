from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q
from django.core.paginator import Paginator
from parametre.models import Operateur
from article.models import Article
from commande.models import Commande, EtatCommande, EnumEtatCmd
from client.models import Client
import csv
import io
import zipfile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.utils.encoding import smart_str
from django.utils import timezone
import json
from datetime import datetime, timedelta

@staff_member_required
@login_required
def page_360(request):
    """Page 360 - Vue d'overview et exportation des données"""
    # Statistiques générales
    total_articles = Article.objects.count()
    total_clients = Client.objects.count()
    total_commandes = Commande.objects.count()
    total_operateurs = Operateur.objects.count()
    
    # Récupérer toutes les commandes avec filtres et pagination - optimisé avec only()
    commandes_360 = Commande.objects.select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'etats__enum_etat',
        'etats__operateur', # Pour récupérer l'opérateur qui a fait l'état
        'paniers__article' # Accéder aux articles via les paniers
    ).only(
        'id', 'num_cmd', 'id_yz', 'date_cmd', 'total_cmd', 'is_upsell',
        'client__nom', 'client__prenom', 'client__numero_tel', 'client__adresse',
        'ville__nom', 'ville__region__nom_region'
    )
    
    # Filtres de recherche
    search = request.GET.get('search')
    if search:
        commandes_360 = commandes_360.filter(
            Q(num_cmd__icontains=search) |
            Q(id_yz__icontains=search) |
            Q(client__nom__icontains=search) |
            Q(client__prenom__icontains=search) |
            Q(client__numero_tel__icontains=search)
        )
    
    date_debut = request.GET.get('date_debut')
    if date_debut:
        commandes_360 = commandes_360.filter(date_cmd__gte=date_debut)
        
    date_fin = request.GET.get('date_fin')
    if date_fin:
        commandes_360 = commandes_360.filter(date_cmd__lte=date_fin)
    
    commandes_360 = commandes_360.order_by('-date_cmd')
    
    # Pagination - 50 commandes par page
    paginator = Paginator(commandes_360, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Préparer les données pour le template
    data_for_template = prepare_commandes_data(page_obj)

    context = {
        'total_articles': total_articles,
        'total_clients': total_clients,
        'total_commandes': total_commandes,
        'total_operateurs': total_operateurs,
        'commandes_data': data_for_template,
        'page_obj': page_obj,
        'search': search,
        'date_debut': date_debut,
        'date_fin': date_fin,
    }
    return render(request, 'parametre/360.html', context)

@staff_member_required
@login_required
def vue_360_realtime_data(request):
    """API pour les données en temps réel de la vue 360"""
    try:
        # Récupérer les paramètres de filtrage
        search = request.GET.get('search')
        date_debut = request.GET.get('date_debut')
        date_fin = request.GET.get('date_fin')
        page = request.GET.get('page', 1)
        
        # Construire la requête avec les mêmes filtres
        commandes_360 = Commande.objects.select_related(
            'client', 'ville', 'ville__region'
        ).prefetch_related(
            'etats__enum_etat', 'etats__operateur', 'paniers__article'
        ).only(
            'id', 'num_cmd', 'id_yz', 'date_cmd', 'total_cmd', 'is_upsell',
            'client__nom', 'client__prenom', 'client__numero_tel', 'client__adresse',
            'ville__nom', 'ville__region__nom_region'
        )
        
        # Appliquer les filtres
        if search:
            commandes_360 = commandes_360.filter(
                Q(num_cmd__icontains=search) |
                Q(id_yz__icontains=search) |
                Q(client__nom__icontains=search) |
                Q(client__prenom__icontains=search) |
                Q(client__numero_tel__icontains=search)
            )
        
        if date_debut:
            commandes_360 = commandes_360.filter(date_cmd__gte=date_debut)
            
        if date_fin:
            commandes_360 = commandes_360.filter(date_cmd__lte=date_fin)
        
        commandes_360 = commandes_360.order_by('-date_cmd')
        
        # Pagination
        paginator = Paginator(commandes_360, 50)
        page_obj = paginator.get_page(page)
        
        # Préparer les données
        commandes_data = prepare_commandes_data(page_obj)
        
        # Statistiques mises à jour
        total_articles = Article.objects.count()
        total_clients = Client.objects.count()
        total_commandes = Commande.objects.count()
        total_operateurs = Operateur.objects.count()
        
        return JsonResponse({
            'success': True,
            'commandes_data': commandes_data,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': page_obj.paginator.num_pages,
                'total_count': page_obj.paginator.count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
                'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
            },
            'statistics': {
                'total_articles': total_articles,
                'total_clients': total_clients,
                'total_commandes': total_commandes,
                'total_operateurs': total_operateurs,
            },
            'filters': {
                'search': search,
                'date_debut': date_debut,
                'date_fin': date_fin,
            },
            'timestamp': timezone.now().isoformat(),
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'timestamp': timezone.now().isoformat(),
        }, status=500)

@staff_member_required
@login_required
def vue_360_statistics_update(request):
    """API pour mettre à jour uniquement les statistiques"""
    try:
        total_articles = Article.objects.count()
        total_clients = Client.objects.count()
        total_commandes = Commande.objects.count()
        total_operateurs = Operateur.objects.count()
        
        return JsonResponse({
            'success': True,
            'statistics': {
                'total_articles': total_articles,
                'total_clients': total_clients,
                'total_commandes': total_commandes,
                'total_operateurs': total_operateurs,
            },
            'timestamp': timezone.now().isoformat(),
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'timestamp': timezone.now().isoformat(),
        }, status=500)

@staff_member_required
@login_required
def vue_360_etats_tracking(request):
    """API pour le suivi en temps réel des états de commande"""
    try:
        # Récupérer les paramètres
        commande_id = request.GET.get('commande_id')
        last_update = request.GET.get('last_update')
        
        if commande_id:
            # Suivi d'une commande spécifique
            try:
                commande = Commande.objects.get(id=commande_id)
                etats = commande.etats.all().order_by('date_debut')
                
                # Vérifier s'il y a eu des modifications depuis la dernière mise à jour
                if last_update:
                    last_update_dt = datetime.fromisoformat(last_update.replace('Z', '+00:00'))
                    etats_recents = etats.filter(date_debut__gt=last_update_dt)
                    
                    if not etats_recents.exists():
                        return JsonResponse({
                            'success': True,
                            'has_changes': False,
                            'timestamp': timezone.now().isoformat(),
                        })
                
                # Préparer les données de suivi
                tracking_data = {
                    'commande_id': commande.id,
                    'num_cmd': commande.num_cmd,
                    'etat_actuel': etats.last().enum_etat.libelle if etats.exists() else "Non définie",
                    'etapes_completes': etats.count(),
                    'derniere_modification': etats.last().date_debut.isoformat() if etats.exists() and etats.last().date_debut else None,
                    'operateur_derniere_modification': etats.last().operateur.mail if etats.exists() and etats.last().operateur else "N/A",
                    'historique_etats': []
                }
                
                for etat in etats:
                    tracking_data['historique_etats'].append({
                        'etat': etat.enum_etat.libelle,
                        'date': etat.date_debut.strftime('%d/%m/%Y %H:%M') if etat.date_debut else "N/A",
                        'operateur': etat.operateur.mail if etat.operateur else "N/A",
                        'commentaire': etat.commentaire or "",
                        'duree': calculate_duration(etat.date_debut, etat.date_fin)
                    })
                
                return JsonResponse({
                    'success': True,
                    'has_changes': True,
                    'tracking_data': tracking_data,
                    'timestamp': timezone.now().isoformat(),
                })
                
            except Commande.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Commande non trouvée',
                    'timestamp': timezone.now().isoformat(),
                }, status=404)
        
        else:
            # Suivi global des modifications récentes
            # Récupérer les états modifiés dans les dernières 5 minutes
            recent_time = timezone.now() - timedelta(minutes=5)
            recent_etats = EtatCommande.objects.filter(
                date_debut__gte=recent_time
            ).select_related('commande', 'enum_etat', 'operateur').order_by('-date_debut')
            
            recent_changes = []
            for etat in recent_etats:
                recent_changes.append({
                    'commande_id': etat.commande.id,
                    'num_cmd': etat.commande.num_cmd,
                    'nouvel_etat': etat.enum_etat.libelle,
                    'operateur': etat.operateur.mail if etat.operateur else "N/A",
                    'date_modification': etat.date_debut.strftime('%d/%m/%Y %H:%M') if etat.date_debut else "N/A",
                    'commentaire': etat.commentaire or ""
                })
            
            return JsonResponse({
                'success': True,
                'recent_changes': recent_changes,
                'timestamp': timezone.now().isoformat(),
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'timestamp': timezone.now().isoformat(),
        }, status=500)

@staff_member_required
@login_required
def vue_360_panier_tracking(request):
    """API pour le suivi en temps réel du panier"""
    try:
        # Récupérer les paramètres
        commande_id = request.GET.get('commande_id')
        last_update = request.GET.get('last_update')
        
        if commande_id:
            # Suivi du panier d'une commande spécifique
            try:
                commande = Commande.objects.get(id=commande_id)
                paniers = commande.paniers.all().select_related('article')
                
                # Vérifier s'il y a eu des modifications depuis la dernière mise à jour
                if last_update:
                    last_update_dt = datetime.fromisoformat(last_update.replace('Z', '+00:00'))
                    paniers_recents = paniers.filter(date_creation__gt=last_update_dt)
                    
                    if not paniers_recents.exists():
                        return JsonResponse({
                            'success': True,
                            'has_changes': False,
                            'timestamp': timezone.now().isoformat(),
                        })
                
                # Préparer les données du panier
                panier_data = {
                    'commande_id': commande.id,
                    'num_cmd': commande.num_cmd,
                    'total_panier': sum(panier.sous_total for panier in paniers),
                    'nombre_articles': paniers.count(),
                    'derniere_modification': timezone.now().isoformat(),
                    'articles': []
                }
                
                for panier in paniers:
                    panier_data['articles'].append({
                        'id': panier.id,
                        'article_nom': panier.article.nom,
                        'article_reference': panier.article.reference,
                        'article_couleur': panier.article.couleur,
                        'article_pointure': panier.article.pointure,
                        'quantite': panier.quantite,
                        'prix_unitaire': panier.article.prix_unitaire,
                        'sous_total': panier.sous_total,
                        'date_ajout': panier.date_creation.isoformat() if hasattr(panier, 'date_creation') else None
                    })
                
                return JsonResponse({
                    'success': True,
                    'has_changes': True,
                    'panier_data': panier_data,
                    'timestamp': timezone.now().isoformat(),
                })
                
            except Commande.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Commande non trouvée',
                    'timestamp': timezone.now().isoformat(),
                }, status=404)
        
        else:
            # Suivi global des modifications de paniers récentes
            # Récupérer les paniers modifiés dans les dernières 5 minutes
            recent_time = timezone.now() - timedelta(minutes=5)
            
            # Note: Cette requête dépend de la structure de votre modèle Panier
            # Ajustez selon vos champs de date de modification
            from commande.models import Panier
            recent_paniers = Panier.objects.filter(
                date_creation__gte=recent_time
            ).select_related('commande', 'article').order_by('-date_creation')
            
            recent_changes = []
            for panier in recent_paniers:
                recent_changes.append({
                    'commande_id': panier.commande.id,
                    'num_cmd': panier.commande.num_cmd,
                    'article_nom': panier.article.nom,
                    'quantite': panier.quantite,
                    'sous_total': panier.sous_total,
                    'date_modification': panier.date_creation.strftime('%d/%m/%Y %H:%M') if hasattr(panier, 'date_creation') else "N/A"
                })
            
            return JsonResponse({
                'success': True,
                'recent_panier_changes': recent_changes,
                'timestamp': timezone.now().isoformat(),
            })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'timestamp': timezone.now().isoformat(),
        }, status=500)

def prepare_commandes_data(commandes_queryset):
    """Prépare les données des commandes pour l'affichage avec suivi des états et panier"""
    data_for_template = []
    for cmd in commandes_queryset:
        # Récupérer tous les états de la commande ordonnés par date
        etats_commande = cmd.etats.all().order_by('date_debut')
        
        # État actuel (le plus récent)
        etat_actuel = etats_commande.last()
        
        # État de confirmation
        confirmation_info = etats_commande.filter(enum_etat__libelle__icontains='Confirmée').first()
        
        # État de préparation
        preparation_info = etats_commande.filter(enum_etat__libelle__icontains='Préparation en cours').first()
        
        # État de livraison
        etat_livraison_obj = etats_commande.filter(enum_etat__libelle__icontains='Livrée').first()
        
        # État de paiement
        etat_paiement_obj = etats_commande.filter(enum_etat__libelle__icontains='Payée').first()
        
        # État de retour
        piece_retournee_obj = etats_commande.filter(enum_etat__libelle__icontains='Retournée').first()
        
        # État d'affectation
        operateur_assigne_obj = etats_commande.filter(enum_etat__libelle__icontains='Affectée').first()

        # Préparer l'historique des états
        historique_etats = []
        for etat in etats_commande:
            historique_etats.append({
                'etat': etat.enum_etat.libelle,
                'date': etat.date_debut.strftime('%d/%m/%Y %H:%M') if etat.date_debut else "N/A",
                'operateur': etat.operateur.mail if etat.operateur else "N/A",
                'commentaire': etat.commentaire or "",
                'duree': calculate_duration(etat.date_debut, etat.date_fin) if etat.date_debut else "N/A"
            })

        # Calculer les métriques de processus
        duree_totale = calculate_total_duration(etats_commande)
        etapes_completes = len(etats_commande)
        etape_actuelle = get_current_step(etat_actuel.enum_etat.libelle if etat_actuel else "Non définie")
        
        # Déterminer le statut du processus
        statut_processus = determine_process_status(etats_commande)
        
        # Récupérer les données du panier en temps réel
        paniers = cmd.paniers.all().select_related('article')
        articles_panier = []
        total_panier = 0
        nombre_articles = 0
        
        for panier in paniers:
            articles_panier.append({
                'nom': panier.article.nom,
                'reference': panier.article.reference,
                'couleur': panier.article.couleur,
                'pointure': panier.article.pointure,
                'quantite': panier.quantite,
                'prix_unitaire': panier.article.prix_unitaire,
                'sous_total': panier.sous_total
            })
            total_panier += panier.sous_total
            nombre_articles += panier.quantite
        
        # Pour le panier: Joindre les noms des articles du panier
        articles_noms = ", ".join([panier.article.nom for panier in paniers]) or "N/A"

        # Opérateur Assigné
        operateur_assigne_nom = operateur_assigne_obj.operateur.mail if operateur_assigne_obj and operateur_assigne_obj.operateur else "N/A"
        
        # Agent Confirmation
        agent_confirmation_nom = confirmation_info.operateur.mail if confirmation_info and confirmation_info.operateur else "N/A"

        # Valeurs par défaut
        etat_paiement = etat_paiement_obj.enum_etat.libelle if etat_paiement_obj else "Non Payé"
        etat_livraison = etat_livraison_obj.enum_etat.libelle if etat_livraison_obj else "En attente"
        piece_retournee = "Oui" if piece_retournee_obj else "Non"
        tarif_livraison = 0.0
        reste_a_payer = cmd.total_cmd
        date_paiement = "N/A"
        observation_livraison = piece_retournee_obj.commentaire if piece_retournee_obj else ""

        data_for_template.append({
            'id': cmd.id,
            'num_cmd': cmd.num_cmd,
            'id_yz': cmd.id_yz,
            'client_nom_prenom': f"{cmd.client.prenom} {cmd.client.nom}" if cmd.client else "N/A",
            'client_telephone': cmd.client.numero_tel if cmd.client else "N/A",
            'client_adresse': cmd.client.adresse if cmd.client else "N/A",
            'ville': cmd.ville.nom if cmd.ville else "N/A",
            'region': cmd.ville.region.nom_region if cmd.ville and cmd.ville.region else "N/A",
            'panier': articles_noms,
            'prix_total_dh': cmd.total_cmd,
            'date_commande': cmd.date_cmd.strftime('%d/%m/%Y') if cmd.date_cmd else "N/A",
            'confirmation_status': confirmation_info.enum_etat.libelle if confirmation_info else "Non Confirmée",
            'date_confirmation': confirmation_info.date_debut.strftime('%d/%m/%Y %H:%M') if confirmation_info and confirmation_info.date_debut else "N/A",
            'observations_confirmation': confirmation_info.commentaire if confirmation_info else "",
            'operateur_assigne': operateur_assigne_nom,
            'agent_confirmation': agent_confirmation_nom,
            'client_fidele': "future qui seras des les tables models plustard dans le projet",
            'upsell_display': "Oui" if cmd.is_upsell else "Non",
            'preparation_status': preparation_info.enum_etat.libelle if preparation_info else "Non Préparée",
            'etat_livraison': etat_livraison,
            'etat_paiement': etat_paiement,
            'tarif_livraison': tarif_livraison,
            'reste_a_payer': reste_a_payer,
            'date_paiement': date_paiement,
            'piece_retournee': piece_retournee,
            'observation_livraison': observation_livraison,
            'last_updated': timezone.now().isoformat(),
            # Nouvelles données pour le suivi des états
            'etat_actuel': etat_actuel.enum_etat.libelle if etat_actuel else "Non définie",
            'etape_actuelle': etape_actuelle,
            'etapes_completes': etapes_completes,
            'duree_totale': duree_totale,
            'statut_processus': statut_processus,
            'historique_etats': historique_etats,
            'derniere_modification': etat_actuel.date_debut.strftime('%d/%m/%Y %H:%M') if etat_actuel and etat_actuel.date_debut else "N/A",
            'operateur_derniere_modification': etat_actuel.operateur.mail if etat_actuel and etat_actuel.operateur else "N/A",
            # Nouvelles données pour le suivi du panier
            'articles_panier': articles_panier,
            'total_panier': total_panier,
            'nombre_articles': nombre_articles,
            'derniere_modification_panier': timezone.now().isoformat()
        })
    
    return data_for_template

def calculate_duration(start_date, end_date):
    """Calcule la durée entre deux dates"""
    if not start_date or not end_date:
        return "N/A"
    
    duration = end_date - start_date
    hours = duration.total_seconds() / 3600
    
    if hours < 1:
        return f"{int(duration.total_seconds() / 60)} min"
    elif hours < 24:
        return f"{hours:.1f} h"
    else:
        days = hours / 24
        return f"{days:.1f} jours"

def calculate_total_duration(etats_commande):
    """Calcule la durée totale du processus"""
    if not etats_commande:
        return "N/A"
    
    premier_etat = etats_commande.first()
    dernier_etat = etats_commande.last()
    
    if premier_etat and dernier_etat and premier_etat.date_debut and dernier_etat.date_debut:
        duration = dernier_etat.date_debut - premier_etat.date_debut
        hours = duration.total_seconds() / 3600
        
        if hours < 1:
            return f"{int(duration.total_seconds() / 60)} min"
        elif hours < 24:
            return f"{hours:.1f} h"
        else:
            days = hours / 24
            return f"{days:.1f} jours"
    
    return "N/A"

def get_current_step(etat_libelle):
    """Détermine l'étape actuelle du processus"""
    etapes = {
        'Nouvelle': 1,
        'Affectée': 2,
        'Confirmée': 3,
        'Préparation en cours': 4,
        'Préparée': 5,
        'Livrée': 6,
        'Payée': 7,
        'Retournée': 8,
        'Annulée': 9
    }
    
    for etat, step in etapes.items():
        if etat.lower() in etat_libelle.lower():
            return step
    
    return 0

def determine_process_status(etats_commande):
    """Détermine le statut global du processus"""
    if not etats_commande:
        return "Non démarré"
    
    dernier_etat = etats_commande.last()
    if not dernier_etat:
        return "Non démarré"
    
    etat_libelle = dernier_etat.enum_etat.libelle.lower()
    
    if 'annulée' in etat_libelle:
        return "Annulé"
    elif 'retournée' in etat_libelle:
        return "Retourné"
    elif 'payée' in etat_libelle:
        return "Terminé"
    elif 'livrée' in etat_libelle:
        return "Livré"
    elif 'préparée' in etat_libelle:
        return "Prêt pour livraison"
    elif 'préparation' in etat_libelle:
        return "En préparation"
    elif 'confirmée' in etat_libelle:
        return "Confirmé"
    elif 'affectée' in etat_libelle:
        return "Affecté"
    else:
        return "En cours"

@staff_member_required
@login_required
def export_all_data_csv(request):
    # Assurez-vous que la méthode de requête est POST
    if request.method == 'POST':
        # Récupérer les filtres pour appliquer les mêmes que dans la vue
        search = request.POST.get('search') or request.GET.get('search')
        date_debut = request.POST.get('date_debut') or request.GET.get('date_debut')
        date_fin = request.POST.get('date_fin') or request.GET.get('date_fin')
        
        # Utiliser une réponse streaming pour de gros volumes
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="export_commandes_avec_paniers_360.csv"'
        
        # Utiliser un writer CSV directement sur la réponse
        writer = csv.writer(response)

        # En-têtes du CSV consolidé avec détails des paniers
        headers = [
            'N°', 'Identifiant Yoozak', 'CLIENT', 'TELEPHONE', 'ADRESSE', 'VILLE', 'REGION',
            'ARTICLE NOM', 'ARTICLE REFERENCE', 'ARTICLE COULEUR', 'ARTICLE POINTURE', 
            'QUANTITE', 'PRIX UNITAIRE', 'SOUS TOTAL ARTICLE',
            'PRIX TOTAL COMMANDE (DH)', 'DATE COMMANDE', 'CONFIRMATION', 'DATE CONFIRMATION',
            'OBSERVATIONS CONFIRMATION', 'OPERATEUR', 'AGENT CONFIRMATION',
            'CLIENT FIDELE', 'UPSELL', 'PREPARATION', 'ETAT LIVRAISON',
            'ETAT PAIEMENT', 'TARIF', 'RESTE A PAYER', 'DATE PAIEMENT', 'PIECE RETOURNEE',
            'OBSERVATION LIVRAISON'
        ]
        writer.writerow(headers)
            
        # Traitement par batch de 1000 commandes pour éviter de surcharger la mémoire
        batch_size = 1000
        offset = 0
        
        while True:
            # Construire la requête avec les mêmes filtres que la vue
            commandes_query = Commande.objects.select_related(
                'client', 
                'ville', 
                'ville__region'
            ).prefetch_related(
                'etats__enum_etat',
                'etats__operateur',
                'paniers__article'
            ).only(
                'id', 'num_cmd', 'id_yz', 'date_cmd', 'total_cmd', 'is_upsell',
                'client__nom', 'client__prenom', 'client__numero_tel', 'client__adresse',
                'ville__nom', 'ville__region__nom_region'
            )
            
            # Appliquer les filtres
            if search:
                commandes_query = commandes_query.filter(
                    Q(num_cmd__icontains=search) |
                    Q(id_yz__icontains=search) |
                    Q(client__nom__icontains=search) |
                    Q(client__prenom__icontains=search) |
                    Q(client__numero_tel__icontains=search)
                )
            
            if date_debut:
                commandes_query = commandes_query.filter(date_cmd__gte=date_debut)
            
            if date_fin:
                commandes_query = commandes_query.filter(date_cmd__lte=date_fin)
            
            commandes_batch = commandes_query.order_by('-date_cmd')[offset:offset + batch_size]
            
            if not commandes_batch:
                break
                
            for cmd in commandes_batch:
                confirmation_info = cmd.etats.filter(enum_etat__libelle__icontains='Confirmée').first()
                preparation_info = cmd.etats.filter(enum_etat__libelle__icontains='Préparation en cours').first()
                
                etat_livraison_obj = cmd.etats.filter(enum_etat__libelle__icontains='Livrée').first()
                etat_paiement_obj = cmd.etats.filter(enum_etat__libelle__icontains='Payée').first()
                piece_retournee_obj = cmd.etats.filter(enum_etat__libelle__icontains='Retournée').first()

                etat_paiement = etat_paiement_obj.enum_etat.libelle if etat_paiement_obj else "Non Payé"
                etat_livraison = etat_livraison_obj.enum_etat.libelle if etat_livraison_obj else "En attente"
                piece_retournee = "Oui" if piece_retournee_obj else "Non"

                operateur_assigne_obj = cmd.etats.filter(enum_etat__libelle__icontains='Affectée').first()
                operateur_assigne_nom = operateur_assigne_obj.operateur.mail if operateur_assigne_obj and operateur_assigne_obj.operateur else "N/A"
                
                agent_confirmation_nom = confirmation_info.operateur.mail if confirmation_info and confirmation_info.operateur else "N/A"

                tarif_livraison = 0.0
                reste_a_payer = cmd.total_cmd
                date_paiement = "N/A"
                observation_livraison = piece_retournee_obj.commentaire if piece_retournee_obj else ""

                # Informations communes à la commande
                commande_info = [
                    cmd.num_cmd,
                    cmd.id_yz,
                    f"{cmd.client.prenom} {cmd.client.nom}" if cmd.client else "N/A",
                    cmd.client.numero_tel if cmd.client else "N/A",
                    cmd.client.adresse if cmd.client else "N/A",
                    cmd.ville.nom if cmd.ville else "N/A",
                    cmd.ville.region.nom_region if cmd.ville and cmd.ville.region else "N/A",
                ]
                
                # Données communes de fin
                commande_fin_info = [
                    cmd.total_cmd,
                    cmd.date_cmd.strftime('%d/%m/%Y') if cmd.date_cmd else "N/A",
                    confirmation_info.enum_etat.libelle if confirmation_info else "Non Confirmée",
                    confirmation_info.date_debut.strftime('%d/%m/%Y %H:%M') if confirmation_info and confirmation_info.date_debut else "N/A",
                    confirmation_info.commentaire if confirmation_info else "",
                    operateur_assigne_nom,
                    agent_confirmation_nom,
                    "future qui seras des les tables models plustard dans le projet", # Client Fidèle
                    "Oui" if cmd.is_upsell else "Non", # UPSELL
                    preparation_info.enum_etat.libelle if preparation_info else "Non Préparée",
                    etat_livraison,
                    etat_paiement,
                    tarif_livraison,
                    reste_a_payer,
                    date_paiement,
                    piece_retournee,
                    observation_livraison,
                ]
                
                # Si la commande a des articles, créer une ligne par article
                paniers = cmd.paniers.all()
                if paniers:
                    for panier in paniers:
                        # Détails de l'article
                        article_info = [
                            panier.article.nom or "N/A",
                            panier.article.reference or "N/A",
                            panier.article.couleur or "N/A",
                            panier.article.pointure or "N/A",
                            panier.quantite,
                            panier.article.prix_unitaire,
                            panier.sous_total,
                        ]
                        
                        # Combiner toutes les informations
                        row_data = commande_info + article_info + commande_fin_info
                        writer.writerow(row_data)
                else:
                    # Si pas d'articles, créer une ligne avec des valeurs N/A
                    article_info = ["N/A", "N/A", "N/A", "N/A", 0, 0, 0]
                    row_data = commande_info + article_info + commande_fin_info
                    writer.writerow(row_data)

            offset += batch_size

        return response
    return redirect('app_admin:page_360')

@staff_member_required
@login_required
def export_all_data_excel(request):
    # Récupérer les filtres
    search = request.GET.get('search')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    
    # Créer un fichier Excel temporaire avec streaming
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="export_commandes_avec_paniers_360.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes avec Paniers"

    # En-têtes Excel avec style - Inclure les détails des paniers
    headers = [
        "N°", "Identifiant Yoozak", "CLIENT", "TELEPHONE", "ADRESSE", "VILLE", "REGION",
        "ARTICLE NOM", "ARTICLE REFERENCE", "ARTICLE COULEUR", "ARTICLE POINTURE", 
        "QUANTITE", "PRIX UNITAIRE", "SOUS TOTAL ARTICLE",
        "PRIX TOTAL COMMANDE (DH)", "DATE COMMANDE", "CONFIRMATION", "DATE CONFIRMATION",
        "OBSERVATIONS CONFIRMATION", "OPERATEUR", "AGENT CONFIRMATION",
        "CLIENT FIDELE", "UPSELL", "PREPARATION", "ETAT LIVRAISON",
        "ETAT PAIEMENT", "TARIF", "RESTE A PAYER", "DATE PAIEMENT", "PIECE RETOURNEE",
        "OBSERVATION LIVRAISON"
    ]
    ws.append(headers)

    # Style des en-têtes
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Traitement par batch pour Excel aussi
    batch_size = 500  # Plus petit pour Excel car plus lourd
    offset = 0
    
    while True:
        # Construire la requête avec les mêmes filtres que la vue
        commandes_query = Commande.objects.select_related(
            'client', 
            'ville', 
            'ville__region'
        ).prefetch_related(
            'etats__enum_etat', 
            'etats__operateur', 
            'paniers__article'
        ).only(
            'id', 'num_cmd', 'id_yz', 'date_cmd', 'total_cmd', 'is_upsell',
            'client__nom', 'client__prenom', 'client__numero_tel', 'client__adresse',
            'ville__nom', 'ville__region__nom_region'
        )
        
        # Appliquer les filtres
        if search:
            commandes_query = commandes_query.filter(
                Q(num_cmd__icontains=search) |
                Q(id_yz__icontains=search) |
                Q(client__nom__icontains=search) |
                Q(client__prenom__icontains=search) |
                Q(client__numero_tel__icontains=search)
            )
        
        if date_debut:
            commandes_query = commandes_query.filter(date_cmd__gte=date_debut)
            
        if date_fin:
            commandes_query = commandes_query.filter(date_cmd__lte=date_fin)
        
        commandes_batch = commandes_query.order_by('-date_cmd')[offset:offset + batch_size]
        
        if not commandes_batch:
            break
            
        for cmd in commandes_batch:
            confirmation_info = cmd.etats.filter(enum_etat__libelle__icontains='Confirmée').first()
            preparation_info = cmd.etats.filter(enum_etat__libelle__icontains='Préparation en cours').first()
            
            # Récupérer les informations de livraison et paiement depuis les EtatsCommande
            etat_livraison_obj = cmd.etats.filter(enum_etat__libelle__icontains='Livrée').first()
            etat_paiement_obj = cmd.etats.filter(enum_etat__libelle__icontains='Payée').first()
            piece_retournee_obj = cmd.etats.filter(enum_etat__libelle__icontains='Retournée').first()

            etat_paiement = etat_paiement_obj.enum_etat.libelle if etat_paiement_obj else "Non Payé"
            etat_livraison = etat_livraison_obj.enum_etat.libelle if etat_livraison_obj else "En attente"
            piece_retournee = "Oui" if piece_retournee_obj else "Non"

            # Opérateur Assigné: Chercher l'opérateur lié à l'état 'Affectée' ou un autre état pertinent
            operateur_assigne_obj = cmd.etats.filter(enum_etat__libelle__icontains='Affectée').first()
            operateur_assigne_nom = operateur_assigne_obj.operateur.mail if operateur_assigne_obj and operateur_assigne_obj.operateur else "N/A"
            
            # Agent Confirmation: Opérateur lié à l'état 'Confirmée'
            agent_confirmation_nom = confirmation_info.operateur.mail if confirmation_info and confirmation_info.operateur else "N/A"

            tarif_livraison = 0.0
            reste_a_payer = cmd.total_cmd
            date_paiement = "N/A"
            observation_livraison = piece_retournee_obj.commentaire if piece_retournee_obj else ""

            # Informations communes à la commande
            commande_info = [
                cmd.num_cmd,
                cmd.id_yz,
                f"{cmd.client.prenom} {cmd.client.nom}" if cmd.client else "N/A",
                cmd.client.numero_tel if cmd.client else "N/A",
                cmd.client.adresse if cmd.client else "N/A",
                cmd.ville.nom if cmd.ville else "N/A",
                cmd.ville.region.nom_region if cmd.ville and cmd.ville.region else "N/A",
            ]
            
            # Données communes de fin
            commande_fin_info = [
                cmd.total_cmd,
                cmd.date_cmd.strftime('%d/%m/%Y') if cmd.date_cmd else "N/A",
                confirmation_info.enum_etat.libelle if confirmation_info else "Non Confirmée",
                confirmation_info.date_debut.strftime('%d/%m/%Y %H:%M') if confirmation_info and confirmation_info.date_debut else "N/A",
                confirmation_info.commentaire if confirmation_info else "",
                operateur_assigne_nom,
                agent_confirmation_nom,
                "future qui seras des les tables models plustard dans le projet", # Client Fidèle
                "Oui" if cmd.is_upsell else "Non", # UPSELL
                preparation_info.enum_etat.libelle if preparation_info else "Non Préparée",
                etat_livraison,
                etat_paiement,
                tarif_livraison,
                reste_a_payer,
                date_paiement,
                piece_retournee,
                observation_livraison,
            ]
            
            # Si la commande a des articles, créer une ligne par article
            paniers = cmd.paniers.all()
            if paniers:
                for panier in paniers:
                    # Détails de l'article
                    article_info = [
                        panier.article.nom or "N/A",
                        panier.article.reference or "N/A",
                        panier.article.couleur or "N/A",
                        panier.article.pointure or "N/A",
                        panier.quantite,
                        panier.article.prix_unitaire,
                        panier.sous_total,
                    ]
                    
                    # Combiner toutes les informations
                    row_data = commande_info + article_info + commande_fin_info
                    ws.append(row_data)
            else:
                # Si pas d'articles, créer une ligne avec des valeurs N/A
                article_info = ["N/A", "N/A", "N/A", "N/A", 0, 0, 0]
                row_data = commande_info + article_info + commande_fin_info
                ws.append(row_data)
            
        offset += batch_size

    # Ajuster la largeur des colonnes
    for i, col in enumerate(headers, 1):
        if i <= 7:  # Colonnes de base commande
            ws.column_dimensions[get_column_letter(i)].width = 20
        elif i <= 14:  # Colonnes articles
            ws.column_dimensions[get_column_letter(i)].width = 15
        else:  # Autres colonnes
            ws.column_dimensions[get_column_letter(i)].width = 18

    wb.save(response)
    return response 