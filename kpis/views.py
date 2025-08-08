from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, Q, F, Min, Max, Subquery, OuterRef
from django.utils import timezone
from datetime import datetime, timedelta
from decimal import Decimal
from functools import wraps
import csv
from io import BytesIO
try:
    import openpyxl
except ImportError:
    openpyxl = None
import logging

from commande.models import Commande, Panier, EtatCommande, Operation, EnumEtatCmd
from article.models import Article
from client.models import Client
from parametre.models import Operateur

logger = logging.getLogger(__name__)

def format_number_fr(number, decimals=0):
    """Formate un nombre selon les standards français (espace comme séparateur de milliers)"""
    if number is None or (isinstance(number, (int, float)) and number == 0):
        return "0"
    
    if isinstance(number, (int, float)):
        if decimals == 0:
            # Pour les entiers, utiliser un format sans décimales
            return f"{number:,.0f}".replace(",", " ")
        else:
            # Pour les décimales
            return f"{number:,.{decimals}f}".replace(",", " ")
    
    # Si c'est déjà une chaîne, la retourner telle quelle
    return str(number)

def api_login_required(view_func):
    """Décorateur qui renvoie du JSON au lieu de rediriger pour les APIs"""
    @wraps(view_func)
    def wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({
                'success': False,
                'message': 'Authentification requise',
                'error': 'non_authentifie'
            }, status=401)
        return view_func(request, *args, **kwargs)
    return wrapped_view

def get_integer_fields():
    """Retourne la liste des champs qui doivent être des entiers"""
    integer_fields = [
        'delai_livraison_defaut',
        'fidelisation_commandes_min', 
        'fidelisation_periode_jours',
        'periode_analyse_standard',
        'delai_livraison_alerte',
        'stock_critique_seuil',
        'stock_ventes_minimum',
        'periode_analyse_defaut'  # Ajouté pour les selects
    ]
    return integer_fields

def validate_integer_value(nom_parametre, valeur):
    """Valide qu'une valeur est un entier pour les champs qui le requièrent"""
    integer_fields = get_integer_fields()
    
    if nom_parametre in integer_fields:
        # Vérifier que c'est un entier
        if not isinstance(valeur, int) and not (isinstance(valeur, float) and valeur.is_integer()):
            return False, f"Ce champ doit être un nombre entier (pas de décimales)"
        
        # Convertir en entier si c'est un float qui représente un entier
        if isinstance(valeur, float) and valeur.is_integer():
            valeur = int(valeur)
    
    return True, valeur

@login_required
def ventes_data(request):
    """API pour les données de l'onglet Ventes - E-commerce téléphonique Yoozak"""
    try:
        # Dates de référence
        aujourd_hui = timezone.now().date()
        debut_mois = aujourd_hui.replace(day=1)
        mois_precedent = (debut_mois - timedelta(days=1)).replace(day=1)
        debut_30j = aujourd_hui - timedelta(days=30)
        # === KPI 1: CA par Période (Commandes livrées uniquement) ===
        # CA de ce mois (du 1er du mois jusqu'à aujourd'hui) - Commandes livrées
        ca_mois_actuel = Commande.objects.filter(
            date_cmd__gte=debut_mois,
            date_cmd__lte=aujourd_hui,
            etats__enum_etat__libelle__iexact='Livrée'
        ).aggregate(total=Sum('total_cmd'))['total'] or 0
        
        # CA mois précédent (du 1er au dernier jour du mois précédent) - Commandes livrées
        fin_mois_precedent = debut_mois - timedelta(days=1)
        ca_mois_precedent = Commande.objects.filter(
            date_cmd__gte=mois_precedent,
            date_cmd__lte=fin_mois_precedent,
            etats__enum_etat__libelle__iexact='Livrée'
        ).aggregate(total=Sum('total_cmd'))['total'] or 0
        
        # Tendance CA
        if ca_mois_precedent > 0:
            tendance_ca = ((ca_mois_actuel - ca_mois_precedent) / ca_mois_precedent) * 100
        else:
            tendance_ca = 100 if ca_mois_actuel > 0 else 0
        
        # === KPI 2: Panier Moyen (Commandes livrées uniquement) ===
        panier_moyen_mois = Commande.objects.filter(
            date_cmd__gte=debut_mois,
            date_cmd__lte=aujourd_hui,
            etats__enum_etat__libelle__iexact='Livrée'
        ).aggregate(moyenne=Avg('total_cmd'))['moyenne'] or 0
        
        # Panier moyen mois précédent - Commandes livrées
        panier_moyen_precedent = Commande.objects.filter(
            date_cmd__gte=mois_precedent,
            date_cmd__lte=fin_mois_precedent,
            etats__enum_etat__libelle__iexact='Livrée'
        ).aggregate(moyenne=Avg('total_cmd'))['moyenne'] or 0
        
        # Tendance panier moyen
        if panier_moyen_precedent > 0:
            tendance_panier = ((panier_moyen_mois - panier_moyen_precedent) / panier_moyen_precedent) * 100
        else:
            tendance_panier = 100 if panier_moyen_mois > 0 else 0
        # === KPI 3: Nombre de Commandes (Commandes livrées uniquement) ===
        nb_commandes_mois = Commande.objects.filter(
            date_cmd__gte=debut_mois,
            date_cmd__lte=aujourd_hui,
            etats__enum_etat__libelle__iexact='Livrée'
        ).count()
        
        nb_commandes_mois_precedent = Commande.objects.filter(
            date_cmd__gte=mois_precedent,
            date_cmd__lte=fin_mois_precedent,
            etats__enum_etat__libelle__iexact='Livrée'
        ).count()
        
        # Tendance nombre de commandes
        if nb_commandes_mois_precedent > 0:
            tendance_commandes = ((nb_commandes_mois - nb_commandes_mois_precedent) / nb_commandes_mois_precedent) * 100
        else:
            tendance_commandes = 100 if nb_commandes_mois > 0 else 0
        
        # === KPIs Secondaires ===
        
        # Top 5 Modèles par CA (ce mois) - Basé sur les commandes livrées
        top_modeles = (Article.objects
            .filter(
                paniers__commande__date_cmd__gte=debut_mois, 
                paniers__commande__date_cmd__lte=aujourd_hui,
                paniers__commande__etats__enum_etat__libelle__iexact='Livrée',
                paniers__commande__etats__date_fin__isnull=True  # État actuel
            )
            .annotate(
                ca_total=Sum('paniers__sous_total'),
                quantite_vendue=Sum('paniers__quantite')
            )
            .order_by('-ca_total')[:5]
        )
        
        # Recherche du top modèle et top région
        top_modele = top_modeles.first() if top_modeles else None
        
        # Top Région par CA (ce mois)
        # Filtrer d'abord les commandes avec des régions valides
        top_regions = (Commande.objects
            .filter(date_cmd__gte=debut_mois, date_cmd__lte=aujourd_hui)
            .exclude(etats__enum_etat__libelle__iexact='Annulée')
            .exclude(ville__region__nom_region__isnull=True)  # Exclure les régions nulles
            .exclude(ville__region__nom_region__exact='')     # Exclure les régions vides
            .values('ville__region__nom_region')
            .annotate(
                ca_total=Sum('total_cmd'),
                nb_commandes=Count('id')
            )
            .order_by('-ca_total')[:5]
        )
        
        top_region = top_regions.first() if top_regions else None
        
        # Vérifier s'il y a des commandes sans région pour diagnostic
        commandes_sans_region = Commande.objects.filter(
            date_cmd__gte=debut_mois,
            date_cmd__lte=aujourd_hui,
            ville__region__nom_region__isnull=True
        ).exclude(etats__enum_etat__libelle__iexact='Annulée').count()
        
        # Commande maximale (ce mois) - Basée sur les commandes livrées
        commande_max = Commande.objects.filter(
            date_cmd__gte=debut_mois,
            date_cmd__lte=aujourd_hui,
            etats__enum_etat__libelle__iexact='Livrée'
        ).aggregate(max_cmd=Max('total_cmd'))['max_cmd'] or 0
        
        # Répartition par Catégorie
        ventes_par_categorie = (Article.objects
            .filter(paniers__commande__date_cmd__gte=debut_30j)
            .exclude(paniers__commande__etats__enum_etat__libelle__iexact='Annulée')
            .values('categorie')
            .annotate(
                ca_total=Sum('paniers__sous_total'),
                quantite=Sum('paniers__quantite')
            )
            .order_by('-ca_total')
        )
        
        # Répartition Géographique (Top 5 villes)
        ventes_par_ville = (Commande.objects
            .filter(date_cmd__gte=debut_30j)
            .exclude(etats__enum_etat__libelle__iexact='Annulée')
            .values('ville__nom', 'ville__region__nom_region')
            .annotate(
                ca_total=Sum('total_cmd'),
                nb_commandes=Count('id')
            )
            .order_by('-ca_total')[:5]
        )
        
        # Performance Opérateurs (Confirmation)
        performance_operateurs = []
        if Operation.objects.filter(date_operation__gte=debut_30j).exists():
            operateurs_stats = (Operation.objects
                .filter(date_operation__gte=debut_30j, type_operation__in=['APPEL', 'Appel Whatsapp'])
                .values('operateur__nom', 'operateur__prenom')
                .annotate(
                    nb_appels=Count('id'),
                    commandes_confirmees=Count('commande__etats__enum_etat__libelle', 
                                            filter=Q(commande__etats__enum_etat__libelle__iexact='Confirmée'))
                )
                .order_by('-nb_appels')[:3]
            )
            
            for stat in operateurs_stats:
                taux = (stat['commandes_confirmees'] / stat['nb_appels'] * 100) if stat['nb_appels'] > 0 else 0
                performance_operateurs.append({
                    'nom': f"{stat['operateur__prenom']} {stat['operateur__nom']}",
                    'nb_appels': stat['nb_appels'],
                    'taux_confirmation': round(taux, 1)
                })
        
        # Réponse JSON
        data = {
            'success': True,
            'timestamp': timezone.now().isoformat(),
            'kpis_principaux': {
                'ca_periode': {
                    'valeur': float(ca_mois_actuel),
                    'valeur_formatee': format_number_fr(ca_mois_actuel),
                    'tendance': round(tendance_ca, 1),
                    'unite': 'DH',
                    'label': 'CA Total',
                    'sub_value': f"Ce mois vs {format_number_fr(ca_mois_precedent)} DH mois dernier"
                },
                'panier_moyen': {
                    'valeur': float(panier_moyen_mois) if panier_moyen_mois else 0,
                    'valeur_formatee': format_number_fr(panier_moyen_mois) if panier_moyen_mois else "0",
                    'tendance': round(tendance_panier, 1),
                    'unite': 'DH',
                    'label': 'Panier Moyen',
                    'sub_value': f"Ce mois vs {format_number_fr(panier_moyen_precedent)} DH mois dernier"
                },
                'nb_commandes': {
                    'valeur': nb_commandes_mois,
                    'valeur_formatee': format_number_fr(nb_commandes_mois),
                    'tendance': round(tendance_commandes, 1),
                    'unite': 'commandes',
                    'label': 'Nb Commandes',
                    'sub_value': f"Ce mois vs {format_number_fr(nb_commandes_mois_precedent)} mois dernier"
                }
            },
            'kpis_secondaires': {
                'top_modele': {
                    'nom': top_modele.nom if top_modele else 'Aucun',
                    'ca': float(top_modele.ca_total) if top_modele and top_modele.ca_total else 0,
                    'ca_formate': format_number_fr(top_modele.ca_total) if top_modele and top_modele.ca_total else "0",
                    'pourcentage': round((float(top_modele.ca_total) / ca_mois_actuel * 100), 1) if top_modele and top_modele.ca_total and ca_mois_actuel > 0 else 0
                },
                'top_region': {
                    'nom': (top_region['ville__region__nom_region'] 
                           if top_region and top_region['ville__region__nom_region'] 
                           else 'Données géographiques manquantes sur les commandes'),
                    'ca': float(top_region['ca_total']) if top_region else 0,
                    'ca_formate': (format_number_fr(top_region['ca_total']) 
                                  if top_region 
                                  else ""),
                    'pourcentage': (round((float(top_region['ca_total']) / ca_mois_actuel * 100), 1) 
                                   if top_region and ca_mois_actuel > 0 
                                   else 0),
                    'est_donnees_manquantes': not bool(top_region),
                    'affichage_simple': not bool(top_region)
                },
                'commande_max': {
                    'valeur': float(commande_max),
                    'valeur_formatee': format_number_fr(commande_max)
                },
                'top_modeles': [
                    {
                        'nom': article.nom,
                        'ca': float(article.ca_total) if article.ca_total else 0,
                        'quantite': article.quantite_vendue or 0,
                        'couleur': getattr(article, 'couleur', '#3b82f6'),
                        'reference': getattr(article, 'reference', 'N/A')
                    } for article in top_modeles
                ]
            }
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'message': 'Erreur lors du chargement des données Ventes'
        }, status=500)

@login_required
def evolution_ca_data(request):
    """API pour l'évolution du CA sur une période donnée"""
    try:
        # Paramètres
        periode = request.GET.get('period', '30j')
        jours_map = {'7j': 7, '30j': 30, '90j': 90}
        nb_jours = jours_map.get(periode, 30)
        
        # Dates avec timezone
        fin_date = timezone.now()
        debut_date = fin_date - timedelta(days=nb_jours)
        
        # Calcul des données réelles basées sur les commandes livrées
        commandes_par_jour = Commande.objects.filter(
            date_cmd__gte=debut_date,
            date_cmd__lte=fin_date,
            etats__enum_etat__libelle__iexact='Livrée',
            etats__date_fin__isnull=False
        ).extra(select={'date_seule': 'DATE(date_cmd)'}).values('date_seule').annotate(
            ca_jour=Sum('total_cmd')
        ).order_by('date_seule')
        
        # Construction des données de réponse
        evolution_data = []
        ca_par_jour = {}
        
        for cmd in commandes_par_jour:
            date_str = cmd['date_seule']
            if isinstance(date_str, str):
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                date_obj = date_str
            ca_par_jour[date_obj] = float(cmd['ca_jour'] or 0)
          # Remplir tous les jours de la période avec les données réelles uniquement
        date_courante = debut_date.date()
        while date_courante <= fin_date.date():
            ca_jour = ca_par_jour.get(date_courante, 0)  # 0 si pas de ventes ce jour-là
            
            evolution_data.append({
                'date': date_courante.strftime('%Y-%m-%d'),
                'date_formatee': date_courante.strftime('%d/%m'),
                'ca': ca_jour,
                'ca_formate': f"{ca_jour:,.0f} DH" if ca_jour > 0 else "0 DH"
            })
            
            date_courante += timedelta(days=1)
          # Calculs de tendance basés sur les données réelles uniquement
        ca_total = sum(d['ca'] for d in evolution_data)
        ca_moyen = ca_total / len(evolution_data) if evolution_data else 0
        
        # Tendance : comparaison entre la première et la dernière semaine (si suffisamment de données)
        tendance = 0
        if len(evolution_data) >= 14:
            # Comparer première et dernière semaine
            premiere_semaine = evolution_data[:7]
            derniere_semaine = evolution_data[-7:]
            
            ca_debut = sum(d['ca'] for d in premiere_semaine) / 7
            ca_fin = sum(d['ca'] for d in derniere_semaine) / 7
            
            if ca_debut > 0:
                tendance = ((ca_fin - ca_debut) / ca_debut * 100)
        elif len(evolution_data) >= 7:
            # Si moins de 14 jours, comparer première et dernière moitié
            milieu = len(evolution_data) // 2
            premiere_moitie = evolution_data[:milieu]
            derniere_moitie = evolution_data[milieu:]
            
            ca_debut = sum(d['ca'] for d in premiere_moitie) / len(premiere_moitie) if premiere_moitie else 0
            ca_fin = sum(d['ca'] for d in derniere_moitie) / len(derniere_moitie) if derniere_moitie else 0
            
            if ca_debut > 0:
                tendance = ((ca_fin - ca_debut) / ca_debut * 100)
          # Statistiques supplémentaires pour l'analyse
        jours_avec_ventes = len([d for d in evolution_data if d['ca'] > 0])
        ca_max_jour = max([d['ca'] for d in evolution_data]) if evolution_data else 0
        ca_min_jour = min([d['ca'] for d in evolution_data if d['ca'] > 0]) if jours_avec_ventes > 0 else 0
        
        response_data = {
            'success': True,
            'periode': periode,
            'evolution': evolution_data,
            'resume': {
                'ca_total': ca_total,
                'ca_moyen': ca_moyen,
                'ca_max_jour': ca_max_jour,
                'ca_min_jour': ca_min_jour,
                'tendance': round(tendance, 2),
                'nb_jours': nb_jours,
                'jours_avec_ventes': jours_avec_ventes,
                'taux_activite': round((jours_avec_ventes / nb_jours * 100), 1) if nb_jours > 0 else 0
            },
            'timestamp': timezone.now().isoformat()
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'message': 'Erreur lors du calcul de l\'évolution du CA'
        }, status=500)

@login_required
def top_modeles_data(request):
    """API pour les données du top modèles par CA"""
    try:
        # Paramètres
        limite = int(request.GET.get('limit', 10))
        periode_jours = int(request.GET.get('days', 30))
        
        # Dates
        fin_date = timezone.now()
        debut_date = fin_date - timedelta(days=periode_jours)
        
        # Calcul des ventes par article/modèle basé sur les commandes livrées
        top_modeles = Article.objects.annotate(
            ca_total=Sum(
                'paniers__sous_total',
                filter=Q(
                    paniers__commande__date_cmd__gte=debut_date,
                    paniers__commande__date_cmd__lte=fin_date,
                    paniers__commande__etats__enum_etat__libelle__iexact='Livrée',
                    paniers__commande__etats__date_fin__isnull=False
                )
            ),
            nb_ventes=Count(
                'paniers',
                filter=Q(
                    paniers__commande__date_cmd__gte=debut_date,
                    paniers__commande__date_cmd__lte=fin_date,
                    paniers__commande__etats__enum_etat__libelle__iexact='Livrée',
                    paniers__commande__etats__date_fin__isnull=False
                )
            )
        ).filter(
            ca_total__isnull=False,
            ca_total__gt=0
        ).order_by('-ca_total')[:limite]
          # Préparation des données (uniquement les données réelles)
        modeles_data = []
        couleurs = [
            '#3b82f6', '#ef4444', '#10b981', '#f59e0b', '#8b5cf6',
            '#ec4899', '#6b7280', '#14b8a6', '#f97316', '#84cc16'
        ]
        
        for i, article in enumerate(top_modeles):
            modeles_data.append({
                'nom': article.nom,
                'reference': article.reference,
                'ca': float(article.ca_total or 0),
                'ca_formate': f"{article.ca_total or 0:,.0f} DH",
                'nb_ventes': article.nb_ventes or 0,
                'prix_moyen': float(article.ca_total / article.nb_ventes) if article.nb_ventes > 0 else 0,
                'couleur': couleurs[i % len(couleurs)]
            })
        
        # Statistiques
        ca_total = sum(m['ca'] for m in modeles_data)
        ca_moyen = ca_total / len(modeles_data) if modeles_data else 0
        
        response_data = {
            'success': True,
            'modeles': modeles_data,
            'stats': {
                'ca_total': ca_total,
                'ca_moyen': ca_moyen,
                'nb_modeles': len(modeles_data),
                'periode_jours': periode_jours
            },
            'timestamp': timezone.now().isoformat()
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'message': 'Erreur lors du chargement du top modèles'
        }, status=500)

@login_required
def performance_regions_data(request):
    """API pour les données de performance par région"""
    try:
        # Période de référence (30 derniers jours par défaut)
        periode = request.GET.get('period', '30j')
        aujourd_hui = timezone.now().date()
        
        if periode == '7j':
            debut_periode = aujourd_hui - timedelta(days=7)
        elif periode == '90j':
            debut_periode = aujourd_hui - timedelta(days=90)
        else:  # 30j par défaut
            debut_periode = aujourd_hui - timedelta(days=30)
        
        # Récupérer les données par région
        regions_data = Commande.objects.filter(
            date_cmd__gte=debut_periode,
            date_cmd__lte=aujourd_hui,
            ville__isnull=False,
            ville__region__isnull=False
        ).exclude(
            etats__enum_etat__libelle__iexact='Annulée',
            etats__date_fin__isnull=True
        ).values(
            'ville__region__nom_region'
        ).annotate(
            ca_total=Sum('total_cmd'),
            nb_commandes=Count('id'),
            ca_moyen=Avg('total_cmd')
        ).order_by('-ca_total')
          # Calculer le total général pour les pourcentages
        total_ca_general = sum(region['ca_total'] or 0 for region in regions_data)
        
        # Gérer le cas où il n'y a pas de données
        if not regions_data:
            return JsonResponse({
                'success': True,
                'regions': [],
                'stats': {
                    'total_regions': 0,
                    'ca_total_general': 0,
                    'ca_total_format': '0 DH',
                    'nb_commandes_total': 0,
                    'ca_moyen_general': 0
                },
                'periode': periode,
                'message': 'Aucune donnée disponible pour cette période',
                'empty': True
            })
        
        # Formater les données pour le frontend
        regions_formattees = []
        
        for i, region in enumerate(regions_data):
            ca_total = region['ca_total'] or 0
            pourcentage = (ca_total / total_ca_general * 100) if total_ca_general > 0 else 0
            
            # Couleurs pour différencier les régions
            couleurs = [
                '#3b82f6',  # Bleu
                '#10b981',  # Vert
                '#f59e0b',  # Orange
                '#ef4444',  # Rouge
                '#8b5cf6',  # Violet
                '#06b6d4',  # Cyan
                '#84cc16',  # Lime
                '#f97316',  # Orange foncé
            ]
            
            regions_formattees.append({
                'nom_region': region['ville__region__nom_region'],
                'ca_total': float(ca_total),
                'ca_total_format': f"{ca_total/1000:.0f}K DH" if ca_total >= 1000 else f"{ca_total:.0f} DH",
                'nb_commandes': region['nb_commandes'],
                'ca_moyen': float(region['ca_moyen'] or 0),
                'pourcentage': round(pourcentage, 1),
                'couleur': couleurs[i % len(couleurs)]
            })
        
        # Limiter aux 5 premières régions pour l'affichage
        top_regions = regions_formattees[:5]
        
        # Calculer les statistiques globales
        stats_globales = {
            'total_regions': len(regions_formattees),
            'ca_total_general': float(total_ca_general),
            'ca_total_format': f"{total_ca_general/1000000:.1f}M DH" if total_ca_general >= 1000000 else f"{total_ca_general/1000:.0f}K DH",
            'nb_commandes_total': sum(region['nb_commandes'] for region in regions_formattees),
            'ca_moyen_general': float(total_ca_general / len(regions_formattees)) if regions_formattees else 0
        };
        
        response_data = {
            'success': True,
            'regions': top_regions,
            'stats': stats_globales,
            'periode': periode,
            'message': f'Données chargées pour {len(top_regions)} régions'
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'message': 'Erreur lors du chargement des performances par région'
        }, status=500)

@login_required
def clients_data(request):
    """API pour les données de l'onglet Clients - Analyse comportementale Yoozak"""
    try:
        # Dates de référence
        aujourd_hui = timezone.now().date()
        debut_mois = aujourd_hui.replace(day=1)
        mois_precedent = (debut_mois - timedelta(days=1)).replace(day=1)
        debut_30j = aujourd_hui - timedelta(days=30)
        
        # === KPI 1: Nouveaux Clients ===
        nouveaux_clients_mois = Client.objects.filter(
            date_creation__date__gte=debut_mois,
            date_creation__date__lte=aujourd_hui
        ).count()
        
        # Nouveaux clients mois précédent
        fin_mois_precedent = debut_mois - timedelta(days=1)
        nouveaux_clients_precedent = Client.objects.filter(
            date_creation__date__gte=mois_precedent,
            date_creation__date__lte=fin_mois_precedent
        ).count()
        
        difference_nouveaux = nouveaux_clients_mois - nouveaux_clients_precedent
        
        # Plus besoin de calcul de moyenne journalière - simplification
        
        # === KPI 2: Clients Actifs (30 derniers jours) ===
        # Clients qui ont au moins une commande livrée
        clients_actifs_ids = Commande.objects.filter(
            date_cmd__gte=debut_30j,
            etats__enum_etat__libelle__iexact='Livrée'
        ).values_list('client_id', flat=True).distinct()
        
        clients_actifs_30j = len(set(clients_actifs_ids))
        
        # Clients actifs période précédente (commandes livrées)
        debut_periode_precedente = debut_30j - timedelta(days=30)
        clients_actifs_precedent_ids = Commande.objects.filter(
            date_cmd__gte=debut_periode_precedente,
            date_cmd__lt=debut_30j,
            etats__enum_etat__libelle__iexact='Livrée'
        ).values_list('client_id', flat=True).distinct()
        
        clients_actifs_precedent = len(set(clients_actifs_precedent_ids))
        difference_actifs = clients_actifs_30j - clients_actifs_precedent
        
        # Pourcentage du total
        total_clients = Client.objects.count()
        pourcentage_actifs = (clients_actifs_30j / total_clients * 100) if total_clients > 0 else 0
        
        # === KPI 3: Taux de Retour ===
        # CALCUL CORRECT - basé sur les commandes livrées uniquement
        # On ne peut retourner que ce qui a été livré !
        commandes_livrees_30j = Commande.objects.filter(
            date_cmd__gte=debut_30j,
            etats__enum_etat__libelle__iexact='Livrée'
        ).distinct().count()
        
        # Commandes réellement retournées (avec état "Retournée")
        commandes_retournees = Commande.objects.filter(
            date_cmd__gte=debut_30j,
            etats__enum_etat__libelle__iexact='Retournée'
        ).distinct().count()
        
        taux_retour = (commandes_retournees / commandes_livrees_30j * 100) if commandes_livrees_30j > 0 else 0
        
        # Taux retour période précédente
        commandes_livrees_precedent = Commande.objects.filter(
            date_cmd__gte=debut_periode_precedente,
            date_cmd__lt=debut_30j,
            etats__enum_etat__libelle__iexact='Livrée'
        ).distinct().count()
        
        retours_precedent = Commande.objects.filter(
            date_cmd__gte=debut_periode_precedente,
            date_cmd__lt=debut_30j,
            etats__enum_etat__libelle__iexact='Retournée'
        ).distinct().count()
        
        taux_retour_precedent = (retours_precedent / commandes_livrees_precedent * 100) if commandes_livrees_precedent > 0 else 0
        tendance_retour = taux_retour - taux_retour_precedent
        
        # === KPI 4: Fidélisation Client (90 jours - adapté secteur chaussures) ===
        # Période plus longue car les chaussures se rachètent moins fréquemment
        debut_90j = aujourd_hui - timedelta(days=90)
        debut_90j_precedent = debut_90j - timedelta(days=90)
        
        # Clients actifs sur 90 jours (commandes livrées uniquement)
        clients_actifs_90j_ids = Commande.objects.filter(
            date_cmd__gte=debut_90j,
            etats__enum_etat__libelle__iexact='Livrée'
        ).values_list('client_id', flat=True).distinct()
        clients_actifs_90j = len(set(clients_actifs_90j_ids))
        
        # Clients fidèles (2+ commandes livrées sur 90 jours)
        clients_fideles_90j = 0
        for client_id in clients_actifs_90j_ids:
            nb_commandes_livrees = Commande.objects.filter(
                client_id=client_id,
                date_cmd__gte=debut_90j,
                etats__enum_etat__libelle__iexact='Livrée'
            ).count()
            if nb_commandes_livrees >= 2:
                clients_fideles_90j += 1
        
        # Calcul du taux de fidélisation
        taux_fidelisation = (clients_fideles_90j / clients_actifs_90j * 100) if clients_actifs_90j > 0 else 0
        
        # Tendance fidélisation (vs période précédente 90j) - commandes livrées
        clients_actifs_90j_precedent_ids = Commande.objects.filter(
            date_cmd__gte=debut_90j_precedent,
            date_cmd__lt=debut_90j,
            etats__enum_etat__libelle__iexact='Livrée'
        ).values_list('client_id', flat=True).distinct()
        clients_actifs_90j_precedent = len(set(clients_actifs_90j_precedent_ids))
        
        clients_fideles_90j_precedent = 0
        for client_id in clients_actifs_90j_precedent_ids:
            nb_commandes_livrees_precedent = Commande.objects.filter(
                client_id=client_id,
                date_cmd__gte=debut_90j_precedent,
                date_cmd__lt=debut_90j,
                etats__enum_etat__libelle__iexact='Livrée'
            ).count()
            if nb_commandes_livrees_precedent >= 2:
                clients_fideles_90j_precedent += 1
        
        taux_fidelisation_precedent = (clients_fideles_90j_precedent / clients_actifs_90j_precedent * 100) if clients_actifs_90j_precedent > 0 else 0
        tendance_fidelisation = taux_fidelisation - taux_fidelisation_precedent
        
        # === ANALYSES DÉTAILLÉES ===
        
        # Top Clients VIP (par CA) - basé sur commandes livrées
        top_clients_data = []
        commandes_avec_ca = Commande.objects.filter(
            date_cmd__gte=debut_30j,
            etats__enum_etat__libelle__iexact='Livrée'
        ).values('client_id').annotate(
            ca_total=Sum('total_cmd'),
            nb_commandes=Count('id')
        ).order_by('-ca_total')[:5]
        
        for client_data in commandes_avec_ca:
            try:
                client = Client.objects.get(id=client_data['client_id'])
                ca_total = float(client_data['ca_total']) if client_data['ca_total'] else 0
                # Formatage français avec espaces comme séparateurs de milliers
                ca_total_format = f"{ca_total:,.0f}".replace(',', ' ') + " DH" if ca_total > 0 else "0 DH"
                
                top_clients_data.append({
                    'nom': f"{client.prenom} {client.nom[0]}." if client.nom else "Client anonyme",
                    'ca_total': ca_total,
                    'ca_total_format': ca_total_format,
                    'nb_commandes': client_data['nb_commandes']
                })
            except Client.DoesNotExist:
                continue
          # Performance mensuelle
        # Commandes du mois en cours (livrées uniquement)
        commandes_mois_actuel = Commande.objects.filter(
            date_cmd__gte=debut_mois,
            date_cmd__lte=aujourd_hui,
            etats__enum_etat__libelle__iexact='Livrée'
        ).count()
        
        # CA moyen par client actif - CALCUL CORRECT basé sur commandes livrées
        # Utiliser le CA total de TOUS les clients actifs, pas seulement le top 5
        ca_total_tous_clients_actifs = Commande.objects.filter(
            date_cmd__gte=debut_30j,
            client_id__in=clients_actifs_ids,
            etats__enum_etat__libelle__iexact='Livrée'
        ).aggregate(ca_total=Sum('total_cmd'))['ca_total'] or 0
        
        ca_moyen_par_client = ca_total_tous_clients_actifs / clients_actifs_30j if clients_actifs_30j > 0 else 0
        
        # Segmentation comportementale adaptée (basée sur 90 jours pour plus de pertinence) - commandes livrées
        clients_reguliers = sum(1 for client_id in clients_actifs_90j_ids 
                              if Commande.objects.filter(
                                  client_id=client_id, 
                                  date_cmd__gte=debut_90j,
                                  etats__enum_etat__libelle__iexact='Livrée'
                              ).count() >= 3)
        
        clients_nouveaux_testeurs = Client.objects.filter(
            date_creation__date__gte=debut_90j
        ).filter(
            id__in=clients_actifs_90j_ids
        ).count()
        
        clients_occasionnels = sum(1 for client_id in clients_actifs_90j_ids 
                                 if Commande.objects.filter(
                                     client_id=client_id, 
                                     date_cmd__gte=debut_90j,
                                     etats__enum_etat__libelle__iexact='Livrée'
                                 ).count() == 2)
        
        clients_vip = len(top_clients_data)
        total_clients_analyse = clients_actifs_90j
        
        # Réponse JSON structurée
        data = {
            'success': True,
            'timestamp': timezone.now().isoformat(),
            'kpis_principaux': {
                'nouveaux_clients': {
                    'valeur': nouveaux_clients_mois,
                    'valeur_formatee': str(nouveaux_clients_mois),
                    'tendance': difference_nouveaux,
                    'unite': 'clients',
                    'label': 'Nouveaux Clients',
                    'sub_value': f"ce mois"
                },
                'clients_actifs': {
                    'valeur': clients_actifs_30j,
                    'valeur_formatee': f"{clients_actifs_30j:,}",
                    'tendance': difference_actifs,
                    'unite': 'clients',
                    'label': 'Clients Actifs',
                    'sub_value': f"{pourcentage_actifs:.1f}% du total clients"
                },
                'taux_retour': {
                    'valeur': round(taux_retour, 1),
                    'valeur_formatee': f"{taux_retour:.1f}",
                    'tendance': round(tendance_retour, 1),
                    'unite': '%',
                    'label': 'Taux Retour',
                    'sub_value': f"{commandes_retournees}/{commandes_livrees_30j} retournées",
                    'status': 'good' if taux_retour < 5 else 'warning' if taux_retour < 10 else 'critical'
                },                'satisfaction': {
                    'valeur': round(taux_fidelisation, 1),
                    'valeur_formatee': f"{taux_fidelisation:.1f}",
                    'tendance': round(tendance_fidelisation, 1),
                    'unite': '%',
                    'label': 'Fidélisation',
                    'sub_value': f"{clients_fideles_90j}/{clients_actifs_90j} clients fidèles (90j)",
                    'status': 'excellent' if taux_fidelisation >= 25 else 'good' if taux_fidelisation >= 15 else 'warning' if taux_fidelisation >= 8 else 'critical'
                },
                'fidelisation': {
                    'valeur': round(taux_fidelisation, 1),
                    'valeur_formatee': f"{taux_fidelisation:.1f}",
                    'tendance': round(tendance_fidelisation, 1),
                    'unite': '%',
                    'label': 'Fidélisation',
                    'sub_value': f"{clients_fideles_90j}/{clients_actifs_90j} clients fidèles (90j)"
                }
            },
            'analyses_detaillees': {
                'top_clients_vip': top_clients_data,
                'performance_mensuelle': {
                    'commandes_mois': commandes_mois_actuel,
                    'ca_par_client': round(ca_moyen_par_client, 2)
                },
                'segmentation': {
                    'acheteurs_reguliers': round((clients_reguliers / total_clients_analyse * 100), 1) if total_clients_analyse > 0 else 0,
                    'nouveaux_testeurs': round((clients_nouveaux_testeurs / total_clients_analyse * 100), 1) if total_clients_analyse > 0 else 0,
                    'clients_occasionnels': round((clients_occasionnels / total_clients_analyse * 100), 1) if total_clients_analyse > 0 else 0,
                    'vip_premium': round((clients_vip / total_clients_analyse * 100), 1) if total_clients_analyse > 0 else 0
                }
            },
            'stats_globales': {
                'total_clients': total_clients,
                'clients_actifs_30j': clients_actifs_30j,
                'taux_activite': round((clients_actifs_30j / total_clients * 100), 1) if total_clients > 0 else 0,
                'commandes_total_30j': commandes_livrees_30j,
                'panier_moyen_clients': round(ca_moyen_par_client, 2)  # Utiliser le CA moyen calculé correctement
            },
            'empty': total_clients == 0
        }
        
        return JsonResponse(data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'message': 'Erreur lors du chargement des données Clients',
            'empty': True
        }, status=500)

@login_required
def vue_quantitative_data(request):
    """API pour les données de l'onglet État des commandes"""
    try:
        # Obtenir la date actuelle et la période demandée
        aujourd_hui = timezone.now().date()
        period = request.GET.get('period', 'aujourd_hui')  # Par défaut aujourd'hui
        
        # Calculer la date de début selon la période
        if period == 'aujourd_hui':
            date_debut = aujourd_hui
            date_fin = aujourd_hui
            jours = 1
        elif period == 'ce_mois':
            # Premier jour du mois actuel jusqu'à aujourd'hui
            date_debut = aujourd_hui.replace(day=1)
            date_fin = aujourd_hui
            jours = (date_fin - date_debut).days + 1
        elif period == 'cette_annee':
            # Premier jour de l'année actuelle jusqu'à aujourd'hui
            date_debut = aujourd_hui.replace(month=1, day=1)
            date_fin = aujourd_hui
            jours = (date_fin - date_debut).days + 1
        elif period.startswith('custom:'):
            # Période personnalisée format: custom:YYYY-MM-DD:YYYY-MM-DD
            try:
                parts = period.split(':')
                if len(parts) == 3:
                    from datetime import datetime
                    date_debut = datetime.strptime(parts[1], '%Y-%m-%d').date()
                    date_fin = datetime.strptime(parts[2], '%Y-%m-%d').date()
                    jours = (date_fin - date_debut).days + 1
                else:
                    raise ValueError("Format de période personnalisée invalide")
            except (ValueError, IndexError) as e:
                return JsonResponse({
                    'success': False,
                    'message': f'Format de période personnalisée invalide: {str(e)}',
                    'error': 'invalid_custom_period'
                }, status=400)
        else:
            # Fallback vers l'ancien système pour compatibilité
            period_days = {
                '7j': 7,
                '30j': 30,
                '90j': 90,
                '180j': 180,
                '365j': 365
            }
            
            jours = period_days.get(period, 30)  # Défaut à 30 jours si période inconnue
            date_debut = aujourd_hui - timedelta(days=jours)
            date_fin = aujourd_hui
        
        # Récupérer toutes les commandes dans la période avec leur état actuel (dernier état non terminé)
        # On utilise une sous-requête pour obtenir l'état le plus récent pour chaque commande
        commandes_avec_etat = Commande.objects.filter(
            etats__date_debut__date__gte=date_debut,  # Filtrer sur la date de l'état plutôt que la date de commande
            etats__date_debut__date__lte=date_fin,
            etats__date_fin__isnull=True  # États actuels (non terminés)
        ).select_related('etats__enum_etat').values(
            'id',
            'num_cmd',
            'etats__enum_etat__libelle'
        )
        
        # Initialiser les compteurs pour tous les états requis
        etats_compteurs = {
            'recue': 0,
            'affectee': 0, 
            'non_affectee': 0,
            'erronnee': 0,
            'doublon': 0,
            'en_cours_confirmation': 0,
            'confirmee': 0,
            'en_cours_preparation': 0,
            'preparee': 0,  # État intermédiaire à identifier
            'en_cours_livraison': 0,
            'livree': 0,
            'retournee': 0
        }
        
        # Mapping des libellés de la base vers nos clés standardisées
        mapping_etats = {
            'Non affectée': 'non_affectee',
            'Affectée': 'affectee',
            'En cours de confirmation': 'en_cours_confirmation',
            'Confirmée': 'confirmee',
            'Erronée': 'erronnee',
            'Doublon': 'doublon',
            'En préparation': 'en_cours_preparation',
            'En livraison': 'en_cours_livraison',
            'Livrée': 'livree',
            'Retournée': 'retournee',
            'Reçue': 'recue',  # Si cet état existe
            'Préparée': 'preparee'  # Si cet état existe
        }
        
        # Compter les commandes par état
        for commande in commandes_avec_etat:
            libelle_etat = commande['etats__enum_etat__libelle']
            if libelle_etat in mapping_etats:
                cle_etat = mapping_etats[libelle_etat]
                etats_compteurs[cle_etat] += 1
        
        # Pour les commandes sans état défini dans la période, les considérer comme "reçues"
        commandes_sans_etat = Commande.objects.filter(
            etats__isnull=True
        ).count()
        
        etats_compteurs['recue'] += commandes_sans_etat
        
        # Calculer le total des commandes
        total_commandes = sum(etats_compteurs.values())
        
        # Note: jours est déjà calculé plus haut dans la fonction
        
        # Ajouter des statistiques supplémentaires
        stats_supplementaires = {
            'commandes_en_cours': (
                etats_compteurs['affectee'] + 
                etats_compteurs['en_cours_confirmation'] + 
                etats_compteurs['en_cours_preparation'] + 
                etats_compteurs['en_cours_livraison']
            ),
            'commandes_problematiques': (
                etats_compteurs['erronnee'] + 
                etats_compteurs['doublon'] + 
                etats_compteurs['retournee']
            ),
            'commandes_completees': etats_compteurs['livree']
        }
        
        # Données de réponse
        response_data = {
            'success': True,
            'data': {
                'etats_commandes': etats_compteurs,
                'stats_supplementaires': stats_supplementaires,
                'total_commandes': total_commandes,
                'periode': {
                    'libelle': period,
                    'jours': jours,
                    'date_debut': date_debut.isoformat(),
                    'date_fin': date_fin.isoformat()
                },
                'derniere_maj': timezone.now().isoformat()
            }
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du calcul des états de commandes : {str(e)}',
            'error': 'erreur_calcul_etats'
        }, status=500)

@login_required
def performance_operateurs_data(request):
    """
    API pour les données de l'onglet Performance Opérateurs
    Version optimisée pour réduire les requêtes.
    """
    logger.info("Début du chargement des données de performance des opérateurs.")
    try:
        export_format = request.GET.get('export')
        logger.debug(f"Format d'export demandé : {export_format}")
        
        # Date de référence pour les calculs sur 30 jours
        date_limite_30j = timezone.now() - timedelta(days=30)

        # 1. Requête principale pour agréger toutes les métriques par opérateur
        operateurs_stats = Operateur.objects.filter(type_operateur='CONFIRMATION').annotate(
            # --- Métriques sur l'état ACTUEL ---
            commands_affected=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Affectée', etats_modifies__date_fin__isnull=True),
                distinct=True
            ),
            commands_in_progress=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='En cours de confirmation', etats_modifies__date_fin__isnull=True),
                distinct=True
            ),

            # --- Métriques HISTORIQUES (tous les temps) ---
            commands_confirmed=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée'),
                distinct=True
            ),

            # --- Métriques FINANCIERES sur commandes confirmées ---
            panier_moyen=Avg(
                'etats_modifies__commande__total_cmd',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée')
            ),
            panier_min=Min(
                'etats_modifies__commande__total_cmd',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée')
            ),
            panier_max=Max(
                'etats_modifies__commande__total_cmd',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée')
            ),
            upsell_count=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée', etats_modifies__commande__is_upsell=True),
                distinct=True
            ),
            upsell_amount=Sum(
                'etats_modifies__commande__total_cmd',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée', etats_modifies__commande__is_upsell=True)
            ),
            
            # --- Métriques sur les OPERATIONS (30 derniers jours) ---
            total_actions_30j=Count(
                'operations', # Le related_name sur Operation est 'operations'
                filter=Q(operations__date_operation__gte=date_limite_30j),
                distinct=True
            ),
            commands_confirmed_30j=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée', etats_modifies__date_debut__gte=date_limite_30j),
                distinct=True
            ),

        ).values(
            'id', 'nom', 'user__username',
            'commands_affected', 'commands_in_progress', 'commands_confirmed',
            'panier_moyen', 'panier_min', 'panier_max',
            'upsell_count', 'upsell_amount',
            'total_actions_30j', 'commands_confirmed_30j'
        )

        operateurs_data = []
        total_commandes_affectees_global = 0
        total_commandes_confirmees_global = 0
        
        for op_stat in operateurs_stats:
            # Calculs post-requête en Python
            total_commandes_traitees = op_stat['commands_affected'] + op_stat['commands_in_progress'] + op_stat['commands_confirmed']
            taux_confirmation = (op_stat['commands_confirmed'] / total_commandes_traitees * 100) if total_commandes_traitees > 0 else 0
            
            operations_par_commande_30j = (op_stat['total_actions_30j'] / op_stat['commands_confirmed_30j']) if op_stat['commands_confirmed_30j'] > 0 else 0

            operateurs_data.append({
                'id': op_stat['id'],
                'nom': op_stat['nom'],
                'username': op_stat.get('user__username', 'N/A'),
                'commands_affected': op_stat['commands_affected'],
                'commands_in_progress': op_stat['commands_in_progress'],
                'commands_confirmed': op_stat['commands_confirmed'],
                'confirmation_rate': round(taux_confirmation, 1),
                'average_basket': float(op_stat['panier_moyen'] or 0),
                'min_basket': float(op_stat['panier_min'] or 0),
                'max_basket': float(op_stat['panier_max'] or 0),
                'upsell_count': op_stat['upsell_count'],
                'upsell_amount': float(op_stat['upsell_amount'] or 0),
                'total_actions': op_stat['total_actions_30j'], # Simplifié aux 30j
                'operations_per_command_30d': round(operations_par_commande_30j, 1),
                # Les métriques de temps réel sont gérées par une autre API
                'avg_confirmation_time_minutes': 0, 
                'avg_arrival_to_confirmation_minutes': 0,
            })
            total_commandes_affectees_global += op_stat['commands_affected'] + op_stat['commands_in_progress']
            total_commandes_confirmees_global += op_stat['commands_confirmed']

        # Calculer le taux de confirmation global
        total_traitees_global = total_commandes_affectees_global + total_commandes_confirmees_global
        taux_confirmation_global = (total_commandes_confirmees_global / total_traitees_global * 100) if total_traitees_global > 0 else 0

        global_metrics = {
            'commands_assigned': total_commandes_affectees_global,
            'confirmations': total_commandes_confirmees_global,
            'global_confirmation_rate': round(taux_confirmation_global, 1),
            'active_operators': operateurs_stats.count()
        }
        
        operateurs_data.sort(key=lambda x: x['commands_confirmed'], reverse=True)
        
        if export_format == 'excel':
            logger.info("Export Excel demandé.")
            # Pour l'export Excel, on ne retourne pas de JSON, on laisse la fonction d'export gérer
            pass

        logger.info(f"Envoi de la réponse JSON pour {len(operateurs_data)} opérateurs.")
        return JsonResponse({
            'success': True,
            'operators': operateurs_data,
            'global_metrics': global_metrics
        })

    except Exception as e:
        logger.error("Erreur dans performance_operateurs_data: %s", str(e), exc_info=True)
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'message': f'Erreur serveur optimisée : {str(e)}',
            'error': 'server_error'
        }, status=500)
        
@login_required
def export_performance_operateurs_excel(request):
    """
    Exporte les données de performance des opérateurs en Excel.
    """
    try:
        logger.info("Début de l'export Excel des performances opérateurs")
        
        # Récupérer l'ID de l'opérateur si spécifié
        operator_id = request.GET.get('operator_id')
        
        # Date de référence pour les calculs sur 30 jours
        date_limite_30j = timezone.now() - timedelta(days=30)

        # 1. Requête principale pour agréger toutes les métriques par opérateur
        operateurs_query = Operateur.objects.filter(type_operateur='CONFIRMATION')
        
        # Si un opérateur spécifique est demandé, filtrer
        if operator_id:
            operateurs_query = operateurs_query.filter(id=operator_id)
            logger.info(f"Export spécifique pour l'opérateur ID: {operator_id}")
        
        operateurs_stats = operateurs_query.annotate(
            # --- Métriques sur l'état ACTUEL ---
            commands_affected=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Affectée', etats_modifies__date_fin__isnull=True),
                distinct=True
            ),
            commands_in_progress=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='En cours de confirmation', etats_modifies__date_fin__isnull=True),
                distinct=True
            ),

            # --- Métriques HISTORIQUES (tous les temps) ---
            commands_confirmed=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée'),
                distinct=True
            ),

            # --- Métriques FINANCIERES sur commandes confirmées ---
            panier_moyen=Avg(
                'etats_modifies__commande__total_cmd',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée')
            ),
            panier_min=Min(
                'etats_modifies__commande__total_cmd',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée')
            ),
            panier_max=Max(
                'etats_modifies__commande__total_cmd',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée')
            ),
            upsell_count=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée', etats_modifies__commande__is_upsell=True),
                distinct=True
            ),
            upsell_amount=Sum(
                'etats_modifies__commande__total_cmd',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée', etats_modifies__commande__is_upsell=True)
            ),
            
            # --- Métriques sur les OPERATIONS (30 derniers jours) ---
            total_actions_30j=Count(
                'operations', # Le related_name sur Operation est 'operations'
                filter=Q(operations__date_operation__gte=date_limite_30j),
                distinct=True
            ),
            commands_confirmed_30j=Count(
                'etats_modifies__commande',
                filter=Q(etats_modifies__enum_etat__libelle__iexact='Confirmée', etats_modifies__date_debut__gte=date_limite_30j),
                distinct=True
            ),

        ).values(
            'id', 'nom', 'user__username',
            'commands_affected', 'commands_in_progress', 'commands_confirmed',
            'panier_moyen', 'panier_min', 'panier_max',
            'upsell_count', 'upsell_amount',
            'total_actions_30j', 'commands_confirmed_30j'
        )

        operateurs_data = []
        total_commandes_affectees_global = 0
        total_commandes_confirmees_global = 0
        
        for op_stat in operateurs_stats:
            # Calculs post-requête en Python
            total_commandes_traitees = op_stat['commands_affected'] + op_stat['commands_in_progress'] + op_stat['commands_confirmed']
            taux_confirmation = (op_stat['commands_confirmed'] / total_commandes_traitees * 100) if total_commandes_traitees > 0 else 0
            
            operations_par_commande_30j = (op_stat['total_actions_30j'] / op_stat['commands_confirmed_30j']) if op_stat['commands_confirmed_30j'] > 0 else 0

            operateurs_data.append({
                'id': op_stat['id'],
                'nom': op_stat['nom'],
                'username': op_stat.get('user__username', 'N/A'),
                'commands_affected': op_stat['commands_affected'],
                'commands_in_progress': op_stat['commands_in_progress'],
                'commands_confirmed': op_stat['commands_confirmed'],
                'confirmation_rate': round(taux_confirmation, 1),
                'average_basket': float(op_stat['panier_moyen'] or 0),
                'min_basket': float(op_stat['panier_min'] or 0),
                'max_basket': float(op_stat['panier_max'] or 0),
                'upsell_count': op_stat['upsell_count'],
                'upsell_amount': float(op_stat['upsell_amount'] or 0),
                'total_actions': op_stat['total_actions_30j'], # Simplifié aux 30j
                'operations_per_command_30d': round(operations_par_commande_30j, 1),
                # Les métriques de temps réel sont gérées par une autre API
                'avg_confirmation_time_minutes': 0, 
                'avg_arrival_to_confirmation_minutes': 0,
            })
            total_commandes_affectees_global += op_stat['commands_affected'] + op_stat['commands_in_progress']
            total_commandes_confirmees_global += op_stat['commands_confirmed']

        # Calculer le taux de confirmation global
        total_traitees_global = total_commandes_affectees_global + total_commandes_confirmees_global
        taux_confirmation_global = (total_commandes_confirmees_global / total_traitees_global * 100) if total_traitees_global > 0 else 0

        global_metrics = {
            'commands_assigned': total_commandes_affectees_global,
            'confirmations': total_commandes_confirmees_global,
            'global_confirmation_rate': round(taux_confirmation_global, 1),
            'active_operators': operateurs_stats.count()
        }
        
        operateurs_data.sort(key=lambda x: x['commands_confirmed'], reverse=True)
        
        logger.info(f"Données récupérées: {len(operateurs_data)} opérateurs")

        # Vérifier que openpyxl est disponible
        if not openpyxl:
            return HttpResponse("Module openpyxl non disponible. Installez-le avec: pip install openpyxl", status=500)

        # Créer un nouveau classeur Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        
        if operator_id and len(operateurs_data) == 1:
            # Si c'est un export spécifique à un opérateur
            operateur = operateurs_data[0]
            ws.title = f"Performance {operateur['username']}"
        else:
            # Export global de tous les opérateurs
            ws.title = "Performance Opérateurs"

        # Styles pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Styles pour les données
        data_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-têtes des colonnes
        headers = [
            'Opérateur',
            'Nom d\'utilisateur',
            'Commandes Affectées',
            'Commandes En Cours',
            'Commandes Confirmées',
            'Taux de Confirmation (%)',
            'Panier Moyen (MAD)',
            'Panier Min (MAD)',
            'Panier Max (MAD)',
            'Upsells (Nombre)',
            'Montant Upsells (MAD)',
            'Actions 30j',
            'Opérations/Commande 30j'
        ]

        # Écrire les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Écrire les données des opérateurs
        for row, operateur in enumerate(operateurs_data, 2):
            ws.cell(row=row, column=1, value=operateur['nom']).border = border
            ws.cell(row=row, column=2, value=operateur['username']).border = border
            ws.cell(row=row, column=3, value=operateur['commands_affected']).border = border
            ws.cell(row=row, column=4, value=operateur['commands_in_progress']).border = border
            ws.cell(row=row, column=5, value=operateur['commands_confirmed']).border = border
            ws.cell(row=row, column=6, value=operateur['confirmation_rate']).border = border
            ws.cell(row=row, column=7, value=operateur['average_basket']).border = border
            ws.cell(row=row, column=8, value=operateur['min_basket']).border = border
            ws.cell(row=row, column=9, value=operateur['max_basket']).border = border
            ws.cell(row=row, column=10, value=operateur['upsell_count']).border = border
            ws.cell(row=row, column=11, value=operateur['upsell_amount']).border = border
            ws.cell(row=row, column=12, value=operateur['total_actions']).border = border
            ws.cell(row=row, column=13, value=operateur['operations_per_command_30d']).border = border

        # Ajouter une ligne de métriques globales
        row_global = len(operateurs_data) + 3
        ws.cell(row=row_global, column=1, value="MÉTRIQUES GLOBALES").font = Font(bold=True)
        ws.cell(row=row_global + 1, column=1, value="Commandes Assignées").border = border
        ws.cell(row=row_global + 1, column=2, value=global_metrics.get('commands_assigned', 0)).border = border
        ws.cell(row=row_global + 2, column=1, value="Confirmations").border = border
        ws.cell(row=row_global + 2, column=2, value=global_metrics.get('confirmations', 0)).border = border
        ws.cell(row=row_global + 3, column=1, value="Taux de Confirmation Global (%)").border = border
        ws.cell(row=row_global + 3, column=2, value=global_metrics.get('global_confirmation_rate', 0)).border = border
        ws.cell(row=row_global + 4, column=1, value="Opérateurs Actifs").border = border
        ws.cell(row=row_global + 4, column=2, value=global_metrics.get('active_operators', 0)).border = border

        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Créer la réponse HTTP
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Générer le nom de fichier avec la date
        from datetime import datetime
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if operator_id and len(operateurs_data) == 1:
            # Si c'est un export spécifique à un opérateur
            operateur = operateurs_data[0]
            filename = f"performance_operateur_{operateur['username']}_{date_str}.xlsx"
        else:
            # Export global de tous les opérateurs
            filename = f"performance_operateurs_{date_str}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Export Excel terminé avec succès: {filename}")
        return response

    except Exception as e:
        logger.error(f"Erreur lors de l'export Excel: {str(e)}", exc_info=True)
        return HttpResponse(f"Erreur lors de l'export Excel: {str(e)}", status=500)

@api_login_required
def operator_history_data(request):
    """API pour récupérer l'historique récent d'un opérateur"""
    try:
        # Paramètres de la requête
        operator_id = request.GET.get('operator_id')
        period = request.GET.get('period', 'today')
        limit = int(request.GET.get('limit', 5))
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not operator_id:
            return JsonResponse({
                'success': False,
                'message': 'ID opérateur manquant',
                'error': 'missing_operator_id'
            }, status=400)
        
        # Vérifier que l'opérateur existe
        try:
            operateur = Operateur.objects.get(id=operator_id)
        except Operateur.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Opérateur introuvable',
                'error': 'operator_not_found'
            }, status=404)
        
        # Calculer les dates selon la période (même logique que performance_operateurs_data)
        aujourd_hui = timezone.now().date()
        
        if period == 'custom' and start_date and end_date:
            try:
                date_debut = datetime.strptime(start_date, '%Y-%m-%d').date()
                date_fin = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': 'Format de date invalide',
                    'error': 'invalid_date_format'
                }, status=400)
        else:
            # Périodes prédéfinies
            if period == 'today':
                date_debut = aujourd_hui
                date_fin = aujourd_hui
            elif period == 'yesterday':
                hier = aujourd_hui - timedelta(days=1)
                date_debut = hier
                date_fin = hier
            elif period == 'week':
                date_debut = aujourd_hui - timedelta(days=7)
                date_fin = aujourd_hui
            elif period == 'month':
                date_debut = aujourd_hui.replace(day=1)
                date_fin = aujourd_hui
            elif period == 'quarter':
                # Début du trimestre actuel
                mois_actuel = aujourd_hui.month
                if mois_actuel <= 3:
                    debut_trimestre = aujourd_hui.replace(month=1, day=1)
                elif mois_actuel <= 6:
                    debut_trimestre = aujourd_hui.replace(month=4, day=1)
                elif mois_actuel <= 9:
                    debut_trimestre = aujourd_hui.replace(month=7, day=1)
                else:
                    debut_trimestre = aujourd_hui.replace(month=10, day=1)
                date_debut = debut_trimestre
                date_fin = aujourd_hui
            elif period == 'year':
                date_debut = aujourd_hui.replace(month=1, day=1)
                date_fin = aujourd_hui
            else:
                # Défaut aujourd'hui
                date_debut = aujourd_hui
                date_fin = aujourd_hui
        
        # Récupérer les opérations récentes de l'opérateur
        operations = Operation.objects.filter(
            operateur=operateur,
            date_operation__date__gte=date_debut,
            date_operation__date__lte=date_fin
        ).select_related('commande').order_by('-date_operation')[:limit]
        
        # Formater les données
        history_data = []
        for operation in operations:
            history_data.append({
                'id': operation.id,
                'type_operation': operation.get_type_operation_display(),
                'commande_num': operation.commande.num_cmd,
                'date_operation': operation.date_operation.isoformat(),
                'conclusion': operation.conclusion[:100] + '...' if len(operation.conclusion) > 100 else operation.conclusion,
                'status': 'Terminé'  # Peut être étendu selon la logique métier
            })
        
        return JsonResponse({
            'success': True,
            'history': history_data,
            'operator': {
                'id': operateur.id,
                'nom': operateur.nom,
                'type': operateur.type_operateur
            },
            'period_info': {
                'start_date': date_debut.isoformat(),
                'end_date': date_fin.isoformat(),
                'period': period
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la récupération de l\'historique : {str(e)}',
            'error': 'erreur_historique'
        }, status=500)

@login_required
def dashboard(request):
    """Page principale du dashboard KPIs"""
    selected_period = request.GET.get('period', 'aujourd_hui')  # Valeur par défaut
    return render(request, 'kpis/dashboard.html', {
        'selected_period': selected_period
    })

@api_login_required
def operator_realtime_times_data(request):
    """
    API pour les métriques de temps en temps réel des opérateurs de confirmation.
    Version optimisée pour éviter les requêtes N+1.
    """
    logger.info("Début du calcul des temps réels pour les opérateurs (version optimisée).")
    try:
        # 1. Récupérer tous les états pertinents pour les opérateurs de confirmation en une seule requête.
        etats_pertinents = EtatCommande.objects.filter(
            operateur__type_operateur='CONFIRMATION',
            enum_etat__libelle__in=['En cours de confirmation', 'Confirmée']
        ).select_related('operateur', 'commande', 'enum_etat', 'operateur__user').order_by('commande_id', 'date_debut')

        # Dictionnaires pour stocker les données intermédiaires
        commandes_data = {}  # {commande_id: {'en_cours': datetime, 'confirmee': datetime, 'operateur': operateur}}
        
        for etat in etats_pertinents:
            cmd_id = etat.commande_id
            if cmd_id not in commandes_data:
                commandes_data[cmd_id] = {'operateur': etat.operateur, 'date_arrivee': etat.commande.last_sync_date or etat.commande.date_creation}

            if etat.enum_etat.libelle == 'En cours de confirmation':
                commandes_data[cmd_id]['en_cours'] = etat.date_debut
            elif etat.enum_etat.libelle == 'Confirmée':
                commandes_data[cmd_id]['confirmee'] = etat.date_debut
                # On s'assure de garder l'opérateur qui a confirmé
                commandes_data[cmd_id]['operateur'] = etat.operateur

        # Dictionnaires pour agréger les temps par opérateur
        operateurs_temps = {}  # {operateur_id: {'temps_conf_total': timedelta, 'nb_conf': int, ...}}

        aujourd_hui = timezone.now().date()

        for cmd_id, data in commandes_data.items():
            if 'en_cours' in data and 'confirmee' in data:
                op = data['operateur']
                if op.id not in operateurs_temps:
                    operateurs_temps[op.id] = {
                        'operateur': op,
                        'temps_confirmation_total': timedelta(), 'nb_conf': 0,
                        'temps_arrivee_total': timedelta(), 'nb_arrivee': 0,
                        'temps_conf_jour_total': timedelta(), 'nb_conf_jour': 0,
                        'nb_commandes_confirmees_aujourd_hui': 0
                    }
                
                # Temps de confirmation (en_cours -> confirmee)
                temps_confirmation = data['confirmee'] - data['en_cours']
                if temps_confirmation > timedelta(0):
                    operateurs_temps[op.id]['temps_confirmation_total'] += temps_confirmation
                    operateurs_temps[op.id]['nb_conf'] += 1

                # Temps d'arrivée (creation/sync -> confirmee)
                if data.get('date_arrivee'):
                    temps_arrivee = data['confirmee'] - data['date_arrivee']
                    if temps_arrivee > timedelta(0):
                        operateurs_temps[op.id]['temps_arrivee_total'] += temps_arrivee
                        operateurs_temps[op.id]['nb_arrivee'] += 1

                # Métriques pour aujourd'hui
                if data['confirmee'].date() == aujourd_hui:
                    operateurs_temps[op.id]['nb_commandes_confirmees_aujourd_hui'] += 1
                    if temps_confirmation > timedelta(0):
                        operateurs_temps[op.id]['temps_conf_jour_total'] += temps_confirmation
                        operateurs_temps[op.id]['nb_conf_jour'] += 1
        
        realtime_data = []
        # Récupérer tous les opérateurs pour inclure ceux sans activité
        all_operators = Operateur.objects.filter(type_operateur='CONFIRMATION').select_related('user')

        for op in all_operators:
            data = operateurs_temps.get(op.id)
            if data:
                temps_moyen_conf = (data['temps_confirmation_total'].total_seconds() / 60 / data['nb_conf']) if data['nb_conf'] > 0 else 0
                temps_moyen_arrivee = (data['temps_arrivee_total'].total_seconds() / 60 / data['nb_arrivee']) if data['nb_arrivee'] > 0 else 0
                temps_moyen_conf_jour = (data['temps_conf_jour_total'].total_seconds() / 60 / data['nb_conf_jour']) if data['nb_conf_jour'] > 0 else 0

                realtime_data.append({
                    'operateur_id': op.id,
                    'operateur_nom': op.nom,
                    'operateur_username': op.user.username if op.user else 'N/A',
                    'temps_confirmation_global_minutes': round(temps_moyen_conf, 1),
                    'temps_arrivee_confirmation_global_minutes': round(temps_moyen_arrivee, 1),
                    'nb_commandes_confirmees_total': data['nb_conf'],
                    'temps_confirmation_aujourd_hui_minutes': round(temps_moyen_conf_jour, 1),
                    'nb_commandes_confirmees_aujourd_hui': data['nb_commandes_confirmees_aujourd_hui'],
                    'last_update': timezone.now().isoformat()
                })
            else:
                # Opérateur sans activité
                realtime_data.append({
                    'operateur_id': op.id,
                    'operateur_nom': op.nom,
                    'operateur_username': op.user.username if op.user else 'N/A',
                    'temps_confirmation_global_minutes': 0,
                    'temps_arrivee_confirmation_global_minutes': 0,
                    'nb_commandes_confirmees_total': 0,
                    'temps_confirmation_aujourd_hui_minutes': 0,
                    'nb_commandes_confirmees_aujourd_hui': 0,
                    'last_update': timezone.now().isoformat()
                })

        logger.info(f"Calcul des temps réels terminé pour {len(realtime_data)} opérateurs.")
        return JsonResponse({
            'success': True,
            'realtime_data': realtime_data,
            'timestamp': timezone.now().isoformat()
        })
        
    except Exception as e:
        logger.error("Erreur dans operator_realtime_times_data: %s", str(e), exc_info=True)
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du calcul des temps en temps réel : {str(e)}',
            'error': 'erreur_calcul_temps_realtime'
        }, status=500)

@login_required
def export_performance_operateurs_csv(request):
    """Export CSV des performances des opérateurs de confirmation"""
    from parametre.models import Operateur
    from commande.models import Commande
    import datetime
    
    # Récupérer la période depuis les paramètres de requête
    period = request.GET.get('period', 'aujourd_hui')
    
    # Déterminer les dates de début et fin en fonction de la période
    today = timezone.now().date()
    if period == 'aujourd_hui':
        date_debut = today
        date_fin = today
    elif period == 'ce_mois':
        date_debut = today.replace(day=1)
        date_fin = today
    elif period == 'cette_annee':
        date_debut = today.replace(month=1, day=1)
        date_fin = today
    else:  # période personnalisée
        try:
            date_debut = datetime.datetime.strptime(request.GET.get('date_debut'), '%Y-%m-%d').date()
            date_fin = datetime.datetime.strptime(request.GET.get('date_fin'), '%Y-%m-%d').date()
        except (TypeError, ValueError):
            date_debut = today
            date_fin = today
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="performance_operateurs.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Opérateur', 'Commandes confirmées', 'Temps moyen confirmation (min)', 'Temps moyen arrivée→confirmation (min)'
    ])
    operateurs = Operateur.objects.filter(type_operateur='CONFIRMATION')
    for operateur in operateurs:
        commandes_confirmees = Commande.objects.filter(
            etats__operateur=operateur,
            etats__enum_etat__libelle__iexact='Confirmée',
            etats__date_debut__date__gte=date_debut,
            etats__date_debut__date__lte=date_fin
        ).distinct()
        total_confirmation = datetime.timedelta()
        total_arrivee = datetime.timedelta()
        n_conf = 0
        n_arr = 0
        for cmd in commandes_confirmees:
            etat_conf = cmd.etats.filter(enum_etat__libelle__iexact='Confirmée', operateur=operateur).order_by('date_debut').first()
            etat_en_cours = cmd.etats.filter(enum_etat__libelle__iexact='En cours de confirmation', operateur=operateur).order_by('date_debut').first()
            if etat_conf and etat_en_cours:
                delta = etat_conf.date_debut - etat_en_cours.date_debut
                total_confirmation += delta
                n_conf += 1
            if etat_conf:
                arrivee = cmd.last_sync_date or cmd.date_creation
                delta = etat_conf.date_debut - arrivee
                total_arrivee += delta
                n_arr += 1
        avg_conf = (total_confirmation.total_seconds() / 60 / n_conf) if n_conf else 0
        avg_arr = (total_arrivee.total_seconds() / 60 / n_arr) if n_arr else 0
        writer.writerow([
            f"{operateur.prenom} {operateur.nom}",
            commandes_confirmees.count(),
            round(avg_conf, 2),
            round(avg_arr, 2)
        ])
    return response

@login_required
def export_etat_commandes_csv(request):
    """Export CSV du suivi de l'état des commandes"""
    try:
        # Récupérer la période depuis les paramètres de requête
        period = request.GET.get('period', 'aujourd_hui')
        
        # Déterminer les dates de début et fin en fonction de la période
        aujourd_hui = timezone.now().date()
        if period == 'aujourd_hui':
            date_debut = aujourd_hui
            date_fin = aujourd_hui
        elif period == 'ce_mois':
            date_debut = aujourd_hui.replace(day=1)
            date_fin = aujourd_hui
        elif period == 'cette_annee':
            date_debut = aujourd_hui.replace(month=1, day=1)
            date_fin = aujourd_hui
        else:  # période personnalisée
            try:
                from datetime import datetime
                date_debut = datetime.strptime(request.GET.get('date_debut'), '%Y-%m-%d').date()
                date_fin = datetime.strptime(request.GET.get('date_fin'), '%Y-%m-%d').date()
            except (TypeError, ValueError):
                date_debut = aujourd_hui
                date_fin = aujourd_hui
        
        # Récupérer les données via la fonction existante
        from django.test import RequestFactory
        factory = RequestFactory()
        test_request = factory.get(f'/kpis/api/vue-quantitative/?period={period}')
        test_request.user = request.user
        
        json_response = vue_quantitative_data(test_request)
        if json_response.status_code != 200:
            return HttpResponse("Erreur lors de la génération des données.", status=500)
        
        import json
        content = json.loads(json_response.content)
        etats_data = content.get('etats', {})
        
        # Créer la réponse CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="etat_commandes_{period}.csv"'
        writer = csv.writer(response)
        
        # En-têtes enrichies
        writer.writerow([
            'État', 
            'Nombre de commandes', 
            'Pourcentage', 
            'Valeur totale (MAD)', 
            'Panier moyen (MAD)',
            'Opérateur principal',
            'Dernière activité'
        ])
        
        # Récupérer des données enrichies
        etats_enrichis = {}
        
        # Liste complète de tous les états possibles
        tous_etats = [
            'Non affectée',
            'Affectée', 
            'En cours de confirmation',
            'Confirmée',
            'Erronée',
            'Doublon',
            'En préparation',
            'Préparée',
            'En livraison',
            'Livrée',
            'Retournée',
            'Reçue'
        ]
        
        # Traiter tous les états, même ceux avec 0 commandes
        for etat in tous_etats:
            nombre = etats_data.get(etat, 0)
            
            if nombre > 0:
                # Récupérer les commandes pour cet état
                commandes_etat = Commande.objects.filter(
                    etats__enum_etat__libelle__iexact=etat,
                    etats__date_debut__date__gte=date_debut,
                    etats__date_debut__date__lte=date_fin,
                    etats__date_fin__isnull=True
                ).select_related('etats__operateur')
                
                # Calculer les métriques
                valeur_totale = commandes_etat.aggregate(total=Sum('total_cmd'))['total'] or 0
                panier_moyen = commandes_etat.aggregate(moyen=Avg('total_cmd'))['moyen'] or 0
                
                # Opérateur principal
                operateur_principal = commandes_etat.values('etats__operateur__nom').annotate(
                    count=Count('id')
                ).order_by('-count').first()
                
                # Dernière activité
                derniere_activite = commandes_etat.aggregate(
                    derniere=Max('etats__date_debut')
                )['derniere']
                
                etats_enrichis[etat] = {
                    'nombre': nombre,
                    'valeur_totale': valeur_totale,
                    'panier_moyen': panier_moyen,
                    'operateur_principal': operateur_principal['etats__operateur__nom'] if operateur_principal else 'N/A',
                    'derniere_activite': derniere_activite.strftime('%d/%m/%Y %H:%M') if derniere_activite else 'N/A'
                }
            else:
                etats_enrichis[etat] = {
                    'nombre': 0,
                    'valeur_totale': 0,
                    'panier_moyen': 0,
                    'operateur_principal': 'N/A',
                    'derniere_activite': 'N/A'
                }
        
        # Données enrichies
        total_commandes = sum(etats_data.values())
        total_valeur = sum(data['valeur_totale'] for data in etats_enrichis.values())
        
        for etat, data in etats_enrichis.items():
            pourcentage = (data['nombre'] / total_commandes * 100) if total_commandes > 0 else 0
            writer.writerow([
                etat, 
                data['nombre'], 
                f"{pourcentage:.1f}%",
                format_number_fr(data['valeur_totale']),
                format_number_fr(data['panier_moyen']),
                data['operateur_principal'],
                data['derniere_activite']
            ])
        
        # Ligne de total
        total_panier_moyen = total_valeur / total_commandes if total_commandes > 0 else 0
        writer.writerow([
            'TOTAL', 
            total_commandes, 
            '100%',
            format_number_fr(total_valeur),
            format_number_fr(total_panier_moyen),
            '',
            ''
        ])
        
        return response
        
    except Exception as e:
        logger.error(f"Erreur lors de l'export CSV: {str(e)}", exc_info=True)
        return HttpResponse(f"Erreur lors de l'export CSV: {str(e)}", status=500)

@login_required
def export_etat_commandes_excel(request):
    """Export Excel du suivi de l'état des commandes"""
    try:
        # Récupérer la période depuis les paramètres de requête
        period = request.GET.get('period', 'aujourd_hui')
        
        # Déterminer les dates de début et fin en fonction de la période
        aujourd_hui = timezone.now().date()
        if period == 'aujourd_hui':
            date_debut = aujourd_hui
            date_fin = aujourd_hui
        elif period == 'ce_mois':
            date_debut = aujourd_hui.replace(day=1)
            date_fin = aujourd_hui
        elif period == 'cette_annee':
            date_debut = aujourd_hui.replace(month=1, day=1)
            date_fin = aujourd_hui
        else:  # période personnalisée
            try:
                from datetime import datetime
                date_debut = datetime.strptime(request.GET.get('date_debut'), '%Y-%m-%d').date()
                date_fin = datetime.strptime(request.GET.get('date_fin'), '%Y-%m-%d').date()
            except (TypeError, ValueError):
                date_debut = aujourd_hui
                date_fin = aujourd_hui
        
        # Récupérer les données via la fonction existante
        from django.test import RequestFactory
        factory = RequestFactory()
        test_request = factory.get(f'/kpis/api/vue-quantitative/?period={period}')
        test_request.user = request.user
        
        json_response = vue_quantitative_data(test_request)
        if json_response.status_code != 200:
            return HttpResponse("Erreur lors de la génération des données.", status=500)
        
        import json
        content = json.loads(json_response.content)
        etats_data = content.get('etats', {})
        
        # Vérifier que openpyxl est disponible
        if not openpyxl:
            return HttpResponse("Module openpyxl non disponible. Installez-le avec: pip install openpyxl", status=500)

        # Créer un nouveau classeur Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "État des Commandes"

        # Styles pour l'en-tête
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Styles pour les données
        data_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-têtes des colonnes enrichies
        headers = [
            'État', 
            'Nombre de commandes', 
            'Pourcentage', 
            'Valeur totale (MAD)', 
            'Panier moyen (MAD)',
            'Temps moyen traitement (min)',
            'Opérateur principal',
            'Dernière activité'
        ]
        
        # Écrire les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Calculer le total et récupérer des données enrichies
        total_commandes = sum(etats_data.values())
        
        # Récupérer des données supplémentaires pour chaque état
        etats_enrichis = {}
        
        # Liste complète de tous les états possibles
        tous_etats = [
            'Non affectée',
            'Affectée', 
            'En cours de confirmation',
            'Confirmée',
            'Erronée',
            'Doublon',
            'En préparation',
            'Préparée',
            'En livraison',
            'Livrée',
            'Retournée',
            'Reçue'
        ]
        
        # Traiter tous les états, même ceux avec 0 commandes
        for etat in tous_etats:
            nombre = etats_data.get(etat, 0)
            
            if nombre > 0:
                # Récupérer les commandes pour cet état
                commandes_etat = Commande.objects.filter(
                    etats__enum_etat__libelle__iexact=etat,
                    etats__date_debut__date__gte=date_debut,
                    etats__date_debut__date__lte=date_fin,
                    etats__date_fin__isnull=True
                ).select_related('etats__operateur')
                
                # Calculer les métriques
                valeur_totale = commandes_etat.aggregate(total=Sum('total_cmd'))['total'] or 0
                panier_moyen = commandes_etat.aggregate(moyen=Avg('total_cmd'))['moyen'] or 0
                
                # Opérateur principal (celui qui a traité le plus de commandes)
                operateur_principal = commandes_etat.values('etats__operateur__nom').annotate(
                    count=Count('id')
                ).order_by('-count').first()
                
                # Dernière activité
                derniere_activite = commandes_etat.aggregate(
                    derniere=Max('etats__date_debut')
                )['derniere']
                
                # Temps moyen de traitement (simplifié)
                temps_moyen = 0  # À calculer plus précisément si nécessaire
                
                etats_enrichis[etat] = {
                    'nombre': nombre,
                    'valeur_totale': valeur_totale,
                    'panier_moyen': panier_moyen,
                    'temps_moyen': temps_moyen,
                    'operateur_principal': operateur_principal['etats__operateur__nom'] if operateur_principal else 'N/A',
                    'derniere_activite': derniere_activite.strftime('%d/%m/%Y %H:%M') if derniere_activite else 'N/A'
                }
            else:
                etats_enrichis[etat] = {
                    'nombre': 0,
                    'valeur_totale': 0,
                    'panier_moyen': 0,
                    'temps_moyen': 0,
                    'operateur_principal': 'N/A',
                    'derniere_activite': 'N/A'
                }
        
        # Écrire les données enrichies
        row = 2
        for etat, data in etats_enrichis.items():
            pourcentage = (data['nombre'] / total_commandes * 100) if total_commandes > 0 else 0
            
            ws.cell(row=row, column=1, value=etat).border = border
            ws.cell(row=row, column=2, value=data['nombre']).border = border
            ws.cell(row=row, column=3, value=f"{pourcentage:.1f}%").border = border
            ws.cell(row=row, column=4, value=format_number_fr(data['valeur_totale'])).border = border
            ws.cell(row=row, column=5, value=format_number_fr(data['panier_moyen'])).border = border
            ws.cell(row=row, column=6, value=f"{data['temps_moyen']:.1f}").border = border
            ws.cell(row=row, column=7, value=data['operateur_principal']).border = border
            ws.cell(row=row, column=8, value=data['derniere_activite']).border = border
            row += 1

        # Ligne de total
        total_valeur = sum(data['valeur_totale'] for data in etats_enrichis.values())
        total_panier_moyen = total_valeur / total_commandes if total_commandes > 0 else 0
        
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value=total_commandes).font = Font(bold=True)
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=3, value="100%").font = Font(bold=True)
        ws.cell(row=row, column=3).border = border
        ws.cell(row=row, column=4, value=format_number_fr(total_valeur)).font = Font(bold=True)
        ws.cell(row=row, column=4).border = border
        ws.cell(row=row, column=5, value=format_number_fr(total_panier_moyen)).font = Font(bold=True)
        ws.cell(row=row, column=5).border = border
        ws.cell(row=row, column=6, value="").font = Font(bold=True)
        ws.cell(row=row, column=6).border = border
        ws.cell(row=row, column=7, value="").font = Font(bold=True)
        ws.cell(row=row, column=7).border = border
        ws.cell(row=row, column=8, value="").font = Font(bold=True)
        ws.cell(row=row, column=8).border = border

        # Ajouter des métriques globales
        metrics_row = row + 2
        ws.cell(row=metrics_row, column=1, value="MÉTRIQUES GLOBALES").font = Font(bold=True, size=14)
        
        # Calculer des métriques supplémentaires
        commandes_en_cours = sum(data['nombre'] for etat, data in etats_enrichis.items() 
                               if 'en_cours' in etat.lower() or 'affectée' in etat)
        commandes_completees = sum(data['nombre'] for etat, data in etats_enrichis.items() 
                                 if 'livrée' in etat or 'confirmée' in etat)
        commandes_problematiques = sum(data['nombre'] for etat, data in etats_enrichis.items() 
                                     if 'erronée' in etat or 'doublon' in etat or 'retournée' in etat)
        
        ws.cell(row=metrics_row + 1, column=1, value="Commandes en cours").border = border
        ws.cell(row=metrics_row + 1, column=2, value=commandes_en_cours).border = border
        ws.cell(row=metrics_row + 2, column=1, value="Commandes complétées").border = border
        ws.cell(row=metrics_row + 2, column=2, value=commandes_completees).border = border
        ws.cell(row=metrics_row + 3, column=1, value="Commandes problématiques").border = border
        ws.cell(row=metrics_row + 3, column=2, value=commandes_problematiques).border = border
        ws.cell(row=metrics_row + 4, column=1, value="Taux de complétion").border = border
        taux_completion = (commandes_completees / total_commandes * 100) if total_commandes > 0 else 0
        ws.cell(row=metrics_row + 4, column=2, value=f"{taux_completion:.1f}%").border = border

        # Ajouter des informations sur la période
        info_row = metrics_row + 6
        ws.cell(row=info_row, column=1, value="Période d'analyse").font = Font(bold=True)
        ws.cell(row=info_row, column=2, value=f"Du {date_debut} au {date_fin}").border = border
        ws.cell(row=info_row + 1, column=1, value="Durée (jours)").font = Font(bold=True)
        ws.cell(row=info_row + 1, column=2, value=(date_fin - date_debut).days + 1).border = border
        ws.cell(row=info_row + 2, column=1, value="Généré le").font = Font(bold=True)
        ws.cell(row=info_row + 2, column=2, value=timezone.now().strftime('%d/%m/%Y à %H:%M')).border = border
        
        # Ajuster la largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Créer la réponse HTTP
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Générer le nom de fichier avec la date
        from datetime import datetime
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"etat_commandes_{period}_{date_str}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    except Exception as e:
        logger.error(f"Erreur lors de l'export Excel: {str(e)}", exc_info=True)
        return HttpResponse(f"Erreur lors de l'export Excel: {str(e)}", status=500)
