from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.db.models import Count, Q, Sum, F, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.paginator import Paginator

import json
from parametre.models import Operateur
from commande.models import Commande, EtatCommande, EnumEtatCmd, Operation, Panier
from django.urls import reverse

import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import base64
import csv

from article.models import Article, MouvementStock
from commande.models import Envoi
from .forms import ArticleForm, AjusterStockForm
from .utils import creer_mouvement_stock

# D√©corateur pour g√©rer les erreurs AJAX
def handle_ajax_errors(view_func):
    def wrapper(request, *args, **kwargs):
        print(f"üîç D√©corateur handle_ajax_errors appel√© pour {view_func.__name__}")
        try:
            result = view_func(request, *args, **kwargs)
            print(f"üîç Fonction {view_func.__name__} termin√©e avec succ√®s")
            return result
        except Exception as e:
            print(f"‚ùå Erreur globale dans {view_func.__name__}: {str(e)}")
            import traceback
            print(f"‚ùå Traceback: {traceback.format_exc()}")
            
            # Si c'est une requ√™te AJAX, retourner JSON
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json':
                print(f"üîç Retour de r√©ponse JSON pour erreur AJAX")
                return JsonResponse({"success": False, "error": f"Erreur serveur: {str(e)}"})
            
            # Sinon, laisser Django g√©rer l'erreur normalement
            print(f"üîç Relance de l'exception pour gestion normale")
            raise
    return wrapper

# Create your views here.


@login_required
def home_view(request):
    """Page d'accueil avec statistiques pour les op√©rateurs de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil op√©rateur n'existe pas.")
        return redirect("login")

    # Date de r√©f√©rence
    today = timezone.now().date()
    # D√©but de la semaine (lundi)
    start_of_week = today - timedelta(days=today.weekday())

    # R√©cup√©rer les √©tats n√©cessaires
    try:
        etat_confirmee = EnumEtatCmd.objects.get(libelle__iexact="Confirm√©e")
        etat_en_preparation = EnumEtatCmd.objects.get(libelle__iexact="En pr√©paration")
        etat_preparee = EnumEtatCmd.objects.get(libelle__iexact="Pr√©par√©e")
    except EnumEtatCmd.DoesNotExist as e:
        messages.error(request, f"√âtat manquant dans le syst√®me: {str(e)}")
        return redirect("login")

    # 1. Commandes √† pr√©parer (√† imprimer et affect√©es √† cet op√©rateur)
    commandes_a_preparer = (
        Commande.objects.filter(
            etats__enum_etat__libelle="En pr√©paration",
        etats__operateur=operateur_profile,
            etats__date_fin__isnull=True,
        )
        .distinct()
        .count()
    )
    
    # 2. Commandes pr√©par√©es aujourd'hui par cet op√©rateur
    commandes_preparees = (
        EtatCommande.objects.filter(
            enum_etat__libelle="Pr√©par√©e",
        date_debut__date=today,
            operateur=operateur_profile,
        )
        .distinct()
        .count()
    )

    # 3. Commandes en cours de pr√©paration
    commandes_en_cours = (
        Commande.objects.filter(
        etats__enum_etat=etat_en_preparation,
        etats__date_fin__isnull=True,
            etats__operateur=operateur_profile,
        )
        .distinct()
        .count()
    )

    # 4. Performance de l'op√©rateur aujourd'hui
    ma_performance = (
        EtatCommande.objects.filter(
            enum_etat=etat_preparee, date_debut__date=today, operateur=operateur_profile
        )
        .distinct()
        .count()
    )

    # === Calculs suppl√©mentaires pour le tableau de bord ===
    # Commandes pr√©par√©es aujourd'hui (toutes)
    commandes_preparees_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation, date_fin__date=today
    ).count()

    # Commandes pr√©par√©es cette semaine (toutes)
    commandes_preparees_week = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date__gte=start_of_week,
        date_fin__date__lte=today,
    ).count()

    # Commandes actuellement en pr√©paration (toutes)
    commandes_en_preparation = Commande.objects.filter(
        etats__enum_etat=etat_en_preparation, etats__date_fin__isnull=True
    ).count()

    # Performance de l'op√©rateur aujourd'hui (commandes pr√©par√©es par lui)
    ma_performance_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation, date_fin__date=today, operateur=operateur_profile
    ).count()

    # Valeur totale (DH) des commandes pr√©par√©es aujourd'hui
    commandes_ids_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation, date_fin__date=today
    ).values_list("commande_id", flat=True)
    valeur_preparees_today = (
        Commande.objects.filter(id__in=commandes_ids_today).aggregate(
            total=Sum("total_cmd")
        )["total"]
        or 0
    )

    # Articles populaires (semaine en cours)
    commandes_ids_week = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date__gte=start_of_week,
        date_fin__date__lte=today,
    ).values_list("commande_id", flat=True)
    articles_populaires = (
        Panier.objects.filter(commande_id__in=commandes_ids_week)
        .values("article__nom", "article__reference")
        .annotate(
            total_quantite=Sum("quantite"),
            total_commandes=Count("commande", distinct=True),
        )
        .order_by("-total_quantite")[:5]
    )

    # Activit√© r√©cente (5 derni√®res pr√©parations de l'op√©rateur)
    activite_recente = (
        EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        operateur=operateur_profile,
            date_fin__isnull=False,
        )
        .select_related("commande", "commande__client")
        .order_by("-date_fin")[:5]
    )

    # Pr√©parer les statistiques
    stats = {
        "commandes_a_preparer": commandes_a_preparer,
        "commandes_preparees": commandes_preparees,
        "commandes_en_cours": commandes_en_cours,
        "ma_performance": ma_performance,
        # Ajout des nouvelles statistiques
        "commandes_preparees_today": commandes_preparees_today,
        "commandes_preparees_week": commandes_preparees_week,
        "commandes_en_preparation": commandes_en_preparation,
        "ma_performance_today": ma_performance_today,
        "valeur_preparees_today": valeur_preparees_today,
        "articles_populaires": articles_populaires,
        "activite_recente": activite_recente,
    }

    context = {
        "page_title": "Tableau de Bord",
        "page_subtitle": "Interface Op√©rateur de Pr√©paration",
        "profile": operateur_profile,
        "stats": stats,
        "total_commandes": commandes_a_preparer,  # Ajout du total des commandes √† pr√©parer
    }
    return render(request, "composant_generale/operatPrepa/home.html", context)


@login_required
def liste_prepa(request):
    """Liste des commandes √† pr√©parer pour les op√©rateurs de pr√©paration"""
    from commande.models import Operation
    
    try:
        operateur_profile = request.user.profil_operateur
        
        # V√©rifier que l'utilisateur est un op√©rateur de pr√©paration
        if not operateur_profile.is_preparation:
            messages.error(
                request,
                "Acc√®s non autoris√©. Vous n'√™tes pas un op√©rateur de pr√©paration.",
            )
            return redirect("login")
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil op√©rateur n'existe pas.")
        return redirect("login")

    # R√©cup√©rer les commandes dont l'√©tat ACTUEL est "√Ä imprimer" ou "En pr√©paration" et qui sont affect√©es √† cet op√©rateur
    # On cherche les commandes qui ont un √©tat "√Ä imprimer" ou "En pr√©paration" actif (sans date_fin) avec cet op√©rateur
    from django.db.models import Q, Max
    
    # D√©finir le type de filtre en premier
    filter_type = request.GET.get("filter", "all")
    
    if filter_type == "livrees_partiellement":
        # Filtrer les commandes qui ont √©t√© livr√©es partiellement et sont maintenant en pr√©paration
        commandes_affectees = []
        commandes_base = (
            Commande.objects.filter(
                Q(etats__enum_etat__libelle="√Ä imprimer")
                | Q(etats__enum_etat__libelle="En pr√©paration")
                | Q(etats__enum_etat__libelle="Collect√©e")
                | Q(etats__enum_etat__libelle="Emball√©e"),
            etats__operateur=operateur_profile,
                etats__date_fin__isnull=True,
            )
            .select_related("client", "ville", "ville__region")
            .prefetch_related("paniers__article", "etats")
            .distinct()
        )
        
        for commande in commandes_base:
            etats_commande = commande.etats.all().order_by("date_debut")
            etat_prepa_actuel = None
            
            # Trouver l'√©tat actuel de pr√©paration
            for etat in etats_commande:
                if (
                    etat.enum_etat.libelle
                    in ["√Ä imprimer", "En pr√©paration", "Collect√©e", "Emball√©e"]
                    and not etat.date_fin
                ):
                    etat_prepa_actuel = etat
                    break
            
            if etat_prepa_actuel:
                # V√©rifier si la commande a un historique de livraison partielle
                has_partially_delivered_history = False
                for etat in etats_commande:
                    if (
                        etat.enum_etat.libelle == "Livr√©e Partiellement"
                        and etat.date_fin
                        and etat.date_fin < etat_prepa_actuel.date_debut
                    ):
                        has_partially_delivered_history = True
                        break
                
                if has_partially_delivered_history:
                    commandes_affectees.append(commande)
    elif filter_type == "renvoyees_logistique":
        # Pour les commandes renvoy√©es par la logistique, ne pas exclure les √©tats probl√©matiques
        # car on veut inclure les commandes avec op√©ration de renvoi m√™me si elles ont des √©tats ult√©rieurs
        commandes_affectees = (
            Commande.objects.filter(
                Q(etats__enum_etat__libelle="√Ä imprimer")
                | Q(etats__enum_etat__libelle="En pr√©paration")
                | Q(etats__enum_etat__libelle="Collect√©e")
                | Q(etats__enum_etat__libelle="Emball√©e"),
            etats__operateur=operateur_profile,
                etats__date_fin__isnull=True,  # √âtat actif (en cours)
            )
            .select_related("client", "ville", "ville__region")
            .prefetch_related("paniers__article", "etats")
            .distinct()
        )
    elif filter_type == "retournees":
        # Obsol√®te: rediriger vers la page d√©di√©e
        return redirect("Prepacommande:commandes_retournees")

    elif filter_type == "affectees_supervision":
        # Filtrer les commandes affect√©es par les superviseurs de pr√©paration
        commandes_affectees = []
        commandes_base = (
            Commande.objects.filter(
                Q(etats__enum_etat__libelle="√Ä imprimer")
                | Q(etats__enum_etat__libelle="En pr√©paration")
                | Q(etats__enum_etat__libelle="Collect√©e")
                | Q(etats__enum_etat__libelle="Emball√©e"),
                etats__operateur=operateur_profile,
                etats__date_fin__isnull=True,
            )
            .select_related("client", "ville", "ville__region")
            .prefetch_related("paniers__article", "etats")
            .distinct()
        )

        for commande in commandes_base:
            etats_commande = commande.etats.all().order_by("date_debut")
            etat_prepa_actuel = None

            # Trouver l'√©tat actuel de pr√©paration
            for etat in etats_commande:
                if (
                    etat.enum_etat.libelle
                    in ["√Ä imprimer", "En pr√©paration", "Collect√©e", "Emball√©e"]
                    and not etat.date_fin
                ):
                    etat_prepa_actuel = etat
                    break

            if etat_prepa_actuel:
                # V√©rifier s'il y a des √©tats ult√©rieurs probl√©matiques
                a_etats_ult√©rieurs_problematiques = False
                for etat in etats_commande:
                    if (
                        etat.date_debut > etat_prepa_actuel.date_debut
                        and etat.enum_etat.libelle
                        in [
                            "Livr√©e",
                            "Livr√©e Partiellement",
                            "Pr√©par√©e",
                            "En cours de livraison",
                        ]
                    ):
                        a_etats_ult√©rieurs_problematiques = True
                        break

                if a_etats_ult√©rieurs_problematiques:
                    continue

                # V√©rifier les op√©rations de renvoi
                operation_renvoi = Operation.objects.filter(
                    commande=commande, type_operation="RENVOI_PREPARATION"
                ).first()

                if operation_renvoi:
                    continue  # Exclure les commandes renvoy√©es par logistique

                # V√©rifier si c'est une commande de renvoi cr√©√©e lors d'une livraison partielle
                if commande.num_cmd and commande.num_cmd.startswith("RENVOI-"):
                    # Chercher la commande originale
                    num_cmd_original = commande.num_cmd.replace("RENVOI-", "")
                    commande_originale = Commande.objects.filter(
                        num_cmd=num_cmd_original,
                        etats__enum_etat__libelle="Livr√©e Partiellement",
                    ).first()

                    if commande_originale:
                        continue  # Exclure les commandes de renvoi livraison partielle

                # V√©rifier l'historique pour renvoi depuis livraison
                has_return_from_delivery = False
                for etat in reversed(etats_commande):
                    if etat.date_fin and etat.date_fin < etat_prepa_actuel.date_debut:
                        if etat.enum_etat.libelle in [
                            "En cours de livraison",
                            "Livr√©e Partiellement",
                        ]:
                            has_return_from_delivery = True
                            break

                if not has_return_from_delivery:
                    # V√©rifier si la commande a √©t√© affect√©e par un superviseur
                    # Chercher les op√©rations d'affectation par supervision
                    operation_affectation_supervision = Operation.objects.filter(
                        commande=commande,
                        type_operation__in=[
                            "AFFECTATION_SUPERVISION",
                            "REAFFECTATION_SUPERVISION",
                        ],
                    ).first()

                    if operation_affectation_supervision:
                        commandes_affectees.append(commande)
    else:  # filter_type == 'all'
        # Pour "Toutes les commandes", afficher toutes les commandes des 5 autres onglets
        commandes_affectees = []
        commandes_base = (
            Commande.objects.filter(
                Q(etats__enum_etat__libelle="√Ä imprimer")
                | Q(etats__enum_etat__libelle="En pr√©paration")
                | Q(etats__enum_etat__libelle="Collect√©e")
                | Q(etats__enum_etat__libelle="Emball√©e"),
            etats__operateur=operateur_profile,
                etats__date_fin__isnull=True,
            )
            .select_related("client", "ville", "ville__region")
            .prefetch_related("paniers__article", "etats")
            .distinct()
        )
        
        for commande in commandes_base:
            etats_commande = commande.etats.all().order_by("date_debut")
            etat_prepa_actuel = None
            
            # Trouver l'√©tat actuel de pr√©paration
            for etat in etats_commande:
                if (
                    etat.enum_etat.libelle
                    in ["√Ä imprimer", "En pr√©paration", "Collect√©e", "Emball√©e"]
                    and not etat.date_fin
                ):
                    etat_prepa_actuel = etat
                    break
            
            if etat_prepa_actuel:
                # V√©rifier s'il y a des √©tats ult√©rieurs probl√©matiques
                a_etats_ult√©rieurs_problematiques = False
                for etat in etats_commande:
                    if (
                        etat.date_debut > etat_prepa_actuel.date_debut
                        and etat.enum_etat.libelle
                        in [
                            "Livr√©e",
                            "Livr√©e Partiellement",
                            "Pr√©par√©e",
                            "En cours de livraison",
                        ]
                    ):
                        a_etats_ult√©rieurs_problematiques = True
                        break
                
                if a_etats_ult√©rieurs_problematiques:
                    continue
                
                # Inclure toutes les commandes valides (pas d'exclusion bas√©e sur le type)
                commandes_affectees.append(commande)
    
    # Pour les commandes renvoy√©es par la logistique, respecter l'affectation sp√©cifique √† chaque op√©rateur
    if filter_type == "renvoyees_logistique":
        # Filtrer seulement les commandes renvoy√©es par la logistique ET affect√©es √† cet op√©rateur sp√©cifique
        commandes_filtrees = []
        for commande in commandes_affectees:
            from commande.models import Operation
            
            # V√©rifier que la commande n'a pas d'√©tats ult√©rieurs probl√©matiques
            etats_commande = commande.etats.all().order_by("date_debut")
            etat_actuel = None
            
            # Trouver l'√©tat actuel (En pr√©paration)
            for etat in etats_commande:
                if (
                    etat.enum_etat.libelle
                    in ["√Ä imprimer", "En pr√©paration", "Collect√©e", "Emball√©e"]
                    and not etat.date_fin
                ):
                    etat_actuel = etat
                    break
            
            if etat_actuel:
                # V√©rifier les op√©rations de tra√ßabilit√© EN PREMIER
                operation_renvoi = Operation.objects.filter(
                    commande=commande, type_operation="RENVOI_PREPARATION"
                ).first()

                # Si il y a une op√©ration de renvoi explicite, inclure la commande
                # m√™me si elle a des √©tats ult√©rieurs probl√©matiques
                if operation_renvoi:
                    commandes_filtrees.append(commande)
                    continue

                # V√©rifier si c'est une commande de renvoi cr√©√©e lors d'une livraison partielle
                if commande.num_cmd and commande.num_cmd.startswith("RENVOI-"):
                    # Chercher la commande originale
                    num_cmd_original = commande.num_cmd.replace("RENVOI-", "")
                    commande_originale = Commande.objects.filter(
                        num_cmd=num_cmd_original,
                        etats__enum_etat__libelle="Livr√©e Partiellement",
                    ).first()

                    if commande_originale:
                        commandes_filtrees.append(commande)
                        continue

                # Sinon, v√©rifier s'il y a des √©tats ult√©rieurs probl√©matiques
                a_etats_ult√©rieurs_problematiques = False
                for etat in etats_commande:
                    if (
                        etat.date_debut > etat_actuel.date_debut
                        and etat.enum_etat.libelle
                        in [
                            "Livr√©e",
                            "Livr√©e Partiellement",
                            "Pr√©par√©e",
                            "En cours de livraison",
                        ]
                    ):
                        a_etats_ult√©rieurs_problematiques = True
                        break

                if a_etats_ult√©rieurs_problematiques:
                    continue  # Ignorer cette commande

                # V√©rifier l'historique des √©tats de la commande
                # Trouver l'√©tat pr√©c√©dent
                for etat in reversed(etats_commande):
                    if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                        if etat.enum_etat.libelle == "En cours de livraison":
                            commandes_filtrees.append(commande)
                            break
        
        commandes_affectees = commandes_filtrees
    
    # Calculer les statistiques par type de commande
    stats_par_type = {
        "renvoyees_logistique": 0,
        "livrees_partiellement": 0,
        "retournees": 0,
        "affectees_admin": 0,
    }
    
    # Pour chaque commande, ajouter l'√©tat pr√©c√©dent pour comprendre d'o√π elle vient
    for commande in commandes_affectees:
        # R√©cup√©rer tous les √©tats de la commande dans l'ordre chronologique
        etats_commande = commande.etats.all().order_by("date_debut")
        
        # Trouver l'√©tat actuel (√Ä imprimer, En pr√©paration, Collect√©e, ou Emball√©e)
        etat_actuel = None
        for etat in etats_commande:
            if (
                etat.enum_etat.libelle
                in ["√Ä imprimer", "En pr√©paration", "Collect√©e", "Emball√©e"]
                and not etat.date_fin
            ):
                etat_actuel = etat
                break
        
        if etat_actuel:
            # Trouver l'√©tat pr√©c√©dent (le dernier √©tat termin√© avant l'√©tat actuel)
            etat_precedent = None
            for etat in reversed(etats_commande):
                if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                    if etat.enum_etat.libelle not in [
                        "√Ä imprimer",
                        "En pr√©paration",
                        "Collect√©e",
                        "Emball√©e",
                    ]:
                        etat_precedent = etat
                        break
            
            commande.etat_precedent = etat_precedent
            
            # Trouver l'√©tat de confirmation (le premier √©tat "Confirm√©e")
            etat_confirmation = None
            for etat in etats_commande:
                if etat.enum_etat.libelle == "Confirm√©e":
                    etat_confirmation = etat
                    break
            
            commande.etat_confirmation = etat_confirmation
    
    # Si aucune commande trouv√©e avec la m√©thode stricte, essayer une approche plus large
    if isinstance(commandes_affectees, list):
        has_commandes = len(commandes_affectees) > 0
    else:
        has_commandes = commandes_affectees.exists()
    
    if not has_commandes:
        # Chercher toutes les commandes qui ont √©t√© affect√©es √† cet op√©rateur pour la pr√©paration
        # et qui n'ont pas encore d'√©tat "Pr√©par√©e" ou "En cours de livraison"
        commandes_affectees = (
            Commande.objects.filter(
                Q(etats__enum_etat__libelle="√Ä imprimer")
                | Q(etats__enum_etat__libelle="En pr√©paration")
                | Q(etats__enum_etat__libelle="Collect√©e")
                | Q(etats__enum_etat__libelle="Emball√©e"),
                etats__operateur=operateur_profile,
            )
            .exclude(
            # Exclure les commandes qui ont d√©j√† un √©tat ult√©rieur actif
                Q(
                    etats__enum_etat__libelle__in=[
                        "Pr√©par√©e",
                        "En cours de livraison",
                        "Livr√©e",
                        "Annul√©e",
                    ],
                    etats__date_fin__isnull=True,
                )
            )
            .select_related("client", "ville", "ville__region")
            .prefetch_related("paniers__article", "etats")
            .distinct()
        )
        
        # Pour chaque commande, ajouter l'√©tat pr√©c√©dent pour comprendre d'o√π elle vient
        for commande in commandes_affectees:
            # R√©cup√©rer tous les √©tats de la commande dans l'ordre chronologique
            etats_commande = commande.etats.all().order_by("date_debut")
            
            # Trouver l'√©tat actuel (√Ä imprimer, En pr√©paration, Collect√©e, ou Emball√©e)
            etat_actuel = None
            for etat in etats_commande:
                if (
                    etat.enum_etat.libelle
                    in ["√Ä imprimer", "En pr√©paration", "Collect√©e", "Emball√©e"]
                    and not etat.date_fin
                ):
                    etat_actuel = etat
                    break
            
            if etat_actuel:
                # Trouver l'√©tat pr√©c√©dent (le dernier √©tat termin√© avant l'√©tat actuel)
                etat_precedent = None
                for etat in reversed(etats_commande):
                    if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                        if etat.enum_etat.libelle not in [
                            "√Ä imprimer",
                            "En pr√©paration",
                            "Collect√©e",
                            "Emball√©e",
                        ]:
                            etat_precedent = etat
                            break
                
                commande.etat_precedent = etat_precedent
                
                # Trouver l'√©tat de confirmation (le premier √©tat "Confirm√©e")
                etat_confirmation = None
                for etat in etats_commande:
                    if etat.enum_etat.libelle == "Confirm√©e":
                        etat_confirmation = etat
                        break
                
                commande.etat_confirmation = etat_confirmation
    
    # Suppression du filtre 'nouvelles' car redondant avec l'affectation automatique
    # Suppression du filtre 'renvoyees_preparation' car non n√©cessaire
    
    # Recherche
    search_query = request.GET.get("search", "")
    if search_query:
        if isinstance(commandes_affectees, list):
            # Si c'est une liste (apr√®s filtrage)
            commandes_affectees = [
                cmd
                for cmd in commandes_affectees
                if search_query.lower() in str(cmd.id_yz).lower()
                or search_query.lower() in (cmd.num_cmd or "").lower()
                or search_query.lower() in cmd.client.nom.lower()
                or search_query.lower() in cmd.client.prenom.lower()
                or search_query.lower() in (cmd.client.numero_tel or "").lower()
            ]
        else:
            # Si c'est un QuerySet
            commandes_affectees = commandes_affectees.filter(
                Q(id_yz__icontains=search_query)
                | Q(num_cmd__icontains=search_query)
                | Q(client__nom__icontains=search_query)
                | Q(client__prenom__icontains=search_query)
                | Q(client__numero_tel__icontains=search_query)
            ).distinct()
    
    # Tri par date d'affectation (plus r√©centes en premier)
    if isinstance(commandes_affectees, list):
        commandes_affectees.sort(
            key=lambda x: x.etats.filter(date_fin__isnull=True).first().date_debut
            if x.etats.filter(date_fin__isnull=True).first()
            else timezone.now(),
            reverse=True,
        )
    else:
        commandes_affectees = commandes_affectees.order_by("-etats__date_debut")

    # G√©n√©rer les codes-barres pour chaque commande
    code128 = barcode.get_barcode_class("code128")
    for commande in commandes_affectees:
        if commande.id_yz:
            barcode_instance = code128(str(commande.id_yz), writer=ImageWriter())
            buffer = BytesIO()
            barcode_instance.write(
                buffer, options={"write_text": False, "module_height": 10.0}
            )
            barcode_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
            commande.barcode_base64 = barcode_base64
        else:
            commande.barcode_base64 = None
    
    # Statistiques
    if isinstance(commandes_affectees, list):
        total_affectees = len(commandes_affectees)
        valeur_totale = sum(cmd.total_cmd or 0 for cmd in commandes_affectees)
        
        # Commandes urgentes (affect√©es depuis plus de 1 jour)
        date_limite_urgence = timezone.now() - timedelta(days=1)
        commandes_urgentes = sum(
            1
            for cmd in commandes_affectees
            if cmd.etats.filter(date_debut__lt=date_limite_urgence).exists()
        )
    else:
        total_affectees = commandes_affectees.count()
        valeur_totale = (
            commandes_affectees.aggregate(total=Sum("total_cmd"))["total"] or 0
        )
        
        # Commandes urgentes (affect√©es depuis plus de 1 jour)
        date_limite_urgence = timezone.now() - timedelta(days=1)
        commandes_urgentes = commandes_affectees.filter(
            etats__date_debut__lt=date_limite_urgence
        ).count()
    
    # Statistiques par type pour les onglets
    stats_par_type = {
        "renvoyees_logistique": 0,
        "livrees_partiellement": 0,
        "retournees": 0,
        "affectees_supervision": 0,
    }
    
    # Recalculer les statistiques pour tous les types
    # D'abord, r√©cup√©rer toutes les commandes affect√©es √† cet op√©rateur (sans filtre)
    toutes_commandes = (
        Commande.objects.filter(
            Q(etats__enum_etat__libelle="√Ä imprimer")
            | Q(etats__enum_etat__libelle="En pr√©paration")
            | Q(etats__enum_etat__libelle="Collect√©e")
            | Q(etats__enum_etat__libelle="Emball√©e"),
        etats__operateur=operateur_profile,
            etats__date_fin__isnull=True,  # √âtat actif (en cours)
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("paniers__article", "etats")
        .distinct()
    )
    
    for cmd in toutes_commandes:
        # V√©rifier si c'est une commande renvoy√©e par la logistique
        operation_renvoi = Operation.objects.filter(
            commande=cmd, type_operation="RENVOI_PREPARATION"
        ).first()
        
        if operation_renvoi:
            stats_par_type["renvoyees_logistique"] += 1
            continue
        
        # V√©rifier si c'est une commande de renvoi cr√©√©e lors d'une livraison partielle
        if cmd.num_cmd and cmd.num_cmd.startswith("RENVOI-"):
            # Chercher la commande originale
            num_cmd_original = cmd.num_cmd.replace("RENVOI-", "")
            commande_originale = Commande.objects.filter(
                num_cmd=num_cmd_original,
                etats__enum_etat__libelle="Livr√©e Partiellement",
            ).first()
            
            if commande_originale:
                stats_par_type["renvoyees_logistique"] += 1
                continue
        
        # V√©rifier l'√©tat pr√©c√©dent
        etats_commande = cmd.etats.all().order_by("date_debut")
        etat_actuel = None
        
        # Trouver l'√©tat actuel
        for etat in etats_commande:
            if (
                etat.enum_etat.libelle
                in ["√Ä imprimer", "En pr√©paration", "Collect√©e", "Emball√©e"]
                and not etat.date_fin
            ):
                etat_actuel = etat
                break
        
        if etat_actuel:
            # V√©rifier s'il y a des √©tats ult√©rieurs probl√©matiques
            a_etats_ult√©rieurs_problematiques = False
            for etat in etats_commande:
                if (
                    etat.date_debut > etat_actuel.date_debut
                    and etat.enum_etat.libelle
                    in [
                        "Livr√©e",
                        "Livr√©e Partiellement",
                        "Pr√©par√©e",
                        "En cours de livraison",
                    ]
                ):
                    a_etats_ult√©rieurs_problematiques = True
                    break
            
            # Si il y a des √©tats ult√©rieurs probl√©matiques, ignorer cette commande
            if a_etats_ult√©rieurs_problematiques:
                continue
            
            # Trouver l'√©tat pr√©c√©dent
            for etat in reversed(etats_commande):
                if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                    if etat.enum_etat.libelle == "En cours de livraison":
                        stats_par_type["renvoyees_logistique"] += 1
                        break
                    elif etat.enum_etat.libelle == "Livr√©e Partiellement":
                        stats_par_type["livrees_partiellement"] += 1
                        break
            
    # Recalculer le compteur des livraisons partielles en utilisant la m√™me logique que la vue s√©par√©e
    # Chercher les commandes de renvoi cr√©√©es lors de livraisons partielles
    commandes_renvoi_livraison_partielle = Commande.objects.filter(
        num_cmd__startswith="RENVOI-",
        etats__enum_etat__libelle="En pr√©paration",
        etats__operateur=operateur_profile,
        etats__date_fin__isnull=True,
    ).distinct()
    
    livrees_partiellement_count = 0
    for commande_renvoi in commandes_renvoi_livraison_partielle:
        # Extraire le num√©ro de commande original
        num_cmd_original = commande_renvoi.num_cmd.replace("RENVOI-", "")
        
        # V√©rifier que la commande originale a √©t√© livr√©e partiellement
        commande_originale = Commande.objects.filter(
            num_cmd=num_cmd_original, etats__enum_etat__libelle="Livr√©e Partiellement"
        ).first()
        
        if commande_originale:
            livrees_partiellement_count += 1
    
    # Mettre √† jour le compteur avec la valeur correcte
    stats_par_type["livrees_partiellement"] = livrees_partiellement_count
    

    


    # Calculer les statistiques pour les commandes affect√©es par supervision
    commandes_affectees_supervision = 0
    for cmd in toutes_commandes:
        # V√©rifier si c'est une commande renvoy√©e par la logistique
        operation_renvoi = Operation.objects.filter(
            commande=cmd, type_operation="RENVOI_PREPARATION"
        ).first()

        if operation_renvoi:
            continue  # D√©j√† compt√©e dans renvoyees_logistique

        # V√©rifier si c'est une commande de renvoi cr√©√©e lors d'une livraison partielle
        if cmd.num_cmd and cmd.num_cmd.startswith("RENVOI-"):
            # Chercher la commande originale
            num_cmd_original = cmd.num_cmd.replace("RENVOI-", "")
            commande_originale = Commande.objects.filter(
                num_cmd=num_cmd_original,
                etats__enum_etat__libelle="Livr√©e Partiellement",
            ).first()

            if commande_originale:
                continue  # D√©j√† compt√©e dans livrees_partiellement

        # V√©rifier l'historique des √©tats
        etats_commande = cmd.etats.all().order_by("date_debut")
        etat_actuel = None

        # Trouver l'√©tat actuel
        for etat in etats_commande:
            if (
                etat.enum_etat.libelle
                in ["√Ä imprimer", "En pr√©paration", "Collect√©e", "Emball√©e"]
                and not etat.date_fin
            ):
                etat_actuel = etat
                break

        if etat_actuel:
            # V√©rifier s'il y a des √©tats ult√©rieurs probl√©matiques
            a_etats_ult√©rieurs_problematiques = False
            for etat in etats_commande:
                if (
                    etat.date_debut > etat_actuel.date_debut
                    and etat.enum_etat.libelle
                    in [
                        "Livr√©e",
                        "Livr√©e Partiellement",
                        "Pr√©par√©e",
                        "En cours de livraison",
                    ]
                ):
                    a_etats_ult√©rieurs_problematiques = True
                    break

            if a_etats_ult√©rieurs_problematiques:
                continue

            # V√©rifier l'historique pour renvoi depuis livraison
            has_return_from_delivery = False
            for etat in reversed(etats_commande):
                if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                    if etat.enum_etat.libelle == "En cours de livraison":
                        has_return_from_delivery = True
                        break
                    elif etat.enum_etat.libelle == "Livr√©e Partiellement":
                        has_return_from_delivery = True
                        break

            if not has_return_from_delivery:
                # V√©rifier si la commande a √©t√© affect√©e par un superviseur
                operation_affectation_supervision = Operation.objects.filter(
                    commande=cmd,
                    type_operation__in=[
                        "AFFECTATION_SUPERVISION",
                        "REAFFECTATION_SUPERVISION",
                    ],
                ).first()

                if operation_affectation_supervision:
                    commandes_affectees_supervision += 1

    stats_par_type["affectees_supervision"] = commandes_affectees_supervision

    # Pour l'onglet "Toutes les commandes", le total doit √™tre la somme des 3 autres onglets
    if filter_type == "all":
        total_affectees = (
            stats_par_type["renvoyees_logistique"]
            + stats_par_type["livrees_partiellement"]
            + stats_par_type["affectees_supervision"]
        )

    # V√©rifier si c'est une requ√™te AJAX pour les statistiques
    if request.GET.get("ajax") == "stats":
        from django.http import JsonResponse

        return JsonResponse(
            {
                "affectees_supervision": stats_par_type.get("affectees_supervision", 0),
                "total_affectees": total_affectees,
                "renvoyees_logistique": stats_par_type.get("renvoyees_logistique", 0),
            }
        )

    context = {
        "page_title": "Mes Commandes √† Pr√©parer",
        "page_subtitle": f"Vous avez {total_affectees} commande(s) affect√©e(s)",
        "commandes_affectees": commandes_affectees,
        "search_query": search_query,
        "filter_type": filter_type,
        "stats": {
            "total_affectees": total_affectees,
            "valeur_totale": valeur_totale,
            "commandes_urgentes": commandes_urgentes,
        },
        "stats_par_type": stats_par_type,
        "operateur_profile": operateur_profile,
        "api_produits_url_base": reverse(
            "Prepacommande:api_commande_produits", args=[99999999]
        ),
    }
    return render(request, "Prepacommande/liste_prepa.html", context)


@login_required
def commandes_en_preparation(request):
    """Liste des commandes en cours de pr√©paration pour les op√©rateurs de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        
        # V√©rifier que l'utilisateur est un op√©rateur de pr√©paration
        if not operateur_profile.is_preparation:
            messages.error(
                request,
                "Acc√®s non autoris√©. Vous n'√™tes pas un op√©rateur de pr√©paration.",
            )
            return redirect("login")
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil op√©rateur n'existe pas.")
        return redirect("login")

    # R√©cup√©rer les commandes dont l'√©tat ACTUEL est "En pr√©paration" et qui sont affect√©es √† cet op√©rateur
    commandes_en_preparation = (
        Commande.objects.filter(
            etats__enum_etat__libelle="En pr√©paration",
        etats__operateur=operateur_profile,
            etats__date_fin__isnull=True,  # √âtat actif (en cours)
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("paniers__article", "etats")
        .distinct()
    )

    context = {
        "page_title": "Commandes en Pr√©paration",
        "page_subtitle": "Interface Op√©rateur de Pr√©paration",
        "profile": operateur_profile,
        "commandes": commandes_en_preparation,
        "active_tab": "en_preparation",
    }
    return render(request, "Prepacommande/commandes_en_preparation.html", context)


@login_required
def commandes_livrees_partiellement(request):
    """Liste des commandes livr√©es partiellement renvoy√©es en pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        
        # V√©rifier que l'utilisateur est un op√©rateur de pr√©paration
        if not operateur_profile.is_preparation:
            messages.error(
                request,
                "Acc√®s non autoris√©. Vous n'√™tes pas un op√©rateur de pr√©paration.",
            )
            return redirect("login")
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil op√©rateur n'existe pas.")
        return redirect("login")

    # Nouvel objectif: r√©cup√©rer les commandes livr√©es partiellement
    # qui ont √©t√© pr√©par√©es (envoy√©es √† la logistique) par cet op√©rateur
    # L'op√©rateur de pr√©paration cr√©e l'√©tat 'En pr√©paration', pas 'Pr√©par√©e'
    commandes_livrees_partiellement_qs = (
        Commande.objects.filter(etats__enum_etat__libelle="Livr√©e Partiellement")
        .filter(
            etats__enum_etat__libelle="En pr√©paration",
            etats__operateur=operateur_profile,
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("paniers__article", "etats")
        .distinct()
    )
    
    # Pour chaque commande originale, essayer de retrouver une √©ventuelle commande de renvoi associ√©e
    commandes_livrees_partiellement = []
    for commande_originale in commandes_livrees_partiellement_qs:
        commande_renvoi = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande_originale.num_cmd}",
            client=commande_originale.client,
        ).first()
        if commande_renvoi:
            commande_originale.commande_renvoi = commande_renvoi
        commandes_livrees_partiellement.append(commande_originale)

    # Les commandes sont d√©j√† filtr√©es et pertinentes
    commandes_filtrees = commandes_livrees_partiellement

    # Pour chaque commande, r√©cup√©rer les d√©tails de la livraison partielle
    for commande in commandes_livrees_partiellement:
        # Trouver l'√©tat "Livr√©e Partiellement" le plus r√©cent
        etat_livraison_partielle = (
            commande.etats.filter(enum_etat__libelle="Livr√©e Partiellement")
            .order_by("-date_debut")
            .first()
        )
        
        if etat_livraison_partielle:
            commande.date_livraison_partielle = etat_livraison_partielle.date_debut
            commande.commentaire_livraison_partielle = (
                etat_livraison_partielle.commentaire
            )
            commande.operateur_livraison = etat_livraison_partielle.operateur
            
            # Le statut est toujours "Renvoy√©e en pr√©paration" car nous ne r√©cup√©rons que les commandes avec renvoi
            commande.statut_actuel = "Renvoy√©e en pr√©paration"
            
            # Ajouter les informations de la commande de renvoi
            if hasattr(commande, "commande_renvoi"):
                commande.commande_renvoi_id = commande.commande_renvoi.id
                commande.commande_renvoi_num = commande.commande_renvoi.num_cmd
                commande.commande_renvoi_id_yz = commande.commande_renvoi.id_yz

    context = {
        "page_title": "Commandes Livr√©es Partiellement",
        "page_subtitle": "Interface Op√©rateur de Pr√©paration",
        "profile": operateur_profile,
        "commandes_livrees_partiellement": commandes_livrees_partiellement,
        "commandes_count": len(commandes_livrees_partiellement),
        "active_tab": "livrees_partiellement",
    }
    return render(
        request, "Prepacommande/commandes_livrees_partiellement.html", context
    )


@login_required
def commandes_retournees(request):
    """Liste des commandes retourn√©es (√©tat actuel 'Retourn√©e') pr√©par√©es initialement par l'op√©rateur courant"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(
                request,
                "Acc√®s non autoris√©. Vous n'√™tes pas un op√©rateur de pr√©paration.",
            )
            return redirect("login")
    
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil op√©rateur n'existe pas.")
        return redirect("login")

    # Commandes dont l'√©tat ACTUEL est 'Retourn√©e' et qui ont √©t√© pr√©par√©es par l'op√©rateur courant
    # L'op√©rateur de pr√©paration cr√©e l'√©tat 'En pr√©paration', pas 'Pr√©par√©e'
    commandes_qs = (
        Commande.objects.filter(
            etats__enum_etat__libelle="Retourn√©e", etats__date_fin__isnull=True
        )  # √âtat actuel
        .filter(
            etats__enum_etat__libelle="En pr√©paration",
            etats__operateur=operateur_profile,
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("paniers__article", "etats")
        .distinct()
    )

    commandes = list(commandes_qs)

    # Enrichir avec m√©ta d'√©tat 'Retourn√©e'
    for commande in commandes:
        etat_retour = (
            commande.etats.filter(enum_etat__libelle="Retourn√©e")
            .order_by("-date_debut")
            .first()
        )
        if etat_retour:
            commande.date_retournee = etat_retour.date_debut
            commande.commentaire_retournee = etat_retour.commentaire
            commande.operateur_retour = etat_retour.operateur

    context = {
        "page_title": "Commandes Retourn√©es",
        "page_subtitle": "Commandes renvoy√©es √† la pr√©paration (√©tat 'Retourn√©e')",
        "profile": operateur_profile,
        "commandes_retournees": commandes,
        "commandes_count": len(commandes),
        "active_tab": "retournees",
    }
    return render(request, "Prepacommande/commandes_retournees.html", context)


@login_required
def traiter_commande_retournee_api(request, commande_id):
    """API pour traiter une commande retourn√©e et g√©rer la r√©incr√©mentation du stock"""
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "M√©thode non autoris√©e"})
    
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            return JsonResponse({"success": False, "message": "Acc√®s non autoris√©"})
        
        # R√©cup√©rer la commande
        commande = get_object_or_404(Commande, id=commande_id)
        
        # V√©rifier que la commande est bien retourn√©e et pr√©par√©e par cet op√©rateur
        if not commande.etats.filter(
            enum_etat__libelle="Retourn√©e", date_fin__isnull=True
        ).exists():
            return JsonResponse({"success": False, "message": "Commande non retourn√©e"})
        
        if not commande.etats.filter(
            enum_etat__libelle="En pr√©paration", operateur=operateur_profile
        ).exists():
            return JsonResponse(
                {"success": False, "message": "Commande non pr√©par√©e par vous"}
            )
        
        # R√©cup√©rer les donn√©es de la requ√™te
        data = json.loads(request.body)
        type_traitement = data.get("type_traitement")
        etat_stock = data.get("etat_stock")
        commentaire = data.get("commentaire")
        
        if not all([type_traitement, etat_stock, commentaire]):
            return JsonResponse({"success": False, "message": "Donn√©es manquantes"})
        
        # Traitement selon le type
        with transaction.atomic():
            if type_traitement == "repreparer":
                # Ne pas changer l'√©tat de la commande, elle reste "Retourn√©e"
                # Seulement r√©incr√©menter le stock si les produits sont en bon √©tat
                if etat_stock == "bon":
                    for panier in commande.paniers.all():
                        article = panier.article
                        quantite = panier.quantite
                        
                        # R√©incr√©menter le stock disponible
                        article.qte_disponible += quantite
                        article.save()
                        
                        # Cr√©er un mouvement de stock pour tracer
                        from article.models import MouvementStock

                        MouvementStock.objects.create(
                            article=article,
                            quantite=quantite,
                            type_mouvement="entree",
                            commentaire=f"R√©incr√©mentation - Commande retourn√©e {commande.id_yz} - Produits en bon √©tat - {commentaire}",
                            operateur=operateur_profile,
                            qte_apres_mouvement=article.qte_disponible,
                        )
                
                message = f"Stock r√©incr√©ment√©: {'Oui' if etat_stock == 'bon' else 'Non'}. Commande reste en √©tat 'Retourn√©e'."
                
                return JsonResponse(
                    {"success": True, "message": message, "commande_id": commande.id}
                )

            return JsonResponse(
                {"success": True, "message": message, "commande_id": commande.id}
            )
            
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Erreur: {str(e)}"})


@login_required
def profile_view(request):
    try:
        operateur_profile = request.user.profil_operateur
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil op√©rateur n'existe pas.")
        return redirect("login")  # Ou une page d'erreur appropri√©e

    context = {
        "page_title": "Mon Profil",
        "page_subtitle": "G√©rez les informations de votre profil",
        "profile": operateur_profile,
    }
    return render(request, "Prepacommande/profile.html", context)


@login_required
def modifier_profile_view(request):
    try:
        operateur_profile = request.user.profil_operateur
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil op√©rateur n'existe pas.")
        return redirect("login")

    if request.method == "POST":
        # R√©cup√©rer les donn√©es du formulaire
        nom = request.POST.get("nom")
        prenom = request.POST.get("prenom")
        mail = request.POST.get("mail")
        telephone = request.POST.get("telephone")
        adresse = request.POST.get("adresse")

        # Validation minimale (vous pouvez ajouter plus de validations ici)
        if not nom or not prenom or not mail:
            messages.error(request, "Nom, pr√©nom et email sont requis.")
            context = {
                "page_title": "Modifier Profil",
                "page_subtitle": "Mettez √† jour vos informations personnelles",
                "profile": operateur_profile,
                "form_data": request.POST,  # Pour pr√©-remplir le formulaire
            }
            return render(request, "Prepacommande/modifier_profile.html", context)
        
        # Mettre √† jour l'utilisateur Django
        request.user.first_name = prenom
        request.user.last_name = nom
        request.user.email = mail
        request.user.save()

        # Mettre √† jour le profil de l'op√©rateur
        operateur_profile.nom = nom
        operateur_profile.prenom = prenom
        operateur_profile.mail = mail
        operateur_profile.telephone = telephone
        operateur_profile.adresse = adresse

        # G√©rer l'image si elle est fournie
        if "photo" in request.FILES:
            operateur_profile.photo = request.FILES["photo"]
        
        operateur_profile.save()

        messages.success(request, "Votre profil a √©t√© mis √† jour avec succ√®s.")
        return redirect(
            "Prepacommande:profile"
        )  # Rediriger vers la page de profil apr√®s succ√®s
    else:
        context = {
            "page_title": "Modifier Profil",
            "page_subtitle": "Mettez √† jour vos informations personnelles",
            "profile": operateur_profile,
        }
        return render(request, "Prepacommande/modifier_profile.html", context)


@login_required
def changer_mot_de_passe_view(request):
    """Page de changement de mot de passe pour l'op√©rateur de pr√©paration - D√©sactiv√©e"""
    return redirect("Prepacommande:profile")


@login_required
def detail_prepa(request, pk):
    """Vue d√©taill√©e pour la pr√©paration d'une commande sp√©cifique"""
    try:
        operateur_profile = request.user.profil_operateur
        
        # V√©rifier que l'utilisateur est un op√©rateur de pr√©paration
        if not operateur_profile.is_preparation:
            messages.error(
                request,
                "Acc√®s non autoris√©. Vous n'√™tes pas un op√©rateur de pr√©paration.",
            )
            return redirect("login")
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil op√©rateur n'existe pas.")
        return redirect("login")

    # R√©cup√©rer la commande sp√©cifique
    try:
        commande = (
            Commande.objects.select_related("client", "ville", "ville__region")
            .prefetch_related(
                "paniers__article", "etats__enum_etat", "etats__operateur"
            )
            .get(id=pk)
        )
    except Commande.DoesNotExist:
        messages.error(request, "La commande demand√©e n'existe pas.")
        return redirect("Prepacommande:liste_prepa")

    # V√©rifier que la commande est bien affect√©e √† cet op√©rateur pour la pr√©paration
    etat_preparation = commande.etats.filter(
        Q(enum_etat__libelle="√Ä imprimer")
        | Q(enum_etat__libelle="En pr√©paration")
        | Q(enum_etat__libelle="Collect√©e")
        | Q(enum_etat__libelle="Emball√©e"),
        operateur=operateur_profile,
    ).first()
    
    if not etat_preparation:
        messages.error(
            request, "Cette commande ne vous est pas affect√©e pour la pr√©paration."
        )
        return redirect("Prepacommande:liste_prepa")

    # R√©cup√©rer les paniers (articles) de la commande
    paniers = commande.paniers.all().select_related("article")
    
    # Initialiser les variables pour les cas de livraison partielle/renvoi
    articles_livres = []
    articles_renvoyes = []
    is_commande_livree_partiellement = False
    commande_renvoi = None  # Initialiser √† None
    commande_originale = None  # Initialiser √† None
    etat_articles_renvoyes = {}  # Initialiser √† un dictionnaire vide

    # Ajouter le prix unitaire, le total de chaque ligne, et l'URL d'image si disponible
    articles_image_urls = {}
    for panier in paniers:
        panier.prix_unitaire = (
            panier.sous_total / panier.quantite if panier.quantite > 0 else 0
        )
        panier.total_ligne = panier.sous_total
        image_url = None
        try:
            # Prot√©ger l'acc√®s √† .url si aucun fichier n'est associ√©
            if getattr(panier.article, "image", None):
                # Certains backends l√®vent une exception si l'image n'existe pas
                if hasattr(panier.article.image, "url"):
                    image_url = panier.article.image.url
        except Exception:
            image_url = None
        # Rendre accessible directement dans le template via panier.article.image_url
        setattr(panier.article, "image_url", image_url)
        # Egalement disponible dans le context si n√©cessaire
        if getattr(panier.article, "id", None) is not None:
            articles_image_urls[panier.article.id] = image_url
    
    # R√©cup√©rer tous les √©tats de la commande pour afficher l'historique
    etats_commande = (
        commande.etats.all()
        .select_related("enum_etat", "operateur")
        .order_by("date_debut")
    )
    
    # D√©terminer l'√©tat actuel
    etat_actuel = etats_commande.filter(date_fin__isnull=True).first()
    
    # R√©cup√©rer l'√©tat pr√©c√©dent pour comprendre d'o√π vient la commande
    etat_precedent = None
    if etat_actuel:
        # Trouver l'√©tat pr√©c√©dent (le dernier √©tat termin√© avant l'√©tat actuel)
        for etat in reversed(etats_commande):
            if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                if etat.enum_etat.libelle not in [
                    "√Ä imprimer",
                    "En pr√©paration",
                    "Collect√©e",
                    "Emball√©e",
                ]:
                    etat_precedent = etat
                    break
    
    # Analyser les articles pour les commandes livr√©es partiellement
    articles_livres = []
    articles_renvoyes = []
    is_commande_livree_partiellement = False
    
    # Import pour JSON
    import json

    # R√©cup√©rer l'√©tat des articles renvoy√©s depuis l'op√©ration de livraison partielle (si elle existe)
    etat_articles_renvoyes = {}
    operation_livraison_partielle = None
    
    # Cas 1: La commande actuelle est la commande originale livr√©e partiellement
    if etat_actuel and etat_actuel.enum_etat.libelle == "Livr√©e Partiellement":
        is_commande_livree_partiellement = True
        # Les articles dans cette commande sont ceux qui ont √©t√© livr√©s partiellement
        for panier in paniers:
            articles_livres.append(
                {
                    "article": panier.article,
                    "quantite_livree": panier.quantite,
                    "prix": panier.article.prix_unitaire,
                    "sous_total": panier.sous_total,
                }
            )
        
        # Chercher la commande de renvoi associ√©e
        commande_renvoi = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande.num_cmd}", client=commande.client
        ).first()
        
        # La commande source pour les articles renvoy√©s est la commande actuelle
        operation_livraison_partielle = (
            commande.operations.filter(type_operation="LIVRAISON_PARTIELLE")
            .order_by("-date_operation")
            .first()
        )

    # Cas 2: La commande actuelle est une commande de renvoi suite √† une livraison partielle
    elif etat_precedent and etat_precedent.enum_etat.libelle == "Livr√©e Partiellement":
        is_commande_livree_partiellement = True
        # Chercher la commande originale qui a √©t√© livr√©e partiellement
        commande_originale = Commande.objects.filter(
            num_cmd=commande.num_cmd.replace("RENVOI-", ""), client=commande.client
        ).first()
        
        # La commande source pour les articles renvoy√©s est la commande originale
        if commande_originale:
            operation_livraison_partielle = (
                commande_originale.operations.filter(
                    type_operation="LIVRAISON_PARTIELLE"
                )
                .order_by("-date_operation")
                .first()
            )

    # Si une op√©ration de livraison partielle est trouv√©e, extraire les √©tats des articles renvoy√©s
    if operation_livraison_partielle:
        try:
            details = json.loads(operation_livraison_partielle.conclusion)
            if "recap_articles_renvoyes" in details:
                for item in details["recap_articles_renvoyes"]:
                    etat_articles_renvoyes[item["article_id"]] = item["etat"]
        except Exception:
            pass

    # Populer articles_renvoyes si c'est une commande de renvoi ou si elle a une commande de renvoi associ√©e
    if is_commande_livree_partiellement:
        # Si la commande actuelle est une commande de renvoi (Cas 2)
        if commande.num_cmd and commande.num_cmd.startswith("RENVOI-"):
            for panier_renvoi in paniers:
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id, "bon")
                articles_renvoyes.append(
                    {
                        "article": panier_renvoi.article,
                        "quantite": panier_renvoi.quantite,
                        "prix": panier_renvoi.article.prix_unitaire,
                        "sous_total": panier_renvoi.sous_total,
                        "etat": etat,
                    }
                )
        # Si la commande actuelle est la commande originale livr√©e partiellement et qu'une commande de renvoi existe (Cas 1)
        elif commande_renvoi:
            for panier_renvoi in commande_renvoi.paniers.all():
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id, "bon")
                articles_renvoyes.append(
                    {
                        "article": panier_renvoi.article,
                        "quantite": panier_renvoi.quantite,
                        "prix": panier_renvoi.article.prix_unitaire,
                        "sous_total": panier_renvoi.sous_total,
                        "etat": etat,
                    }
                )

    # Pour les articles livr√©s, on lit l'op√©ration de livraison partielle sur la commande originale
    if is_commande_livree_partiellement and commande_originale:
        operation_livraison_partielle_for_livres = (
            commande_originale.operations.filter(type_operation="LIVRAISON_PARTIELLE")
            .order_by("-date_operation")
            .first()
        )
        if operation_livraison_partielle_for_livres:
            try:
                details = json.loads(
                    operation_livraison_partielle_for_livres.conclusion
                )
                if "articles_livres" in details:
                    for article_livre in details["articles_livres"]:
                        article = Article.objects.filter(
                            id=article_livre.get("article_id")
                        ).first()
                        if article:
                            articles_livres.append(
                                {
                                    "article": article,
                                    "quantite_livree": article_livre.get("quantite", 0),
                                    "prix": article.prix_unitaire,
                                    "sous_total": article.prix_unitaire
                                    * article_livre.get("quantite", 0),
                                }
                            )
            except Exception:
                pass
    
    # Calculer le total des articles
    total_articles = sum(panier.total_ligne for panier in paniers)
    
    # R√©cup√©rer les op√©rations associ√©es √† la commande
    operations = commande.operations.select_related("operateur").order_by(
        "-date_operation"
    )
    
    # G√©n√©rer le code-barres pour la commande
    code128 = barcode.get_barcode_class("code128")
    barcode_instance = code128(str(commande.id_yz), writer=ImageWriter())
    buffer = BytesIO()
    barcode_instance.write(buffer, options={"write_text": False, "module_height": 10.0})
    barcode_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    commande_barcode = f"data:image/png;base64,{barcode_base64}"

    # Gestion des actions POST (marquer comme pr√©par√©e, etc.)
    if request.method == "POST":
        action = request.POST.get("action")
        
        # Action 'commencer_preparation' supprim√©e car les commandes passent maintenant 
        # directement en "En pr√©paration" lors de l'affectation
        
        if action == "marquer_collectee":
            with transaction.atomic():
                # Marquer l'√©tat 'En pr√©paration' comme termin√©
                etat_en_preparation, created = EnumEtatCmd.objects.get_or_create(
                    libelle="En pr√©paration"
                )
                
                etat_actuel = EtatCommande.objects.filter(
                    commande=commande,
                    enum_etat=etat_en_preparation,
                    date_fin__isnull=True,
                ).first()
                
                if etat_actuel:
                    etat_actuel.date_fin = timezone.now()
                    etat_actuel.operateur = operateur_profile
                    etat_actuel.save()
                
                # Cr√©er le nouvel √©tat 'Collect√©e'
                etat_collectee, created = EnumEtatCmd.objects.get_or_create(
                    libelle="Collect√©e"
                )
                EtatCommande.objects.create(
                    commande=commande,
                    enum_etat=etat_collectee,
                    operateur=operateur_profile,
                )
                
                # Log de l'op√©ration
                Operation.objects.create(
                    commande=commande,
                    type_operation="ARTICLES_COLLECTES",
                    operateur=operateur_profile,
                    conclusion=f"Articles collect√©s par {operateur_profile.nom_complet}.",
                )

            messages.success(
                request,
                f"La commande {commande.id_yz} a bien √©t√© marqu√©e comme collect√©e.",
            )
            return redirect("Prepacommande:detail_prepa", pk=commande.pk)

        elif action == "marquer_emballee":
            with transaction.atomic():
                # Marquer l'√©tat 'Collect√©e' comme termin√©
                etat_collectee, created = EnumEtatCmd.objects.get_or_create(
                    libelle="Collect√©e"
                )

                etat_actuel = EtatCommande.objects.filter(
                    commande=commande, enum_etat=etat_collectee, date_fin__isnull=True
                ).first()

                if etat_actuel:
                    etat_actuel.date_fin = timezone.now()
                    etat_actuel.operateur = operateur_profile
                    etat_actuel.save()

                # Cr√©er le nouvel √©tat 'Emball√©e'
                etat_emballee, created = EnumEtatCmd.objects.get_or_create(
                    libelle="Emball√©e"
                )
                EtatCommande.objects.create(
                    commande=commande,
                    enum_etat=etat_emballee,
                    operateur=operateur_profile,
                )

                # Log de l'op√©ration
                Operation.objects.create(
                    commande=commande,
                    type_operation="COMMANDE_EMBALLEE",
                    operateur=operateur_profile,
                    conclusion=f"Commande emball√©e par {operateur_profile.nom_complet}. Notification envoy√©e au superviseur.",
                )

                # TODO: Ajouter ici la notification au superviseur
                # Pour l'instant, on peut utiliser les messages Django ou cr√©er un syst√®me de notification

            messages.success(
                request,
                f"La commande {commande.id_yz} a bien √©t√© marqu√©e comme emball√©e. Le superviseur a √©t√© notifi√©.",
            )
            return redirect("Prepacommande:detail_prepa", pk=commande.pk)

        elif action == "signaler_probleme":
            with transaction.atomic():
                # 1. Terminer l'√©tat "En pr√©paration" actuel
                etat_en_preparation_enum = get_object_or_404(
                    EnumEtatCmd, libelle="En pr√©paration"
                )
                etat_actuel = EtatCommande.objects.filter(
                    commande=commande,
                    enum_etat=etat_en_preparation_enum,
                    date_fin__isnull=True,
                ).first()

                if etat_actuel:
                    etat_actuel.date_fin = timezone.now()
                    etat_actuel.commentaire = "Probl√®me signal√© par le pr√©parateur."
                    etat_actuel.save()

                # 2. Trouver l'op√©rateur de confirmation d'origine
                operateur_confirmation_origine = None
                etats_precedents = commande.etats.select_related("operateur").order_by(
                    "-date_debut"
                )
                
                for etat in etats_precedents:
                    if etat.operateur and etat.operateur.is_confirmation:
                        operateur_confirmation_origine = etat.operateur
                        break
                
                # 3. Cr√©er l'√©tat "Retour Confirmation" et l'affecter
                etat_retour_enum, _ = EnumEtatCmd.objects.get_or_create(
                    libelle="Retour Confirmation",
                    defaults={"ordre": 25, "couleur": "#D97706"},
                )
                
                EtatCommande.objects.create(
                    commande=commande,
                    enum_etat=etat_retour_enum,
                    operateur=operateur_confirmation_origine,  # Affectation directe
                    date_debut=timezone.now(),
                    commentaire="Retourn√© par la pr√©paration pour v√©rification.",
                )

                # 4. Log et message de succ√®s
                if operateur_confirmation_origine:
                    log_conclusion = f"Probl√®me signal√© par {operateur_profile.nom_complet}. Commande retourn√©e et affect√©e √† l'op√©rateur {operateur_confirmation_origine.nom_complet}."
                    messages.success(
                        request,
                        f"La commande {commande.id_yz} a √©t√© retourn√©e √† {operateur_confirmation_origine.nom_complet} pour v√©rification.",
                    )
                else:
                    log_conclusion = f"Probl√®me signal√© par {operateur_profile.nom_complet}. Op√©rateur d'origine non trouv√©, commande renvoy√©e au pool de confirmation."
                    messages.warning(
                        request,
                        f"La commande {commande.id_yz} a √©t√© renvoy√©e au pool de confirmation (op√©rateur d'origine non trouv√©).",
                    )

                Operation.objects.create(
                    commande=commande,
                    type_operation="PROBLEME_SIGNAL√â",
                    operateur=operateur_profile,
                    conclusion=log_conclusion,
                )

            return redirect("Prepacommande:liste_prepa")
    
    # Avant le return render dans detail_prepa
    commande_renvoi_id = None
    if commande_renvoi:
        commande_renvoi_id = commande_renvoi.id
    
    context = {
        "page_title": f"Pr√©paration Commande {commande.id_yz}",
        "page_subtitle": f"D√©tails de la commande et √©tapes de pr√©paration",
        "commande": commande,
        "paniers": paniers,
        "etats_commande": etats_commande,
        "etat_actuel": etat_actuel,
        "etat_precedent": etat_precedent,
        "etat_preparation": etat_preparation,
        "total_articles": total_articles,
        "operations": operations,
        "commande_barcode": commande_barcode,
        "is_commande_livree_partiellement": is_commande_livree_partiellement,
        "articles_livres": articles_livres,
        "articles_renvoyes": articles_renvoyes,
        "articles_image_urls": articles_image_urls,
        # Variables de debug/informations suppl√©mentaires
        "commande_originale": commande_originale,
        "commande_renvoi": commande_renvoi,
        "etat_articles_renvoyes": etat_articles_renvoyes,
        "commande_renvoi_id": commande_renvoi_id,
    }
    return render(request, "Prepacommande/detail_prepa.html", context)


@login_required
def api_commande_produits(request, commande_id):
    """API pour r√©cup√©rer les produits d'une commande pour les √©tiquettes"""
    try:
        # R√©cup√©rer la commande. La s√©curit√© est d√©j√† g√©r√©e par la page
        # qui appelle cette API, qui ne liste que les commandes autoris√©es.
        commande = Commande.objects.get(id=commande_id)
        
        # R√©cup√©rer tous les produits de la commande
        paniers = commande.paniers.all().select_related("article")
        
        # Construire la liste des produits
        produits_list = []
        for panier in paniers:
            if panier.article:
                # Format: "NOM REFERENCE , POINTURE"
                produit_info = f"{panier.article.nom or ''} {panier.article.reference or ''}".strip()
                if panier.article.pointure:
                    produit_info += f" , {panier.article.pointure}"
                
                # Ajouter la quantit√© si elle est sup√©rieure √† 1
                if panier.quantite > 1:
                    produit_info += (
                        f" (x{panier.quantite})"  # Mettre la quantit√© entre parenth√®ses
                    )
                produits_list.append(produit_info)
        
        # Joindre tous les produits en une seule cha√Æne, en utilisant " + " comme s√©parateur
        produits_text = (
            " + ".join(produits_list) if produits_list else "PRODUITS NON SP√âCIFI√âS"
        )

        return JsonResponse(
            {
                "success": True,
                "produits": produits_text,
                "nombre_articles": len(produits_list),
            }
        )
        
    except Commande.DoesNotExist:
        return JsonResponse({"success": False, "message": "Commande non trouv√©e"})
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Erreur: {str(e)}"})


# API api_changer_etat_preparation supprim√©e car les commandes passent maintenant 
# directement de "Confirm√©e" √† "En pr√©paration" lors de l'affectation


@login_required
@handle_ajax_errors
def modifier_commande_prepa(request, commande_id):
    """Page de modification compl√®te d'une commande pour les op√©rateurs de pr√©paration"""
    print(f"üîç ===== D√âBUT modifier_commande_prepa =====")
    print(f"üîç M√©thode: {request.method}")
    print(f"üîç Commande ID: {commande_id}")
    print(f"üîç URL: {request.path}")
    
    import json
    from commande.models import Commande, Operation
    from parametre.models import Ville
    
    print(f"üîç R√©cup√©ration de l'op√©rateur")
    # R√©cup√©rer l'op√©rateur
    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
        print(f"‚úÖ Op√©rateur trouv√©: {operateur.user.username}")
    except Operateur.DoesNotExist:
        print(f"‚ùå Op√©rateur non trouv√© pour l'utilisateur: {request.user.username}")
        messages.error(request, "Profil d'op√©rateur de pr√©paration non trouv√©.")
        return redirect("login")
    
    print(f"üîç R√©cup√©ration de la commande: {commande_id}")
    # R√©cup√©rer la commande
    commande = get_object_or_404(Commande, id=commande_id)
    print(f"‚úÖ Commande trouv√©e: {commande.id_yz}")
    
    print(f"üîç V√©rification de l'√©tat de pr√©paration")
    # V√©rifier que la commande est affect√©e √† cet op√©rateur pour la pr√©paration
    etat_preparation = commande.etats.filter(
        Q(enum_etat__libelle="√Ä imprimer") | Q(enum_etat__libelle="En pr√©paration"),
        operateur=operateur,
        date_fin__isnull=True,
    ).first()
    
    print(f"üîç √âtat de pr√©paration trouv√©: {etat_preparation}")
    
    if not etat_preparation:
        print(f"‚ùå Commande non affect√©e √† l'op√©rateur pour la pr√©paration")
        messages.error(
            request, "Cette commande ne vous est pas affect√©e pour la pr√©paration."
        )
        return redirect("Prepacommande:liste_prepa")
    
    print(f"‚úÖ Commande affect√©e √† l'op√©rateur pour la pr√©paration")

    if request.method == "POST":
        print(f"üîç Requ√™te POST re√ßue")
        print(f"üîç POST data: {dict(request.POST)}")
        try:
            # ================ GESTION DES ACTIONS AJAX SP√âCIFIQUES ================
            action = request.POST.get("action")
            
            # V√©rifier si c'est une requ√™te AJAX
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json'
            
            # Forcer le traitement AJAX pour les actions sp√©cifiques
            if action in ['add_article', 'delete_article', 'update_quantity', 'replace_article', 'update_operation', 'add_operation', 'update_commande_info']:
                is_ajax = True
                print(f"üîÑ Traitement AJAX forc√© pour l'action: {action}")
            
            print(f"üîç Action d√©tect√©e: {action}")
            print(f"üîç Requ√™te AJAX: {is_ajax}")
            print(f"üîç Content-Type: {request.content_type}")
            print(f"üîç Headers: {dict(request.headers)}")
            
            # Si c'est une action AJAX mais pas d√©tect√©e comme telle, forcer le traitement
            if action and action in ['add_article', 'delete_article', 'update_quantity', 'replace_article', 'update_operation', 'add_operation', 'update_commande_info'] and not is_ajax:
                print(f"‚ö†Ô∏è Action AJAX non d√©tect√©e, for√ßage du traitement AJAX pour: {action}")
                is_ajax = True
            
            if action == "add_article":
                print(f"üîÑ Traitement de l'action add_article (AJAX: {is_ajax})")
                print(f"üîç D√©but du traitement add_article")
                
                # Ajouter un nouvel article imm√©diatement
                from article.models import Article, Variante
                from commande.models import Panier
                
                # Support both parameter names for backward compatibility
                article_id = request.POST.get("articleId") or request.POST.get("article_id")
                quantite = int(request.POST.get("quantite", 1))
                variante_id = request.POST.get("varianteId") or request.POST.get("variante_id")
                
                print(f"üîç Param√®tres re√ßus: article_id={article_id}, quantite={quantite}, variante_id={variante_id}")
                print(f"üîç Type des param√®tres: article_id={type(article_id)}, quantite={type(quantite)}, variante_id={type(variante_id)}")
                
                try:
                    print(f"üîç D√©but du bloc try pour add_article")
                    print(f"üîç V√©rification des variantes et articles")
                    # V√©rifier si une variante sp√©cifique a √©t√© s√©lectionn√©e
                    if variante_id:
                        print(f"üîç Recherche de variante sp√©cifique: {variante_id}")
                        try:
                            variante = Variante.objects.get(id=variante_id)
                            article = variante.article  # Utiliser l'article parent de la variante
                            print(f"üîÑ Variante sp√©cifique trouv√©e: ID={variante.id}, Article parent: {article.nom}")
                        except Variante.DoesNotExist:
                            print(f"‚ùå Variante sp√©cifi√©e non trouv√©e: {variante_id}")
                            return JsonResponse({"success": False, "error": "Variante non trouv√©e"})
                    else:
                        print(f"üîç Pas de variante sp√©cifi√©e, recherche d'article: {article_id}")
                        # V√©rifier si l'article_id est une variante ou un article
                        try:
                            # Essayer de trouver une variante d'abord
                            variante = Variante.objects.get(id=article_id)
                            article = variante.article  # Utiliser l'article parent de la variante
                            print(f"üîÑ Variante trouv√©e: ID={variante.id}, Article parent: {article.nom}")
                        except Variante.DoesNotExist:
                            print(f"üîç Pas une variante, recherche d'article normal: {article_id}")
                            # Si ce n'est pas une variante, c'est un article normal
                            article = Article.objects.get(id=article_id)
                            variante = None
                            print(f"üîÑ Article normal trouv√©: ID={article.id}, Nom: {article.nom}")
                    
                    print(f"üîç V√©rification du panier existant")
                    # V√©rifier si l'article existe d√©j√† dans la commande
                    panier_existant = Panier.objects.filter(
                        commande=commande, article=article
                    ).first()
                    
                    print(f"üîç Panier existant trouv√©: {panier_existant is not None}")
                    
                    if panier_existant:
                        print(f"üîç Mise √† jour du panier existant")
                        # Si l'article existe d√©j√†, mettre √† jour la quantit√©
                        panier_existant.quantite += quantite
                        panier_existant.save()
                        panier = panier_existant
                        print(
                            f"üîÑ Article existant mis √† jour: ID={article.id}, nouvelle quantit√©={panier.quantite}"
                        )
                    else:
                        print(f"üîç Cr√©ation d'un nouveau panier")
                        # Si l'article n'existe pas, cr√©er un nouveau panier
                        panier = Panier.objects.create(
                            commande=commande,
                            article=article,
                            quantite=quantite,
                            sous_total=0,  # Sera recalcul√© apr√®s
                        )
                        print(
                            f"‚ûï Nouvel article ajout√©: ID={article.id}, quantit√©={quantite}"
                        )
                    
                    print(f"üîç Gestion des variantes et calcul des prix")
                    # Si c'√©tait une variante, stocker l'information de la variante
                    if variante:
                        print(f"üîç Stockage des informations de variante")
                        # Vous pouvez ajouter un champ personnalis√© au panier pour stocker la variante
                        # Ou utiliser un syst√®me de commentaires pour stocker cette information
                        panier.commentaire = f"Variante s√©lectionn√©e: {variante.couleur} - {variante.pointure}"
                        panier.save()
                        print(f"üìù Variante stock√©e: {variante.couleur} - {variante.pointure}")
                    
                    print(f"üîç V√©rification si article est upsell")
                    # Recalculer le compteur apr√®s ajout (logique de confirmation)
                    if (
                        article.isUpsell
                        and hasattr(article, "prix_upsell_1")
                        and article.prix_upsell_1 is not None
                    ):
                        print(f"üîç Calcul du compteur upsell")
                        # Compter la quantit√© totale d'articles upsell (apr√®s ajout)
                        total_quantite_upsell = (
                            commande.paniers.filter(article__isUpsell=True).aggregate(
                                total=Sum("quantite")
                            )["total"]
                            or 0
                        )
                        
                        print(f"üîç Total quantit√© upsell: {total_quantite_upsell}")
                        
                        # Le compteur ne s'incr√©mente qu'√† partir de 2 unit√©s d'articles upsell
                        # 0-1 unit√©s upsell ‚Üí compteur = 0
                        # 2+ unit√©s upsell ‚Üí compteur = total_quantite_upsell - 1
                        if total_quantite_upsell >= 2:
                            commande.compteur = total_quantite_upsell - 1
                        else:
                            commande.compteur = 0
                        
                        print(f"üîç Nouveau compteur: {commande.compteur}")
                        commande.save()
                        
                        print(f"üîç Recalcul des totaux upsell")
                        # Recalculer TOUS les articles de la commande avec le nouveau compteur
                        commande.recalculer_totaux_upsell()
                    else:
                        print(f"üîç Calcul des prix pour article normal")
                        # Pour les articles normaux, juste calculer le sous-total
                        from commande.templatetags.commande_filters import (
                            get_prix_upsell_avec_compteur,
                        )

                        print(f"üîç R√©cup√©ration du prix unitaire")
                        prix_unitaire = get_prix_upsell_avec_compteur(
                            article, commande.compteur
                        )
                        print(f"üîç Prix unitaire: {prix_unitaire}")
                        sous_total = prix_unitaire * panier.quantite
                        print(f"üîç Sous-total: {sous_total}")
                        panier.sous_total = float(sous_total)
                        panier.save()
                        print(f"üîç Panier sauvegard√© avec sous-total: {panier.sous_total}")
                    
                    # Recalculer le total de la commande avec frais de livraison
                    total_articles = (
                        commande.paniers.aggregate(total=Sum("sous_total"))["total"]
                        or 0
                    )
                    frais_livraison = (
                        commande.ville.frais_livraison if commande.ville else 0
                    )
                    commande.total_cmd = float(
                        total_articles
                    )  # + float(frais_livraison)
                    commande.save()
                    
                    # Calculer les statistiques upsell pour la r√©ponse
                    articles_upsell = commande.paniers.filter(article__isUpsell=True)
                    total_quantite_upsell = (
                        articles_upsell.aggregate(total=Sum("quantite"))["total"] or 0
                    )
                    
                    # D√©terminer si c'√©tait un ajout ou une mise √† jour
                    message = (
                        "Article ajout√© avec succ√®s"
                        if not panier_existant
                        else f"Quantit√© mise √† jour ({panier.quantite})"
                    )
                    
                    # Pr√©parer les donn√©es de l'article pour le frontend
                    article_data = {
                        "panier_id": panier.id,
                        "nom": article.nom,
                        "reference": article.reference,
                        "couleur_fr": variante.couleur if variante else (article.couleur or ""),
                        "couleur_ar": variante.couleur if variante else (article.couleur or ""),
                        "pointure": variante.pointure if variante else (article.pointure or ""),
                        "quantite": panier.quantite,
                        "prix": panier.sous_total / panier.quantite,  # Prix unitaire
                        "sous_total": panier.sous_total,
                        "is_upsell": article.isUpsell,
                        "isUpsell": article.isUpsell,
                        "phase": article.phase,
                        "qte_disponible": article.get_total_qte_disponible(),
                        "has_promo_active": article.has_promo_active if hasattr(article, 'has_promo_active') else False,
                        "description": article.description or "",
                        "variante_info": f"{variante.couleur} - {variante.pointure}" if variante else None,
                    }

                    print(f"üîç Pr√©paration de la r√©ponse JSON")
                    print(f"‚úÖ Action add_article termin√©e avec succ√®s, retour de la r√©ponse JSON")
                    return JsonResponse(
                        {
                            "success": True,
                            "message": message,
                            "article_id": panier.id,
                            "total_commande": float(commande.total_cmd),
                            "nb_articles": commande.paniers.count(),
                            "compteur": commande.compteur,
                            "was_update": panier_existant is not None,
                            "new_quantity": panier.quantite,
                            "article_data": article_data,
                            "articles_count": commande.paniers.count(),
                            "sous_total_articles": float(
                                sum(p.sous_total for p in commande.paniers.all())
                            ),
                            "articles_upsell": articles_upsell.count(),
                            "quantite_totale_upsell": total_quantite_upsell,
                        }
                    )
                    
                except Article.DoesNotExist:
                    print(f"‚ùå Article non trouv√©: {article_id}")
                    return JsonResponse(
                        {"success": False, "error": "Article non trouv√©"}
                    )
                except (ValueError, TypeError) as e:
                    print(f"‚ùå Erreur de type dans l'action add_article: {str(e)}")
                    return JsonResponse({"success": False, "error": f"Erreur de type: {str(e)}"})
                except Exception as e:
                    print(f"‚ùå Erreur g√©n√©rale dans l'action add_article: {str(e)}")
                    import traceback
                    print(f"‚ùå Traceback: {traceback.format_exc()}")
                    return JsonResponse({"success": False, "error": f"Erreur lors de l'ajout: {str(e)}"})
            
            elif action == "replace_article":
                # Remplacer un article existant
                from article.models import Article
                from commande.models import Panier
                
                ancien_article_id = request.POST.get("ancien_article_id")
                nouvel_article_id = request.POST.get("nouvel_article_id")
                nouvelle_quantite = int(request.POST.get("nouvelle_quantite", 1))
                
                try:
                    # Supprimer l'ancien panier et d√©cr√©menter le compteur
                    ancien_panier = Panier.objects.get(
                        id=ancien_article_id, commande=commande
                    )
                    ancien_article = ancien_panier.article
                    
                    # Sauvegarder les infos avant suppression
                    ancien_etait_upsell = ancien_article.isUpsell
                    
                    # Supprimer l'ancien panier
                    ancien_panier.delete()
                    
                    # Cr√©er le nouveau panier
                    nouvel_article = Article.objects.get(id=nouvel_article_id)
                    
                    # Recalculer le compteur apr√®s remplacement (logique de confirmation)
                    total_quantite_upsell = (
                        commande.paniers.filter(article__isUpsell=True).aggregate(
                            total=Sum("quantite")
                        )["total"]
                        or 0
                    )
                    
                    # Ajouter la quantit√© si le nouvel article est upsell
                    if nouvel_article.isUpsell:
                        total_quantite_upsell += nouvelle_quantite
                    
                    # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                    if total_quantite_upsell >= 2:
                        commande.compteur = total_quantite_upsell - 1
                    else:
                        commande.compteur = 0
                    
                    commande.save()
                    
                    # Recalculer TOUS les articles de la commande avec le nouveau compteur
                    commande.recalculer_totaux_upsell()
                    
                    # Calculer le sous-total selon le compteur de la commande
                    from commande.templatetags.commande_filters import (
                        get_prix_upsell_avec_compteur,
                    )

                    prix_unitaire = get_prix_upsell_avec_compteur(
                        nouvel_article, commande.compteur
                    )
                    sous_total = prix_unitaire * nouvelle_quantite
                    
                    nouveau_panier = Panier.objects.create(
                        commande=commande,
                        article=nouvel_article,
                        quantite=nouvelle_quantite,
                        sous_total=float(sous_total),
                    )
                    
                    # Recalculer le total de la commande avec frais de livraison
                    total_commande = (
                        commande.paniers.aggregate(total=Sum("sous_total"))["total"]
                        or 0
                    )
                    frais_livraison = (
                        commande.ville.frais_livraison if commande.ville else 0
                    )
                    commande.total_cmd = float(
                        total_commande
                    )  # + float(frais_livraison)
                    commande.save()
                    
                    # Calculer les statistiques upsell pour la r√©ponse
                    articles_upsell = commande.paniers.filter(article__isUpsell=True)
                    total_quantite_upsell = (
                        articles_upsell.aggregate(total=Sum("quantite"))["total"] or 0
                    )

                    return JsonResponse(
                        {
                            "success": True,
                            "message": "Article remplac√© avec succ√®s",
                            "nouvel_article_id": nouveau_panier.id,
                            "total_commande": float(commande.total_cmd),
                            "nb_articles": commande.paniers.count(),
                            "compteur": commande.compteur,
                            "articles_upsell": articles_upsell.count(),
                            "quantite_totale_upsell": total_quantite_upsell,
                            "sous_total_articles": float(commande.sous_total_articles),
                        }
                    )
                    
                except Panier.DoesNotExist:
                    return JsonResponse(
                        {"success": False, "error": "Article original non trouv√©"}
                    )
                except Article.DoesNotExist:
                    return JsonResponse(
                        {"success": False, "error": "Nouvel article non trouv√©"}
                    )
                except Exception as e:
                    return JsonResponse({"success": False, "error": str(e)})
            
            elif action == "delete_article":
                # Supprimer un article
                from commande.models import Panier
                
                panier_id = request.POST.get("panier_id")
                
                try:
                    panier = Panier.objects.get(id=panier_id, commande=commande)
                    
                    # Sauvegarder l'info avant suppression
                    etait_upsell = panier.article.isUpsell
                    
                    # Supprimer l'article
                    panier.delete()
                    
                    # Recalculer le compteur apr√®s suppression (logique de confirmation)
                    if etait_upsell:
                        # Compter la quantit√© totale d'articles upsell restants (apr√®s suppression)
                        total_quantite_upsell = (
                            commande.paniers.filter(article__isUpsell=True).aggregate(
                                total=Sum("quantite")
                            )["total"]
                            or 0
                        )
                        
                        # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                        if total_quantite_upsell >= 2:
                            commande.compteur = total_quantite_upsell - 1
                        else:
                            commande.compteur = 0
                        
                        commande.save()
                        
                        # Recalculer TOUS les articles de la commande avec le nouveau compteur
                        commande.recalculer_totaux_upsell()
                    
                    # Recalculer le total de la commande avec frais de livraison
                    total_commande = (
                        commande.paniers.aggregate(total=Sum("sous_total"))["total"]
                        or 0
                    )
                    frais_livraison = (
                        commande.ville.frais_livraison if commande.ville else 0
                    )
                    commande.total_cmd = float(
                        total_commande
                    )  # + float(frais_livraison)
                    commande.save()
                    
                    # Calculer les statistiques upsell pour la r√©ponse
                    articles_upsell = commande.paniers.filter(article__isUpsell=True)
                    total_quantite_upsell = (
                        articles_upsell.aggregate(total=Sum("quantite"))["total"] or 0
                    )

                    return JsonResponse(
                        {
                            "success": True,
                            "message": "Article supprim√© avec succ√®s",
                            "total_commande": float(commande.total_cmd),
                            "nb_articles": commande.paniers.count(),
                            "compteur": commande.compteur,
                            "articles_upsell": articles_upsell.count(),
                            "quantite_totale_upsell": total_quantite_upsell,
                            "sous_total_articles": float(commande.sous_total_articles),
                        }
                    )
                    
                except Panier.DoesNotExist:
                    return JsonResponse(
                        {"success": False, "error": "Article non trouv√©"}
                    )
                except Exception as e:
                    return JsonResponse({"success": False, "error": str(e)})
            
            elif action == "update_operation":
                # Mettre √† jour une op√©ration existante
                try:
                    from commande.models import Operation
                    import logging

                    logger = logging.getLogger(__name__)
                    
                    operation_id = request.POST.get("operation_id")
                    nouveau_commentaire = request.POST.get(
                        "nouveau_commentaire", ""
                    ).strip()
                    
                    if not operation_id or not nouveau_commentaire:
                        return JsonResponse(
                            {
                                "success": False,
                                "error": "ID op√©ration et commentaire requis",
                            }
                        )
                    
                    # R√©cup√©rer et mettre √† jour l'op√©ration
                    operation = Operation.objects.get(
                        id=operation_id, commande=commande
                    )
                    operation.conclusion = nouveau_commentaire
                    operation.operateur = (
                        operateur  # Mettre √† jour l'op√©rateur qui modifie
                    )
                    operation.save()
                    
                    return JsonResponse(
                        {
                            "success": True,
                            "message": "Op√©ration mise √† jour avec succ√®s",
                            "operation_id": operation_id,
                            "nouveau_commentaire": nouveau_commentaire,
                        }
                    )
                    
                except Operation.DoesNotExist:
                    return JsonResponse(
                        {"success": False, "error": "Op√©ration non trouv√©e"}
                    )
                except Exception as e:
                    return JsonResponse({"success": False, "error": str(e)})
            
            elif action == "add_operation":
                # Ajouter une nouvelle op√©ration
                try:
                    from commande.models import Operation
                    
                    type_operation = request.POST.get("type_operation", "").strip()
                    commentaire = request.POST.get("commentaire", "").strip()
                    
                    if not type_operation or not commentaire:
                        return JsonResponse(
                            {
                                "success": False,
                                "error": "Type d'op√©ration et commentaire requis",
                            }
                        )
                    
                    # Cr√©er la nouvelle op√©ration
                    operation = Operation.objects.create(
                        commande=commande,
                        type_operation=type_operation,
                        conclusion=commentaire,
                        operateur=operateur,
                    )

                    return JsonResponse(
                        {
                            "success": True,
                            "message": "Op√©ration ajout√©e avec succ√®s",
                            "operation_id": operation.id,
                            "type_operation": type_operation,
                            "commentaire": commentaire,
                        }
                    )
                    
                except Exception as e:
                    return JsonResponse({"success": False, "error": str(e)})
            
            elif action == "modifier_quantites_multiple":
                # Modifier plusieurs quantit√©s d'articles en une fois
                try:
                    from commande.models import Panier
                    import json
                    
                    modifications_json = request.POST.get("modifications", "[]")
                    modifications = json.loads(modifications_json)
                    
                    if not modifications:
                        return JsonResponse(
                            {"success": False, "error": "Aucune modification fournie"}
                        )
                    
                    # Appliquer les modifications
                    for mod in modifications:
                        panier_id = mod.get("panier_id")
                        nouvelle_quantite = mod.get("nouvelle_quantite", 0)
                        
                        try:
                            panier = Panier.objects.get(id=panier_id, commande=commande)
                            
                            if nouvelle_quantite <= 0:
                                # Supprimer l'article si quantit√© = 0
                                panier.delete()
                            else:
                                # Mettre √† jour la quantit√© et le sous-total
                                panier.quantite = nouvelle_quantite
                                panier.sous_total = float(
                                    panier.article.prix_unitaire * nouvelle_quantite
                                )
                                panier.save()
                                
                        except Panier.DoesNotExist:
                            continue  # Ignorer les paniers non trouv√©s
                    
                    # Recalculer le total de la commande
                    total_commande = (
                        commande.paniers.aggregate(total=Sum("sous_total"))["total"]
                        or 0
                    )
                    commande.total_cmd = float(total_commande)
                    commande.save()
                    
                    # Cr√©er une op√©ration pour consigner la modification
                    Operation.objects.create(
                        commande=commande,
                        type_operation="MODIFICATION_QUANTITES",
                        conclusion=f"Modification en masse des quantit√©s d'articles par l'op√©rateur de pr√©paration.",
                        operateur=operateur,
                    )

                    return JsonResponse(
                        {
                            "success": True,
                            "message": f"{len(modifications)} quantit√©(s) modifi√©e(s) avec succ√®s",
                            "total_commande": float(commande.total_cmd),
                            "nb_articles": commande.paniers.count(),
                        }
                    )
                    
                except json.JSONDecodeError:
                    return JsonResponse(
                        {"success": False, "error": "Format de donn√©es invalide"}
                    )
                except Exception as e:
                    return JsonResponse({"success": False, "error": str(e)})
            
            elif action == "modifier_quantite_directe":
                # Modifier directement la quantit√© d'un article
                try:
                    from commande.models import Panier
                    
                    panier_id = request.POST.get("panier_id")
                    nouvelle_quantite = int(request.POST.get("nouvelle_quantite", 0))
                    
                    print(
                        f"üîÑ Modification quantit√© directe - Panier ID: {panier_id}, Nouvelle quantit√©: {nouvelle_quantite}"
                    )
                    
                    if nouvelle_quantite < 0:
                        return JsonResponse(
                            {
                                "success": False,
                                "error": "La quantit√© ne peut pas √™tre n√©gative",
                            }
                        )
                    
                    try:
                        panier = Panier.objects.get(id=panier_id, commande=commande)
                        ancienne_quantite = panier.quantite
                        nouveau_sous_total = 0
                        
                        print(
                            f"üì¶ Article trouv√©: {panier.article.nom}, Ancienne quantit√©: {ancienne_quantite}"
                        )
                        
                        # V√©rifier que le panier appartient bien √† cette commande
                        if panier.commande.id != commande.id:
                            print(
                                f"‚ùå ERREUR: Le panier {panier_id} n'appartient pas √† la commande {commande.id}"
                            )
                            return JsonResponse(
                                {
                                    "success": False,
                                    "error": "Article non trouv√© dans cette commande",
                                }
                            )
                        
                        if nouvelle_quantite == 0:
                            # Supprimer l'article si quantit√© = 0
                            panier.delete()
                            message = "Article supprim√© avec succ√®s"
                        else:
                            # Mettre √† jour la quantit√© et le sous-total avec la logique compl√®te de prix
                            panier.quantite = nouvelle_quantite
                            
                            # Recalculer le compteur si c'est un article upsell
                            if panier.article.isUpsell:
                                # Sauvegarder d'abord la nouvelle quantit√©
                                panier.save()
                                
                                # Compter la quantit√© totale d'articles upsell apr√®s modification
                                total_quantite_upsell = (
                                    commande.paniers.filter(
                                        article__isUpsell=True
                                    ).aggregate(total=Sum("quantite"))["total"]
                                    or 0
                                )

                                print(
                                    f"üîÑ Calcul du compteur upsell: total_quantite_upsell = {total_quantite_upsell}"
                                )
                                
                                # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                                if total_quantite_upsell >= 2:
                                    commande.compteur = total_quantite_upsell - 1
                                else:
                                    commande.compteur = 0
                                
                                print(
                                    f"‚úÖ Nouveau compteur calcul√©: {commande.compteur}"
                                )
                                commande.save()
                                
                                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                                commande.recalculer_totaux_upsell()
                            else:
                                # Pour les articles normaux, mettre √† jour la quantit√© et calculer le sous-total
                                from commande.templatetags.commande_filters import (
                                    get_prix_upsell_avec_compteur,
                                )

                                prix_unitaire = get_prix_upsell_avec_compteur(
                                    panier.article, commande.compteur
                                )
                                panier.quantite = nouvelle_quantite
                                panier.sous_total = float(
                                    prix_unitaire * nouvelle_quantite
                                )
                                panier.save()
                            
                            nouveau_sous_total = panier.sous_total
                            message = "Quantit√© modifi√©e avec succ√®s"
                            
                            print(
                                f"‚úÖ Quantit√© mise √† jour: {ancienne_quantite} ‚Üí {nouvelle_quantite}, Nouveau sous-total: {nouveau_sous_total}"
                            )
                            
                            # V√©rifier que la mise √† jour a bien √©t√© sauvegard√©e
                            panier.refresh_from_db()
                            if panier.quantite != nouvelle_quantite:
                                print(
                                    f"‚ùå ERREUR: La quantit√© n'a pas √©t√© sauvegard√©e correctement. Attendu: {nouvelle_quantite}, Actuel: {panier.quantite}"
                                )
                            else:
                                print(
                                    f"‚úÖ V√©rification OK: Quantit√© sauvegard√©e: {panier.quantite}"
                                )
                            
                    except Panier.DoesNotExist:
                        return JsonResponse(
                            {"success": False, "error": "Article non trouv√©"}
                        )
                    
                    # Recalculer le total de la commande avec frais de livraison
                    total_articles = (
                        commande.paniers.aggregate(total=Sum("sous_total"))["total"]
                        or 0
                    )
                    frais_livraison = (
                        commande.ville.frais_livraison if commande.ville else 0
                    )
                    commande.total_cmd = float(
                        total_articles
                    )  # + float(frais_livraison)
                    commande.save()
                    
                    # Calculer les statistiques upsell pour la r√©ponse
                    articles_upsell = commande.paniers.filter(article__isUpsell=True)
                    total_quantite_upsell = (
                        articles_upsell.aggregate(total=Sum("quantite"))["total"] or 0
                    )
                    
                    # Cr√©er une op√©ration pour consigner la modification
                    Operation.objects.create(
                        commande=commande,
                        type_operation="MODIFICATION_QUANTITE",
                        conclusion=f"Quantit√© d'article modifi√©e de {ancienne_quantite} √† {nouvelle_quantite}.",
                        operateur=operateur,
                    )

                    return JsonResponse(
                        {
                            "success": True,
                            "message": message,
                            "sous_total": float(nouveau_sous_total),
                            "sous_total_articles": float(total_articles),
                            "total_commande": float(commande.total_cmd),
                            "frais_livraison": float(frais_livraison),
                            "nb_articles": commande.paniers.count(),
                            "compteur": commande.compteur,
                            "articles_upsell": articles_upsell.count(),
                            "quantite_totale_upsell": total_quantite_upsell,
                        }
                    )
                    
                except ValueError:
                    return JsonResponse(
                        {"success": False, "error": "Quantit√© invalide"}
                    )
                except Exception as e:
                    return JsonResponse({"success": False, "error": str(e)})
            
            elif action == "update_commande_info":
                # Mettre √† jour les informations de base de la commande
                try:
                    # R√©cup√©rer les donn√©es du formulaire
                    nouvelle_adresse = request.POST.get("adresse", "").strip()
                    nouvelle_ville_id = request.POST.get("ville_id")
                    
                    # Mettre √† jour l'adresse
                    if nouvelle_adresse:
                        commande.adresse = nouvelle_adresse
                    
                    # Mettre √† jour la ville si fournie
                    if nouvelle_ville_id:
                        try:
                            nouvelle_ville = Ville.objects.get(id=nouvelle_ville_id)
                            commande.ville = nouvelle_ville
                        except Ville.DoesNotExist:
                            return JsonResponse(
                                {"success": False, "error": "Ville non trouv√©e"}
                            )
                    
                    commande.save()
                    
                    # Cr√©er une op√©ration pour consigner la modification
                    Operation.objects.create(
                        commande=commande,
                        type_operation="MODIFICATION_PREPA",
                        conclusion=f"La commande a √©t√© modifi√©e par l'op√©rateur.",
                        operateur=operateur,
                    )

                    messages.success(
                        request, f"Commande {commande.id_yz} mise √† jour avec succ√®s."
                    )
                    return redirect("Prepacommande:detail_prepa", pk=commande.id)
                    
                except Exception as e:
                    return JsonResponse({"success": False, "error": str(e)})
            
            else:
                # Traitement du formulaire principal (non-AJAX)
                with transaction.atomic():
                    # Mettre √† jour les informations du client
                    client = commande.client
                    client.nom = request.POST.get("client_nom", client.nom).strip()
                    client.prenom = request.POST.get(
                        "client_prenom", client.prenom
                    ).strip()
                    client.numero_tel = request.POST.get(
                        "client_telephone", client.numero_tel
                    ).strip()
                    client.save()

                    # Mettre √† jour les informations de base de la commande
                    nouvelle_adresse = request.POST.get("adresse", "").strip()
                    nouvelle_ville_id = request.POST.get("ville_id")
                    
                    if nouvelle_adresse:
                        commande.adresse = nouvelle_adresse
                    
                    if nouvelle_ville_id:
                        try:
                            nouvelle_ville = Ville.objects.get(id=nouvelle_ville_id)
                            commande.ville = nouvelle_ville
                        except Ville.DoesNotExist:
                            messages.error(request, "Ville s√©lectionn√©e non trouv√©e.")
                            return redirect(
                                "Prepacommande:modifier_commande",
                                commande_id=commande.id,
                            )
                    
                    commande.save()

                    # Cr√©er une op√©ration pour consigner la modification
                    Operation.objects.create(
                        commande=commande,
                        type_operation="MODIFICATION_PREPA",
                        conclusion=f"La commande a √©t√© modifi√©e par l'op√©rateur.",
                        operateur=operateur,
                    )

                    messages.success(
                        request,
                        f"Les modifications de la commande {commande.id_yz} ont √©t√© enregistr√©es avec succ√®s.",
                    )
                    return redirect("Prepacommande:detail_prepa", pk=commande.id)
                
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification: {str(e)}")
            return redirect("Prepacommande:modifier_commande", commande_id=commande.id)
    
    # R√©cup√©rer les donn√©es pour l'affichage
    paniers = commande.paniers.all().select_related("article")
    operations = (
        commande.operations.all()
        .select_related("operateur")
        .order_by("-date_operation")
    )
    villes = Ville.objects.all().order_by("nom")
    
    # Calculer le total des articles
    total_articles = sum(panier.sous_total for panier in paniers)
    
    # V√©rifier si c'est une commande renvoy√©e par la logistique
    operation_renvoi = operations.filter(type_operation="RENVOI_PREPARATION").first()
    is_commande_renvoyee = operation_renvoi is not None
    
    # Initialiser les variables pour les cas de livraison partielle/renvoi
    articles_livres = []
    articles_renvoyes = []
    is_commande_livree_partiellement = False
    commande_renvoi_obj = None  # Variable pour la commande de renvoi trouv√©e
    commande_originale_obj = None  # Variable pour la commande originale trouv√©e
    etat_articles_renvoyes = (
        {}
    )  # Dictionnaire pour stocker l'√©tat des articles renvoy√©s (article_id -> etat)
    operation_livraison_partielle_source = (
        None  # Op√©ration source pour les d√©tails de livraison partielle
    )

    # R√©cup√©rer l'√©tat actuel de la commande
    etat_actuel = commande.etats.filter(date_fin__isnull=True).first()
    etat_precedent = None
    
    if etat_actuel:
        # Trouver l'√©tat pr√©c√©dent
        etats_precedents = commande.etats.all().order_by("-date_debut")
        for etat in etats_precedents:
            if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                if etat.enum_etat.libelle not in ["√Ä imprimer", "En pr√©paration"]:
                    etat_precedent = etat
                    break
    
    # NOUVELLE LOGIQUE POUR D√âTECTER LA LIVRAISON PARTIELLE ET LES ARTICLES RENVOY√âS
    # Une commande est consid√©r√©e comme "livr√©e partiellement" dans le contexte de modification
    # si elle-m√™me a √©t√© livr√©e partiellement ou si c'est une commande de RENVOI associ√©e √† une livraison partielle.
    
    if commande.num_cmd and commande.num_cmd.startswith("RENVOI-"):
        # C'est une commande de renvoi. On cherche la commande originale.
        num_cmd_original = commande.num_cmd.replace("RENVOI-", "")
        commande_originale_obj = Commande.objects.filter(
            num_cmd=num_cmd_original, client=commande.client
        ).first()
        
        if commande_originale_obj:
            # V√©rifier si la commande originale a bien √©t√© livr√©e partiellement
            if commande_originale_obj.etats.filter(
                enum_etat__libelle="Livr√©e Partiellement"
            ).exists():
                is_commande_livree_partiellement = True
                operation_livraison_partielle_source = (
                    commande_originale_obj.operations.filter(
                        type_operation="LIVRAISON_PARTIELLE"
                    )
                    .order_by("-date_operation")
                    .first()
                )
                commande_renvoi_obj = commande  # Dans ce cas, la commande actuelle est la commande de renvoi

    elif etat_actuel and etat_actuel.enum_etat.libelle == "Livr√©e Partiellement":
        # La commande actuelle est l'originale qui a √©t√© livr√©e partiellement
        is_commande_livree_partiellement = True
        operation_livraison_partielle_source = (
            commande.operations.filter(type_operation="LIVRAISON_PARTIELLE")
            .order_by("-date_operation")
            .first()
        )
        # Chercher une commande de renvoi associ√©e si elle existe
        commande_renvoi_obj = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande.num_cmd}", client=commande.client
        ).first()

    # Si une op√©ration de livraison partielle est trouv√©e, extraire les √©tats des articles renvoy√©s
    if operation_livraison_partielle_source:
        try:
            details = json.loads(operation_livraison_partielle_source.conclusion)
            if "recap_articles_renvoyes" in details:
                for item in details["recap_articles_renvoyes"]:
                    etat_articles_renvoyes[item["article_id"]] = item["etat"]
            
            # Populer articles_livres √† partir de la conclusion de l'op√©ration de livraison partielle
            if "articles_livres" in details:
                for article_livre in details["articles_livres"]:
                    article_obj = Article.objects.filter(
                        id=article_livre.get("article_id")
                    ).first()
                    if article_obj:
                        articles_livres.append(
                            {
                                "article": article_obj,
                                "quantite_livree": article_livre.get("quantite", 0),
                                "prix": article_obj.prix_unitaire,
                                "sous_total": article_obj.prix_unitaire
                                * article_livre.get("quantite", 0),
                            }
                        )
        except Exception as e:
            print(
                f"DEBUG: Erreur lors du parsing des d√©tails de l'op√©ration de livraison partielle: {e}"
            )
            pass



    # Populer articles_renvoyes si c'est une commande de renvoi ou si elle a une commande de renvoi associ√©e
    if is_commande_livree_partiellement:
        # Si la commande actuelle est une commande de renvoi (celle que nous modifions)
        if commande.num_cmd and commande.num_cmd.startswith("RENVOI-"):
            # Les paniers de la commande actuelle sont les articles renvoy√©s
            for panier_renvoi in paniers:
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id)
                if etat is None:
                    etat = "inconnu"
                    print(
                        f"ALERTE: √âtat inconnu pour l'article ID {panier_renvoi.article.id} dans la commande {commande.id_yz}"
                    )
                articles_renvoyes.append(
                    {
                        "article": panier_renvoi.article,
                        "quantite": panier_renvoi.quantite,
                        "prix": panier_renvoi.article.prix_unitaire,
                        "sous_total": panier_renvoi.sous_total,
                        "etat": etat,
                    }
                )
        # Si la commande actuelle est la commande originale livr√©e partiellement (Cas 1 initial)
        elif commande_renvoi_obj:
            # Les paniers de la commande de renvoi associ√©e sont les articles renvoy√©s
            for panier_renvoi in commande_renvoi_obj.paniers.all():
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id)
                if etat is None:
                    etat = "inconnu"
                    print(
                        f"ALERTE: √âtat inconnu pour l'article ID {panier_renvoi.article.id} dans la commande {commande_renvoi_obj.id_yz}"
                    )
                articles_renvoyes.append(
                    {
                        "article": panier_renvoi.article,
                        "quantite": panier_renvoi.quantite,
                        "prix": panier_renvoi.article.prix_unitaire,
                        "sous_total": panier_renvoi.sous_total,
                        "etat": etat,
                    }
                )
    
    # DEBUG: Afficher le contenu de articles_renvoyes apr√®s peuplement
    print(
        f"DEBUG (modifier_commande_prepa): articles_renvoyes APRES POPULATION: {articles_renvoyes}"
    )

    # Cr√©er un map pour acc√©der facilement aux articles renvoy√©s par leur ID dans le template
    # articles_renvoyes_map = {item['article'].id: item for item in articles_renvoyes}

    # Pour les articles livr√©s, on lit l'op√©ration de livraison partielle sur la commande originale
    # C'est pertinent uniquement si la commande actuelle est la commande de renvoi
    # if is_commande_livree_partiellement and commande.num_cmd and commande.num_cmd.startswith('RENVOI-') and commande_originale_obj:
    #     operation_livraison_partielle_for_livres = commande_originale_obj.operations.filter(
    #         type_operation='LIVRAISON_PARTIELLE'
    #     ).order_by('-date_operation').first()
    #     if operation_livraison_partielle_for_livres:
    #         try:
    #             details = json.loads(operation_livraison_partielle_for_livres.conclusion)
    #             if 'articles_livres' in details:
    #                 for article_livre in details['articles_livres']:
    #                     article = Article.objects.filter(id=article_livre.get('article_id')).first()
    #                     if article:
    #                         articles_livres.append({
    #                             'article': article,
    #                             'quantite_livree': article_livre.get('quantite', 0),
    #                             'prix': article.prix_unitaire,
    #                             'sous_total': article.prix_unitaire * article_livre.get('quantite', 0)
    #                         })
    #         except Exception:
    #             pass

    context = {
        "page_title": "Modifier Commande " + str(commande.id_yz),
        "page_subtitle": "Modification des d√©tails de la commande en pr√©paration",
        "commande": commande,
        "paniers": paniers,
        "villes": villes,
        "total_articles": total_articles,
        "is_commande_renvoyee": is_commande_renvoyee,
        "operation_renvoi": operation_renvoi,
        "is_commande_livree_partiellement": is_commande_livree_partiellement,
        "articles_livres": articles_livres,
        "articles_renvoyes": articles_renvoyes,
        # Variables de debug/informations suppl√©mentaires
        "commande_originale": commande_originale_obj,
        "commande_renvoi": commande_renvoi_obj,
        "etat_articles_renvoyes": etat_articles_renvoyes,
        # 'articles_renvoyes_map': articles_renvoyes_map, # Retir√© car plus n√©cessaire
    }
    
    # Gestion d'erreur pour les requ√™tes AJAX non trait√©es
    if request.method == "POST":
        action = request.POST.get("action")
        if action and action not in ["add_article", "delete_article", "update_quantity", "replace_article", "update_operation", "add_operation", "update_commande_info"]:
            print(f"‚ö†Ô∏è Action non reconnue: {action}")
            return JsonResponse({"success": False, "message": f"Action non reconnue: {action}"})
        
        # Si c'est une action AJAX mais qu'aucune r√©ponse JSON n'a √©t√© retourn√©e, retourner une erreur
        if action and action in ["add_article", "delete_article", "update_quantity", "replace_article", "update_operation", "add_operation", "update_commande_info"]:
            print(f"‚ö†Ô∏è Action AJAX non trait√©e: {action}")
            return JsonResponse({"success": False, "message": f"Action non trait√©e: {action}"})
    
    print(f"üîç ===== FIN modifier_commande_prepa (RENDER) =====")
    return render(request, "Prepacommande/modifier_commande.html", context)


@login_required
def ajouter_article_commande_prepa(request, commande_id):
    """Ajouter un article √† la commande en pr√©paration"""
    print("üîÑ ===== D√âBUT ajouter_article_commande_prepa =====")
    print(f"üì¶ M√©thode HTTP: {request.method}")
    print(f"üì¶ Commande ID: {commande_id}")
    print(f"üì¶ User: {request.user}")
    print(f"üì¶ POST data: {dict(request.POST)}")
    print(f"üì¶ Headers: {dict(request.headers)}")
    
    if request.method != 'POST':
        print("‚ùå M√©thode non autoris√©e")
        return JsonResponse({'error': 'M√©thode non autoris√©e'}, status=405)
    
    try:
        # R√©cup√©rer l'op√©rateur de pr√©paration
        operateur = Operateur.objects.get(user=request.user, type_operateur="PREPARATION")
        print(f"‚úÖ Op√©rateur trouv√©: {operateur.id} - Type: {operateur.type_operateur}")
    except Operateur.DoesNotExist:
        print("‚ùå Profil d'op√©rateur de pr√©paration non trouv√©")
        return JsonResponse({'error': 'Profil d\'op√©rateur de pr√©paration non trouv√©.'}, status=403)
    
    try:
        with transaction.atomic():
            print("üîß D√©but de la transaction atomique")
            commande = Commande.objects.select_for_update().get(id=commande_id)
            print(f"‚úÖ Commande trouv√©e: {commande.id} - ID YZ: {commande.id_yz}")
            
            # V√©rifier que la commande est affect√©e √† cet op√©rateur pour la pr√©paration
            etat_preparation = commande.etats.filter(
                Q(enum_etat__libelle="√Ä imprimer") | Q(enum_etat__libelle="En pr√©paration"),
                operateur=operateur,
                date_fin__isnull=True,
            ).first()
            
            if not etat_preparation:
                print("‚ùå Commande non affect√©e √† l'op√©rateur pour la pr√©paration")
                return JsonResponse({'error': 'Cette commande ne vous est pas affect√©e pour la pr√©paration.'}, status=403)
            
            print(f"‚úÖ √âtat de pr√©paration trouv√©: {etat_preparation.enum_etat.libelle}")
            
            # Support both parameter names for backward compatibility
            article_id = request.POST.get("articleId") or request.POST.get("article_id")
            quantite = int(request.POST.get("quantite", 1))
            variante_id = request.POST.get("varianteId") or request.POST.get("variante_id")

            print("[AJOUT VARIANTE] entr√©e:", {
                'commande_id': commande_id,
                'operateur': getattr(operateur, 'id', None),
                'article_id': article_id,
                'quantite': quantite,
                'variante_id': variante_id,
            })
            
            if not article_id or quantite <= 0:
                print(f"‚ùå Donn√©es invalides: article_id={article_id}, quantite={quantite}")
                return JsonResponse({'error': 'Donn√©es invalides'}, status=400)

            print(f"‚úÖ Donn√©es re√ßues: article_id={article_id}, quantite={quantite}, variante_id={variante_id}")
            article = Article.objects.get(id=article_id)
            print(f"‚úÖ Article trouv√©: {article.id} - {article.nom}")
            
            # Handle variant if provided
            variante_obj = None
            if variante_id:
                try:
                    from article.models import Variante
                    variante_obj = Variante.objects.get(id=int(variante_id), article=article)
                    print("[AJOUT VARIANTE] variante trouv√©e:", {
                        'id': variante_obj.id,
                        'couleur': getattr(variante_obj, 'couleur', None),
                        'pointure': getattr(variante_obj, 'pointure', None),
                        'qte_disponible_avant': variante_obj.qte_disponible,
                    })
                except Exception:
                    variante_obj = None
                    print("[AJOUT VARIANTE] variante introuvable ou invalide", variante_id)
            
            # D√©cr√©menter le stock et cr√©er un mouvement
            print("[AJOUT VARIANTE] cr√©ation mouvement stock", {
                'article': article.id,
                'quantite': quantite,
                'type': 'sortie',
                'variante': getattr(variante_obj, 'id', None),
            })
            creer_mouvement_stock(
                article=article, quantite=quantite, type_mouvement='sortie',
                commande=commande, operateur=operateur,
                commentaire=f'Ajout article pendant pr√©paration cmd {commande.id_yz}',
                variante=variante_obj
            )
            
            # V√©rifier si l'article existe d√©j√† dans la commande
            panier_existant = Panier.objects.filter(
                commande=commande, article=article
            ).first()
            
            print("[AJOUT VARIANTE] filtre panier:", {
                'commande': commande.id,
                'article': article.id,
                'existant': getattr(panier_existant, 'id', None)
            })
            
            if panier_existant:
                # Si l'article existe d√©j√†, mettre √† jour la quantit√©
                panier_existant.quantite += quantite
                panier_existant.save()
                panier = panier_existant
                print(f"[AJOUT VARIANTE] üîÑ panier existant {panier.id} mis √† jour, nouvelle_quantite={panier.quantite}")
            else:
                # Si l'article n'existe pas, cr√©er un nouveau panier
                panier = Panier.objects.create(
                    commande=commande,
                    article=article,
                    quantite=quantite,
                    sous_total=0,
                )
                print(f"[AJOUT VARIANTE] ‚ûï nouveau panier cr√©√© id={panier.id}, article={article.id}, quantite={quantite}")
            
            # Si c'√©tait une variante, stocker l'information de la variante
            if variante_obj:
                panier.commentaire = f"Variante s√©lectionn√©e: {variante_obj.couleur} - {variante_obj.pointure}"
                panier.save()
                print(f"üìù Variante stock√©e: {variante_obj.couleur} - {variante_obj.pointure}")
            
            # Recalculer le compteur apr√®s ajout (logique de confirmation)
            if article.isUpsell and hasattr(article, 'prix_upsell_1') and article.prix_upsell_1 is not None:
                # Compter la quantit√© totale d'articles upsell (apr√®s ajout)
                total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                    total=Sum('quantite')
                )['total'] or 0
                
                # Le compteur ne s'incr√©mente qu'√† partir de 2 unit√©s d'articles upsell
                # 0-1 unit√©s upsell ‚Üí compteur = 0
                # 2+ unit√©s upsell ‚Üí compteur = total_quantite_upsell - 1
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()
            else:
                # Pour les articles normaux, juste calculer le sous-total
                from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                prix_unitaire = get_prix_upsell_avec_compteur(article, commande.compteur)
                sous_total = prix_unitaire * panier.quantite
                panier.sous_total = float(sous_total)
                panier.save()
            
            # Recalculer le total
            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()
            print("[AJOUT VARIANTE] totaux mis √† jour:", {
                'commande_id': commande.id,
                'total_cmd': commande.total_cmd,
                'articles_count': commande.paniers.count(),
            })
            
            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = articles_upsell.aggregate(
                total=Sum('quantite')
            )['total'] or 0
            
            # D√©terminer si c'√©tait un ajout ou une mise √† jour
            message = 'Article ajout√© avec succ√®s' if not panier_existant else f'Quantit√© mise √† jour ({panier.quantite})'
            
            # Pr√©parer les donn√©es de l'article pour le frontend
            article_data = {
                'panier_id': panier.id,
                'nom': article.nom,
                'reference': article.reference,
                'couleur_fr': variante_obj.couleur if variante_obj else (article.couleur or ""),
                'couleur_ar': variante_obj.couleur if variante_obj else (article.couleur or ""),
                'pointure': variante_obj.pointure if variante_obj else (article.pointure or ""),
                'quantite': panier.quantite,
                'prix': panier.sous_total / panier.quantite,  # Prix unitaire
                'sous_total': panier.sous_total,
                'is_upsell': article.isUpsell,
                'isUpsell': article.isUpsell,
                'phase': article.phase,
                'qte_disponible': article.get_total_qte_disponible(),
                'has_promo_active': article.has_promo_active if hasattr(article, 'has_promo_active') else False,
                'description': article.description or "",
                'variante_info': f"{variante_obj.couleur} - {variante_obj.pointure}" if variante_obj else None,
            }
            
            response_data = {
                'success': True, 
                'message': message,
                'article_id': panier.id,
                'total_commande': float(commande.total_cmd),
                'nb_articles': commande.paniers.count(),
                'compteur': commande.compteur,
                'was_update': panier_existant is not None,
                'new_quantity': panier.quantite,
                'article_data': article_data,
                'articles_count': commande.paniers.count(),
                'sous_total_articles': float(sum(p.sous_total for p in commande.paniers.all())),
                'articles_upsell': articles_upsell.count(),
                'quantite_totale_upsell': total_quantite_upsell
            }
            
            print("‚úÖ ===== SUCC√àS ajouter_article_commande_prepa =====")
            print(f"üì¶ R√©ponse: {response_data}")
            
            return JsonResponse(response_data)
            
    except Article.DoesNotExist:
        print("‚ùå Article non trouv√©")
        return JsonResponse({'success': False, 'error': 'Article non trouv√©'}, status=404)
    except Exception as e:
        print(f"‚ùå ===== ERREUR ajouter_article_commande_prepa =====")
        print(f"‚ùå Exception: {str(e)}")
        import traceback
        print(f"‚ùå Traceback: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': f'Erreur interne: {str(e)}'}, status=500)


@login_required
def api_articles_disponibles_prepa(request):
    """API pour r√©cup√©rer les articles disponibles pour les op√©rateurs de pr√©paration"""
    from article.models import Article
    
    try:
        # V√©rifier que l'utilisateur est un op√©rateur de pr√©paration
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"success": False, "message": "Acc√®s non autoris√©"})
    
    search_query = request.GET.get("search", "")
    filter_type = request.GET.get("filter", "tous")
    
    # R√©cup√©rer les articles actifs
    articles = Article.objects.filter(actif=True)
    
    # Appliquer les filtres selon le type
    if filter_type == "disponible":
        # Filtrer les articles qui ont au moins une variante avec stock > 0
        articles = articles.filter(variantes__qte_disponible__gt=0, variantes__actif=True).distinct()
    elif filter_type == "upsell":
        articles = articles.filter(isUpsell=True)
    elif filter_type == "liquidation":
        articles = articles.filter(phase="LIQUIDATION")
    elif filter_type == "test":
        articles = articles.filter(phase="EN_TEST")
    
    # Recherche textuelle
    if search_query:
        articles = articles.filter(
            Q(nom__icontains=search_query)
            | Q(reference__icontains=search_query)
            | Q(couleur__icontains=search_query)
            | Q(pointure__icontains=search_query)
            | Q(description__icontains=search_query)
        )
    
    # Compter les articles par type pour les statistiques
    stats = {
        "tous": Article.objects.filter(actif=True).count(),
        "disponible": Article.objects.filter(actif=True, variantes__qte_disponible__gt=0, variantes__actif=True).distinct().count(),
        "upsell": Article.objects.filter(actif=True, isUpsell=True).count(),
        "liquidation": Article.objects.filter(actif=True, phase="LIQUIDATION").count(),
        "test": Article.objects.filter(actif=True, phase="EN_TEST").count(),
    }
    
    # Compter les articles en promotion en utilisant une approche diff√©rente
    # Chercher les articles qui ont un prix actuel diff√©rent du prix unitaire
    articles_promo_count = Article.objects.filter(
        actif=True, prix_actuel__lt=F("prix_unitaire")
    ).count()
    stats["promo"] = articles_promo_count
    
    # Filtrer les articles en promotion si n√©cessaire
    if filter_type == "promo":
        articles = articles.filter(prix_actuel__lt=F("prix_unitaire"))
    
    # Limiter les r√©sultats
    articles = articles[:50]
    
    articles_data = []
    for article in articles:
        # Prix √† afficher (prix actuel si diff√©rent du prix unitaire)
        prix_affichage = float(article.prix_actuel or article.prix_unitaire)
        prix_original = float(article.prix_unitaire)
        has_reduction = prix_affichage < prix_original
        
        # D√©terminer le type d'article pour l'affichage
        article_type = "normal"
        type_icon = "fas fa-box"
        type_color = "text-gray-600"
        
        if article.isUpsell:
            article_type = "upsell"
            type_icon = "fas fa-arrow-up"
            type_color = "text-purple-600"
        elif article.phase == "LIQUIDATION":
            article_type = "liquidation"
            type_icon = "fas fa-money-bill-wave"
            type_color = "text-red-600"
        elif article.phase == "EN_TEST":
            article_type = "test"
            type_icon = "fas fa-flask"
            type_color = "text-yellow-600"
        
        # V√©rifier si l'article est en promotion (prix actuel < prix unitaire)
        if has_reduction:
            article_type = "promo"
            type_icon = "fas fa-fire"
            type_color = "text-orange-600"

        # R√©cup√©rer les variantes de l'article
        variantes_data = []
        for variante in article.variantes.filter(actif=True):
            variantes_data.append({
                "id": variante.id,
                "couleur": str(variante.couleur) if variante.couleur else "",
                "pointure": str(variante.pointure) if variante.pointure else "",
                "prix_actuel": float(variante.prix_actuel or variante.prix_unitaire),
                "prix": float(variante.prix_unitaire),
                "qte_disponible": variante.qte_disponible,
                "stock": variante.qte_disponible,
                "reference_variante": variante.reference_variante or "",
                "actif": variante.actif,
            })

        articles_data.append(
            {
                "id": article.id,
                "nom": article.nom,
                "reference": article.reference or "",
                "couleur": str(article.couleur) if article.couleur else "",
                "pointure": str(article.pointure) if article.pointure else "",
                "description": article.description or "",
                "prix": prix_affichage,
                "prix_original": prix_original,
                "has_reduction": has_reduction,
                "reduction_pourcentage": round(
                    ((prix_original - prix_affichage) / prix_original) * 100, 0
                )
                if has_reduction
                else 0,
                "qte_disponible": article.get_total_qte_disponible(),
                "article_type": article_type,
                "type_icon": type_icon,
                "type_color": type_color,
                "phase": article.phase,
                "isUpsell": article.isUpsell,
                "display_text": f"{article.nom} - {str(article.couleur) if article.couleur else ''} - {str(article.pointure) if article.pointure else ''} ({prix_affichage} DH)",
                "variantes_all": variantes_data,
            }
        )

    return JsonResponse(
        {
            "success": True,
            "articles": articles_data,
            "stats": stats,
            "filter_applied": filter_type,
        }
    )


@login_required
def api_panier_commande_prepa(request, commande_id):
    """API pour r√©cup√©rer le panier d'une commande pour les op√©rateurs de pr√©paration"""
    try:
        # V√©rifier que l'utilisateur est un op√©rateur de pr√©paration
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"success": False, "message": "Acc√®s non autoris√©"})
    
    # R√©cup√©rer la commande
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({"success": False, "message": "Commande non trouv√©e"})
    
    # V√©rifier que la commande est affect√©e √† cet op√©rateur
    etat_preparation = commande.etats.filter(
        Q(enum_etat__libelle="√Ä imprimer") | Q(enum_etat__libelle="En pr√©paration"),
        operateur=operateur,
        date_fin__isnull=True,
    ).first()
    
    if not etat_preparation:
        return JsonResponse({"success": False, "message": "Commande non affect√©e"})
    
    # R√©cup√©rer les paniers
    paniers = commande.paniers.all().select_related("article")
    
    paniers_data = []
    for panier in paniers:
        paniers_data.append(
            {
                "id": panier.id,
                "article_id": panier.article.id,
                "article_nom": panier.article.nom,
                "article_reference": panier.article.reference or "",
                "article_couleur": panier.article.couleur,
                "article_pointure": panier.article.pointure,
                "quantite": panier.quantite,
                "prix_unitaire": float(panier.article.prix_unitaire),
                "sous_total": float(panier.sous_total),
                "display_text": f"{panier.article.nom} - {panier.article.couleur} - {panier.article.pointure}",
            }
        )

    return JsonResponse(
        {
            "success": True,
            "paniers": paniers_data,
            "total_commande": float(commande.total_cmd),
            "nb_articles": len(paniers_data),
        }
    )


@login_required
def imprimer_tickets_preparation(request):
    """
    Vue pour imprimer les tickets de pr√©paration SANS changer l'√©tat des commandes.
    Permet d'imprimer ou de r√©imprimer des tickets pour les commandes en pr√©paration.
    """
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            return HttpResponse("Acc√®s non autoris√©.", status=403)
    except Operateur.DoesNotExist:
        return HttpResponse("Profil op√©rateur non trouv√©.", status=403)

    commande_ids_str = request.GET.get("ids")
    if not commande_ids_str:
        return HttpResponse("Aucun ID de commande fourni.", status=400)

    try:
        commande_ids = [int(id) for id in commande_ids_str.split(",") if id.isdigit()]
    except ValueError:
        return HttpResponse("IDs de commande invalides.", status=400)
    
    # R√©cup√©rer les commandes en pr√©paration affect√©es √† cet op√©rateur
    commandes = Commande.objects.filter(
        id__in=commande_ids,
        etats__operateur=operateur_profile,
        etats__enum_etat__libelle="En pr√©paration",
        etats__date_fin__isnull=True,
    ).distinct()

    if not commandes.exists():
        return HttpResponse(
            "Aucune commande en pr√©paration trouv√©e pour cet op√©rateur.", status=404
        )

    # G√©n√©ration du code-barres pour chaque commande (sans transition d'√©tat)
    code128 = barcode.get_barcode_class("code128")
    
    for commande in commandes:
        # G√©n√©rer le code-barres uniquement si pas d√©j√† pr√©sent
        if not hasattr(commande, "barcode_base64") or not commande.barcode_base64:
            barcode_instance = code128(str(commande.id_yz), writer=ImageWriter())
            buffer = BytesIO()
            barcode_instance.write(
                buffer,
                options={
                    "write_text": False,
                    "module_height": 15.0,
                    "module_width": 0.3,
                },
            )
            barcode_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
            commande.barcode_base64 = barcode_base64
        
        # D√©finir la date de pr√©paration pour l'affichage (sans sauvegarder en DB)
        if not hasattr(commande, "date_preparation") or not commande.date_preparation:
            commande.date_preparation = timezone.now()

    context = {
        "commandes": commandes,
        "is_reprint": True,  # Indicateur pour diff√©rencier des impressions initiales
    }
    
    return render(request, "Prepacommande/tickets_preparation.html", context)

# === NOUVELLES FONCTIONNALIT√âS : GESTION DE STOCK ===

# === FONCTION SUPPRIM√âE : GESTION DE STOCK D√âPLAC√âE VERS ADMIN ===
# @login_required
# def ajuster_stock(request, article_id):
#     """Ajuster le stock d'un article - Service de pr√©paration"""
#     try:
#         operateur_profile = request.user.profil_operateur
#         if not operateur_profile.is_preparation:
#             messages.error(request, "Acc√®s non autoris√©.")
#             return redirect('login')
#     except Operateur.DoesNotExist:
#         messages.error(request, "Profil op√©rateur non trouv√©.")
#         return redirect('login')
#     
#     article = get_object_or_404(Article, pk=article_id)
#     
#     if request.method == 'POST':
#         form = AjusterStockForm(request.POST)
#         if form.is_valid():
#             type_mouvement = form.cleaned_data['type_mouvement']
#             quantite = form.cleaned_data['quantite']
#             commentaire = form.cleaned_data['commentaire']
#             
#             try:
#                 print(f"üîß Ajustement stock - Article: {article.nom}, Type: {type_mouvement}, Quantit√©: {quantite}")
#                 print(f"üîß Stock avant ajustement: {article.qte_disponible}")
#                 
#                 mouvement = creer_mouvement_stock(
#                     article=article,
#                     quantite=quantite,
#                     type_mouvement=type_mouvement,
#                     operateur=operateur_profile,
#                     commentaire=commentaire
#                 )
#                 
#                 # Recharger l'article pour voir le stock mis √† jour
#                 article.refresh_from_db()
#                 print(f"‚úÖ Stock apr√®s ajustement: {article.qte_disponible}")
#                 
#                 if mouvement:
#                     messages.success(request, f"Le stock de l'article '{article.nom}' a √©t√© ajust√© avec succ√®s. Nouveau stock: {article.qte_disponible}")
#                 else:
#                     messages.warning(request, "L'ajustement n'a pas pu √™tre effectu√©.")
#                     
#                 return redirect('Prepacommande:detail_article', article_id=article.id)
#             except Exception as e:
#                 print(f"‚ùå Erreur dans ajuster_stock: {str(e)}")
#                 import traceback
#                 traceback.print_exc()
#                 messages.error(request, f"Une erreur est survenue lors de l'ajustement du stock : {e}")
# 
#     else:
#         form = AjusterStockForm()
# 
#     mouvements_recents = article.mouvements.order_by('-date_mouvement')[:10]
# 
#     context = {
#         'form': form,
#         'article': article,
#         'mouvements_recents': mouvements_recents,
#         'page_title': f"Ajuster le Stock - {article.nom}",
#     }
#     return render(request, 'Prepacommande/stock/ajuster_stock.html', context)
    pass

# === FONCTION SUPPRIM√âE : GESTION DE STOCK D√âPLAC√âE VERS ADMIN ===
# @login_required
# def detail_article(request, article_id):
#     """Afficher les d√©tails d'un article sp√©cifique - Service de pr√©paration"""
    article = get_object_or_404(Article, pk=article_id)
    
    # Calculer la valeur totale du stock
    valeur_stock = (
        article.prix_actuel * article.qte_disponible if article.prix_actuel else 0
    )
    
    # R√©cup√©rer le dernier mouvement de stock pour cet article
    dernier_mouvement = article.mouvements.order_by("-date_mouvement").first()

    context = {
        "article": article,
        "valeur_stock": valeur_stock,
        "dernier_mouvement": dernier_mouvement,
        "page_title": f"D√©tail de l'article : {article.nom}",
        "page_subtitle": "Informations compl√®tes sur l'article",
    }
    return render(request, "Prepacommande/stock/detail_article.html", context)

# === FONCTION SUPPRIM√âE : GESTION DE STOCK D√âPLAC√âE VERS ADMIN ===
# @login_required
# def liste_articles(request):
#     """Afficher la liste des articles avec filtres et statistiques - Service de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Acc√®s non autoris√©.")
            return redirect("login")
    except Operateur.DoesNotExist:
        messages.error(request, "Profil op√©rateur non trouv√©.")
        return redirect("login")
    
    # Calcul des statistiques globales (avant tout filtrage)
    articles_qs = Article.objects.all()
    articles_total = articles_qs.count()
    articles_actifs = articles_qs.filter(actif=True).count()
    articles_inactifs = articles_qs.filter(actif=False).count()
    articles_rupture = articles_qs.filter(qte_disponible__lte=0).count()
    
    # Articles cr√©√©s aujourd'hui
    today = timezone.now().date()
    articles_crees_aujourd_hui = articles_qs.filter(date_creation__date=today).count()

    # R√©cup√©ration des articles pour la liste, filtr√©e
    articles_list = Article.objects.all()
    
    # Filtres de recherche am√©lior√©s
    query = request.GET.get("q", "").strip()
    categorie_filter = request.GET.get("categorie", "").strip()
    statut_filter = request.GET.get("statut", "").strip()
    stock_filter = request.GET.get("stock", "").strip()
    prix_min = request.GET.get("prix_min", "").strip()
    prix_max = request.GET.get("prix_max", "").strip()
    couleur_filter = request.GET.get("couleur", "").strip()
    phase_filter = request.GET.get("phase", "").strip()
    tri = request.GET.get("tri", "date_creation").strip()
    
    # Recherche textuelle intelligente
    if query:
        articles_list = articles_list.filter(
            Q(nom__icontains=query)
            | Q(reference__icontains=query)
            | Q(description__icontains=query)
            | Q(categorie__icontains=query)
            | Q(couleur__icontains=query)
        )
    
    # Filtre par cat√©gorie
    if categorie_filter:
        articles_list = articles_list.filter(categorie__icontains=categorie_filter)
    
    # Filtre par statut
    if statut_filter:
        if statut_filter == "actif":
            articles_list = articles_list.filter(actif=True)
        elif statut_filter == "inactif":
            articles_list = articles_list.filter(actif=False)
    
    # Filtre par niveau de stock
    if stock_filter:
        if stock_filter == "rupture":
            articles_list = articles_list.filter(qte_disponible__lte=0)
        elif stock_filter == "faible":
            articles_list = articles_list.filter(
                qte_disponible__gt=0, qte_disponible__lte=10
            )
        elif stock_filter == "normal":
            articles_list = articles_list.filter(
                qte_disponible__gt=10, qte_disponible__lte=50
            )
        elif stock_filter == "eleve":
            articles_list = articles_list.filter(qte_disponible__gt=50)
    
    # Filtre par prix
    if prix_min:
        try:
            prix_min_val = float(prix_min.replace(",", "."))
            articles_list = articles_list.filter(prix_unitaire__gte=prix_min_val)
        except (ValueError, TypeError):
            pass
    
    if prix_max:
        try:
            prix_max_val = float(prix_max.replace(",", "."))
            articles_list = articles_list.filter(prix_unitaire__lte=prix_max_val)
        except (ValueError, TypeError):
            pass
    
    # Filtre par couleur
    if couleur_filter:
        articles_list = articles_list.filter(couleur__icontains=couleur_filter)
    
    # Filtre par phase
    if phase_filter:
        articles_list = articles_list.filter(phase=phase_filter)
    
    # Tri des r√©sultats
    if tri == "nom":
        articles_list = articles_list.order_by("nom")
    elif tri == "prix_asc":
        articles_list = articles_list.order_by("prix_unitaire")
    elif tri == "prix_desc":
        articles_list = articles_list.order_by("-prix_unitaire")
    elif tri == "stock_asc":
        articles_list = articles_list.order_by("qte_disponible")
    elif tri == "stock_desc":
        articles_list = articles_list.order_by("-qte_disponible")
    elif tri == "date_creation":
        articles_list = articles_list.order_by("-date_creation")
    elif tri == "reference":
        articles_list = articles_list.order_by("reference")
    else:
        articles_list = articles_list.order_by("-date_creation")
    
    # R√©cup√©ration des valeurs uniques pour les filtres
    categories_uniques = (
        Article.objects.values_list("categorie", flat=True)
        .distinct()
        .exclude(categorie__isnull=True)
        .exclude(categorie__exact="")
    )
    couleurs_uniques = (
        Article.objects.values_list("couleur", flat=True)
        .distinct()
        .exclude(couleur__isnull=True)
        .exclude(couleur__exact="")
    )
    phases_uniques = (
        Article.objects.values_list("phase", flat=True)
        .distinct()
        .exclude(phase__isnull=True)
        .exclude(phase__exact="")
    )

    # Pagination
    paginator = Paginator(articles_list, 12)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    context = {
        "articles": page_obj,
        "categories_uniques": categories_uniques,
        "couleurs_uniques": couleurs_uniques,
        "phases_uniques": phases_uniques,
        "articles_total": articles_total,
        "articles_actifs": articles_actifs,
        "articles_inactifs": articles_inactifs,
        "articles_rupture": articles_rupture,
        "articles_crees_aujourd_hui": articles_crees_aujourd_hui,
        "page_title": "Liste des Articles",
        "page_subtitle": "Inventaire complet et gestion du stock",
        "request": request,
        "query": query,
        "current_filters": {
            "categorie": categorie_filter,
            "statut": statut_filter,
            "stock": stock_filter,
            "prix_min": prix_min,
            "prix_max": prix_max,
            "couleur": couleur_filter,
            "phase": phase_filter,
            "tri": tri,
        },
    }
    return render(request, "Prepacommande/stock/liste_articles.html", context)

# === FONCTION SUPPRIM√âE : GESTION DE STOCK D√âPLAC√âE VERS ADMIN ===
# @login_required
# def mouvements_stock(request):
#     """Vue pour afficher l'historique des mouvements de stock - Service de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Acc√®s non autoris√©.")
            return redirect("login")
    except Operateur.DoesNotExist:
        messages.error(request, "Profil op√©rateur non trouv√©.")
        return redirect("login")
    
    from article.models import MouvementStock
    
    # R√©cup√©ration de tous les mouvements
    mouvements_list = MouvementStock.objects.select_related(
        "article", "operateur"
    ).order_by("-date_mouvement")
    
    # Filtres de recherche
    article_filter = request.GET.get("article", "").strip()
    type_filter = request.GET.get("type", "").strip()
    date_filter = request.GET.get("date_range", "").strip()
    
    # Filtre par article (nom ou r√©f√©rence)
    if article_filter:
        mouvements_list = mouvements_list.filter(
            Q(article__nom__icontains=article_filter)
            | Q(article__reference__icontains=article_filter)
        )
    
    # Filtre par type de mouvement
    if type_filter:
        if type_filter == "entree":
            mouvements_list = mouvements_list.filter(type_mouvement="entree")
        elif type_filter == "sortie":
            mouvements_list = mouvements_list.filter(type_mouvement="sortie")
        elif type_filter == "ajustement":
            mouvements_list = mouvements_list.filter(
                type_mouvement__in=["ajustement_pos", "ajustement_neg"]
            )
    
    # Filtre par date
    if date_filter:
        try:
            date_obj = datetime.strptime(date_filter, "%Y-%m-%d").date()
            mouvements_list = mouvements_list.filter(date_mouvement__date=date_obj)
        except ValueError:
            pass
    
    # Pagination
    paginator = Paginator(mouvements_list, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    # Statistiques rapides
    total_mouvements = mouvements_list.count()
    mouvements_aujourd_hui = MouvementStock.objects.filter(
        date_mouvement__date=timezone.now().date()
    ).count()
    
    context = {
        "mouvements": page_obj,
        "total_mouvements": total_mouvements,
        "mouvements_aujourd_hui": mouvements_aujourd_hui,
        "page_title": "Mouvements de Stock",
        "current_filters": {
            "article": article_filter,
            "type": type_filter,
            "date_range": date_filter,
        },
    }
    return render(request, "Prepacommande/stock/mouvements_stock.html", context)

# === FONCTION SUPPRIM√âE : GESTION DE STOCK D√âPLAC√âE VERS ADMIN ===
# @login_required
# def alertes_stock(request):
#     """Vue pour afficher les alertes de stock - Service de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Acc√®s non autoris√©.")
            return redirect("login")
    except Operateur.DoesNotExist:
        messages.error(request, "Profil op√©rateur non trouv√©.")
        return redirect("login")
    
    from article.models import MouvementStock
    
    # Param√®tres de seuils
    SEUIL_RUPTURE = 0
    SEUIL_STOCK_FAIBLE = 10
    SEUIL_A_COMMANDER = 20
    
    # R√©cup√©ration de tous les articles actifs
    articles_actifs = Article.objects.filter(actif=True)
    
    # Filtres par niveau d'alerte
    filtre_alerte = request.GET.get("filtre", "tous")
    
    if filtre_alerte == "rupture":
        articles_alerte = articles_actifs.filter(qte_disponible__lte=SEUIL_RUPTURE)
    elif filtre_alerte == "faible":
        articles_alerte = articles_actifs.filter(
            qte_disponible__gt=SEUIL_RUPTURE, qte_disponible__lte=SEUIL_STOCK_FAIBLE
        )
    elif filtre_alerte == "a_commander":
        articles_alerte = articles_actifs.filter(
            qte_disponible__gt=SEUIL_STOCK_FAIBLE, qte_disponible__lte=SEUIL_A_COMMANDER
        )
    else:
        articles_alerte = articles_actifs.filter(qte_disponible__lte=SEUIL_A_COMMANDER)
    
    # Tri des r√©sultats
    tri = request.GET.get("tri", "stock_asc")
    if tri == "stock_asc":
        articles_alerte = articles_alerte.order_by("qte_disponible")
    elif tri == "stock_desc":
        articles_alerte = articles_alerte.order_by("-qte_disponible")
    elif tri == "nom":
        articles_alerte = articles_alerte.order_by("nom")
    elif tri == "reference":
        articles_alerte = articles_alerte.order_by("reference")
    elif tri == "categorie":
        articles_alerte = articles_alerte.order_by("categorie")
    else:
        articles_alerte = articles_alerte.order_by("qte_disponible")
    
    # Statistiques d√©taill√©es
    stats = {
        "total_articles": articles_actifs.count(),
        "rupture_stock": articles_actifs.filter(
            qte_disponible__lte=SEUIL_RUPTURE
        ).count(),
        "stock_faible": articles_actifs.filter(
            qte_disponible__gt=SEUIL_RUPTURE, qte_disponible__lte=SEUIL_STOCK_FAIBLE
        ).count(),
        "a_commander": articles_actifs.filter(
            qte_disponible__gt=SEUIL_STOCK_FAIBLE, qte_disponible__lte=SEUIL_A_COMMANDER
        ).count(),
        "stock_ok": articles_actifs.filter(
            qte_disponible__gt=SEUIL_A_COMMANDER
        ).count(),
    }
    
    # Alertes critiques
    alertes_critiques = articles_actifs.filter(
        qte_disponible__lte=SEUIL_RUPTURE
    ).order_by("qte_disponible")[:5]
    
    # Analyse par cat√©gorie
    categories_alertes = (
        articles_actifs.values("categorie")
        .annotate(
            total=Count("id"),
            rupture=Count("id", filter=Q(qte_disponible__lte=SEUIL_RUPTURE)),
            faible=Count(
                "id",
                filter=Q(
                    qte_disponible__gt=SEUIL_RUPTURE,
                    qte_disponible__lte=SEUIL_STOCK_FAIBLE,
                ),
            ),
            a_commander=Count(
                "id",
                filter=Q(
                    qte_disponible__gt=SEUIL_STOCK_FAIBLE,
                    qte_disponible__lte=SEUIL_A_COMMANDER,
                ),
            ),
            stock_moyen=Avg("qte_disponible"),
            valeur_stock=Sum("qte_disponible"),
        )
        .exclude(categorie__isnull=True)
        .exclude(categorie__exact="")
        .order_by("-rupture", "-faible")
    )
    
    # Historique des mouvements r√©cents
    mouvements_recents = (
        MouvementStock.objects.filter(
        article__in=articles_alerte,
            date_mouvement__gte=timezone.now() - timedelta(days=30),
        )
        .select_related("article", "operateur")
        .order_by("-date_mouvement")[:10]
    )
    
    # Suggestions d'actions
    suggestions = []
    
    if stats["rupture_stock"] > 0:
        suggestions.append(
            {
                "type": "danger",
                "titre": "Rupture de Stock Critique",
                "message": f'{stats["rupture_stock"]} article(s) en rupture totale n√©cessitent un r√©approvisionnement imm√©diat.',
                "action": "Contacter les fournisseurs",
                "icone": "fas fa-exclamation-triangle",
            }
        )

    if stats["stock_faible"] > 0:
        suggestions.append(
            {
                "type": "warning",
                "titre": "Stock Faible",
                "message": f'{stats["stock_faible"]} article(s) ont un stock faible. Planifier les commandes.',
                "action": "Pr√©parer les commandes",
                "icone": "fas fa-exclamation-circle",
            }
        )

    if stats["a_commander"] > 0:
        suggestions.append(
            {
                "type": "info",
                "titre": "√Ä Commander Bient√¥t",
                "message": f'{stats["a_commander"]} article(s) devront √™tre command√©s prochainement.',
                "action": "Surveiller l'√©volution",
                "icone": "fas fa-info-circle",
            }
        )
    
    # Pagination
    paginator = Paginator(articles_alerte, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    context = {
        "articles": page_obj,
        "stats": stats,
        "alertes_critiques": alertes_critiques,
        "categories_alertes": categories_alertes,
        "mouvements_recents": mouvements_recents,
        "suggestions": suggestions,
        "filtre_actuel": filtre_alerte,
        "tri_actuel": tri,
        "seuils": {
            "rupture": SEUIL_RUPTURE,
            "faible": SEUIL_STOCK_FAIBLE,
            "a_commander": SEUIL_A_COMMANDER,
        },
        "page_title": "Alertes Stock",
        "page_subtitle": "Articles n√©cessitant une attention imm√©diate",
    }
    return render(request, "Prepacommande/stock/alertes_stock.html", context)

# === FONCTION SUPPRIM√âE : GESTION DE STOCK D√âPLAC√âE VERS ADMIN ===
# @login_required
# def statistiques_stock(request):
#     """Vue pour afficher les statistiques de stock - Service de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Acc√®s non autoris√©.")
            return redirect("login")
    except Operateur.DoesNotExist:
        messages.error(request, "Profil op√©rateur non trouv√©.")
        return redirect("login")
    
    from article.models import MouvementStock
    
    # Param√®tres de filtrage
    periode = int(request.GET.get("periode", 30))
    categorie_filter = request.GET.get("categorie", "")
    
    # Date de d√©but selon la p√©riode
    date_debut = timezone.now() - timedelta(days=periode)
    
    # Articles de base
    articles_qs = Article.objects.filter(actif=True)
    
    # Filtrage par cat√©gorie si sp√©cifi√©
    if categorie_filter:
        articles_qs = articles_qs.filter(categorie=categorie_filter)
    
    # Valeur totale du stock
    valeur_stock = (
        articles_qs.aggregate(
            valeur_totale=Sum(F("qte_disponible") * F("prix_unitaire"))
        )["valeur_totale"]
        or 0
    )
    
    # Nombre total d'articles en stock
    articles_en_stock = articles_qs.filter(qte_disponible__gt=0).count()
    
    # Articles par niveau de stock
    stats_niveaux = articles_qs.aggregate(
        total_articles=Count("id"),
        rupture=Count("id", filter=Q(qte_disponible=0)),
        stock_faible=Count(
            "id", filter=Q(qte_disponible__gt=0, qte_disponible__lte=10)
        ),
        stock_normal=Count(
            "id", filter=Q(qte_disponible__gt=10, qte_disponible__lte=50)
        ),
        stock_eleve=Count("id", filter=Q(qte_disponible__gt=50)),
    )
    
    # Taux de rupture
    taux_rupture = (
        (stats_niveaux["rupture"] / stats_niveaux["total_articles"] * 100)
        if stats_niveaux["total_articles"] > 0
        else 0
    )
    
    # Statistiques par cat√©gorie
    stats_categories = (
        articles_qs.values("categorie")
        .annotate(
            total_articles=Count("id"),
            stock_total=Sum("qte_disponible"),
            valeur_totale=Sum(F("qte_disponible") * F("prix_unitaire")),
            prix_moyen=Avg("prix_unitaire"),
            stock_moyen=Avg("qte_disponible"),
            articles_rupture=Count("id", filter=Q(qte_disponible=0)),
            articles_faible=Count(
                "id", filter=Q(qte_disponible__gt=0, qte_disponible__lte=10)
            ),
        )
        .exclude(categorie__isnull=True)
        .exclude(categorie__exact="")
        .order_by("-valeur_totale")
    )
    
    # Top articles
    top_articles_valeur = (
        articles_qs.annotate(valeur_stock=F("qte_disponible") * F("prix_unitaire"))
        .filter(qte_disponible__gt=0)
        .order_by("-valeur_stock")[:10]
    )

    top_articles_quantite = articles_qs.filter(qte_disponible__gt=0).order_by(
        "-qte_disponible"
    )[:10]
    
    # Mouvements de stock
    mouvements_periode = MouvementStock.objects.filter(
        date_mouvement__gte=date_debut, article__in=articles_qs
    ).select_related("article")

    mouvements_sortie = (
        mouvements_periode.filter(
            type_mouvement__in=["sortie", "ajustement_neg"]
        ).aggregate(total_sorties=Sum("quantite"))["total_sorties"]
        or 0
    )
    
    rotation_stock = (mouvements_sortie / valeur_stock * 100) if valeur_stock > 0 else 0
    
    # √âvolution temporelle
    evolution_donnees = []
    nb_semaines = min(periode // 7, 12)
    
    for i in range(nb_semaines):
        date_fin = timezone.now() - timedelta(days=i * 7)
        valeur_semaine = (
            articles_qs.aggregate(valeur=Sum(F("qte_disponible") * F("prix_unitaire")))[
                "valeur"
            ]
            or 0
        )

        evolution_donnees.append(
            {"date": date_fin.strftime("%d/%m"), "valeur": float(valeur_semaine)}
        )
    
    evolution_donnees.reverse()
    
    # Alertes
    alertes = []
    
    if stats_niveaux["rupture"] > 0:
        alertes.append(
            {
                "type": "danger",
                "titre": "Articles en Rupture",
                "message": f'{stats_niveaux["rupture"]} article(s) en rupture de stock',
                "valeur": stats_niveaux["rupture"],
            }
        )
    
    if taux_rupture > 10:
        alertes.append(
            {
                "type": "warning",
                "titre": "Taux de Rupture √âlev√©",
                "message": f"Taux de rupture de {taux_rupture:.1f}% (seuil recommand√©: 5%)",
                "valeur": f"{taux_rupture:.1f}%",
            }
        )
    
    if rotation_stock < 2:
        alertes.append(
            {
                "type": "info",
                "titre": "Rotation Faible",
                "message": "La rotation du stock est faible, optimisation possible",
                "valeur": f"{rotation_stock:.1f}",
            }
        )
    
    # Donn√©es pour graphiques
    categories_chart_data = {
        "labels": [cat["categorie"] for cat in stats_categories],
        "values": [float(cat["valeur_totale"] or 0) for cat in stats_categories],
        "colors": ["#FF6384", "#36A2EB", "#FFCE56", "#4BC0C0", "#9966FF", "#FF9F40"],
    }
    
    top_articles_chart_data = {
        "labels": [art.nom[:20] for art in top_articles_valeur[:5]],
        "values": [
            float((art.qte_disponible or 0) * (art.prix_unitaire or 0))
            for art in top_articles_valeur[:5]
        ],
    }

    categories_disponibles = (
        Article.objects.filter(actif=True)
        .values_list("categorie", flat=True)
        .distinct()
        .exclude(categorie__isnull=True)
        .exclude(categorie__exact="")
        .order_by("categorie")
    )
    
    context = {
        "page_title": "Statistiques Stock",
        "page_subtitle": "Analyse de la performance et de la valeur de l'inventaire",
        "valeur_stock": valeur_stock,
        "articles_en_stock": articles_en_stock,
        "rotation_stock": rotation_stock,
        "taux_rupture": taux_rupture,
        "stats_niveaux": stats_niveaux,
        "stats_categories": stats_categories,
        "top_articles_valeur": top_articles_valeur,
        "top_articles_quantite": top_articles_quantite,
        "evolution_donnees": evolution_donnees,
        "alertes": alertes,
        "categories_chart_data": categories_chart_data,
        "top_articles_chart_data": top_articles_chart_data,
        "categories_disponibles": categories_disponibles,
        "periode_actuelle": periode,
        "categorie_actuelle": categorie_filter,
    }
    return render(request, "Prepacommande/stock/statistiques_stock.html", context)

# === FONCTION SUPPRIM√âE : GESTION DE STOCK D√âPLAC√âE VERS ADMIN ===
# @login_required
# def creer_article(request):
#     """Cr√©er un nouvel article - Service de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Acc√®s non autoris√©.")
            return redirect("login")
    except Operateur.DoesNotExist:
        messages.error(request, "Profil op√©rateur non trouv√©.")
        return redirect("login")
    
    if request.method == "POST":
        # R√©cup√©ration des donn√©es
        nom = request.POST.get("nom")
        reference = request.POST.get("reference")
        categorie = request.POST.get("categorie")
        couleur = request.POST.get("couleur")
        pointure_str = request.POST.get("pointure", "").strip()
        phase = request.POST.get("phase")
        prix_str = request.POST.get("prix_unitaire", "").strip().replace(",", ".")
        description = request.POST.get("description")
        qte_disponible_str = request.POST.get("qte_disponible", "0").strip()
        actif = "actif" in request.POST
        image = request.FILES.get("image")

        if not all([nom, reference, categorie, prix_str]):
            messages.error(
                request,
                "Veuillez remplir tous les champs obligatoires (Nom, R√©f√©rence, Cat√©gorie, Prix).",
            )
        else:
            try:
                prix_unitaire = float(prix_str)
                qte_disponible = int(qte_disponible_str) if qte_disponible_str else 0
                pointure = pointure_str if pointure_str else None

                article = Article.objects.create(
                    nom=nom,
                    reference=reference,
                    categorie=categorie,
                    couleur=couleur,
                    pointure=pointure,
                    phase=phase,
                    prix_unitaire=prix_unitaire,
                    description=description,
                    qte_disponible=qte_disponible,
                    actif=actif,
                    image=image,
                )
                messages.success(
                    request, f"L'article '{article.nom}' a √©t√© cr√©√© avec succ√®s."
                )
                return redirect("Prepacommande:liste_articles")
            except (ValueError, TypeError):
                messages.error(
                    request, "Le prix et la quantit√© doivent √™tre des nombres valides."
                )

    context = {
        "article_phases": Article.PHASE_CHOICES,
        "page_title": "Cr√©er un Nouvel Article",
        "page_subtitle": "Ajouter un article au catalogue",
    }
    return render(request, "Prepacommande/stock/creer_article.html", context)

# === FONCTION SUPPRIM√âE : GESTION DE STOCK D√âPLAC√âE VERS ADMIN ===
# @login_required
# def modifier_article(request, article_id):
#     """Modifier un article existant - Service de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Acc√®s non autoris√©.")
            return redirect("login")
    except Operateur.DoesNotExist:
        messages.error(request, "Profil op√©rateur non trouv√©.")
        return redirect("login")
    
    article = get_object_or_404(Article, pk=article_id)
    
    if request.method == "POST":
        form = ArticleForm(request.POST, request.FILES, instance=article)
        if form.is_valid():
            # Sauvegarder l'article
            article_modifie = form.save()
            
            # G√©rer l'ajustement de stock optionnel
            type_mouvement_stock = request.POST.get("type_mouvement_stock")
            quantite_ajustement = request.POST.get("quantite_ajustement")
            commentaire_ajustement = request.POST.get("commentaire_ajustement", "")
            
            if type_mouvement_stock and quantite_ajustement:
                try:
                    quantite = int(quantite_ajustement)
                    if quantite > 0:
                        print(
                            f"üîß Ajustement stock via modification - Article: {article_modifie.nom}, Type: {type_mouvement_stock}, Quantit√©: {quantite}"
                        )
                        print(
                            f"üîß Stock avant ajustement: {article_modifie.qte_disponible}"
                        )
                        
                        mouvement = creer_mouvement_stock(
                            article=article_modifie,
                            quantite=quantite,
                            type_mouvement=type_mouvement_stock,
                            operateur=operateur_profile,
                            commentaire=f"Ajustement via modification article. {commentaire_ajustement}".strip(),
                        )
                        
                        # Recharger l'article pour voir le stock mis √† jour
                        article_modifie.refresh_from_db()
                        print(
                            f"‚úÖ Stock apr√®s ajustement: {article_modifie.qte_disponible}"
                        )
                        
                        if mouvement:
                            messages.success(
                                request,
                                f"L'article '{article_modifie.nom}' a √©t√© modifi√© avec succ√®s. Stock ajust√© : nouveau stock = {article_modifie.qte_disponible} unit√©s.",
                            )
                        else:
                            messages.warning(
                                request,
                                f"L'article '{article_modifie.nom}' a √©t√© modifi√© avec succ√®s, mais l'ajustement de stock a √©chou√©.",
                            )
                    else:
                        messages.warning(
                            request,
                            f"L'article '{article_modifie.nom}' a √©t√© modifi√© avec succ√®s, mais la quantit√© d'ajustement doit √™tre positive.",
                        )
                except (ValueError, TypeError) as e:
                    print(f"‚ùå Erreur lors de l'ajustement de stock: {str(e)}")
                    messages.warning(
                        request,
                        f"L'article '{article_modifie.nom}' a √©t√© modifi√© avec succ√®s, mais l'ajustement de stock a √©chou√© (quantit√© invalide).",
                    )
                except Exception as e:
                    print(f"‚ùå Erreur lors de l'ajustement de stock: {str(e)}")
                    import traceback

                    traceback.print_exc()
                    messages.warning(
                        request,
                        f"L'article '{article_modifie.nom}' a √©t√© modifi√© avec succ√®s, mais l'ajustement de stock a √©chou√©.",
                    )
            else:
                messages.success(
                    request,
                    f"L'article '{article_modifie.nom}' a √©t√© modifi√© avec succ√®s.",
                )
                
            return redirect(
                "Prepacommande:detail_article", article_id=article_modifie.id
            )
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        form = ArticleForm(instance=article)

    context = {
        "form": form,
        "article": article,
        "page_title": "Modifier l'Article",
        "page_subtitle": f"Mise √† jour de {article.nom}",
    }
    return render(request, "Prepacommande/stock/modifier_article.html", context)


# === NOUVELLES FONCTIONNALIT√âS : R√âPARTITION AUTOMATIQUE ===


def get_operateur_display_name(operateur):
    """Fonction helper pour obtenir le nom d'affichage d'un op√©rateur"""
    if not operateur:
        return "Op√©rateur inconnu"
    
    if hasattr(operateur, "nom_complet") and operateur.nom_complet:
        return operateur.nom_complet
    elif operateur.nom and operateur.prenom:
        return f"{operateur.prenom} {operateur.nom}"
    elif operateur.nom:
        return operateur.nom
    elif hasattr(operateur, "user") and operateur.user:
        return operateur.user.username
    else:
        return "Op√©rateur inconnu"


# === VUES DE R√âPARTITION SUPPRIM√âES (D√âPLAC√âES VERS ADMIN) ===
# Les vues de r√©partition ont √©t√© d√©plac√©es vers l'interface admin
# car ce sont les administrateurs qui s'en occupent maintenant

# === FONCTIONS UTILITAIRES DE R√âPARTITION SUPPRIM√âES (D√âPLAC√âES VERS ADMIN) ===

# === VUES DE GESTION DES ENVOIS ===


@login_required
def etats_livraison(request):
    """Gestion des √©tats de livraison - Service de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Acc√®s non autoris√©.")
            return redirect("login")
    except Operateur.DoesNotExist:
        messages.error(request, "Profil op√©rateur non trouv√©.")
        return redirect("login")
    
    from parametre.models import Region
    
    # Filtres
    date_debut = request.GET.get("date_debut")
    date_fin = request.GET.get("date_fin")
    region_id = request.GET.get("region")
    statut = request.GET.get("statut")
    
    # Base queryset
    commandes = (
        Commande.objects.filter(
            etats__enum_etat__libelle__in=[
                "En pr√©paration",
                "Pr√™te",
                "En cours de livraison",
                "Livr√©e",
            ],
            etats__date_fin__isnull=True,
        )
        .select_related("ville__region", "client")
        .prefetch_related("etats__enum_etat", "etats__operateur__user")
        .distinct()
    )
    
    # Appliquer les filtres
    if date_debut:
        commandes = commandes.filter(date_creation__gte=date_debut)
    if date_fin:
        commandes = commandes.filter(date_creation__lte=date_fin)
    if region_id:
        commandes = commandes.filter(ville__region_id=region_id)
    if statut:
        commandes = commandes.filter(
            etats__enum_etat__libelle=statut, etats__date_fin__isnull=True
        )
    
    # Statistiques
    stats = {
        "total_commandes": commandes.count(),
        "en_preparation": commandes.filter(
            etats__enum_etat__libelle="En pr√©paration", etats__date_fin__isnull=True
        ).count(),
        "pretes": commandes.filter(
            etats__enum_etat__libelle="Pr√™te", etats__date_fin__isnull=True
        ).count(),
        "livrees": commandes.filter(
            etats__enum_etat__libelle="Livr√©e", etats__date_fin__isnull=True
        ).count(),
    }
    
    # Pagination
    from django.core.paginator import Paginator

    paginator = Paginator(commandes, 50)
    page_number = request.GET.get("page")
    commandes = paginator.get_page(page_number)
    
    regions = Region.objects.all()
    
    context = {
        "commandes": commandes,
        "regions": regions,
        "stats": stats,
    }

    return render(request, "Prepacommande/etats_livraison.html", context)


@login_required
def export_envois(request):
    """Export des envois journaliers - Service de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Acc√®s non autoris√©.")
            return redirect("login")
    except Operateur.DoesNotExist:
        messages.error(request, "Profil op√©rateur non trouv√©.")
        return redirect("login")
    
    from parametre.models import Region, Operateur
    from django.utils import timezone
    import datetime
    
    # Date par d√©faut : aujourd'hui
    today = timezone.now().date()
    date_envoi = request.GET.get("date_envoi", today)
    region_id = request.GET.get("region")
    livreur_id = request.GET.get("livreur")
    
    # Obtenir tous les livreurs (op√©rateurs de livraison)
    livreurs = Operateur.objects.filter(is_livraison=True, actif=True)
    regions = Region.objects.all()
    
    # Simuler des envois (√† remplacer par votre mod√®le Envoi)
    envois = []
    
    # Commandes PR√âPAR√âES √† √™tre envoy√©es
    commandes_pretes = Commande.objects.filter(
        etats__enum_etat__libelle="Pr√©par√©e", etats__date_fin__isnull=True
    ).select_related("ville__region")
    
    if region_id:
        commandes_pretes = commandes_pretes.filter(ville__region_id=region_id)
    
    # Statistiques
    stats = {
        "total_envois": len(envois),
        "total_commandes": 0,
        "commandes_pretes": commandes_pretes.count(),
        "livreurs_actifs": livreurs.filter(actif=True).count(),
    }
    
    context = {
        "envois": envois,
        "commandes_pretes": commandes_pretes,
        "livreurs": livreurs,
        "regions": regions,
        "stats": stats,
        "today": today,
    }

    return render(request, "Prepacommande/export_envois.html", context)


@login_required
def creer_envoi(request):
    """Cr√©er un nouvel envoi"""
    if request.method == "POST":
        try:
            livreur_id = request.POST.get("livreur")
            region_id = request.POST.get("region")
            notes = request.POST.get("notes", "")
            commandes_selectionnees = request.POST.get(
                "commandes_selectionnees", ""
            ).split(",")
            
            # Ici vous devriez cr√©er l'objet Envoi
            # envoi = Envoi.objects.create(
            #     livreur_id=livreur_id,
            #     region_id=region_id if region_id else None,
            #     notes=notes,
            #     date_creation=timezone.now()
            # )
            
            # Associer les commandes √† l'envoi
            # for commande_id in commandes_selectionnees:
            #     if commande_id:
            #         commande = Commande.objects.get(id=commande_id)
            #         commande.envoi = envoi
            #         commande.save()
            
            return JsonResponse({"success": True, "message": "Envoi cr√©√© avec succ√®s"})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})
    
    return JsonResponse({"success": False, "message": "M√©thode non autoris√©e"})


@login_required
def details_envoi(request, envoi_id):
    """Afficher les d√©tails d'un envoi"""
    # Ici vous devriez r√©cup√©rer l'envoi par son ID
    # envoi = get_object_or_404(Envoi, id=envoi_id)
    
    # Pour l'exemple, retourner un contenu HTML simple
    html_content = f"""
    <div class="p-3">
        <h6>Envoi ENV-{envoi_id}</h6>
        <p><strong>Statut:</strong> En cours</p>
        <p><strong>Date cr√©ation:</strong> {timezone.now().strftime('%d/%m/%Y %H:%M')}</p>
        <p><strong>Commandes associ√©es:</strong> 0</p>
    </div>
    """
    
    return HttpResponse(html_content)


# === VUES D'EXPORT ET D'IMPRESSION ===


@login_required
def details_region_view(request):
    """Vue d√©taill√©e pour afficher les commandes par r√©gion - Service de pr√©paration"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Acc√®s non autoris√©.")
            return redirect("login")
    except Operateur.DoesNotExist:
        messages.error(request, "Profil op√©rateur non trouv√©.")
        return redirect("login")
    
    from parametre.models import Region, Ville
    
    # R√©cup√©rer les param√®tres de filtrage
    region_name = request.GET.get("region")
    ville_name = request.GET.get("ville")
    
    # Base queryset pour toutes les commandes en traitement
    commandes_reparties = (
        Commande.objects.filter(
            etats__enum_etat__libelle__in=[
                "Confirm√©e",
                "√Ä imprimer",
                "Pr√©par√©e",
                "En cours de livraison",
            ],
        etats__date_fin__isnull=True,
        ville__isnull=False,  # Exclure les commandes sans ville
            ville__region__isnull=False,  # Exclure les commandes sans r√©gion
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("etats__operateur", "etats__enum_etat", "paniers__article")
        .distinct()
    )
    
    # Appliquer les filtres
    if region_name:
        commandes_reparties = commandes_reparties.filter(
            ville__region__nom_region=region_name
        )
    if ville_name:
        commandes_reparties = commandes_reparties.filter(ville__nom=ville_name)
    
    # Statistiques par ville dans la r√©gion/ville filtr√©e
    stats_par_ville = (
        commandes_reparties.values(
            "ville__id", "ville__nom", "ville__region__nom_region"
        )
        .annotate(nb_commandes=Count("id"), total_montant=Sum("total_cmd"))
        .order_by("ville__region__nom_region", "ville__nom")
    )
    
    # Statistiques des commandes PR√âPAR√âES par ville dans la r√©gion/ville filtr√©e
    commandes_preparees = Commande.objects.filter(
        etats__enum_etat__libelle="Pr√©par√©e",
        etats__date_fin__isnull=True,
        ville__isnull=False,
        ville__region__isnull=False,
    ).select_related("ville", "ville__region")
    
    # Appliquer les m√™mes filtres que pour les commandes en traitement
    if region_name:
        commandes_preparees = commandes_preparees.filter(
            ville__region__nom_region=region_name
        )
    if ville_name:
        commandes_preparees = commandes_preparees.filter(ville__nom=ville_name)
    
    stats_preparees_par_ville = (
        commandes_preparees.values("ville__nom", "ville__region__nom_region")
        .annotate(nb_commandes_preparees=Count("id"))
        .order_by("ville__region__nom_region", "ville__nom")
    )
    
    # Cr√©er un dictionnaire pour un acc√®s rapide
    preparees_par_ville = {
        (stat["ville__nom"], stat["ville__region__nom_region"]): stat[
            "nb_commandes_preparees"
        ]
        for stat in stats_preparees_par_ville
    }
    
    # Calculer les totaux
    total_commandes = commandes_reparties.count()
    total_montant = commandes_reparties.aggregate(total=Sum("total_cmd"))["total"] or 0
    
    # D√©finir le titre selon le filtre appliqu√©
    if region_name:
        page_title = f"D√©tails - {region_name}"
        page_subtitle = f"Commandes en traitement dans la r√©gion {region_name}"
    elif ville_name:
        page_title = f"D√©tails - {ville_name}"
        page_subtitle = f"Commandes en traitement √† {ville_name}"
    else:
        page_title = "D√©tails par R√©gion"
        page_subtitle = "R√©partition d√©taill√©e des commandes en traitement"
    
    context = {
        "operateur": operateur_profile,
        "commandes_reparties": commandes_reparties,
        "stats_par_ville": stats_par_ville,
        "preparees_par_ville": preparees_par_ville,
        "total_commandes": total_commandes,
        "total_montant": total_montant,
        "region_name": region_name,
        "ville_name": ville_name,
        "page_title": page_title,
        "page_subtitle": page_subtitle,
    }

    return render(request, "Prepacommande/details_region.html", context)


@login_required
def imprimer_commande(request, commande_id):
    """
    Imprime une commande sp√©cifique.
    """
    commande = get_object_or_404(Commande, id=commande_id)
    # Assurez-vous que l'op√©rateur a le droit de voir cette commande si n√©cessaire
    return render(
        request, "Prepacommande/impression_commande.html", {"commande": commande}
    )


@login_required 
def exporter_etats_pdf(request):
    """
    Exporte l'√©tat actuel des livraisons en PDF.
    """
    # Votre logique d'exportation PDF ici
    return HttpResponse(
        "Export PDF des √©tats de livraison √† impl√©menter.", content_type="text/plain"
    )


@login_required
def imprimer_envoi(request, envoi_id):
    """
    Imprime les d√©tails d'un envoi.
    """
    envoi = get_object_or_404(Envoi, id=envoi_id)
    return render(request, "Prepacommande/impression_envoi.html", {"envoi": envoi})


@login_required
def exporter_envoi(request, envoi_id, format):
    """
    Exporte un envoi dans un format sp√©cifique (CSV/PDF).
    """
    envoi = get_object_or_404(Envoi, id=envoi_id)
    if format == "csv":
        # Logique d'export CSV
        return HttpResponse(
            f"Export CSV de l'envoi {envoi_id}", content_type="text/csv"
        )
    elif format == "pdf":
        # Logique d'export PDF
        return HttpResponse(
            f"Export PDF de l'envoi {envoi_id}", content_type="application/pdf"
        )
    return HttpResponse("Format non support√©", status=400)


@login_required
def exporter_envois_journaliers(request):
    """
    Exporte tous les envois du jour.
    """
    # Votre logique d'exportation ici
    return HttpResponse(
        "Export des envois journaliers √† impl√©menter.", content_type="text/plain"
    )


@login_required
def rafraichir_articles_commande_prepa(request, commande_id):
    """Rafra√Æchir la section des articles de la commande en pr√©paration"""
    print(f"üîÑ Rafra√Æchissement des articles pour la commande {commande_id}")
    
    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse(
            {"error": "Profil d'op√©rateur de pr√©paration non trouv√©."}, status=403
        )
    
    try:
        commande = Commande.objects.get(id=commande_id)
        
        # V√©rifier que la commande est affect√©e √† cet op√©rateur
        etat_preparation = commande.etats.filter(
            operateur=operateur,
            enum_etat__libelle__in=["En pr√©paration", "√Ä imprimer"],
            date_fin__isnull=True,
        ).first()
        
        if not etat_preparation:
            return JsonResponse(
                {"error": "Cette commande ne vous est pas affect√©e."}, status=403
            )
        
        # G√©n√©rer le HTML directement pour √©viter les erreurs de template
        paniers = commande.paniers.select_related("article").all()
        html_rows = []
        
        for panier in paniers:
            # Utiliser le filtre Django pour calculer le prix selon la logique upsell
            from commande.templatetags.commande_filters import (
                get_prix_upsell_avec_compteur,
            )
            
            prix_calcule = get_prix_upsell_avec_compteur(
                panier.article, commande.compteur
            )
            
            # D√©terminer l'affichage selon le type d'article
            if panier.article.isUpsell:
                prix_class = "text-orange-600"
                upsell_badge = f'<span class="inline-flex items-center px-2 py-1 rounded-full text-xs font-bold bg-orange-100 text-orange-700 ml-2"><i class="fas fa-arrow-up mr-1"></i>Upsell</span>'
                
                # D√©terminer le niveau d'upsell affich√©
                if commande.compteur >= 4 and panier.article.prix_upsell_4 is not None:
                    niveau_upsell = 4
                elif (
                    commande.compteur >= 3 and panier.article.prix_upsell_3 is not None
                ):
                    niveau_upsell = 3
                elif (
                    commande.compteur >= 2 and panier.article.prix_upsell_2 is not None
                ):
                    niveau_upsell = 2
                elif (
                    commande.compteur >= 1 and panier.article.prix_upsell_1 is not None
                ):
                    niveau_upsell = 1
                else:
                    niveau_upsell = 0
                
                if niveau_upsell > 0:
                    upsell_info = f'<div class="text-xs text-orange-600 flex items-center justify-center gap-1"><i class="fas fa-arrow-up"></i>Upsell Niveau {niveau_upsell}</div>'
                else:
                    upsell_info = f'<div class="text-xs text-orange-600 flex items-center justify-center gap-1"><i class="fas fa-arrow-up"></i>Upsell (Prix normal)</div>'
            else:
                # Prix normal
                prix_class = "text-gray-700"
                upsell_badge = ""
                upsell_info = ""
            
            sous_total = prix_calcule * panier.quantite
            
            html_row = f"""
            <tr data-panier-id="{panier.id}" data-article-id="{panier.article.id}">
                <td class="px-4 py-3">
                    <div class="flex items-center">
                        <div class="flex-1">
                            <div class="font-medium text-gray-900 flex items-center">
                                {panier.article.nom}
                                {upsell_badge}
                            </div>
                            <div class="text-sm text-gray-500">
                                R√©f: {panier.article.reference or 'N/A'}
                            </div>
                        </div>
                    </div>
                </td>
                <td class="px-4 py-3 text-center">
                    <div class="flex items-center justify-center space-x-2">
                        <button type="button" onclick="modifierQuantite({panier.id}, -1)" 
                                class="w-8 h-8 bg-gray-200 hover:bg-gray-300 rounded text-sm transition-colors">-</button>
                        <input type="number" id="quantite-{panier.id}" value="{panier.quantite}" min="1" 
                               class="w-16 p-2 border rounded-lg text-center" 
                               onchange="modifierQuantiteDirecte({panier.id}, this.value)">
                        <button type="button" onclick="modifierQuantite({panier.id}, 1)" 
                                class="w-8 h-8 bg-gray-200 hover:bg-gray-300 rounded text-sm transition-colors">+</button>
                    </div>
                </td>
                <td class="px-4 py-3 text-center">
                    <div class="font-medium {prix_class}" id="prix-unitaire-{panier.id}">
                        {prix_calcule:.2f} DH
                    </div>
                    {upsell_info}
                </td>
                <td class="px-4 py-3 text-center">
                    <div class="font-bold text-gray-900">
                        {sous_total:.2f} DH
                    </div>
                </td>
                <td class="px-4 py-3 text-center">
                    <button type="button" onclick="supprimerArticle({panier.id})" 
                            class="px-3 py-2 bg-red-500 hover:bg-red-600 text-white text-sm rounded-lg transition-colors">
                        <i class="fas fa-trash mr-1"></i>Supprimer
                    </button>
                </td>
            </tr>
            """
            html_rows.append(html_row)
        
        html = "".join(html_rows)
        
        # Calculer les statistiques upsell
        articles_upsell = commande.paniers.filter(article__isUpsell=True)
        total_quantite_upsell = (
            articles_upsell.aggregate(total=Sum("quantite"))["total"] or 0
        )
        
        response_data = {
            "success": True,
            "html": html,
            "articles_count": commande.paniers.count(),
            "total_commande": float(commande.total_cmd),
            "sous_total_articles": float(commande.sous_total_articles),
            "compteur": commande.compteur,
            "articles_upsell": articles_upsell.count(),
            "quantite_totale_upsell": total_quantite_upsell,
        }

        print(
            f"‚úÖ Rafra√Æchissement termin√© - Articles: {response_data['articles_count']}, Compteur: {response_data['compteur']}"
        )
        return JsonResponse(response_data)
        
    except Commande.DoesNotExist:
        print(f"‚ùå Commande {commande_id} non trouv√©e")
        return JsonResponse({"error": "Commande non trouv√©e"}, status=404)
    except Exception as e:
        print(f"‚ùå Erreur lors du rafra√Æchissement: {str(e)}")
        import traceback

        traceback.print_exc()
        return JsonResponse({"error": f"Erreur interne: {str(e)}"}, status=500)


@login_required
def ajouter_article_commande_prepa(request, commande_id):
    """Ajouter un article √† la commande en pr√©paration"""
    if request.method != "POST":
        return JsonResponse({"error": "M√©thode non autoris√©e"}, status=405)
    
    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"error": "Profil d'op√©rateur non trouv√©."}, status=403)
    
    try:
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # V√©rifier que la commande est bien en pr√©paration pour cet op√©rateur
            etat_preparation = commande.etats.filter(
                operateur=operateur,
                enum_etat__libelle__in=["En pr√©paration", "√Ä imprimer"],
                date_fin__isnull=True,
            ).first()
            
            if not etat_preparation:
                return JsonResponse(
                    {"error": "Cette commande n'est pas en pr√©paration pour vous."},
                    status=403,
                )
            
            # Support both parameter names for backward compatibility
            article_id = request.POST.get("articleId") or request.POST.get("article_id")
            quantite = int(request.POST.get("quantite", 1))
            variante_id = request.POST.get("varianteId")
            
            if not article_id or quantite <= 0:
                return JsonResponse({"error": "Donn√©es invalides"}, status=400)

            article = Article.objects.get(id=article_id)
            
            # Handle variant if provided
            variante = None
            if variante_id:
                try:
                    from article.models import VarianteArticle
                    variante = VarianteArticle.objects.get(id=variante_id, article=article)
                except VarianteArticle.DoesNotExist:
                    return JsonResponse({"error": "Variante non trouv√©e"}, status=404)
            
            # D√©cr√©menter le stock et cr√©er un mouvement
            creer_mouvement_stock(
                article=article,
                quantite=quantite,
                type_mouvement="sortie",
                commande=commande,
                operateur=operateur,
                commentaire=f"Ajout article pendant pr√©paration cmd {commande.id_yz}",
            )
            
            # V√©rifier si l'article existe d√©j√† dans la commande
            panier_existant = Panier.objects.filter(
                commande=commande, article=article
            ).first()
            
            if panier_existant:
                # Si l'article existe d√©j√†, mettre √† jour la quantit√©
                panier_existant.quantite += quantite
                panier_existant.save()
                panier = panier_existant
                print(
                    f"üîÑ Article existant mis √† jour: ID={article.id}, nouvelle quantit√©={panier.quantite}"
                )
            else:
                # Si l'article n'existe pas, cr√©er un nouveau panier
                panier = Panier.objects.create(
                    commande=commande,
                    article=article,
                    quantite=quantite,
                    sous_total=0,  # Sera recalcul√© apr√®s
                )
                print(f"‚ûï Nouvel article ajout√©: ID={article.id}, quantit√©={quantite}")
            
            # Recalculer le compteur apr√®s ajout (logique de confirmation)
            if (
                article.isUpsell
                and hasattr(article, "prix_upsell_1")
                and article.prix_upsell_1 is not None
            ):
                # Compter la quantit√© totale d'articles upsell (apr√®s ajout)
                total_quantite_upsell = (
                    commande.paniers.filter(article__isUpsell=True).aggregate(
                        total=Sum("quantite")
                    )["total"]
                    or 0
                )
                
                # Le compteur ne s'incr√©mente qu'√† partir de 2 unit√©s d'articles upsell
                # 0-1 unit√©s upsell ‚Üí compteur = 0
                # 2+ unit√©s upsell ‚Üí compteur = total_quantite_upsell - 1
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()
            else:
                # Pour les articles normaux, juste calculer le sous-total
                from commande.templatetags.commande_filters import (
                    get_prix_upsell_avec_compteur,
                )

                prix_unitaire = get_prix_upsell_avec_compteur(
                    article, commande.compteur
                )
                sous_total = prix_unitaire * panier.quantite
                panier.sous_total = float(sous_total)
                panier.save()
            
            # Recalculer le total
            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()
            
            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = (
                articles_upsell.aggregate(total=Sum("quantite"))["total"] or 0
            )
            
            # D√©terminer si c'√©tait un ajout ou une mise √† jour
            message = (
                "Article ajout√©"
                if not panier_existant
                else f"Quantit√© mise √† jour ({panier.quantite})"
            )
            
            # Pr√©parer les donn√©es de l'article pour le frontend
            article_data = {
                "panier_id": panier.id,
                "article_id": article.id,
                "nom": article.nom,
                "reference": article.reference,
                "couleur_fr": article.couleur or "",
                "couleur_ar": article.couleur or "",
                "pointure": article.pointure or "",
                "quantite": panier.quantite,
                "prix": panier.sous_total / panier.quantite,  # Prix unitaire
                "sous_total": panier.sous_total,
                "is_upsell": article.isUpsell,
                "compteur": commande.compteur,
                "description": article.description or "",
            }

            return JsonResponse(
                {
                    "success": True,
                    "message": message,
                    "compteur": commande.compteur,
                    "total_commande": float(commande.total_cmd),
                    "was_update": panier_existant is not None,
                    "new_quantity": panier.quantite,
                    "article_data": article_data,
                    "articles_count": commande.paniers.count(),
                    "sous_total_articles": float(
                        sum(p.sous_total for p in commande.paniers.all())
                    ),
                    "articles_upsell": articles_upsell.count(),
                    "quantite_totale_upsell": total_quantite_upsell,
                }
            )
            
    except Article.DoesNotExist:
        return JsonResponse({"error": "Article non trouv√©"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"Erreur interne: {str(e)}"}, status=500)


@login_required
def modifier_quantite_article_prepa(request, commande_id):
    """Modifier la quantit√© d'un article dans la commande en pr√©paration"""
    if request.method != "POST":
        return JsonResponse({"error": "M√©thode non autoris√©e"}, status=405)

    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"error": "Profil d'op√©rateur non trouv√©."}, status=403)

    try:
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # V√©rifier l'affectation
            if not commande.etats.filter(
                operateur=operateur,
                enum_etat__libelle__in=["En pr√©paration", "√Ä imprimer"],
                date_fin__isnull=True,
            ).exists():
                return JsonResponse({"error": "Commande non affect√©e."}, status=403)

            panier_id = request.POST.get("panier_id")
            nouvelle_quantite = int(request.POST.get("quantite", 1))

            panier = Panier.objects.get(id=panier_id, commande=commande)
            ancienne_quantite = panier.quantite
            article = panier.article
            difference = nouvelle_quantite - ancienne_quantite

            if difference > 0:
                creer_mouvement_stock(
                    article,
                    difference,
                    "sortie",
                    commande,
                    operateur,
                    f"Ajustement qt√© cmd {commande.id_yz}",
                )
            elif difference < 0:
                creer_mouvement_stock(
                    article,
                    abs(difference),
                    "entree",
                    commande,
                    operateur,
                    f"Ajustement qt√© cmd {commande.id_yz}",
                )

            panier.quantite = nouvelle_quantite
            
            # Recalculer le compteur si c'est un article upsell
            if article.isUpsell:
                # Compter la quantit√© totale d'articles upsell apr√®s modification
                total_quantite_upsell = (
                    commande.paniers.filter(article__isUpsell=True).aggregate(
                        total=Sum("quantite")
                    )["total"]
                    or 0
                )
                
                # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()
            else:
                # Pour les articles normaux, juste calculer le sous-total
                from commande.templatetags.commande_filters import (
                    get_prix_upsell_avec_compteur,
                )

                prix_unitaire = get_prix_upsell_avec_compteur(
                    article, commande.compteur
                )
                panier.sous_total = prix_unitaire * nouvelle_quantite
            panier.save()

            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()

            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = (
                articles_upsell.aggregate(total=Sum("quantite"))["total"] or 0
            )

            return JsonResponse(
                {
                    "success": True,
                    "message": "Quantit√© modifi√©e",
                    "compteur": commande.compteur,
                    "articles_upsell": articles_upsell.count(),
                    "quantite_totale_upsell": total_quantite_upsell,
                    "total_commande": float(commande.total_cmd),
                    "sous_total_articles": float(commande.sous_total_articles),
                }
            )

    except Panier.DoesNotExist:
        return JsonResponse({"error": "Panier non trouv√©"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"Erreur interne: {str(e)}"}, status=500)


@login_required
def supprimer_article_commande_prepa(request, commande_id):
    """Supprimer un article de la commande en pr√©paration"""
    if request.method != "POST":
        return JsonResponse({"error": "M√©thode non autoris√©e"}, status=405)

    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"error": "Profil d'op√©rateur non trouv√©."}, status=403)

    try:
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # V√©rifier l'affectation
            if not commande.etats.filter(
                operateur=operateur,
                enum_etat__libelle__in=["En pr√©paration", "√Ä imprimer"],
                date_fin__isnull=True,
            ).exists():
                return JsonResponse({"error": "Commande non affect√©e."}, status=403)

            panier_id = request.POST.get("panier_id")
            panier = Panier.objects.get(id=panier_id, commande=commande)
            quantite_supprimee = panier.quantite
            article = panier.article
            
            creer_mouvement_stock(
                article,
                quantite_supprimee,
                "entree",
                commande,
                operateur,
                f"Suppression article cmd {commande.id_yz}",
            )
            
            # Sauvegarder l'info avant suppression
            etait_upsell = panier.article.isUpsell
            
            # Supprimer l'article
            panier.delete()

            # Recalculer le compteur apr√®s suppression (logique de confirmation)
            if etait_upsell:
                # Compter la quantit√© totale d'articles upsell restants (apr√®s suppression)
                total_quantite_upsell = (
                    commande.paniers.filter(article__isUpsell=True).aggregate(
                        total=Sum("quantite")
                    )["total"]
                    or 0
                )
                
                # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()

            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()

            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = (
                articles_upsell.aggregate(total=Sum("quantite"))["total"] or 0
            )

            return JsonResponse(
                {
                    "success": True,
                    "message": "Article supprim√©",
                    "compteur": commande.compteur,
                    "articles_upsell": articles_upsell.count(),
                    "quantite_totale_upsell": total_quantite_upsell,
                    "total_commande": float(commande.total_cmd),
                    "sous_total_articles": float(commande.sous_total_articles),
                }
            )

    except Panier.DoesNotExist:
        return JsonResponse({"error": "Panier non trouv√©"}, status=404)
    except Exception as e:
        return JsonResponse({"error": f"Erreur interne: {str(e)}"}, status=500)


# === VUES DE R√âPARTITION SUPPRIM√âES (D√âPLAC√âES VERS ADMIN) ===
# Les vues de r√©partition ont √©t√© d√©plac√©es vers l'interface admin
# car ce sont les administrateurs qui s'en occupent maintenant


@login_required
def api_panier_commande_livraison(request, commande_id):
    """API pour r√©cup√©rer le panier d'une commande pour les op√©rateurs de livraison"""
    try:
        # V√©rifier que l'utilisateur est un op√©rateur de livraison
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="LOGISTIQUE"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"success": False, "message": "Acc√®s non autoris√©"})
    
    # R√©cup√©rer la commande
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({"success": False, "message": "Commande non trouv√©e"})
    
    # R√©cup√©rer les paniers de la commande
    paniers = commande.paniers.all().select_related("article")
    
    paniers_data = []
    for panier in paniers:
        paniers_data.append(
            {
                "id": panier.id,
                "article_id": panier.article.id,
                "nom": panier.article.nom,
                "reference": panier.article.reference,
                "couleur": panier.article.couleur,
                "pointure": panier.article.pointure,
                "prix_unitaire": float(panier.article.prix_unitaire),
                "quantite": panier.quantite,
                "sous_total": float(panier.sous_total),
                "qte_disponible": panier.article.qte_disponible,
            }
        )

    return JsonResponse(
        {
            "success": True,
            "paniers": paniers_data,
            "total_commande": float(commande.total_cmd),
        }
    )


@login_required
def api_articles_commande_livree_partiellement(request, commande_id):
    """API pour r√©cup√©rer les d√©tails des articles d'une commande livr√©e partiellement"""
    import json
    from article.models import Article
    from commande.models import Commande, EtatCommande, EnumEtatCmd, Operation
    from parametre.models import Operateur

    try:
        # V√©rifier que l'utilisateur est un op√©rateur de pr√©paration
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"success": False, "message": "Acc√®s non autoris√©"})
    
    # R√©cup√©rer la commande
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({"success": False, "message": "Commande non trouv√©e"})
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "message": f"Erreur lors de la r√©cup√©ration de la commande: {str(e)}",
            }
        )
    
    # Analyser les articles pour les commandes livr√©es partiellement
    articles_livres = []
    articles_renvoyes = []
    
    # R√©cup√©rer tous les √©tats de la commande
    etats_commande = (
        commande.etats.all()
        .select_related("enum_etat", "operateur")
        .order_by("date_debut")
    )
    
    # D√©terminer l'√©tat actuel
    etat_actuel = etats_commande.filter(date_fin__isnull=True).first()
    
    # R√©cup√©rer l'√©tat pr√©c√©dent pour comprendre d'o√π vient la commande
    etat_precedent = None
    if etat_actuel:
        # Trouver l'√©tat pr√©c√©dent (le dernier √©tat termin√© avant l'√©tat actuel)
        for etat in reversed(etats_commande):
            if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                if etat.enum_etat.libelle not in ["√Ä imprimer", "En pr√©paration"]:
                    etat_precedent = etat
                    break
    
    # V√©rifier si c'est une commande livr√©e partiellement
    if etat_actuel and etat_actuel.enum_etat.libelle == "Livr√©e Partiellement":
        # Les articles dans cette commande sont ceux qui ont √©t√© livr√©s partiellement
        for panier in commande.paniers.all():
            articles_livres.append(
                {
                    "article_id": panier.article.id,
                    "nom": panier.article.nom,
                    "reference": panier.article.reference,
                    "couleur": panier.article.couleur,
                    "pointure": panier.article.pointure,
                    "quantite_livree": panier.quantite,
                    "prix": float(panier.article.prix_unitaire),
                    "sous_total": float(panier.sous_total),
                }
            )
        
        # Chercher la commande de renvoi associ√©e
        commande_renvoi = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande.num_cmd}", client=commande.client
        ).first()
        
        if commande_renvoi:
            # R√©cup√©rer l'√©tat des articles renvoy√©s depuis l'op√©ration de livraison partielle
            etat_articles_renvoyes = {}
            operation_livraison_partielle = (
                commande.operations.filter(type_operation="LIVRAISON_PARTIELLE")
                .order_by("-date_operation")
                .first()
            )
            if operation_livraison_partielle:
                try:
                    details = json.loads(operation_livraison_partielle.conclusion)
                    if "recap_articles_renvoyes" in details:
                        for item in details["recap_articles_renvoyes"]:
                            etat_articles_renvoyes[item["article_id"]] = item["etat"]
                except Exception:
                    pass
            if commande_renvoi:
                for panier_renvoi in commande_renvoi.paniers.all():
                    etat = etat_articles_renvoyes.get(panier_renvoi.article.id, "bon")
                    articles_renvoyes.append(
                        {
                            "article_id": panier_renvoi.article.id,
                            "nom": panier_renvoi.article.nom,
                            "reference": panier_renvoi.article.reference,
                            "couleur": panier_renvoi.article.couleur,
                            "pointure": panier_renvoi.article.pointure,
                            "quantite": panier_renvoi.quantite,
                            "prix": float(panier_renvoi.article.prix_unitaire),
                            "sous_total": float(panier_renvoi.sous_total),
                            "etat": etat,
                        }
                    )
    # V√©rifier si c'est une commande renvoy√©e apr√®s livraison partielle
    elif etat_precedent and etat_precedent.enum_etat.libelle == "Livr√©e Partiellement":
        # Chercher la commande originale qui a √©t√© livr√©e partiellement
        commande_originale = Commande.objects.filter(
            num_cmd=commande.num_cmd.replace("RENVOI-", ""), client=commande.client
        ).first()
        # R√©cup√©rer l'√©tat des articles renvoy√©s depuis l'op√©ration de livraison partielle
        etat_articles_renvoyes = {}
        if commande_originale:
            operation_livraison_partielle = (
                commande_originale.operations.filter(
                    type_operation="LIVRAISON_PARTIELLE"
                )
                .order_by("-date_operation")
                .first()
            )
            if operation_livraison_partielle:
                try:
                    details = json.loads(operation_livraison_partielle.conclusion)
                    if "recap_articles_renvoyes" in details:
                        for item in details["recap_articles_renvoyes"]:
                            etat_articles_renvoyes[item["article_id"]] = item["etat"]
                except Exception:
                    pass
        if commande_originale:
            # Les articles dans cette commande de renvoi sont ceux qui ont √©t√© renvoy√©s
            for panier in paniers:
                etat = etat_articles_renvoyes.get(panier.article.id, "bon")
                articles_renvoyes.append(
                    {
                        "article": panier.article,
                        "quantite": panier.quantite,
                        "prix": panier.article.prix_unitaire,
                        "sous_total": panier.sous_total,
                        "etat": etat,
                    }
                )
    
    # R√©cup√©rer les d√©tails de la livraison partielle
    date_livraison_partielle = None
    commentaire_livraison_partielle = None
    operateur_livraison = None
    
    if etat_actuel and etat_actuel.enum_etat.libelle == "Livr√©e Partiellement":
        date_livraison_partielle = etat_actuel.date_debut
        commentaire_livraison_partielle = etat_actuel.commentaire
        operateur_livraison = etat_actuel.operateur
    elif etat_precedent and etat_precedent.enum_etat.libelle == "Livr√©e Partiellement":
        date_livraison_partielle = etat_precedent.date_debut
        commentaire_livraison_partielle = etat_precedent.commentaire
        operateur_livraison = etat_precedent.operateur
    
    try:
        return JsonResponse(
            {
                "success": True,
                "commande": {
                    "id": commande.id,
                    "id_yz": commande.id_yz,
                    "num_cmd": commande.num_cmd,
                    "total_cmd": float(commande.total_cmd),
                    "date_livraison_partielle": date_livraison_partielle.isoformat()
                    if date_livraison_partielle
                    else None,
                    "commentaire_livraison_partielle": commentaire_livraison_partielle,
                    "operateur_livraison": {
                        "nom": operateur_livraison.nom_complet
                        if operateur_livraison
                        else None,
                        "email": operateur_livraison.mail
                        if operateur_livraison
                        else None,
                    }
                    if operateur_livraison
                    else None,
                },
                "articles_livres": articles_livres,
                "articles_renvoyes": articles_renvoyes,
                "total_articles_livres": len(articles_livres),
                "total_articles_renvoyes": len(articles_renvoyes),
            }
        )
    except Exception as e:
        print(f"Erreur lors de la g√©n√©ration de la r√©ponse JSON: {e}")
        return JsonResponse(
            {
                "success": False,
                "message": f"Erreur lors de la g√©n√©ration de la r√©ponse: {str(e)}",
            }
        )


@login_required
def export_commandes_consolidees_csv(request):
    """
    Export CSV consolid√© : chaque commande sur une seule ligne avec articles regroup√©s
    """
    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"error": "Acc√®s non autoris√©"}, status=403)
    
    # R√©cup√©rer les filtres
    region_name = request.GET.get("region")
    ville_name = request.GET.get("ville")
    
    # Construire la requ√™te de base - UNIQUEMENT les commandes PR√âPAR√âES
    commandes_query = (
        Commande.objects.filter(
            etats__enum_etat__libelle="Pr√©par√©e", etats__date_fin__isnull=True
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("paniers__article")
        .distinct()
    )

    commandes = commandes_query.order_by("-date_cmd")
    
    # Cr√©er la r√©ponse CSV
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    filename = f"commandes_consolidees_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    # √âcrire l'en-t√™te BOM pour Excel
    response.write("\ufeff")
    
    writer = csv.writer(response, delimiter=";")
    
    # En-t√™tes
    headers = [
        "N¬∞ Commande",
        "Client",
        "T√©l√©phone",
        "Ville",
        "R√©gion",
        "Articles et Quantit√©s",
        "Prix Total (DH)",
        "Adresse",
        "√âtat",
    ]
    writer.writerow(headers)
    
    # Traiter chaque commande
    for commande in commandes:
        # Construire la liste des articles avec quantit√©s
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = (
            ", ".join(articles_list) if articles_list else "Aucun article"
        )
        
        # √âtat actuel de la commande
        etat_actuel = (
            commande.etat_actuel.enum_etat.libelle
            if commande.etat_actuel
            else "Non d√©fini"
        )
        
        # √âcrire la ligne
        row = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}"
            if commande.client
            else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region
            if commande.ville and commande.ville.region
            else "N/A",
            articles_consolides,
            f"{commande.total_cmd:.2f}" if commande.total_cmd else "0.00",
            commande.adresse or "N/A",
            etat_actuel,
        ]
        writer.writerow(row)
    
    return response


@login_required
def export_commandes_consolidees_excel(request):
    """
    Export Excel consolid√© : chaque commande sur une seule ligne avec articles regroup√©s
    """
    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"error": "Acc√®s non autoris√©"}, status=403)
    
    # R√©cup√©rer les filtres
    region_name = request.GET.get("region")
    ville_name = request.GET.get("ville")
    
    # Construire la requ√™te de base - UNIQUEMENT les commandes PR√âPAR√âES
    commandes_query = (
        Commande.objects.filter(
            etats__enum_etat__libelle="Pr√©par√©e", etats__date_fin__isnull=True
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("paniers__article")
        .distinct()
    )
    
    # Appliquer les filtres
    if region_name:
        commandes_query = commandes_query.filter(ville__region__nom_region=region_name)
    if ville_name:
        commandes_query = commandes_query.filter(ville__nom=ville_name)
    
    commandes = commandes_query.order_by("-date_cmd")
    
    # Cr√©er le fichier Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes Consolid√©es"
    
    # En-t√™tes
    headers = [
        "N¬∞ Commande",
        "Client",
        "T√©l√©phone",
        "Ville",
        "R√©gion",
        "Articles et Quantit√©s",
        "Prix Total (DH)",
        "Adresse",
        "√âtat",
    ]
    
    # Ajouter les en-t√™tes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Traiter chaque commande
    for row, commande in enumerate(commandes, 2):
        # Construire la liste des articles avec quantit√©s
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = (
            ", ".join(articles_list) if articles_list else "Aucun article"
        )
        
        # √âtat actuel de la commande
        etat_actuel = (
            commande.etat_actuel.enum_etat.libelle
            if commande.etat_actuel
            else "Non d√©fini"
        )
        
        # Ajouter les donn√©es
        ws.cell(row=row, column=1, value=commande.id_yz or commande.num_cmd)
        ws.cell(
            row=row,
            column=2,
            value=f"{commande.client.prenom} {commande.client.nom}"
            if commande.client
            else "N/A",
        )
        ws.cell(
            row=row,
            column=3,
            value=commande.client.numero_tel if commande.client else "N/A",
        )
        ws.cell(
            row=row, column=4, value=commande.ville.nom if commande.ville else "N/A"
        )
        ws.cell(
            row=row,
            column=5,
            value=commande.ville.region.nom_region
            if commande.ville and commande.ville.region
            else "N/A",
        )
        ws.cell(row=row, column=6, value=articles_consolides)
        ws.cell(
            row=row,
            column=7,
            value=float(commande.total_cmd) if commande.total_cmd else 0.00,
        )
        ws.cell(row=row, column=8, value=commande.adresse or "N/A")
        ws.cell(row=row, column=9, value=etat_actuel)
        
        # Ajuster la hauteur de la ligne pour les articles
        if len(articles_consolides) > 100:
            ws.row_dimensions[row].height = 30
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 25, 15, 15, 15, 50, 15, 40, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Cr√©er la r√©ponse
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"commandes_consolidees_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_region_consolidee_csv(request, region_name):
    """
    Export CSV consolid√© pour une r√©gion sp√©cifique
    """
    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"error": "Acc√®s non autoris√©"}, status=403)
    
    # R√©cup√©rer les commandes PR√âPAR√âES de la r√©gion
    commandes = (
        Commande.objects.filter(
            etats__enum_etat__libelle="Pr√©par√©e",
        etats__date_fin__isnull=True,
            ville__region__nom_region=region_name,
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("paniers__article")
        .distinct()
        .order_by("-date_cmd")
    )
    
    # Cr√©er la r√©ponse CSV
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    filename = f"region_{region_name.lower().replace(' ', '_')}_consolidee_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    # √âcrire l'en-t√™te BOM pour Excel
    response.write("\ufeff")
    
    writer = csv.writer(response, delimiter=";")
    
    # En-t√™tes
    headers = [
        "N¬∞ Commande",
        "Client",
        "T√©l√©phone",
        "Ville",
        "R√©gion",
        "Articles et Quantit√©s",
        "Prix Total (DH)",
        "Adresse",
        "√âtat",
    ]
    writer.writerow(headers)
    
    # Traiter chaque commande
    for commande in commandes:
        # Construire la liste des articles avec quantit√©s
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = (
            ", ".join(articles_list) if articles_list else "Aucun article"
        )
        
        # √âtat actuel de la commande
        etat_actuel = (
            commande.etat_actuel.enum_etat.libelle
            if commande.etat_actuel
            else "Non d√©fini"
        )
        
        # √âcrire la ligne
        row = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}"
            if commande.client
            else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region
            if commande.ville and commande.ville.region
            else "N/A",
            articles_consolides,
            f"{commande.total_cmd:.2f}" if commande.total_cmd else "0.00",
            commande.adresse or "N/A",
            etat_actuel,
        ]
        writer.writerow(row)
    
    return response


@login_required
def export_region_consolidee_excel(request, region_name):
    """
    Export Excel consolid√© pour une r√©gion sp√©cifique
    """
    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"error": "Acc√®s non autoris√©"}, status=403)
    
    # R√©cup√©rer les commandes PR√âPAR√âES de la r√©gion
    commandes = (
        Commande.objects.filter(
            etats__enum_etat__libelle="Pr√©par√©e",
        etats__date_fin__isnull=True,
            ville__region__nom_region=region_name,
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("paniers__article")
        .distinct()
        .order_by("-date_cmd")
    )
    
    # Cr√©er le fichier Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"R√©gion {region_name}"
    
    # En-t√™tes
    headers = [
        "N¬∞ Commande",
        "Client",
        "T√©l√©phone",
        "Ville",
        "R√©gion",
        "Articles et Quantit√©s",
        "Prix Total (DH)",
        "Adresse",
        "√âtat",
    ]
    
    # Ajouter les en-t√™tes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Traiter chaque commande
    for row, commande in enumerate(commandes, 2):
        # Construire la liste des articles avec quantit√©s
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = (
            ", ".join(articles_list) if articles_list else "Aucun article"
        )
        
        # √âtat actuel de la commande
        etat_actuel = (
            commande.etat_actuel.enum_etat.libelle
            if commande.etat_actuel
            else "Non d√©fini"
        )
        
        # Ajouter les donn√©es
        ws.cell(row=row, column=1, value=commande.id_yz or commande.num_cmd)
        ws.cell(
            row=row,
            column=2,
            value=f"{commande.client.prenom} {commande.client.nom}"
            if commande.client
            else "N/A",
        )
        ws.cell(
            row=row,
            column=3,
            value=commande.client.numero_tel if commande.client else "N/A",
        )
        ws.cell(
            row=row, column=4, value=commande.ville.nom if commande.ville else "N/A"
        )
        ws.cell(
            row=row,
            column=5,
            value=commande.ville.region.nom_region
            if commande.ville and commande.ville.region
            else "N/A",
        )
        ws.cell(row=row, column=6, value=articles_consolides)
        ws.cell(
            row=row,
            column=7,
            value=float(commande.total_cmd) if commande.total_cmd else 0.00,
        )
        ws.cell(row=row, column=8, value=commande.adresse or "N/A")
        ws.cell(row=row, column=9, value=etat_actuel)
        
        # Ajuster la hauteur de la ligne pour les articles
        if len(articles_consolides) > 100:
            ws.row_dimensions[row].height = 30
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 25, 15, 15, 15, 50, 15, 40, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Ajouter une feuille de r√©sum√©
    ws_resume = wb.create_sheet("R√©sum√©")
    
    # Statistiques de la r√©gion
    total_commandes = commandes.count()
    total_montant = sum(float(cmd.total_cmd) for cmd in commandes if cmd.total_cmd)
    
    resume_data = [
        ["R√©gion", region_name],
        ["Nombre de commandes", total_commandes],
        ["Montant total", f"{total_montant:.2f} DH"],
        ["Date d'export", timezone.now().strftime("%d/%m/%Y %H:%M")],
    ]
    
    for row, (label, value) in enumerate(resume_data, 1):
        ws_resume.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_resume.cell(row=row, column=2, value=value)
    
    # Ajuster la largeur des colonnes du r√©sum√©
    ws_resume.column_dimensions["A"].width = 25
    ws_resume.column_dimensions["B"].width = 20
    
    # Cr√©er la r√©ponse
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"region_{region_name.lower().replace(' ', '_')}_consolidee_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_ville_consolidee_csv(request, ville_id):
    """
    Export CSV consolid√© pour une ville sp√©cifique
    """
    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"error": "Acc√®s non autoris√©"}, status=403)
    
    # R√©cup√©rer la ville
    try:
        ville = Ville.objects.get(id=ville_id)
    except Ville.DoesNotExist:
        return JsonResponse({"error": "Ville non trouv√©e"}, status=404)
    
    # R√©cup√©rer les commandes PR√âPAR√âES de la ville
    commandes = (
        Commande.objects.filter(
            etats__enum_etat__libelle="Pr√©par√©e",
        etats__date_fin__isnull=True,
            ville=ville,
        )
        .select_related("client", "ville", "ville__region")
        .prefetch_related("paniers__article")
        .distinct()
        .order_by("-date_cmd")
    )
    
    # Cr√©er la r√©ponse CSV
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response[
        "Content-Disposition"
    ] = f'attachment; filename="ville_{ville.nom}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # √âcrire l'en-t√™te BOM pour Excel
    response.write("\ufeff")
    
    writer = csv.writer(response, delimiter=";")
    
    # En-t√™tes
    writer.writerow(
        [
            "N¬∞ Commande",
            "Client",
            "T√©l√©phone",
            "Ville",
            "R√©gion",
            "Articles et Quantit√©s",
            "Prix Total (DH)",
            "Adresse",
            "√âtat",
        ]
    )
    
    # Traiter chaque commande
    for commande in commandes:
        # Construire la liste des articles avec quantit√©s
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = (
            ", ".join(articles_list) if articles_list else "Aucun article"
        )
        
        # √âtat actuel de la commande
        etat_actuel = (
            commande.etat_actuel.enum_etat.libelle
            if commande.etat_actuel
            else "Non d√©fini"
        )
        
        # √âcrire la ligne
        row = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}"
            if commande.client
            else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region
            if commande.ville and commande.ville.region
            else "N/A",
            articles_consolides,
            float(commande.total_cmd) if commande.total_cmd else 0.00,
            commande.adresse or "N/A",
            etat_actuel,
        ]
        writer.writerow(row)
    
    return response


@login_required
def export_ville_consolidee_excel(request, ville_id):
    """
    Export Excel consolid√© pour une ville sp√©cifique
    """
    try:
        ville = get_object_or_404(Ville, id=ville_id)
        
        # R√©cup√©rer toutes les commandes de cette ville
        commandes = (
            Commande.objects.filter(
            ville=ville,
                etat_actuel__enum_etat__libelle__in=[
                    "Confirm√©e",
                    "En pr√©paration",
                    "Pr√™te",
                    "Livr√©e",
                ],
            )
            .select_related(
                "client",
                "ville",
                "ville__region",
                "etat_actuel",
                "etat_actuel__enum_etat",
            )
            .prefetch_related("paniers__article")
            .order_by("date_cmd")
        )
        
        if not commandes.exists():
            messages.warning(
                request, f"Aucune commande trouv√©e pour la ville {ville.nom}"
            )
            return redirect("Prepacommande:liste_prepa")
    
        # Cr√©er le fichier Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
    
        wb = Workbook()
        ws = wb.active
        ws.title = f"Commandes {ville.nom}"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    
        # En-t√™tes
        headers = [
            "ID YZ",
            "Num√©ro",
            "Date",
            "Client",
            "T√©l√©phone",
            "Adresse",
            "Ville",
            "R√©gion",
            "√âtat",
            "Total Articles",
            "Frais Livraison",
            "Total Commande",
            "Compteur Upsell",
            "Articles",
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Donn√©es
        row = 2
        for commande in commandes:
            # Articles d√©taill√©s
            articles_detail = []
            for panier in commande.paniers.all():
                article_info = f"{panier.article.nom} (Qt√©: {panier.quantite}, Prix: {panier.sous_total:.2f} DH)"
                if panier.article.isUpsell:
                    article_info += " [UPSELL]"
                articles_detail.append(article_info)
            
            articles_text = "\n".join(articles_detail)
            
            ws.cell(row=row, column=1, value=commande.id_yz).border = border
            ws.cell(row=row, column=2, value=commande.num_cmd).border = border
            ws.cell(
                row=row, column=3, value=commande.date_cmd.strftime("%d/%m/%Y")
            ).border = border
            ws.cell(
                row=row,
                column=4,
                value=f"{commande.client.nom} {commande.client.prenom}",
            ).border = border
            ws.cell(row=row, column=5, value=commande.client.numero_tel).border = border
            ws.cell(row=row, column=6, value=commande.adresse).border = border
            ws.cell(row=row, column=7, value=commande.ville.nom).border = border
            ws.cell(
                row=row, column=8, value=commande.ville.region.nom_region
            ).border = border
            ws.cell(
                row=row, column=9, value=commande.etat_actuel.enum_etat.libelle
            ).border = border
            ws.cell(
                row=row, column=10, value=float(commande.sous_total_articles)
            ).border = border
            ws.cell(
                row=row, column=11, value=float(commande.ville.frais_livraison or 0)
            ).border = border
            ws.cell(row=row, column=12, value=float(commande.total_cmd)).border = border
            ws.cell(row=row, column=13, value=commande.compteur).border = border
            ws.cell(row=row, column=14, value=articles_text).border = border
            
            # Ajuster la hauteur de ligne pour les articles
            ws.row_dimensions[row].height = max(20, len(articles_detail) * 15)
            
            row += 1
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Largeur sp√©ciale pour la colonne articles
        ws.column_dimensions["N"].width = 40
        
        # Nom du fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"commandes_consolidees_{ville.nom}_{timestamp}.xlsx"
        
        # R√©ponse
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de l'export: {str(e)}")
        return redirect("Prepacommande:liste_prepa")


@login_required
def api_prix_upsell_articles(request, commande_id):
    """API pour r√©cup√©rer les prix upsell mis √† jour des articles"""
    print(f"üîÑ R√©cup√©ration des prix upsell pour la commande {commande_id}")
    
    try:
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse(
            {"error": "Profil d'op√©rateur de pr√©paration non trouv√©."}, status=403
        )
    
    try:
        commande = Commande.objects.get(id=commande_id)
        
        # V√©rifier que la commande est affect√©e √† cet op√©rateur
        etat_preparation = commande.etats.filter(
            operateur=operateur,
            enum_etat__libelle__in=["En pr√©paration", "√Ä imprimer"],
            date_fin__isnull=True,
        ).first()
        
        if not etat_preparation:
            return JsonResponse(
                {"error": "Cette commande ne vous est pas affect√©e."}, status=403
            )
        
        # R√©cup√©rer tous les articles du panier avec leurs prix mis √† jour
        paniers = commande.paniers.select_related("article").all()
        prix_articles = {}
        
        for panier in paniers:
            # Utiliser le filtre Django pour calculer le prix selon la logique upsell
            from commande.templatetags.commande_filters import (
                get_prix_upsell_avec_compteur,
            )
            
            # Debug: Afficher les informations de l'article
            print(f"üîç Article {panier.article.id} ({panier.article.nom}):")
            print(f"   - isUpsell: {panier.article.isUpsell}")
            print(f"   - prix_actuel: {panier.article.prix_actuel}")
            print(f"   - prix_unitaire: {panier.article.prix_unitaire}")
            print(f"   - prix_upsell_1: {panier.article.prix_upsell_1}")
            print(f"   - prix_upsell_2: {panier.article.prix_upsell_2}")
            print(f"   - prix_upsell_3: {panier.article.prix_upsell_3}")
            print(f"   - prix_upsell_4: {panier.article.prix_upsell_4}")
            print(f"   - compteur: {commande.compteur}")
            
            prix_calcule = get_prix_upsell_avec_compteur(
                panier.article, commande.compteur
            )
            print(f"   - prix_calcule: {prix_calcule}")
            
            # Convertir en float pour √©viter les probl√®mes de s√©rialisation JSON
            prix_calcule = float(prix_calcule) if prix_calcule is not None else 0.0
            
            # D√©terminer le type de prix et les informations
            if panier.article.isUpsell:
                prix_type = "upsell"
                
                # D√©terminer le niveau d'upsell affich√©
                if commande.compteur >= 4 and panier.article.prix_upsell_4 is not None:
                    niveau_upsell = 4
                elif (
                    commande.compteur >= 3 and panier.article.prix_upsell_3 is not None
                ):
                    niveau_upsell = 3
                elif (
                    commande.compteur >= 2 and panier.article.prix_upsell_2 is not None
                ):
                    niveau_upsell = 2
                elif (
                    commande.compteur >= 1 and panier.article.prix_upsell_1 is not None
                ):
                    niveau_upsell = 1
                else:
                    niveau_upsell = 0
                
                if niveau_upsell > 0:
                    libelle = f"Upsell Niveau {niveau_upsell}"
                else:
                    libelle = "Upsell (Prix normal)"
                    
                icone = "fas fa-arrow-up"
            else:
                prix_type = "normal"
                libelle = "Prix normal"
                icone = "fas fa-tag"
                niveau_upsell = 0
            
            prix_articles[panier.id] = {
                "prix": prix_calcule,
                "type": prix_type,
                "libelle": libelle,
                "icone": icone,
                "niveau_upsell": niveau_upsell,
                "is_upsell": panier.article.isUpsell,
                "sous_total": float(prix_calcule * panier.quantite),
            }

        return JsonResponse(
            {
                "success": True,
                "prix_articles": prix_articles,
                "compteur": commande.compteur,
                "message": f"Prix mis √† jour pour {len(prix_articles)} articles",
            }
        )
        
    except Commande.DoesNotExist:
        return JsonResponse({"error": "Commande non trouv√©e."}, status=404)
    except Exception as e:
        print(f"‚ùå Erreur lors de la r√©cup√©ration des prix upsell: {e}")
        return JsonResponse({"error": f"Erreur serveur: {str(e)}"}, status=500)


@login_required
def diagnostiquer_compteur(request, commande_id):
    """
    Fonction pour diagnostiquer et corriger le compteur d'une commande
    """
    try:
        commande = get_object_or_404(Commande, id=commande_id)
        
        # Diagnostiquer la situation actuelle
        articles_upsell = commande.paniers.filter(article__isUpsell=True)
        compteur_actuel = commande.compteur
        
        # Calculer la quantit√© totale d'articles upsell
        total_quantite_upsell = (
            articles_upsell.aggregate(total=Sum("quantite"))["total"] or 0
        )
        
        print(f"üîç DIAGNOSTIC Commande {commande.id_yz}:")
        print(f"üìä Compteur actuel: {compteur_actuel}")
        print(f"üì¶ Articles upsell trouv√©s: {articles_upsell.count()}")
        print(f"üî¢ Quantit√© totale d'articles upsell: {total_quantite_upsell}")
        
        if articles_upsell.exists():
            print("üìã Articles upsell dans la commande:")
            for panier in articles_upsell:
                print(
                    f"  - {panier.article.nom} (Qt√©: {panier.quantite}, ID: {panier.article.id}, isUpsell: {panier.article.isUpsell})"
                )
        
        # D√©terminer le compteur correct selon la nouvelle logique :
        # 0-1 unit√©s upsell ‚Üí compteur = 0
        # 2+ unit√©s upsell ‚Üí compteur = total_quantite_upsell - 1
        if total_quantite_upsell >= 2:
            compteur_correct = total_quantite_upsell - 1
        else:
            compteur_correct = 0
        
        print(f"‚úÖ Compteur correct: {compteur_correct}")
        print(
            "üìñ Logique: 0-1 unit√©s upsell ‚Üí compteur=0 | 2+ unit√©s upsell ‚Üí compteur=total_quantit√©-1"
        )
        
        # Corriger si n√©cessaire
        if compteur_actuel != compteur_correct:
            print(f"üîß CORRECTION: {compteur_actuel} -> {compteur_correct}")
            commande.compteur = compteur_correct
            commande.save()
            
            # Recalculer tous les totaux
            commande.recalculer_totaux_upsell()
            
            # Retourner les nouvelles donn√©es
            return JsonResponse(
                {
                    "success": True,
                    "message": f"Compteur corrig√© de {compteur_actuel} vers {compteur_correct}",
                    "ancien_compteur": compteur_actuel,
                    "nouveau_compteur": compteur_correct,
                    "total_commande": float(commande.total_cmd),
                    "articles_upsell": articles_upsell.count(),
                    "quantite_totale_upsell": total_quantite_upsell,
                    "articles_count": commande.paniers.count(),
                    "sous_total_articles": float(commande.sous_total_articles),
                }
            )
        else:
            return JsonResponse(
                {
                    "success": True,
                    "message": "Compteur d√©j√† correct",
                    "compteur": compteur_actuel,
                    "articles_upsell": articles_upsell.count(),
                    "quantite_totale_upsell": total_quantite_upsell,
                    "total_commande": float(commande.total_cmd),
                    "articles_count": commande.paniers.count(),
                    "sous_total_articles": float(commande.sous_total_articles),
                }
            )
            
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


@login_required
def api_changer_etat_commande(request, commande_id):
    """API pour changer l'√©tat d'une commande (Collect√©e, Emball√©e)"""
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "M√©thode non autoris√©e"}, status=405)
    
    try:
        # V√©rifier que l'utilisateur est un op√©rateur de pr√©paration
        operateur = Operateur.objects.get(
            user=request.user, type_operateur="PREPARATION"
        )
    except Operateur.DoesNotExist:
        return JsonResponse({"success": False, "error": "Acc√®s non autoris√©"}, status=403)
    
    try:
        # R√©cup√©rer les donn√©es de la requ√™te
        data = json.loads(request.body)
        nouvel_etat = data.get('nouvel_etat')
        
        if not nouvel_etat:
            return JsonResponse({"success": False, "error": "Nouvel √©tat non sp√©cifi√©"}, status=400)
        
        # V√©rifier que l'√©tat est valide
        etats_valides = ['Collect√©e', 'Emball√©e']
        if nouvel_etat not in etats_valides:
            return JsonResponse({"success": False, "error": f"√âtat invalide. √âtats autoris√©s: {', '.join(etats_valides)}"}, status=400)
        
        # R√©cup√©rer la commande
        commande = Commande.objects.get(id=commande_id)
        
        # V√©rifier que la commande est affect√©e √† cet op√©rateur
        etat_actuel = commande.etats.filter(
            operateur=operateur,
            enum_etat__libelle__in=["En pr√©paration", "√Ä imprimer", "Collect√©e", "Emball√©e"],
            date_fin__isnull=True,
        ).first()
        
        if not etat_actuel:
            return JsonResponse({"success": False, "error": "Cette commande ne vous est pas affect√©e"}, status=403)
        
        # V√©rifier la logique de transition d'√©tat
        etat_actuel_libelle = etat_actuel.enum_etat.libelle
        
        if nouvel_etat == 'Collect√©e':
            if etat_actuel_libelle not in ['En pr√©paration', '√Ä imprimer']:
                return JsonResponse({"success": False, "error": "Impossible de passer √† 'Collect√©e' depuis cet √©tat"}, status=400)
        elif nouvel_etat == 'Emball√©e':
            if etat_actuel_libelle not in ['Collect√©e']:
                return JsonResponse({"success": False, "error": "Impossible de passer √† 'Emball√©e' depuis cet √©tat"}, status=400)
        
        # R√©cup√©rer l'√©tat correspondant
        try:
            nouvel_etat_enum = EnumEtatCmd.objects.get(libelle=nouvel_etat)
        except EnumEtatCmd.DoesNotExist:
            return JsonResponse({"success": False, "error": f"√âtat '{nouvel_etat}' non trouv√© dans le syst√®me"}, status=400)
        
        with transaction.atomic():
            # Terminer l'√©tat actuel
            etat_actuel.date_fin = timezone.now()
            etat_actuel.save()
            
            # Cr√©er le nouvel √©tat
            EtatCommande.objects.create(
                commande=commande,
                enum_etat=nouvel_etat_enum,
                operateur=operateur,
                date_debut=timezone.now(),
                commentaire=f"√âtat chang√© vers {nouvel_etat} par l'op√©rateur de pr√©paration"
            )
            
            # Cr√©er une op√©ration de tra√ßabilit√©
            Operation.objects.create(
                commande=commande,
                type_operation=f"CHANGEMENT_ETAT_{nouvel_etat.upper()}",
                operateur=operateur,
                date_operation=timezone.now(),
                conclusion=json.dumps({
                    "ancien_etat": etat_actuel_libelle,
                    "nouvel_etat": nouvel_etat,
                    "commentaire": f"Changement d'√©tat de {etat_actuel_libelle} vers {nouvel_etat}"
                })
            )
        
        return JsonResponse({
            "success": True,
            "message": f"√âtat chang√© avec succ√®s vers '{nouvel_etat}'",
            "nouvel_etat": nouvel_etat,
            "commande_id": commande_id
        })
        
    except Commande.DoesNotExist:
        return JsonResponse({"success": False, "error": "Commande non trouv√©e"}, status=404)
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Donn√©es JSON invalides"}, status=400)
    except Exception as e:
        return JsonResponse({"success": False, "error": f"Erreur serveur: {str(e)}"}, status=500)
