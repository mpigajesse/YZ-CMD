from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.db.models import Count, Q, Sum, F, Avg
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.db import transaction
from django.core.paginator import Paginator

import json
from parametre.models import Operateur, Ville
from commande.models import Commande, EtatCommande, EnumEtatCmd, Operation, Panier, Envoi
from django.urls import reverse

import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import base64
import csv

from article.models import Article, MouvementStock, VarianteArticle
from commande.models import Envoi
from .forms import ArticleForm, AjusterStockForm
from .utils import creer_mouvement_stock
from .decorators import superviseur_preparation_required, superviseur_only_required, preparation_team_required

# Create your views here.

@superviseur_preparation_required
def home_view(request):
    """Tableau de bord de supervision avec métriques globales et suivi en temps réel"""
    try:
        operateur_profile = request.user.profil_operateur
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')

    # Date de référence
    today = timezone.now().date()
    # Début de la semaine (lundi)
    start_of_week = today - timedelta(days=today.weekday())

    # Récupérer les états nécessaires
    try:
        etat_confirmee = EnumEtatCmd.objects.get(libelle__iexact='Confirmée')
        etat_en_preparation = EnumEtatCmd.objects.get(libelle__iexact='En préparation')
        etat_preparee = EnumEtatCmd.objects.get(libelle__iexact='Préparée')
    except EnumEtatCmd.DoesNotExist as e:
        messages.error(request, f"État manquant dans le système: {str(e)}")
        return redirect('login')

    # STATISTIQUES GLOBALES DE SUPERVISION
    # 1. Commandes à préparer (affectées mais pas encore commencées)
    commandes_a_preparer = Commande.objects.filter(
        etats__enum_etat__libelle='À imprimer',
        etats__date_fin__isnull=True
    ).distinct().count()


    
    # 3. Commandes livrées partiellement (supervision)
    commandes_livrees_partiellement = Commande.objects.filter(
        etats__enum_etat__libelle='Livrée Partiellement',
        etats__date_fin__isnull=True
    ).distinct().count()
    
    # 4. Commandes retournées (supervision)
    commandes_retournees = Commande.objects.filter(
        etats__enum_etat__libelle='Retournée',
        etats__date_fin__isnull=True
    ).distinct().count()

    # 5. Commandes en cours de préparation (tous opérateurs)
    commandes_en_cours = Commande.objects.filter(
        etats__enum_etat=etat_en_preparation,
        etats__date_fin__isnull=True
    ).distinct().count()

    # 4. Performance de l'opérateur aujourd'hui
    ma_performance = EtatCommande.objects.filter(
        enum_etat=etat_preparee,
        date_debut__date=today,
        operateur=operateur_profile
    ).distinct().count()

    # === Calculs supplémentaires pour le tableau de bord ===
    # Commandes préparées aujourd'hui (toutes)
    commandes_preparees_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date=today
    ).count()

    # Commandes préparées cette semaine (toutes)
    commandes_preparees_week = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date__gte=start_of_week,
        date_fin__date__lte=today
    ).count()

    # Commandes actuellement en préparation (toutes)
    commandes_en_preparation = Commande.objects.filter(
        etats__enum_etat=etat_en_preparation,
        etats__date_fin__isnull=True
    ).count()

    # Performance de l'opérateur aujourd'hui (commandes préparées par lui)
    ma_performance_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date=today,
        operateur=operateur_profile
    ).count()

    # Valeur totale (DH) des commandes préparées aujourd'hui
    commandes_ids_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date=today
    ).values_list('commande_id', flat=True)
    valeur_preparees_today = Commande.objects.filter(id__in=commandes_ids_today).aggregate(total=Sum('total_cmd'))['total'] or 0

    # Articles populaires (semaine en cours)
    commandes_ids_week = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date__gte=start_of_week,
        date_fin__date__lte=today
    ).values_list('commande_id', flat=True)
    articles_populaires = Panier.objects.filter(
        commande_id__in=commandes_ids_week
    ).values('article__nom', 'article__reference').annotate(
        total_quantite=Sum('quantite'),
        total_commandes=Count('commande', distinct=True)
    ).order_by('-total_quantite')[:5]

    # Activité récente (5 dernières préparations de l'opérateur)
    activite_recente = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        operateur=operateur_profile,
        date_fin__isnull=False
    ).select_related('commande', 'commande__client').order_by('-date_fin')[:5]

    # STATISTIQUES GLOBALES ADDITIONNELLES
    # Commandes confirmées aujourd'hui
    commandes_confirmees_today = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée',
        etats__date_debut__date=today
    ).distinct().count()
    
    # Total global des commandes actives
    total_commandes_actives = commandes_en_cours + commandes_livrees_partiellement + commandes_retournees
    
    # Performance globale (toutes les préparations terminées aujourd'hui)
    performance_globale_today = EtatCommande.objects.filter(
        enum_etat=etat_en_preparation,
        date_fin__date=today
    ).count()
    
    # Valeur totale traitée aujourd'hui
    valeur_totale_today = Commande.objects.filter(
        etats__enum_etat=etat_en_preparation,
        etats__date_fin__date=today
    ).aggregate(total=Sum('total_cmd'))['total'] or 0
    
    # Taux de réussite (commandes terminées vs commandes reçues)
    taux_reussite = 0
    if commandes_confirmees_today > 0:
        taux_reussite = round((performance_globale_today / commandes_confirmees_today) * 100, 1)
    
    # Alertes de supervision
    alertes = []
    if commandes_retournees > 0:
        alertes.append({
            'type': 'warning',
            'icon': 'fas fa-exclamation-triangle',
            'message': f'{commandes_retournees} commande(s) retournée(s) nécessite(nt) une attention',
            'url': 'Superpreparation:commandes_retournees'
        })
    
    if commandes_livrees_partiellement > 0:
        alertes.append({
            'type': 'info',
            'icon': 'fas fa-info-circle',
            'message': f'{commandes_livrees_partiellement} commande(s) livrée(s) partiellement',
            'url': 'Superpreparation:commandes_livrees_partiellement'
        })
    

        alertes.append({
            'type': 'urgent',
            'icon': 'fas fa-fire',
            
        })
    
    # Préparer les statistiques de supervision
    stats = {
        # Statistiques principales de supervision

        'commandes_a_preparer': commandes_a_preparer,
        'commandes_en_cours': commandes_en_cours,
        'commandes_livrees_partiellement': commandes_livrees_partiellement,
        'commandes_retournees': commandes_retournees,
        'commandes_confirmees_today': commandes_confirmees_today,
        'total_commandes_actives': total_commandes_actives,
        
        # Performance et KPIs
        'performance_globale_today': performance_globale_today,
        'ma_performance_today': ma_performance_today,
        'taux_reussite': taux_reussite,
        'valeur_totale_today': valeur_totale_today,
        'valeur_preparees_today': valeur_preparees_today,
        
        # Données historiques
        'commandes_preparees_today': commandes_preparees_today,
        'commandes_preparees_week': commandes_preparees_week,
        'commandes_en_preparation': commandes_en_preparation,
        
        # Données détaillées
        'articles_populaires': articles_populaires,
        'activite_recente': activite_recente,
        'alertes': alertes
    }

    context = {
        'page_title': 'Centre de Supervision',
        'page_subtitle': f'Tableau de bord global - {total_commandes_actives} commandes actives',
        'profile': operateur_profile,
        'stats': stats,
        'alertes': alertes,
        'is_supervision': True
    }
    return render(request, 'composant_generale/Superpreparation/home.html', context)

@superviseur_preparation_required
def liste_prepa(request):
    """Liste des commandes à préparer pour les opérateurs de préparation"""
    from commande.models import Operation
    
    try:
        operateur_profile = request.user.profil_operateur
        
        # Autoriser superviseur ou équipe préparation
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')

    # Récupérer les commandes dont l'état ACTUEL est "À imprimer" ou "En préparation" (toutes les commandes en préparation)
    # On cherche les commandes qui ont un état "À imprimer" ou "En préparation" actif (sans date_fin)
    from django.db.models import Q, Max
    
    # Définir le type de filtre en premier
    filter_type = request.GET.get('filter', 'all')
    
    if filter_type == 'livrees_partiellement':
        # Filtrer les commandes qui ont été livrées partiellement et sont maintenant en préparation
        commandes_affectees = []
        commandes_base = Commande.objects.filter(
            Q(etats__enum_etat__libelle='À imprimer') | Q(etats__enum_etat__libelle='En préparation'),
            etats__date_fin__isnull=True
        ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats').distinct()
        
        for commande in commandes_base:
            etats_commande = commande.etats.all().order_by('date_debut')
            etat_prepa_actuel = None
            
            # Trouver l'état actuel de préparation
            for etat in etats_commande:
                if etat.enum_etat.libelle in ['À imprimer', 'En préparation'] and not etat.date_fin:
                    etat_prepa_actuel = etat
                    break
            
            if etat_prepa_actuel:
                # Vérifier si la commande a un historique de livraison partielle
                has_partially_delivered_history = False
                for etat in etats_commande:
                    if (etat.enum_etat.libelle == 'Livrée Partiellement' and 
                        etat.date_fin and 
                        etat.date_fin < etat_prepa_actuel.date_debut):
                        has_partially_delivered_history = True
                        break
                
                if has_partially_delivered_history:
                    commandes_affectees.append(commande)
    elif filter_type == 'renvoyees_logistique':
        # Pour les commandes renvoyées par la logistique, ne pas exclure les états problématiques
        # car on veut inclure les commandes avec opération de renvoi même si elles ont des états ultérieurs
        commandes_affectees = Commande.objects.filter(
            Q(etats__enum_etat__libelle='À imprimer') | Q(etats__enum_etat__libelle='En préparation'),
            etats__date_fin__isnull=True  # État actif (en cours)
        ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats').distinct()
    elif filter_type == 'retournees':
        # Obsolète: rediriger vers la page dédiée
        return redirect('Superpreparation:commandes_retournees')

    else:  # filter_type == 'all'
        # Pour "Toutes les commandes", afficher TOUTES les commandes en préparation (quel que soit leur état)
        commandes_affectees = []
        commandes_base = Commande.objects.filter(
            Q(etats__enum_etat__libelle='À imprimer') | 
            Q(etats__enum_etat__libelle='En préparation') |
            Q(etats__enum_etat__libelle='Collectée') |
            Q(etats__enum_etat__libelle='Emballée') |
            Q(etats__enum_etat__libelle='Préparée'),
            etats__date_fin__isnull=True
        ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats').distinct()
        
        for commande in commandes_base:
            etats_commande = commande.etats.all().order_by('date_debut')
            etat_prepa_actuel = None
            
            # Trouver l'état actuel de préparation (peu importe l'état)
            for etat in etats_commande:
                if etat.enum_etat.libelle in ['À imprimer', 'En préparation', 'Collectée', 'Emballée', 'Préparée'] and not etat.date_fin:
                    etat_prepa_actuel = etat
                    break
            
            if etat_prepa_actuel:
                # Initialiser la variable
                a_etats_ultérieurs_problematiques = False
                
                # Vérifier s'il y a des états ultérieurs problématiques (seulement si l'état actuel n'est pas "Préparée")
                if etat_prepa_actuel.enum_etat.libelle != 'Préparée':
                    for etat in etats_commande:
                        if (etat.date_debut > etat_prepa_actuel.date_debut and 
                                etat.enum_etat.libelle in ['Livrée', 'Livrée Partiellement', 'En cours de livraison']):
                            a_etats_ultérieurs_problematiques = True
                            break
                
                if a_etats_ultérieurs_problematiques:
                    continue
                
                # Inclure toutes les commandes valides (tous les états de préparation)
                commandes_affectees.append(commande)
    
    # Pour les commandes renvoyées par la logistique, respecter l'affectation spécifique à chaque opérateur
    if filter_type == 'renvoyees_logistique':
        # Filtrer seulement les commandes renvoyées par la logistique ET affectées à cet opérateur spécifique
        commandes_filtrees = []
        for commande in commandes_affectees:
            from commande.models import Operation
            
            # Vérifier que la commande n'a pas d'états ultérieurs problématiques
            etats_commande = commande.etats.all().order_by('date_debut')
            etat_actuel = None
            
            # Trouver l'état actuel (En préparation)
            for etat in etats_commande:
                if etat.enum_etat.libelle in ['À imprimer', 'En préparation'] and not etat.date_fin:
                    etat_actuel = etat
                    break
            
            if etat_actuel:
                # Vérifier les opérations de traçabilité EN PREMIER
                operation_renvoi = Operation.objects.filter(
                    commande=commande,
                    type_operation='RENVOI_PREPARATION'
                ).first()

                # Si il y a une opération de renvoi explicite, inclure la commande
                # même si elle a des états ultérieurs problématiques
                if operation_renvoi:
                    commandes_filtrees.append(commande)
                    continue

                # Vérifier si c'est une commande de renvoi créée lors d'une livraison partielle
                if commande.num_cmd and commande.num_cmd.startswith('RENVOI-'):
                    # Chercher la commande originale
                    num_cmd_original = commande.num_cmd.replace('RENVOI-', '')
                    commande_originale = Commande.objects.filter(
                        num_cmd=num_cmd_original,
                        etats__enum_etat__libelle='Livrée Partiellement'
                    ).first()

                    if commande_originale:
                        commandes_filtrees.append(commande)
                        continue

                # Sinon, vérifier s'il y a des états ultérieurs problématiques
                a_etats_ultérieurs_problematiques = False
                for etat in etats_commande:
                    if (etat.date_debut > etat_actuel.date_debut and 
                        etat.enum_etat.libelle in ['Livrée', 'Livrée Partiellement', 'Préparée', 'En cours de livraison']):
                        a_etats_ultérieurs_problematiques = True
                        break

                if a_etats_ultérieurs_problematiques:
                    continue  # Ignorer cette commande

                # Vérifier l'historique des états de la commande
                # Trouver l'état précédent
                for etat in reversed(etats_commande):
                    if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                        if etat.enum_etat.libelle == 'En cours de livraison':
                            commandes_filtrees.append(commande)
                            break
        
        commandes_affectees = commandes_filtrees
    
    # Calculer les statistiques par type de commande
    stats_par_type = {
        'renvoyees_logistique': 0,
        'livrees_partiellement': 0,
        'retournees': 0,
        'affectees_admin': 0,
        'affectees_moi': 0,
    }
    
    # Pour chaque commande, ajouter l'état précédent pour comprendre d'où elle vient
    for commande in commandes_affectees:
        # Récupérer tous les états de la commande dans l'ordre chronologique
        etats_commande = commande.etats.all().order_by('date_debut')
        
        # Trouver l'état actuel (tous les états de préparation)
        etat_actuel = None
        for etat in etats_commande:
            if etat.enum_etat.libelle in ['À imprimer', 'En préparation', 'Collectée', 'Emballée', 'Préparée'] and not etat.date_fin:
                etat_actuel = etat
                break
        
        if etat_actuel:
            # Trouver l'état précédent (le dernier état terminé avant l'état actuel)
            etat_precedent = None
            for etat in reversed(etats_commande):
                if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                    if etat.enum_etat.libelle not in ['À imprimer', 'En préparation', 'Collectée', 'Emballée', 'Préparée']:
                        etat_precedent = etat
                        break
            
            commande.etat_precedent = etat_precedent
            
            # Trouver l'état de confirmation : priorité à l'opérateur de CONFIRMATION
            etat_conf_op = commande.etats.filter(
                enum_etat__libelle='Confirmée',
                operateur__type_operateur='CONFIRMATION'
            ).order_by('-date_debut').first()

            if etat_conf_op:
                commande.etat_confirmation = etat_conf_op
            else:
                etat_conf_any = commande.etats.filter(enum_etat__libelle='Confirmée').order_by('date_debut').first()
                commande.etat_confirmation = etat_conf_any
    
            # Ajouter l'état actuel pour l'affichage des étapes de préparation
            commande.etat_actuel_preparation = etat_actuel
    
    # Si aucune commande trouvée avec la méthode stricte, essayer une approche plus large
    if isinstance(commandes_affectees, list):
        has_commandes = len(commandes_affectees) > 0
    else:
        has_commandes = commandes_affectees.exists()
    
    if not has_commandes:
        # Chercher toutes les commandes en préparation (tous les états)
        commandes_affectees = Commande.objects.filter(
            Q(etats__enum_etat__libelle='À imprimer') | 
            Q(etats__enum_etat__libelle='En préparation') |
            Q(etats__enum_etat__libelle='Collectée') |
            Q(etats__enum_etat__libelle='Emballée') |
            Q(etats__enum_etat__libelle='Préparée')
        ).exclude(
            # Exclure seulement les commandes qui ont déjà un état ultérieur actif (livraison, etc.)
            Q(etats__enum_etat__libelle__in=['En cours de livraison', 'Livrée', 'Annulée'], etats__date_fin__isnull=True)
        ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats').distinct()
        
        # Pour chaque commande, ajouter l'état précédent pour comprendre d'où elle vient
        for commande in commandes_affectees:
            # Récupérer tous les états de la commande dans l'ordre chronologique
            etats_commande = commande.etats.all().order_by('date_debut')
            
            # Trouver l'état actuel (tous les états de préparation)
            etat_actuel = None
            for etat in etats_commande:
                if etat.enum_etat.libelle in ['À imprimer', 'En préparation', 'Collectée', 'Emballée', 'Préparée'] and not etat.date_fin:
                    etat_actuel = etat
                    break
            
            if etat_actuel:
                # Trouver l'état précédent (le dernier état terminé avant l'état actuel)
                etat_precedent = None
                for etat in reversed(etats_commande):
                    if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                        if etat.enum_etat.libelle not in ['À imprimer', 'En préparation', 'Collectée', 'Emballée', 'Préparée']:
                            etat_precedent = etat
                            break
                
                commande.etat_precedent = etat_precedent
                
                # Trouver l'état de confirmation : priorité à l'opérateur de CONFIRMATION
                etat_conf_op = commande.etats.filter(
                    enum_etat__libelle='Confirmée',
                    operateur__type_operateur='CONFIRMATION'
                ).order_by('-date_debut').first()

                if etat_conf_op:
                    commande.etat_confirmation = etat_conf_op
                else:
                    etat_conf_any = commande.etats.filter(enum_etat__libelle='Confirmée').order_by('date_debut').first()
                    commande.etat_confirmation = etat_conf_any
                
                # Ajouter l'état actuel pour l'affichage des étapes de préparation
                commande.etat_actuel_preparation = etat_actuel
    
    # Suppression du filtre 'nouvelles' car redondant avec l'affectation automatique
    # Suppression du filtre 'renvoyees_preparation' car non nécessaire
    
    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        if isinstance(commandes_affectees, list):
            # Si c'est une liste (après filtrage)
            commandes_affectees = [cmd for cmd in commandes_affectees if 
                search_query.lower() in str(cmd.id_yz).lower() or
                search_query.lower() in (cmd.num_cmd or '').lower() or
                search_query.lower() in cmd.client.nom.lower() or
                search_query.lower() in cmd.client.prenom.lower() or
                search_query.lower() in (cmd.client.numero_tel or '').lower()
            ]
        else:
            # Si c'est un QuerySet
            commandes_affectees = commandes_affectees.filter(
                Q(id_yz__icontains=search_query) |
                Q(num_cmd__icontains=search_query) |
                Q(client__nom__icontains=search_query) |
                Q(client__prenom__icontains=search_query) |
                Q(client__numero_tel__icontains=search_query)
            ).distinct()
    
    # Tri par date d'affectation (plus récentes en premier)
    if isinstance(commandes_affectees, list):
        commandes_affectees.sort(key=lambda x: x.etats.filter(date_fin__isnull=True).first().date_debut if x.etats.filter(date_fin__isnull=True).first() else timezone.now(), reverse=True)
    else:
        commandes_affectees = commandes_affectees.order_by('-etats__date_debut')

    # Générer les codes-barres pour chaque commande et s'assurer que etat_actuel_preparation est défini
    code128 = barcode.get_barcode_class('code128')
    for commande in commandes_affectees:
        # S'assurer que etat_actuel_preparation est défini
        if not hasattr(commande, 'etat_actuel_preparation') or commande.etat_actuel_preparation is None:
            commande.etat_actuel_preparation = commande.etat_actuel
        
        if commande.id_yz:
            barcode_instance = code128(str(commande.id_yz), writer=ImageWriter())
            buffer = BytesIO()
            barcode_instance.write(buffer, options={'write_text': False, 'module_height': 10.0})
            barcode_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
            commande.barcode_base64 = barcode_base64
        else:
            commande.barcode_base64 = None
    
    # Statistiques
    if isinstance(commandes_affectees, list):
        total_affectees = len(commandes_affectees)
        valeur_totale = sum(cmd.total_cmd or 0 for cmd in commandes_affectees)
        
        # Commandes urgentes (affectées depuis plus de 1 jour)
        date_limite_urgence = timezone.now() - timedelta(days=1)
        commandes_urgentes = sum(1 for cmd in commandes_affectees if 
            cmd.etats.filter(date_debut__lt=date_limite_urgence).exists()
        )
    else:
        total_affectees = commandes_affectees.count()
        valeur_totale = commandes_affectees.aggregate(total=Sum('total_cmd'))['total'] or 0
        
        # Commandes urgentes (affectées depuis plus de 1 jour)
        date_limite_urgence = timezone.now() - timedelta(days=1)
        commandes_urgentes = commandes_affectees.filter(
            etats__date_debut__lt=date_limite_urgence
        ).count()
    
    # Statistiques par type pour les onglets
    stats_par_type = {
        'renvoyees_logistique': 0,
        'livrees_partiellement': 0,
        'retournees': 0,
        'affectees_moi': 0,
    }
    
    # Recalculer les statistiques pour tous les types
    # D'abord, récupérer toutes les commandes en préparation (sans filtre d'opérateur)
    toutes_commandes = Commande.objects.filter(
        Q(etats__enum_etat__libelle='À imprimer') | Q(etats__enum_etat__libelle='En préparation'),
        etats__date_fin__isnull=True  # État actif (en cours)
    ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats').distinct()
    
    for cmd in toutes_commandes:
        # Vérifier si c'est une commande renvoyée par la logistique
        operation_renvoi = Operation.objects.filter(
            commande=cmd,
            type_operation='RENVOI_PREPARATION'
        ).first()
        
        if operation_renvoi:
            stats_par_type['renvoyees_logistique'] += 1
            continue
        
        # Vérifier si c'est une commande de renvoi créée lors d'une livraison partielle
        if cmd.num_cmd and cmd.num_cmd.startswith('RENVOI-'):
            # Chercher la commande originale
            num_cmd_original = cmd.num_cmd.replace('RENVOI-', '')
            commande_originale = Commande.objects.filter(
                num_cmd=num_cmd_original,
                etats__enum_etat__libelle='Livrée Partiellement'
            ).first()
            
            if commande_originale:
                stats_par_type['renvoyees_logistique'] += 1
                continue
        
        # Vérifier l'état précédent
        etats_commande = cmd.etats.all().order_by('date_debut')
        etat_actuel = None
        
        # Trouver l'état actuel
        for etat in etats_commande:
            if etat.enum_etat.libelle in ['À imprimer', 'En préparation'] and not etat.date_fin:
                etat_actuel = etat
                break
        
        if etat_actuel:
            # Vérifier s'il y a des états ultérieurs problématiques
            a_etats_ultérieurs_problematiques = False
            for etat in etats_commande:
                if (etat.date_debut > etat_actuel.date_debut and 
                    etat.enum_etat.libelle in ['Livrée', 'Livrée Partiellement', 'Préparée', 'En cours de livraison']):
                    a_etats_ultérieurs_problematiques = True
                    break
            
            # Si il y a des états ultérieurs problématiques, ignorer cette commande
            if a_etats_ultérieurs_problematiques:
                continue
            
            # Trouver l'état précédent
            for etat in reversed(etats_commande):
                if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                    if etat.enum_etat.libelle == 'En cours de livraison':
                        stats_par_type['renvoyees_logistique'] += 1
                        break
                    elif etat.enum_etat.libelle == 'Livrée Partiellement':
                        stats_par_type['livrees_partiellement'] += 1
                        break
            
    # Recalculer le compteur des livraisons partielles en utilisant la même logique que la vue séparée
    # Chercher les commandes de renvoi créées lors de livraisons partielles
    commandes_renvoi_livraison_partielle = Commande.objects.filter(
        num_cmd__startswith='RENVOI-',
        etats__enum_etat__libelle='En préparation',
        etats__date_fin__isnull=True
    ).distinct()
    
    livrees_partiellement_count = 0
    for commande_renvoi in commandes_renvoi_livraison_partielle:
        # Extraire le numéro de commande original
        num_cmd_original = commande_renvoi.num_cmd.replace('RENVOI-', '')
        
        # Vérifier que la commande originale a été livrée partiellement
        commande_originale = Commande.objects.filter(
            num_cmd=num_cmd_original,
            etats__enum_etat__libelle='Livrée Partiellement'
        ).first()
        
        if commande_originale:
            livrees_partiellement_count += 1
    
    # Mettre à jour le compteur avec la valeur correcte
    stats_par_type['livrees_partiellement'] = livrees_partiellement_count
    

    


    # Compter les commandes affectées par le superviseur connecté
    try:
        superviseur_nom = operateur_profile.nom_complet
    except Exception:
        superviseur_nom = ''
    if superviseur_nom:
        stats_par_type['affectees_moi'] = Commande.objects.filter(
            etats__enum_etat__libelle='En préparation',
            etats__commentaire__icontains=f"Affectée à la préparation par {superviseur_nom}"
        ).distinct().count()
    
    # Pour l'onglet "Toutes les commandes", le total doit être la somme des 2 autres onglets
    if filter_type == 'all':
        total_affectees = (
            stats_par_type['affectees_moi'] +
            stats_par_type['renvoyees_logistique'] +
            stats_par_type['livrees_partiellement']
        )
    
    # ===== Pagination (par onglet) =====
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
        if items_per_page <= 0:
            items_per_page = 10
    except (ValueError, TypeError):
        items_per_page = 10

    from django.core.paginator import Paginator
    paginator = Paginator(commandes_affectees if isinstance(commandes_affectees, list) or hasattr(commandes_affectees, '__iter__') else list(commandes_affectees), items_per_page)
    page_number = request.GET.get('page', 1)
    commandes_page = paginator.get_page(page_number)

    context = {
        'page_title': 'Suivi Général - Préparation',
        'page_subtitle': f'Vue d\'ensemble des commandes en file de préparation ({total_affectees})',
        'commandes_affectees': commandes_page,
        'search_query': search_query,
        'filter_type': filter_type,
        'items_per_page': items_per_page,
        'stats': {
            'total_affectees': total_affectees,
            'valeur_totale': valeur_totale,
            'commandes_urgentes': commandes_urgentes,
        },
        'stats_par_type': stats_par_type,
        'operateur_profile': operateur_profile,
        'api_produits_url_base': reverse('Superpreparation:api_commande_produits', args=[99999999]),
    }
    return render(request, 'Superpreparation/liste_prepa.html', context)

@superviseur_preparation_required
def commandes_preparees(request):
    """Liste des commandes préparées (état final) pour les superviseurs"""
    try:
        operateur_profile = request.user.profil_operateur
        
        # Autoriser superviseur ou équipe préparation
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')

    # Récupérer les commandes dans l'état "Préparée" (état final)
    commandes_preparees = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True  # État actif
    ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats').distinct()

    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_preparees = commandes_preparees.filter(
            Q(id_yz__icontains=search_query) |
            Q(num_cmd__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query)
        ).distinct()

    # Statistiques
    total_preparees = commandes_preparees.count()
    valeur_totale = commandes_preparees.aggregate(total=Sum('total_cmd'))['total'] or 0
    
    # Commandes préparées aujourd'hui
    today = timezone.now().date()
    preparees_today = commandes_preparees.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_debut__date=today
    ).count()

    # Pagination
    items_per_page = request.GET.get('items_per_page', 10)
    try:
        items_per_page = int(items_per_page)
        if items_per_page <= 0:
            items_per_page = 10
    except (ValueError, TypeError):
        items_per_page = 10

    paginator = Paginator(commandes_preparees, items_per_page)
    page_number = request.GET.get('page', 1)
    commandes_page = paginator.get_page(page_number)

    context = {
        'page_title': 'Commandes Préparées',
        'page_subtitle': f'Commandes finalisées et prêtes pour livraison ({total_preparees})',
        'commandes_preparees': commandes_page,
        'search_query': search_query,
        'items_per_page': items_per_page,
        'stats': {
            'total_preparees': total_preparees,
            'valeur_totale': valeur_totale,
            'preparees_today': preparees_today,
        },
        'operateur_profile': operateur_profile,
    }
    return render(request, 'Superpreparation/commandes_preparees.html', context)



@superviseur_preparation_required
def commandes_en_preparation(request):
    """Page de suivi (lecture seule) des commandes en préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        
        # Autoriser superviseur ou équipe préparation
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('Superpreparation:home')

    # Récupérer TOUTES les commandes dont l'état ACTUEL est "En préparation" (pas seulement celles de l'opérateur connecté)
    commandes_en_preparation = Commande.objects.filter(
        etats__enum_etat__libelle='En préparation',
        etats__date_fin__isnull=True  # État actif (en cours)
    ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats__operateur').distinct()

    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_en_preparation = commandes_en_preparation.filter(
            Q(id_yz__icontains=search_query) |
            Q(num_cmd__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query)
        ).distinct()

    # Statistiques
    stats = {
        'total_commandes': commandes_en_preparation.count(),
        'commandes_urgentes': commandes_en_preparation.filter(
            etats__date_debut__lt=timezone.now() - timedelta(days=1)
        ).count(),
        'valeur_totale': commandes_en_preparation.aggregate(total=Sum('total_cmd'))['total'] or 0
    }

    context = {
        'page_title': 'Suivi - Commandes en Préparation',
        'page_subtitle': f'Suivi en temps réel de {commandes_en_preparation.count()} commande(s) en cours de préparation',
        'profile': operateur_profile,
        'commandes': commandes_en_preparation,
        'search_query': search_query,
        'stats': stats,
        'active_tab': 'en_preparation',
        'is_readonly': True,
        'is_tracking_page': True
    }
    return render(request, 'Superpreparation/commandes_en_preparation.html', context)

@superviseur_preparation_required
def commandes_emballees(request):
    """Page de suivi des commandes emballées qui attendent la finalisation par le superviseur"""
    try:
        operateur_profile = request.user.profil_operateur
        
        # Autoriser superviseur ou équipe préparation
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('Superpreparation:home')

    # Récupérer TOUTES les commandes emballées qui attendent la finalisation
    commandes_emballees = Commande.objects.filter(
        etats__enum_etat__libelle='Emballée',
        etats__date_fin__isnull=True  # État actif (en cours)
    ).select_related('client', 'ville', 'ville__region').prefetch_related('paniers__article', 'etats__operateur').distinct()

    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_emballees = commandes_emballees.filter(
            Q(id_yz__icontains=search_query) |
            Q(num_cmd__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query)
        ).distinct()

    # Statistiques
    stats = {
        'total_commandes': commandes_emballees.count(),
        'commandes_urgentes': commandes_emballees.filter(
            etats__date_debut__lt=timezone.now() - timedelta(hours=2)  # Urgent si emballée depuis plus de 2h
        ).count(),
        'valeur_totale': commandes_emballees.aggregate(total=Sum('total_cmd'))['total'] or 0
    }

    context = {
        'page_title': 'Suivi - Commandes Emballées',
        'page_subtitle': f'Suivi en temps réel de {commandes_emballees.count()} commande(s) emballées en attente de finalisation',
        'profile': operateur_profile,
        'commandes': commandes_emballees,
        'search_query': search_query,
        'stats': stats,
        'active_tab': 'emballees',
        'is_readonly': False,  # Les superviseurs peuvent finaliser
        'is_tracking_page': True
    }
    return render(request, 'Superpreparation/commandes_emballees.html', context)

@superviseur_preparation_required
def commandes_livrees_partiellement(request):
    """Page de suivi (lecture seule) des commandes livrées partiellement"""
    try:
        operateur_profile = request.user.profil_operateur
        
        # Autoriser superviseur ou équipe préparation
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')

    # Nouvel objectif: récupérer les commandes livrées partiellement
    # qui ont été préparées (envoyées à la logistique) par cet opérateur
    # L'opérateur de préparation crée l'état 'En préparation', pas 'Préparée'
    commandes_livrees_partiellement_qs = (
        Commande.objects
        .filter(etats__enum_etat__libelle='Livrée Partiellement')
        .filter(etats__enum_etat__libelle='En préparation', etats__operateur=operateur_profile)
        .select_related('client', 'ville', 'ville__region')
        .prefetch_related('paniers__article', 'etats')
        .distinct()
    )
    

    
    # Pour chaque commande originale, essayer de retrouver une éventuelle commande de renvoi associée
    commandes_livrees_partiellement = []
    for commande_originale in commandes_livrees_partiellement_qs:
        commande_renvoi = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande_originale.num_cmd}",
            client=commande_originale.client
        ).first()
        if commande_renvoi:
            commande_originale.commande_renvoi = commande_renvoi
        commandes_livrees_partiellement.append(commande_originale)

    # Les commandes sont déjà filtrées et pertinentes
    commandes_filtrees = commandes_livrees_partiellement

    # Pour chaque commande, récupérer les détails de la livraison partielle
    for commande in commandes_livrees_partiellement:
        # Trouver l'état "Livrée Partiellement" le plus récent
        etat_livraison_partielle = commande.etats.filter(
            enum_etat__libelle='Livrée Partiellement'
        ).order_by('-date_debut').first()
        
        if etat_livraison_partielle:
            commande.date_livraison_partielle = etat_livraison_partielle.date_debut
            commande.commentaire_livraison_partielle = etat_livraison_partielle.commentaire
            commande.operateur_livraison = etat_livraison_partielle.operateur
            
            # Le statut est toujours "Renvoyée en préparation" car nous ne récupérons que les commandes avec renvoi
            commande.statut_actuel = "Renvoyée en préparation"
            
            # Ajouter les informations de la commande de renvoi
            if hasattr(commande, 'commande_renvoi'):
                commande.commande_renvoi_id = commande.commande_renvoi.id
                commande.commande_renvoi_num = commande.commande_renvoi.num_cmd
                commande.commande_renvoi_id_yz = commande.commande_renvoi.id_yz

    context = {
        'page_title': 'Suivi - Commandes Livrées Partiellement',
        'page_subtitle': f'Suivi en temps réel de {len(commandes_livrees_partiellement)} commande(s) livrées partiellement',
        'profile': operateur_profile,
        'commandes_livrees_partiellement': commandes_livrees_partiellement,
        'commandes_count': len(commandes_livrees_partiellement),
        'active_tab': 'livrees_partiellement',
        'is_readonly': True,
        'is_tracking_page': True
    }
    return render(request, 'Superpreparation/commandes_livrees_partiellement.html', context)

@superviseur_preparation_required
def commandes_retournees(request):
    """Page de suivi (lecture seule) des commandes retournées"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
    
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')

    # Commandes dont l'état ACTUEL est 'Retournée' et qui ont été préparées par l'opérateur courant
    # L'opérateur de préparation crée l'état 'En préparation', pas 'Préparée'
    commandes_qs = (
        Commande.objects
        .filter(etats__enum_etat__libelle='Retournée', etats__date_fin__isnull=True)  # État actuel
        .filter(etats__enum_etat__libelle='En préparation', etats__operateur=operateur_profile)
        .select_related('client', 'ville', 'ville__region')
        .prefetch_related('paniers__article', 'etats')
        .distinct()
    )

    commandes = list(commandes_qs)
    


    # Enrichir avec méta d'état 'Retournée'
    for commande in commandes:
        etat_retour = commande.etats.filter(enum_etat__libelle='Retournée').order_by('-date_debut').first()
        if etat_retour:
            commande.date_retournee = etat_retour.date_debut
            commande.commentaire_retournee = etat_retour.commentaire
            commande.operateur_retour = etat_retour.operateur

    context = {
        'page_title': 'Suivi - Commandes Retournées',
        'page_subtitle': f'Suivi en temps réel de {len(commandes)} commande(s) retournées',
        'profile': operateur_profile,
        'commandes_retournees': commandes,
        'commandes_count': len(commandes),
        'active_tab': 'retournees',
        'is_readonly': True,
        'is_tracking_page': True
    }
    return render(request, 'Superpreparation/commandes_retournees.html', context)


@superviseur_preparation_required
def traiter_commande_retournee_api(request, commande_id):
    """API pour traiter une commande retournée et gérer la réincrémentation du stock"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})
    
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            return JsonResponse({'success': False, 'message': "Accès réservé à l'équipe préparation"})
        
        # Récupérer la commande
        commande = get_object_or_404(Commande, id=commande_id)
        
        # Vérifier que la commande est bien retournée et préparée par cet opérateur
        if not commande.etats.filter(
            enum_etat__libelle='Retournée',
            date_fin__isnull=True
        ).exists():
            return JsonResponse({'success': False, 'message': 'Commande non retournée'})
        
        if not commande.etats.filter(
            enum_etat__libelle='En préparation',
            operateur=operateur_profile
        ).exists():
            return JsonResponse({'success': False, 'message': 'Commande non préparée par vous'})
        
        # Récupérer les données de la requête
        data = json.loads(request.body)
        type_traitement = data.get('type_traitement')
        etat_stock = data.get('etat_stock')
        commentaire = data.get('commentaire')
        
        if not all([type_traitement, etat_stock, commentaire]):
            return JsonResponse({'success': False, 'message': 'Données manquantes'})
        
        # Traitement selon le type
        with transaction.atomic():
            if type_traitement == 'repreparer':
                # Ne pas changer l'état de la commande, elle reste "Retournée"
                # Seulement réincrémenter le stock si les produits sont en bon état
                if etat_stock == 'bon':
                    for panier in commande.paniers.all():
                        article = panier.article
                        quantite = panier.quantite
                        
                        # Réincrémenter le stock disponible
                        article.qte_disponible += quantite
                        article.save()
                        
                        # Créer un mouvement de stock pour tracer
                        from article.models import MouvementStock
                        MouvementStock.objects.create(
                            article=article,
                            quantite=quantite,
                            type_mouvement='entree',
                            commentaire=f"Réincrémentation - Commande retournée {commande.id_yz} - Produits en bon état - {commentaire}",
                            operateur=operateur_profile,
                            qte_apres_mouvement=article.qte_disponible
                        )
                
                message = f"Stock réincrémenté: {'Oui' if etat_stock == 'bon' else 'Non'}. Commande reste en état 'Retournée'."
                
                return JsonResponse({
                    'success': True,
                    'message': message,
                    'commande_id': commande.id
                })
            
            return JsonResponse({
                'success': True,
                'message': message,
                'commande_id': commande.id
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})


@superviseur_preparation_required
def profile_view(request):
    try:
        operateur_profile = request.user.profil_operateur
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login') # Ou une page d'erreur appropriée

    context = {
        'page_title': 'Mon Profil',
        'page_subtitle': 'Gérez les informations de votre profil',
        'profile': operateur_profile,
    }
    return render(request, 'Superpreparation/profile.html', context)

@superviseur_preparation_required
def modifier_profile_view(request):
    try:
        operateur_profile = request.user.profil_operateur
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')

    if request.method == 'POST':
        # Récupérer les données du formulaire
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        mail = request.POST.get('mail')
        telephone = request.POST.get('telephone')
        adresse = request.POST.get('adresse')

        # Validation minimale (vous pouvez ajouter plus de validations ici)
        if not nom or not prenom or not mail:
            messages.error(request, "Nom, prénom et email sont requis.")
            context = {
                'page_title': 'Modifier Profil',
                'page_subtitle': 'Mettez à jour vos informations personnelles',
                'profile': operateur_profile,
                'form_data': request.POST, # Pour pré-remplir le formulaire
            }
            return render(request, 'Superpreparation/modifier_profile.html', context)
        
        # Mettre à jour l'utilisateur Django
        request.user.first_name = prenom
        request.user.last_name = nom
        request.user.email = mail
        request.user.save()

        # Mettre à jour le profil de l'opérateur
        operateur_profile.nom = nom
        operateur_profile.prenom = prenom
        operateur_profile.mail = mail
        operateur_profile.telephone = telephone
        operateur_profile.adresse = adresse

        # Gérer l'image si elle est fournie
        if 'photo' in request.FILES:
            operateur_profile.photo = request.FILES['photo']
        
        operateur_profile.save()

        messages.success(request, "Votre profil a été mis à jour avec succès.")
        return redirect('Superpreparation:profile') # Rediriger vers la page de profil après succès
    else:
        context = {
            'page_title': 'Modifier Profil',
            'page_subtitle': 'Mettez à jour vos informations personnelles',
            'profile': operateur_profile,
        }
        return render(request, 'Superpreparation/modifier_profile.html', context)

@superviseur_preparation_required
def changer_mot_de_passe_view(request):
    """Page de changement de mot de passe pour l'opérateur de préparation - Désactivée"""
    return redirect('Superpreparation:profile')

@superviseur_preparation_required
def detail_prepa(request, pk):
    """Vue détaillée pour la préparation d'une commande spécifique"""
    try:
        operateur_profile = request.user.profil_operateur
        
        # Autoriser superviseur ou équipe préparation
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
            
    except Operateur.DoesNotExist:
        messages.error(request, "Votre profil opérateur n'existe pas.")
        return redirect('login')

    # Récupérer la commande spécifique
    try:
        commande = Commande.objects.select_related(
            'client', 'ville', 'ville__region'
        ).prefetch_related(
            'paniers__article', 'paniers__variante', 'etats__enum_etat', 'etats__operateur'
        ).get(id=pk)
    except Commande.DoesNotExist:
        messages.error(request, "La commande demandée n'existe pas.")
        return redirect('Superpreparation:liste_prepa')

    # Récupérer l'état de préparation actuel (peut être affecté à n'importe quel opérateur)
    etat_preparation = commande.etats.filter(
        Q(enum_etat__libelle='À imprimer') | Q(enum_etat__libelle='En préparation'),
        date_fin__isnull=True
    ).first()

    # Récupérer les paniers (articles) de la commande
    paniers = commande.paniers.all().select_related('article', 'variante', 'variante__couleur', 'variante__pointure')
    
    # Initialiser les variables pour les cas de livraison partielle/renvoi
    articles_livres = []
    articles_renvoyes = []
    is_commande_livree_partiellement = False
    commande_renvoi = None # Initialiser à None
    commande_originale = None # Initialiser à None
    etat_articles_renvoyes = {} # Initialiser à un dictionnaire vide

    # Ajouter le prix unitaire, le total de chaque ligne, et l'URL d'image si disponible
    articles_image_urls = {}
    for panier in paniers:
        panier.prix_unitaire = panier.sous_total / panier.quantite if panier.quantite > 0 else 0
        panier.total_ligne = panier.sous_total
        # Préparer affichage variante (couleur/pointure)
        try:
            panier.couleur_display = (
                panier.variante.couleur.nom if getattr(panier, 'variante', None) and getattr(panier.variante, 'couleur', None) else getattr(panier.article, 'couleur', '')
            )
        except Exception:
            panier.couleur_display = getattr(panier.article, 'couleur', '')
        try:
            panier.pointure_display = (
                panier.variante.pointure.pointure if getattr(panier, 'variante', None) and getattr(panier.variante, 'pointure', None) else getattr(panier.article, 'pointure', '')
            )
        except Exception:
            panier.pointure_display = getattr(panier.article, 'pointure', '')
        image_url = None
        try:
            # Protéger l'accès à .url si aucun fichier n'est associé
            if getattr(panier.article, 'image', None):
                # Certains backends lèvent une exception si l'image n'existe pas
                if hasattr(panier.article.image, 'url'):
                    image_url = panier.article.image.url
        except Exception:
            image_url = None
        # Rendre accessible directement dans le template via panier.article.image_url
        setattr(panier.article, 'image_url', image_url)
        # Egalement disponible dans le context si nécessaire
        if getattr(panier.article, 'id', None) is not None:
            articles_image_urls[panier.article.id] = image_url
    
    # Récupérer tous les états de la commande pour afficher l'historique
    etats_commande = commande.etats.all().select_related('enum_etat', 'operateur').order_by('date_debut')
    
    # Déterminer l'état actuel
    etat_actuel = etats_commande.filter(date_fin__isnull=True).first()
    
    # Récupérer l'état précédent pour comprendre d'où vient la commande
    etat_precedent = None
    if etat_actuel:
        # Trouver l'état précédent (le dernier état terminé avant l'état actuel)
        for etat in reversed(etats_commande):
            if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                if etat.enum_etat.libelle not in ['À imprimer', 'En préparation']:
                    etat_precedent = etat
                    break
    
    # Analyser les articles pour les commandes livrées partiellement
    articles_livres = []
    articles_renvoyes = []
    is_commande_livree_partiellement = False
    
    # Import pour JSON
    import json

    # Récupérer l'état des articles renvoyés depuis l'opération de livraison partielle (si elle existe)
    etat_articles_renvoyes = {}
    operation_livraison_partielle = None
    
    # Cas 1: La commande actuelle est la commande originale livrée partiellement
    if etat_actuel and etat_actuel.enum_etat.libelle == 'Livrée Partiellement':
        is_commande_livree_partiellement = True
        # Les articles dans cette commande sont ceux qui ont été livrés partiellement
        for panier in paniers:
            articles_livres.append({
                'article': panier.article,
                'quantite_livree': panier.quantite,
                'prix': panier.article.prix_unitaire,
                'sous_total': panier.sous_total
            })
        
        # Chercher la commande de renvoi associée
        commande_renvoi = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande.num_cmd}",
            client=commande.client
        ).first()
        
        # La commande source pour les articles renvoyés est la commande actuelle
        operation_livraison_partielle = commande.operations.filter(
            type_operation='LIVRAISON_PARTIELLE'
        ).order_by('-date_operation').first()

    # Cas 2: La commande actuelle est une commande de renvoi suite à une livraison partielle
    elif etat_precedent and etat_precedent.enum_etat.libelle == 'Livrée Partiellement':
        is_commande_livree_partiellement = True
        # Chercher la commande originale qui a été livrée partiellement
        commande_originale = Commande.objects.filter(
            num_cmd=commande.num_cmd.replace('RENVOI-', ''),
            client=commande.client
        ).first()
        
        # La commande source pour les articles renvoyés est la commande originale
        if commande_originale:
            operation_livraison_partielle = commande_originale.operations.filter(
                type_operation='LIVRAISON_PARTIELLE'
            ).order_by('-date_operation').first()

    # Si une opération de livraison partielle est trouvée, extraire les états des articles renvoyés
    if operation_livraison_partielle:
        try:
            details = json.loads(operation_livraison_partielle.conclusion)
            if 'recap_articles_renvoyes' in details:
                for item in details['recap_articles_renvoyes']:
                    etat_articles_renvoyes[item['article_id']] = item['etat']
        except Exception:
            pass

    # Populer articles_renvoyes si c'est une commande de renvoi ou si elle a une commande de renvoi associée
    if is_commande_livree_partiellement:
        # Si la commande actuelle est une commande de renvoi (Cas 2)
        if commande.num_cmd and commande.num_cmd.startswith('RENVOI-'):
            for panier_renvoi in paniers:
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id, 'bon')
                articles_renvoyes.append({
                    'article': panier_renvoi.article,
                    'quantite': panier_renvoi.quantite,
                    'prix': panier_renvoi.article.prix_unitaire,
                    'sous_total': panier_renvoi.sous_total,
                    'etat': etat
                })
        # Si la commande actuelle est la commande originale livrée partiellement et qu'une commande de renvoi existe (Cas 1)
        elif commande_renvoi:
            for panier_renvoi in commande_renvoi.paniers.all():
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id, 'bon')
                articles_renvoyes.append({
                    'article': panier_renvoi.article,
                    'quantite': panier_renvoi.quantite,
                    'prix': panier_renvoi.article.prix_unitaire,
                    'sous_total': panier_renvoi.sous_total,
                    'etat': etat
                })

    # Pour les articles livrés, on lit l'opération de livraison partielle sur la commande originale
    if is_commande_livree_partiellement and commande_originale:
        operation_livraison_partielle_for_livres = commande_originale.operations.filter(
            type_operation='LIVRAISON_PARTIELLE'
        ).order_by('-date_operation').first()
        if operation_livraison_partielle_for_livres:
            try:
                details = json.loads(operation_livraison_partielle_for_livres.conclusion)
                if 'articles_livres' in details:
                    for article_livre in details['articles_livres']:
                        article_id = article_livre.get('article_id')
                        if article_id:
                            article = Article.objects.filter(id=article_id).first()
                        if article:
                            articles_livres.append({
                                'article': article,
                                'quantite_livree': article_livre.get('quantite', 0),
                                'prix': article.prix_unitaire,
                                'sous_total': article.prix_unitaire * article_livre.get('quantite', 0)
                            })
            except Exception:
                pass
    
    # Calculer le total des articles
    total_articles = sum(panier.total_ligne for panier in paniers)
    
    # Récupérer les opérations associées à la commande
    operations = commande.operations.select_related('operateur').order_by('-date_operation')
    
    # Générer le code-barres pour la commande
    code128 = barcode.get_barcode_class('code128')
    barcode_instance = code128(str(commande.id_yz), writer=ImageWriter())
    buffer = BytesIO()
    barcode_instance.write(buffer, options={'write_text': False, 'module_height': 10.0})
    barcode_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    commande_barcode = f"data:image/png;base64,{barcode_base64}"

    # Gestion des actions POST (marquer comme préparée, etc.)
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Action 'commencer_preparation' supprimée car les commandes passent maintenant 
        # directement en "En préparation" lors de l'affectation
        
        if action == 'marquer_preparee':
            with transaction.atomic():
                # Marquer l'état 'En préparation' comme terminé
                etat_en_preparation, created = EnumEtatCmd.objects.get_or_create(libelle='En préparation')
                
                etat_actuel = EtatCommande.objects.filter(
                    commande=commande,
                    enum_etat=etat_en_preparation,
                    date_fin__isnull=True
                ).first()
                
                if etat_actuel:
                    etat_actuel.date_fin = timezone.now()
                    etat_actuel.operateur = operateur_profile
                    etat_actuel.save()
                
                # Créer le nouvel état 'Préparée'
                etat_preparee, created = EnumEtatCmd.objects.get_or_create(libelle='Préparée')
                EtatCommande.objects.create(
                    commande=commande,
                    enum_etat=etat_preparee,
                    operateur=operateur_profile
                )
                
                # Log de l'opération
                Operation.objects.create(
                    commande=commande,
                    type_operation='PREPARATION_TERMINEE',
                    operateur=operateur_profile,
                    conclusion=f"Commande marquée comme préparée par {operateur_profile.nom_complet}."
                )
            
            messages.success(request, f"La commande {commande.id_yz} a bien été marquée comme préparée.")
            return redirect('Superpreparation:detail_prepa', pk=commande.pk)

        elif action == 'signaler_probleme':
            with transaction.atomic():
                # 1. Terminer l'état "En préparation" actuel
                etat_en_preparation_enum = get_object_or_404(EnumEtatCmd, libelle='En préparation')
                etat_actuel = EtatCommande.objects.filter(
                    commande=commande,
                    enum_etat=etat_en_preparation_enum,
                    date_fin__isnull=True
                ).first()

                if etat_actuel:
                    etat_actuel.date_fin = timezone.now()
                    etat_actuel.commentaire = "Problème signalé par le préparateur."
                    etat_actuel.save()

                # 2. Trouver l'opérateur de confirmation d'origine
                operateur_confirmation_origine = None
                etats_precedents = commande.etats.select_related('operateur').order_by('-date_debut')
                
                for etat in etats_precedents:
                    if etat.operateur and etat.operateur.is_confirmation:
                        operateur_confirmation_origine = etat.operateur
                        break
                
                # 3. Créer l'état "Retour Confirmation" et l'affecter
                etat_retour_enum, _ = EnumEtatCmd.objects.get_or_create(
                    libelle='Retour Confirmation',
                    defaults={'ordre': 25, 'couleur': '#D97706'}
                )
                
                EtatCommande.objects.create(
                    commande=commande,
                    enum_etat=etat_retour_enum,
                    operateur=operateur_confirmation_origine, # Affectation directe
                    date_debut=timezone.now(),
                    commentaire="Retourné par la préparation pour vérification."
                )

                # 4. Log et message de succès
                if operateur_confirmation_origine:
                    log_conclusion = f"Problème signalé par {operateur_profile.nom_complet}. Commande retournée et affectée à l'opérateur {operateur_confirmation_origine.nom_complet}."
                    messages.success(request, f"La commande {commande.id_yz} a été retournée à {operateur_confirmation_origine.nom_complet} pour vérification.")
                else:
                    log_conclusion = f"Problème signalé par {operateur_profile.nom_complet}. Opérateur d'origine non trouvé, commande renvoyée au pool de confirmation."
                    messages.warning(request, f"La commande {commande.id_yz} a été renvoyée au pool de confirmation (opérateur d'origine non trouvé).")

                Operation.objects.create(
                    commande=commande,
                    type_operation='PROBLEME_SIGNALÉ',
                    operateur=operateur_profile,
                    conclusion=log_conclusion
                )

            return redirect('Superpreparation:liste_prepa')
    
    # Avant le return render dans detail_prepa
    commande_renvoi_id = None
    if commande_renvoi:
        commande_renvoi_id = commande_renvoi.id
    
    context = {
        'page_title': f'Préparation Commande {commande.id_yz}',
        'page_subtitle': f'Détails de la commande et étapes de préparation',
        'commande': commande,
        'paniers': paniers,
        'etats_commande': etats_commande,
        'etat_actuel': etat_actuel,
        'etat_precedent': etat_precedent,
        'etat_preparation': etat_preparation,
        'total_articles': total_articles,
        'operations': operations,
        'commande_barcode': commande_barcode,
        'is_commande_livree_partiellement': is_commande_livree_partiellement,
        'articles_livres': articles_livres,
        'articles_renvoyes': articles_renvoyes,
        'articles_image_urls': articles_image_urls,
        # Variables de debug/informations supplémentaires
        'commande_originale': commande_originale,
        'commande_renvoi': commande_renvoi,
        'etat_articles_renvoyes': etat_articles_renvoyes,
        'commande_renvoi_id': commande_renvoi_id,
    }
    return render(request, 'Superpreparation/detail_prepa.html', context)

# Vues supprimées - fonctionnalités maintenant gérées depuis "Suivi des Commandes Confirmées"
# def etiquette_view(request):
# def etiquettes_articles_view(request):

@superviseur_preparation_required
def api_commandes_confirmees(request):
    """API pour récupérer toutes les commandes confirmées"""
    try:
        # Récupérer toutes les commandes confirmées
        commandes_confirmees = Commande.objects.filter(
            etats__enum_etat__libelle='Confirmée'
        ).select_related('client').distinct()
        
        commandes_data = []
        for commande in commandes_confirmees:
            commandes_data.append({
                'id': commande.id_yz,
                'client_nom': f"{commande.client.prenom} {commande.client.nom}",
                'date_creation': commande.date_creation.strftime('%Y-%m-%d %H:%M:%S'),
                'total': float(commande.total_cmd)
            })
        
        return JsonResponse({
            'success': True,
            'commandes': commandes_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@superviseur_preparation_required
def api_articles_commande(request, commande_id):
    """API pour récupérer les articles d'une commande avec leurs codes-barres"""
    try:
        # Récupérer la commande avec ses paniers
        commande = get_object_or_404(
            Commande.objects.select_related('client').prefetch_related(
                'paniers__article', 'paniers__variante'
            ),
            id_yz=commande_id
        )
        
        # Préparer les données des articles
        articles_data = []
        for panier in commande.paniers.all():
            # Déterminer la référence pour le code-barres
            reference = panier.article.reference
            if panier.variante and panier.variante.reference_variante:
                reference = panier.variante.reference_variante
            
            # Générer le code-barres en utilisant le filtre existant
            try:
                from .templatetags.barcode_filters import barcode_image_url
                barcode_url = barcode_image_url(reference)
            except Exception as e:
                # En cas d'erreur, utiliser une URL vide
                barcode_url = ""
            
            article_data = {
                'nom': panier.article.nom,
                'reference': panier.article.reference,
                'quantite': panier.quantite,
                'sous_total': float(panier.sous_total),
                'barcode_url': barcode_url,
            }
            
            # Ajouter les informations de variante si elle existe
            if panier.variante:
                variante_info = {
                    'reference_variante': panier.variante.reference_variante,
                }
                
                # Ajouter les informations de couleur si elle existe
                if panier.variante.couleur:
                    variante_info['couleur'] = panier.variante.couleur.nom
                
                # Ajouter les informations de pointure si elle existe
                if panier.variante.pointure:
                    variante_info['pointure'] = panier.variante.pointure.pointure
                
                article_data['variante'] = variante_info
            
            articles_data.append(article_data)
        
        return JsonResponse({
            'success': True,
            'articles': articles_data,
            'commande_id': commande.id_yz,
            'client': f"{commande.client.prenom} {commande.client.nom}"
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@superviseur_preparation_required
def api_commande_produits(request, commande_id):
    """API pour récupérer les produits d'une commande pour les étiquettes"""
    try:
        # Récupérer la commande. La sécurité est déjà gérée par la page
        # qui appelle cette API, qui ne liste que les commandes autorisées.
        commande = Commande.objects.get(id=commande_id)
        
        # Récupérer tous les produits de la commande
        paniers = commande.paniers.all().select_related('article')
        
        # Construire la liste des produits
        produits_list = []
        for panier in paniers:
            if panier.article:
                # Format: "NOM REFERENCE , POINTURE"
                produit_info = f"{panier.article.nom or ''} {panier.article.reference or ''}".strip()
                if panier.article.pointure:
                    produit_info += f" , {panier.article.pointure}"
                
                # Ajouter la quantité si elle est supérieure à 1
                if panier.quantite > 1:
                    produit_info += f" (x{panier.quantite})" # Mettre la quantité entre parenthèses
                produits_list.append(produit_info)
        
        # Joindre tous les produits en une seule chaîne, en utilisant " + " comme séparateur
        produits_text = " + ".join(produits_list) if produits_list else "PRODUITS NON SPÉCIFIÉS"
        
        return JsonResponse({
            'success': True,
            'produits': produits_text,
            'nombre_articles': len(produits_list)
        })
        
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})

# API api_changer_etat_preparation supprimée car les commandes passent maintenant 
# directement de "Confirmée" à "En préparation" lors de l'affectation

@superviseur_preparation_required
def modifier_commande_prepa(request, commande_id):
    """Page de modification complète d'une commande pour les opérateurs de préparation"""
    import json
    from commande.models import Commande, Operation
    from parametre.models import Ville
    
    try:
        # Récupérer l'opérateur
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil d'opérateur de préparation non trouvé.")
        return redirect('login')
    
    # Récupérer la commande
    commande = get_object_or_404(Commande, id=commande_id)
    
    # Vérifier que la commande est affectée à cet opérateur pour la préparation
    etat_preparation = commande.etats.filter(
        Q(enum_etat__libelle='À imprimer') | Q(enum_etat__libelle='En préparation'),
        operateur=operateur,
        date_fin__isnull=True
    ).first()
    
    if not etat_preparation:
        messages.error(request, "Cette commande ne vous est pas affectée pour la préparation.")
        return redirect('Superpreparation:liste_prepa')
    
    if request.method == 'POST':
        try:
            # ================ GESTION DES ACTIONS AJAX SPÉCIFIQUES ================
            action = request.POST.get('action')
            
            if action == 'add_article':
                # Ajouter un nouvel article immédiatement
                from article.models import Article
                from commande.models import Panier
                
                article_id = request.POST.get('article_id')
                quantite = int(request.POST.get('quantite', 1))
                
                try:
                    article = Article.objects.get(id=article_id)
                    
                    # Vérifier si l'article existe déjà dans la commande
                    panier_existant = Panier.objects.filter(commande=commande, article=article).first()
                    
                    if panier_existant:
                        # Si l'article existe déjà, mettre à jour la quantité
                        panier_existant.quantite += quantite
                        panier_existant.save()
                        panier = panier_existant
                        print(f"🔄 Article existant mis à jour: ID={article.id}, nouvelle quantité={panier.quantite}")
                    else:
                        # Si l'article n'existe pas, créer un nouveau panier
                        panier = Panier.objects.create(
                            commande=commande,
                            article=article,
                            quantite=quantite,
                            sous_total=0  # Sera recalculé après
                        )
                        print(f"➕ Nouvel article ajouté: ID={article.id}, quantité={quantite}")
                    
                    # Recalculer le compteur après ajout (logique de confirmation)
                    if article.isUpsell and hasattr(article, 'prix_upsell_1') and article.prix_upsell_1 is not None:
                        # Compter la quantité totale d'articles upsell (après ajout)
                        total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                            total=Sum('quantite')
                        )['total'] or 0
                        
                        # Le compteur ne s'incrémente qu'à partir de 2 unités d'articles upsell
                        # 0-1 unités upsell → compteur = 0
                        # 2+ unités upsell → compteur = total_quantite_upsell - 1
                        if total_quantite_upsell >= 2:
                            commande.compteur = total_quantite_upsell - 1
                        else:
                            commande.compteur = 0
                        
                        commande.save()
                        
                        # Recalculer TOUS les articles de la commande avec le nouveau compteur
                        commande.recalculer_totaux_upsell()
                    else:
                        # Pour les articles normaux, juste calculer le sous-total
                        from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                        prix_unitaire = get_prix_upsell_avec_compteur(article, commande.compteur)
                        sous_total = prix_unitaire * panier.quantite
                        panier.sous_total = float(sous_total)
                        panier.save()
                    
                    # Recalculer le total de la commande avec frais de livraison
                    total_articles = commande.paniers.aggregate(
                        total=Sum('sous_total')
                    )['total'] or 0
                    frais_livraison = commande.ville.frais_livraison if commande.ville else 0
                    commande.total_cmd = float(total_articles) #+ float(frais_livraison)
                    commande.save()
                    
                    # Calculer les statistiques upsell pour la réponse
                    articles_upsell = commande.paniers.filter(article__isUpsell=True)
                    total_quantite_upsell = articles_upsell.aggregate(
                        total=Sum('quantite')
                    )['total'] or 0
                    
                    # Déterminer si c'était un ajout ou une mise à jour
                    message = 'Article ajouté avec succès' if not panier_existant else f'Quantité mise à jour ({panier.quantite})'
                    
                    # Préparer les données de l'article pour le frontend
                    article_data = {
                        'panier_id': panier.id,
                        'nom': article.nom,
                        'reference': article.reference,
                        'couleur_fr': article.couleur or '',
                        'couleur_ar': article.couleur or '',
                        'pointure': article.pointure or '',
                        'quantite': panier.quantite,
                        'prix': panier.sous_total / panier.quantite,  # Prix unitaire
                        'sous_total': panier.sous_total,
                        'is_upsell': article.isUpsell,
                        'phase': article.phase,
                        'has_promo_active': article.has_promo_active,
                        'qte_disponible': article.qte_disponible,
                        'description': article.description or ''
                    }
                    
                    return JsonResponse({
                        'success': True,
                        'message': message,
                        'article_id': panier.id,
                        'total_commande': float(commande.total_cmd),
                        'nb_articles': commande.paniers.count(),
                        'compteur': commande.compteur,
                        'was_update': panier_existant is not None,
                        'new_quantity': panier.quantite,
                        'article_data': article_data,
                        'articles_count': commande.paniers.count(),
                        'sous_total_articles': float(sum(p.sous_total for p in commande.paniers.all())),
                        'articles_upsell': articles_upsell.count(),
                        'quantite_totale_upsell': total_quantite_upsell
                    })
                    
                except Article.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'Article non trouvé'
                    })
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    })
            
            elif action == 'replace_article':
                # Remplacer un article existant
                from article.models import Article
                from commande.models import Panier
                
                ancien_article_id = request.POST.get('ancien_article_id')
                nouvel_article_id = request.POST.get('nouvel_article_id')
                nouvelle_quantite = int(request.POST.get('nouvelle_quantite', 1))
                
                try:
                    # Supprimer l'ancien panier et décrémenter le compteur
                    ancien_panier = Panier.objects.get(id=ancien_article_id, commande=commande)
                    ancien_article = ancien_panier.article
                    
                    # Sauvegarder les infos avant suppression
                    ancien_etait_upsell = ancien_article.isUpsell
                    
                    # Supprimer l'ancien panier
                    ancien_panier.delete()
                    
                    # Créer le nouveau panier
                    nouvel_article = Article.objects.get(id=nouvel_article_id)
                    
                    # Recalculer le compteur après remplacement (logique de confirmation)
                    total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                        total=Sum('quantite')
                    )['total'] or 0
                    
                    # Ajouter la quantité si le nouvel article est upsell
                    if nouvel_article.isUpsell:
                        total_quantite_upsell += nouvelle_quantite
                    
                    # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                    if total_quantite_upsell >= 2:
                        commande.compteur = total_quantite_upsell - 1
                    else:
                        commande.compteur = 0
                    
                    commande.save()
                    
                    # Recalculer TOUS les articles de la commande avec le nouveau compteur
                    commande.recalculer_totaux_upsell()
                    
                    # Calculer le sous-total selon le compteur de la commande
                    from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                    prix_unitaire = get_prix_upsell_avec_compteur(nouvel_article, commande.compteur)
                    sous_total = prix_unitaire * nouvelle_quantite
                    
                    nouveau_panier = Panier.objects.create(
                        commande=commande,
                        article=nouvel_article,
                        quantite=nouvelle_quantite,
                        sous_total=float(sous_total)
                    )
                    
                    # Recalculer le total de la commande avec frais de livraison
                    total_commande = commande.paniers.aggregate(
                        total=Sum('sous_total')
                    )['total'] or 0
                    frais_livraison = commande.ville.frais_livraison if commande.ville else 0
                    commande.total_cmd = float(total_commande) #+ float(frais_livraison)
                    commande.save()
                    
                    # Calculer les statistiques upsell pour la réponse
                    articles_upsell = commande.paniers.filter(article__isUpsell=True)
                    total_quantite_upsell = articles_upsell.aggregate(
                        total=Sum('quantite')
                    )['total'] or 0
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'Article remplacé avec succès',
                        'nouvel_article_id': nouveau_panier.id,
                        'total_commande': float(commande.total_cmd),
                        'nb_articles': commande.paniers.count(),
                        'compteur': commande.compteur,
                        'articles_upsell': articles_upsell.count(),
                        'quantite_totale_upsell': total_quantite_upsell,
                        'sous_total_articles': float(commande.sous_total_articles)
                    })
                    
                except Panier.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'Article original non trouvé'
                    })
                except Article.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'Nouvel article non trouvé'
                    })
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    })
            
            
            elif action == 'update_operation':
                # Mettre à jour une opération existante
                try:
                    from commande.models import Operation
                    import logging
                    logger = logging.getLogger(__name__)
                    
                    operation_id = request.POST.get('operation_id')
                    nouveau_commentaire = request.POST.get('nouveau_commentaire', '').strip()
                    
                    if not operation_id or not nouveau_commentaire:
                        return JsonResponse({'success': False, 'error': 'ID opération et commentaire requis'})
                    
                    # Récupérer et mettre à jour l'opération
                    operation = Operation.objects.get(id=operation_id, commande=commande)
                    operation.conclusion = nouveau_commentaire
                    operation.operateur = operateur  # Mettre à jour l'opérateur qui modifie
                    operation.save()
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'Opération mise à jour avec succès',
                        'operation_id': operation_id,
                        'nouveau_commentaire': nouveau_commentaire
                    })
                    
                except Operation.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': 'Opération non trouvée'
                    })
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    })
            
            elif action == 'add_operation':
                # Ajouter une nouvelle opération
                try:
                    from commande.models import Operation
                    
                    type_operation = request.POST.get('type_operation', '').strip()
                    commentaire = request.POST.get('commentaire', '').strip()
                    
                    if not type_operation or not commentaire:
                        return JsonResponse({
                            'success': False,
                            'error': 'Type d\'opération et commentaire requis'
                        })
                    
                    # Créer la nouvelle opération
                    operation = Operation.objects.create(
                        commande=commande,
                        type_operation=type_operation,
                        conclusion=commentaire,
                        operateur=operateur
                    )
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'Opération ajoutée avec succès',
                        'operation_id': operation.id,
                        'type_operation': type_operation,
                        'commentaire': commentaire
                    })
                    
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    })
            
            elif action == 'modifier_quantites_multiple':
                # Modifier plusieurs quantités d'articles en une fois
                try:
                    from commande.models import Panier
                    import json
                    
                    modifications_json = request.POST.get('modifications', '[]')
                    modifications = json.loads(modifications_json)
                    
                    if not modifications:
                        return JsonResponse({
                            'success': False,
                            'error': 'Aucune modification fournie'
                        })
                    
                    # Appliquer les modifications
                    for mod in modifications:
                        panier_id = mod.get('panier_id')
                        nouvelle_quantite = mod.get('nouvelle_quantite', 0)
                        
                        try:
                            panier = Panier.objects.get(id=panier_id, commande=commande)
                            
                            if nouvelle_quantite <= 0:
                                # Supprimer l'article si quantité = 0
                                panier.delete()
                            else:
                                # Mettre à jour la quantité et le sous-total
                                panier.quantite = nouvelle_quantite
                                panier.sous_total = float(panier.article.prix_unitaire * nouvelle_quantite)
                                panier.save()
                                
                        except Panier.DoesNotExist:
                            continue  # Ignorer les paniers non trouvés
                    
                    # Recalculer le total de la commande
                    total_commande = commande.paniers.aggregate(
                        total=Sum('sous_total')
                    )['total'] or 0
                    commande.total_cmd = float(total_commande)
                    commande.save()
                    
                    # Créer une opération pour consigner la modification
                    Operation.objects.create(
                        commande=commande,
                        type_operation='MODIFICATION_QUANTITES',
                        conclusion=f"Modification en masse des quantités d'articles par l'opérateur de préparation.",
                        operateur=operateur
                    )
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'{len(modifications)} quantité(s) modifiée(s) avec succès',
                        'total_commande': float(commande.total_cmd),
                        'nb_articles': commande.paniers.count(),
                    })
                    
                except json.JSONDecodeError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Format de données invalide'
                    })
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    })
            
            elif action == 'modifier_quantite_directe':
                # Modifier directement la quantité d'un article
                try:
                    from commande.models import Panier
                    
                    panier_id = request.POST.get('panier_id')
                    nouvelle_quantite = int(request.POST.get('nouvelle_quantite', 0))
                    
                    print(f"🔄 Modification quantité directe - Panier ID: {panier_id}, Nouvelle quantité: {nouvelle_quantite}")
                    
                    if nouvelle_quantite < 0:
                        return JsonResponse({
                            'success': False,
                            'error': 'La quantité ne peut pas être négative'
                        })
                    
                    try:
                        panier = Panier.objects.get(id=panier_id, commande=commande)
                        ancienne_quantite = panier.quantite
                        nouveau_sous_total = 0
                        
                        print(f"📦 Article trouvé: {panier.article.nom}, Ancienne quantité: {ancienne_quantite}")
                        
                        # Vérifier que le panier appartient bien à cette commande
                        if panier.commande.id != commande.id:
                            print(f"❌ ERREUR: Le panier {panier_id} n'appartient pas à la commande {commande.id}")
                            return JsonResponse({
                                'success': False,
                                'error': 'Article non trouvé dans cette commande'
                            })
                        
                        if nouvelle_quantite == 0:
                            # Supprimer l'article si quantité = 0
                            panier.delete()
                            message = 'Article supprimé avec succès'
                        else:
                            # Mettre à jour la quantité et le sous-total avec la logique complète de prix
                            panier.quantite = nouvelle_quantite
                            
                            # Recalculer le compteur si c'est un article upsell
                            if panier.article.isUpsell:
                                # Sauvegarder d'abord la nouvelle quantité
                                panier.save()
                                
                                # Compter la quantité totale d'articles upsell après modification
                                total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                                    total=Sum('quantite')
                                )['total'] or 0
                                
                                print(f"🔄 Calcul du compteur upsell: total_quantite_upsell = {total_quantite_upsell}")
                                
                                # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                                if total_quantite_upsell >= 2:
                                    commande.compteur = total_quantite_upsell - 1
                                else:
                                    commande.compteur = 0
                                
                                print(f"✅ Nouveau compteur calculé: {commande.compteur}")
                                commande.save()
                                
                                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                                commande.recalculer_totaux_upsell()
                            else:
                                # Pour les articles normaux, mettre à jour la quantité et calculer le sous-total
                                from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                                prix_unitaire = get_prix_upsell_avec_compteur(panier.article, commande.compteur)
                                panier.quantite = nouvelle_quantite
                                panier.sous_total = float(prix_unitaire * nouvelle_quantite)
                                panier.save()
                            
                            nouveau_sous_total = panier.sous_total
                            message = 'Quantité modifiée avec succès'
                            
                            print(f"✅ Quantité mise à jour: {ancienne_quantite} → {nouvelle_quantite}, Nouveau sous-total: {nouveau_sous_total}")
                            
                            # Vérifier que la mise à jour a bien été sauvegardée
                            panier.refresh_from_db()
                            if panier.quantite != nouvelle_quantite:
                                print(f"❌ ERREUR: La quantité n'a pas été sauvegardée correctement. Attendu: {nouvelle_quantite}, Actuel: {panier.quantite}")
                            else:
                                print(f"✅ Vérification OK: Quantité sauvegardée: {panier.quantite}")
                            
                    except Panier.DoesNotExist:
                        return JsonResponse({
                            'success': False,
                            'error': 'Article non trouvé'
                        })
                    
                    # Recalculer le total de la commande avec frais de livraison
                    total_articles = commande.paniers.aggregate(
                        total=Sum('sous_total')
                    )['total'] or 0
                    frais_livraison = commande.ville.frais_livraison if commande.ville else 0
                    commande.total_cmd = float(total_articles) #+ float(frais_livraison)
                    commande.save()
                    
                    # Calculer les statistiques upsell pour la réponse
                    articles_upsell = commande.paniers.filter(article__isUpsell=True)
                    total_quantite_upsell = articles_upsell.aggregate(
                        total=Sum('quantite')
                    )['total'] or 0
                    
                    # Créer une opération pour consigner la modification
                    Operation.objects.create(
                        commande=commande,
                        type_operation='MODIFICATION_QUANTITE',
                        conclusion=f"Quantité d'article modifiée de {ancienne_quantite} à {nouvelle_quantite}.",
                        operateur=operateur
                    )
                    
                    return JsonResponse({
                        'success': True,
                        'message': message,
                        'sous_total': float(nouveau_sous_total),
                        'sous_total_articles': float(total_articles),
                        'total_commande': float(commande.total_cmd),
                        'frais_livraison': float(frais_livraison),
                        'nb_articles': commande.paniers.count(),
                        'compteur': commande.compteur,
                        'articles_upsell': articles_upsell.count(),
                        'quantite_totale_upsell': total_quantite_upsell
                    })
                    
                except ValueError:
                    return JsonResponse({
                        'success': False,
                        'error': 'Quantité invalide'
                    })
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    })
            
            elif action == 'update_commande_info':
                # Mettre à jour les informations de base de la commande
                try:
                    # Récupérer les données du formulaire
                    nouvelle_adresse = request.POST.get('adresse', '').strip()
                    nouvelle_ville_id = request.POST.get('ville_id')
                    
                    # Mettre à jour l'adresse
                    if nouvelle_adresse:
                        commande.adresse = nouvelle_adresse
                    
                    # Mettre à jour la ville si fournie
                    if nouvelle_ville_id:
                        try:
                            nouvelle_ville = Ville.objects.get(id=nouvelle_ville_id)
                            commande.ville = nouvelle_ville
                        except Ville.DoesNotExist:
                            return JsonResponse({
                                'success': False,
                                'error': 'Ville non trouvée'
                            })
                    
                    commande.save()
                    
                    # Créer une opération pour consigner la modification
                    Operation.objects.create(
                        commande=commande,
                        type_operation='MODIFICATION_PREPA',
                        conclusion=f"La commande a été modifiée par l'opérateur.",
                        operateur=operateur
                    )

                    messages.success(request, f"Commande {commande.id_yz} mise à jour avec succès.")
                    return redirect('Superpreparation:detail_prepa', pk=commande.id)
                    
                except Exception as e:
                    return JsonResponse({
                        'success': False,
                        'error': str(e)
                    })
            
            else:
                # Traitement du formulaire principal (non-AJAX)
                with transaction.atomic():
                    # Mettre à jour les informations du client
                    client = commande.client
                    client.nom = request.POST.get('client_nom', client.nom).strip()
                    client.prenom = request.POST.get('client_prenom', client.prenom).strip()
                    client.numero_tel = request.POST.get('client_telephone', client.numero_tel).strip()
                    client.save()

                    # Mettre à jour les informations de base de la commande
                    nouvelle_adresse = request.POST.get('adresse', '').strip()
                    nouvelle_ville_id = request.POST.get('ville_id')
                    
                    if nouvelle_adresse:
                        commande.adresse = nouvelle_adresse
                    
                    if nouvelle_ville_id:
                        try:
                            nouvelle_ville = Ville.objects.get(id=nouvelle_ville_id)
                            commande.ville = nouvelle_ville
                        except Ville.DoesNotExist:
                            messages.error(request, "Ville sélectionnée non trouvée.")
                            return redirect('Superpreparation:modifier_commande', commande_id=commande.id)
                    
                    commande.save()

                    # Créer une opération pour consigner la modification
                    Operation.objects.create(
                        commande=commande,
                        type_operation='MODIFICATION_PREPA',
                        conclusion=f"La commande a été modifiée par l'opérateur.",
                        operateur=operateur
                    )

                    messages.success(request, f"Les modifications de la commande {commande.id_yz} ont été enregistrées avec succès.")
                    return redirect('Superpreparation:detail_prepa', pk=commande.id)
                
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification: {str(e)}")
            return redirect('Superpreparation:modifier_commande', commande_id=commande.id)
    
    # Récupérer les données pour l'affichage
    paniers = commande.paniers.all().select_related('article')
    operations = commande.operations.all().select_related('operateur').order_by('-date_operation')
    villes = Ville.objects.all().order_by('nom')
    
    # Calculer le total des articles
    total_articles = sum(panier.sous_total for panier in paniers)
    
    # Vérifier si c'est une commande renvoyée par la logistique
    operation_renvoi = operations.filter(type_operation='RENVOI_PREPARATION').first()
    is_commande_renvoyee = operation_renvoi is not None
    
    # Initialiser les variables pour les cas de livraison partielle/renvoi
    articles_livres = []
    articles_renvoyes = []
    is_commande_livree_partiellement = False
    commande_renvoi_obj = None # Variable pour la commande de renvoi trouvée
    commande_originale_obj = None # Variable pour la commande originale trouvée
    etat_articles_renvoyes = {} # Dictionnaire pour stocker l'état des articles renvoyés (article_id -> etat)
    operation_livraison_partielle_source = None # Opération source pour les détails de livraison partielle

    # Récupérer l'état actuel de la commande
    etat_actuel = commande.etats.filter(date_fin__isnull=True).first()
    etat_precedent = None
    
    if etat_actuel:
        # Trouver l'état précédent
        etats_precedents = commande.etats.all().order_by('-date_debut')
        for etat in etats_precedents:
            if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                if etat.enum_etat.libelle not in ['À imprimer', 'En préparation']:
                    etat_precedent = etat
                    break
    
    # NOUVELLE LOGIQUE POUR DÉTECTER LA LIVRAISON PARTIELLE ET LES ARTICLES RENVOYÉS
    # Une commande est considérée comme "livrée partiellement" dans le contexte de modification
    # si elle-même a été livrée partiellement ou si c'est une commande de RENVOI associée à une livraison partielle.
    
    if commande.num_cmd and commande.num_cmd.startswith('RENVOI-'):
        # C'est une commande de renvoi. On cherche la commande originale.
        num_cmd_original = commande.num_cmd.replace('RENVOI-', '')
        commande_originale_obj = Commande.objects.filter(num_cmd=num_cmd_original, client=commande.client).first()
        
        if commande_originale_obj:
            # Vérifier si la commande originale a bien été livrée partiellement
            if commande_originale_obj.etats.filter(enum_etat__libelle='Livrée Partiellement').exists():
                is_commande_livree_partiellement = True
                operation_livraison_partielle_source = commande_originale_obj.operations.filter(
                    type_operation='LIVRAISON_PARTIELLE'
                ).order_by('-date_operation').first()
                commande_renvoi_obj = commande # Dans ce cas, la commande actuelle est la commande de renvoi

    elif etat_actuel and etat_actuel.enum_etat.libelle == 'Livrée Partiellement':
        # La commande actuelle est l'originale qui a été livrée partiellement
        is_commande_livree_partiellement = True
        operation_livraison_partielle_source = commande.operations.filter(
            type_operation='LIVRAISON_PARTIELLE'
        ).order_by('-date_operation').first()
        # Chercher une commande de renvoi associée si elle existe
        commande_renvoi_obj = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande.num_cmd}",
            client=commande.client
        ).first()

    # Si une opération de livraison partielle est trouvée, extraire les états des articles renvoyés
    if operation_livraison_partielle_source:
        try:
            details = json.loads(operation_livraison_partielle_source.conclusion)
            if 'recap_articles_renvoyes' in details:
                for item in details['recap_articles_renvoyes']:
                    etat_articles_renvoyes[item['article_id']] = item['etat']
            
            # Populer articles_livres à partir de la conclusion de l'opération de livraison partielle
            if 'articles_livres' in details:
                for article_livre in details['articles_livres']:
                    article_id = article_livre.get('article_id')
                    if article_id:
                        article_obj = Article.objects.filter(id=article_id).first()
                    if article_obj:
                        articles_livres.append({
                            'article': article_obj,
                            'quantite_livree': article_livre.get('quantite', 0),
                            'prix': article_obj.prix_unitaire,
                            'sous_total': article_obj.prix_unitaire * article_livre.get('quantite', 0)
                        })
        except Exception as e:
            print(f"DEBUG: Erreur lors du parsing des détails de l'opération de livraison partielle: {e}")
            pass

    # Populer articles_renvoyes si c'est une commande de renvoi ou si elle a une commande de renvoi associée
    if is_commande_livree_partiellement:
        # Si la commande actuelle est une commande de renvoi (celle que nous modifions)
        if commande.num_cmd and commande.num_cmd.startswith('RENVOI-'):
            # Les paniers de la commande actuelle sont les articles renvoyés
            for panier_renvoi in paniers:
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id)
                if etat is None:
                    etat = 'inconnu'
                    print(f"ALERTE: État inconnu pour l'article ID {panier_renvoi.article.id} dans la commande {commande.id_yz}")
                articles_renvoyes.append({
                    'article': panier_renvoi.article,
                    'quantite': panier_renvoi.quantite,
                    'prix': panier_renvoi.article.prix_unitaire,
                    'sous_total': panier_renvoi.sous_total,
                    'etat': etat
                })
        # Si la commande actuelle est la commande originale livrée partiellement (Cas 1 initial)
        elif commande_renvoi_obj:
            # Les paniers de la commande de renvoi associée sont les articles renvoyés
            for panier_renvoi in commande_renvoi_obj.paniers.all():
                etat = etat_articles_renvoyes.get(panier_renvoi.article.id)
                if etat is None:
                    etat = 'inconnu'
                    print(f"ALERTE: État inconnu pour l'article ID {panier_renvoi.article.id} dans la commande {commande_renvoi_obj.id_yz}")
                articles_renvoyes.append({
                    'article': panier_renvoi.article,
                    'quantite': panier_renvoi.quantite,
                    'prix': panier_renvoi.article.prix_unitaire,
                    'sous_total': panier_renvoi.sous_total,
                    'etat': etat
                })
    
    # DEBUG: Afficher le contenu de articles_renvoyes après peuplement
    print(f"DEBUG (modifier_commande_prepa): articles_renvoyes APRES POPULATION: {articles_renvoyes}")

    # Créer un map pour accéder facilement aux articles renvoyés par leur ID dans le template
    # articles_renvoyes_map = {item['article'].id: item for item in articles_renvoyes}

    # Pour les articles livrés, on lit l'opération de livraison partielle sur la commande originale
    # C'est pertinent uniquement si la commande actuelle est la commande de renvoi
    # if is_commande_livree_partiellement and commande.num_cmd and commande.num_cmd.startswith('RENVOI-') and commande_originale_obj:
    #     operation_livraison_partielle_for_livres = commande_originale_obj.operations.filter(
    #         type_operation='LIVRAISON_PARTIELLE'
    #     ).order_by('-date_operation').first()
    #     if operation_livraison_partielle_for_livres:
    #         try:
    #             details = json.loads(operation_livraison_partielle_for_livres.conclusion)
    #             if 'articles_livres' in details:
    #                 for article_livre in details['articles_livres']:
    #                     article = Article.objects.filter(id=article_livre.get('article_id')).first()
    #                     if article:
    #                         articles_livres.append({
    #                             'article': article,
    #                             'quantite_livree': article_livre.get('quantite', 0),
    #                             'prix': article.prix_unitaire,
    #                             'sous_total': article.prix_unitaire * article_livre.get('quantite', 0)
    #                         })
    #         except Exception:
    #             pass

    context = {
        'page_title': "Modifier Commande " + str(commande.id_yz),
        'page_subtitle': "Modification des détails de la commande en préparation",
        'commande': commande,
        'paniers': paniers,
        'villes': villes,
        'total_articles': total_articles,
        'is_commande_renvoyee': is_commande_renvoyee,
        'operation_renvoi': operation_renvoi,
        'is_commande_livree_partiellement': is_commande_livree_partiellement,
        'articles_livres': articles_livres,
        'articles_renvoyes': articles_renvoyes,
        # Variables de debug/informations supplémentaires
        'commande_originale': commande_originale_obj,
        'commande_renvoi': commande_renvoi_obj,
        'etat_articles_renvoyes': etat_articles_renvoyes,
        # 'articles_renvoyes_map': articles_renvoyes_map, # Retiré car plus nécessaire
    }
    return render(request, 'Superpreparation/modifier_commande.html', context)

@superviseur_preparation_required
def modifier_commande_superviseur(request, commande_id):
    """Page de modification complète d'une commande pour les superviseurs de préparation"""
    import json
    from commande.models import Commande, Operation
    from parametre.models import Ville
    
    try:
        # Accepter PREPARATION et SUPERVISEUR_PREPARATION
        operateur = Operateur.objects.get(user=request.user, actif=True)
        if operateur.type_operateur not in ['PREPARATION', 'SUPERVISEUR_PREPARATION']:
            messages.error(request, "Accès non autorisé. Réservé à l'équipe préparation.")
            return redirect('Superpreparation:home')
    except Operateur.DoesNotExist:
        # Pas de profil: continuer (le décorateur a déjà validé l'accès via groupes)
        operateur = None
    
    # Récupérer la commande avec ses relations
    commande = get_object_or_404(Commande.objects.select_related('client', 'ville', 'ville__region'), id=commande_id)
    
    # Pour les superviseurs, on ne vérifie pas l'affectation spécifique
    # Ils peuvent modifier toutes les commandes en préparation
    
    if request.method == 'POST':
        try:
            # ================ GESTION DES ACTIONS AJAX SPÉCIFIQUES ================
            action = request.POST.get('action')
            
            if action == 'add_article':
                # Ajouter un nouvel article immédiatement
                from article.models import Article
                from commande.models import Panier
                
                article_id = request.POST.get('article_id')
                quantite = int(request.POST.get('quantite', 1))
                
                try:
                    article = Article.objects.get(id=article_id)
                    
                    # Vérifier si l'article existe déjà dans la commande
                    panier_existant = Panier.objects.filter(commande=commande, article=article).first()
                    
                    if panier_existant:
                        # Si l'article existe déjà, mettre à jour la quantité
                        panier_existant.quantite += quantite
                        panier_existant.save()
                        panier = panier_existant
                        print(f"🔄 Article existant mis à jour: ID={article.id}, nouvelle quantité={panier.quantite}")
                    else:
                        # Si l'article n'existe pas, créer un nouveau panier
                        panier = Panier.objects.create(
                            commande=commande,
                            article=article,
                            quantite=quantite,
                            sous_total=0  # Sera recalculé après
                        )
                        print(f"➕ Nouvel article ajouté: ID={article.id}, quantité={quantite}")
                    
                    # Recalculer le compteur après ajout (logique de confirmation)
                    if article.isUpsell and hasattr(article, 'prix_upsell_1') and article.prix_upsell_1 is not None:
                        # Compter la quantité totale d'articles upsell (après ajout)
                        total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                            total=Sum('quantite')
                        )['total'] or 0
                        
                        # Le compteur ne s'incrémente qu'à partir de 2 unités d'articles upsell
                        # 0-1 unités upsell → compteur = 0
                        # 2+ unités upsell → compteur = total_quantite_upsell - 1
                        if total_quantite_upsell >= 2:
                            commande.compteur = total_quantite_upsell - 1
                        else:
                            commande.compteur = 0
                        
                        commande.save()
                        
                        # Recalculer TOUS les articles de la commande avec le nouveau compteur
                        commande.recalculer_totaux_upsell()
                    else:
                        # Pour les articles normaux, juste calculer le sous-total
                        from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                        prix_unitaire = get_prix_upsell_avec_compteur(article, commande.compteur)
                        sous_total = prix_unitaire * panier.quantite
                        panier.sous_total = float(sous_total)
                        panier.save()
                    
                    # Recalculer le total de la commande avec frais de livraison
                    total_articles = commande.paniers.aggregate(
                        total=Sum('sous_total')
                    )['total'] or 0
                    frais_livraison = commande.ville.frais_livraison if commande.ville else 0
                    commande.total_cmd = float(total_articles) #+ float(frais_livraison)
                    commande.save()
                    
                    # Calculer les statistiques upsell pour la réponse
                    articles_upsell = commande.paniers.filter(article__isUpsell=True)
                    total_quantite_upsell = articles_upsell.aggregate(
                        total=Sum('quantite')
                    )['total'] or 0
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'Article {article.nom} ajouté avec succès',
                        'article_id': article.id,
                        'article_nom': article.nom,
                        'quantite': panier.quantite,
                        'sous_total': float(panier.sous_total),
                        'compteur_upsell': commande.compteur,
                        'total_quantite_upsell': total_quantite_upsell,
                        'total_commande': float(commande.total_cmd),
                        'frais_livraison': float(frais_livraison),
                        'total_final': float(commande.total_cmd) + float(frais_livraison)
                    })
                    
                except Article.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Article non trouvé'})
                except Exception as e:
                    print(f"❌ Erreur lors de l'ajout d'article: {e}")
                    return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
            
            elif action in ['modify_quantity', 'modifier_quantite_directe', 'update_article']:
                # Modifier la quantité d'un article existant
                from commande.models import Panier
                
                panier_id = request.POST.get('panier_id')
                # Les différentes actions peuvent envoyer des noms différents
                quantite_str = request.POST.get('nouvelle_quantite') or request.POST.get('quantite') or request.POST.get('new_quantity')
                try:
                    nouvelle_quantite = int(quantite_str) if quantite_str is not None else 1
                except ValueError:
                    nouvelle_quantite = 1
                
                try:
                    panier = Panier.objects.get(id=panier_id, commande=commande)
                    article = panier.article
                    
                    if nouvelle_quantite <= 0:
                        # Supprimer l'article si quantité <= 0
                        panier.delete()
                        message = f'Article {article.nom} supprimé de la commande'
                    else:
                        # Mettre à jour la quantité
                        panier.quantite = nouvelle_quantite
                        
                        # Recalculer le sous-total
                        if article.isUpsell and hasattr(article, 'prix_upsell_1') and article.prix_upsell_1 is not None:
                            # Recalculer le compteur upsell
                            total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                                total=Sum('quantite')
                            )['total'] or 0
                            
                            if total_quantite_upsell >= 2:
                                commande.compteur = total_quantite_upsell - 1
                            else:
                                commande.compteur = 0
                            
                            commande.save()
                            commande.recalculer_totaux_upsell()
                        else:
                            from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                            prix_unitaire = get_prix_upsell_avec_compteur(article, commande.compteur)
                            sous_total = prix_unitaire * panier.quantite
                            panier.sous_total = float(sous_total)
                            panier.save()
                        
                        message = f'Quantité de {article.nom} mise à jour: {nouvelle_quantite}'
                    
                    # Recalculer le total de la commande
                    total_articles = commande.paniers.aggregate(
                        total=Sum('sous_total')
                    )['total'] or 0
                    frais_livraison = commande.ville.frais_livraison if commande.ville else 0
                    commande.total_cmd = float(total_articles) #+ float(frais_livraison)
                    commande.save()
                    
                    return JsonResponse({
                        'success': True,
                        'message': message,
                        'sous_total': float(panier.sous_total) if nouvelle_quantite > 0 else 0,
                        'compteur_upsell': commande.compteur,
                        'total_commande': float(commande.total_cmd),
                        'frais_livraison': float(frais_livraison),
                        'total_final': float(commande.total_cmd) + float(frais_livraison)
                    })
                    
                except Panier.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Article non trouvé dans la commande'})
                except Exception as e:
                    print(f"❌ Erreur lors de la modification de quantité: {e}")
                    return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
            
            elif action == 'delete_article':
                # Supprimer un article de la commande
                from commande.models import Panier
                
                panier_id = request.POST.get('panier_id')
                
                try:
                    panier = Panier.objects.get(id=panier_id, commande=commande)
                    article_nom = panier.article.nom
                    etait_upsell = panier.article.isUpsell  # Sauvegarder avant suppression
                    panier.delete()
                    
                    # Recalculer le compteur upsell si nécessaire
                    if etait_upsell:
                        total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                            total=Sum('quantite')
                        )['total'] or 0
                        
                        if total_quantite_upsell >= 2:
                            commande.compteur = total_quantite_upsell - 1
                        else:
                            commande.compteur = 0
                        
                        commande.save()
                        commande.recalculer_totaux_upsell()
                    
                    # Recalculer le total de la commande
                    total_articles = commande.paniers.aggregate(
                        total=Sum('sous_total')
                    )['total'] or 0
                    frais_livraison = commande.ville.frais_livraison if commande.ville else 0
                    commande.total_cmd = float(total_articles) #+ float(frais_livraison)
                    commande.save()
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'Article {article_nom} supprimé avec succès',
                        'compteur_upsell': commande.compteur,
                        'articles_count': commande.paniers.count(),
                        'total_commande': float(commande.total_cmd),
                        'frais_livraison': float(frais_livraison),
                        'total_final': float(commande.total_cmd) + float(frais_livraison)
                    })
                    
                except Panier.DoesNotExist:
                    return JsonResponse({'success': False, 'message': 'Article non trouvé dans la commande'})
                except Exception as e:
                    print(f"❌ Erreur lors de la suppression d'article: {e}")
                    return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
            
            else:
                # Soumission du formulaire principal (pas d'action AJAX)
                try:
                    nouvelle_adresse = request.POST.get('adresse', '').strip()
                    nouvelle_ville_id = request.POST.get('ville_id')

                    if nouvelle_adresse:
                        commande.adresse = nouvelle_adresse

                    if nouvelle_ville_id:
                        try:
                            nouvelle_ville = Ville.objects.get(id=nouvelle_ville_id)
                            commande.ville = nouvelle_ville
                        except Ville.DoesNotExist:
                            messages.error(request, "Ville sélectionnée non trouvée.")
                            return redirect('Superpreparation:modifier_commande_superviseur', commande_id=commande.id)

                    commande.save()

                    # Journaliser l'opération
                    Operation.objects.create(
                        commande=commande,
                        type_operation='MODIFICATION_PREPA',
                        conclusion=request.POST.get('commentaire_operateur', "Modifications enregistrées par le superviseur."),
                        operateur=operateur
                    )

                    messages.success(request, f"Les modifications de la commande {commande.id_yz} ont été enregistrées avec succès.")
                    return redirect('Superpreparation:detail_prepa', pk=commande.id)
                except Exception as e:
                    messages.error(request, f"Erreur lors de l'enregistrement: {str(e)}")
                    return redirect('Superpreparation:modifier_commande_superviseur', commande_id=commande.id)
                
        except Exception as e:
            print(f"❌ Erreur générale dans modifier_commande_superviseur: {e}")
            return JsonResponse({'success': False, 'message': f'Erreur: {str(e)}'})
    
    # ================ AFFICHAGE DE LA PAGE (GET) ================
    
    # Récupérer les articles de la commande
    paniers = commande.paniers.select_related('article').all()
    
    # Calculer le total des articles
    total_articles = sum(panier.sous_total for panier in paniers)
    
    # Récupérer les informations de livraison
    frais_livraison = commande.ville.frais_livraison if commande.ville else 0
    
    # Récupérer toutes les villes pour le formulaire
    villes = Ville.objects.all().order_by('nom')
    
    # Vérifier si c'est une commande renvoyée
    is_commande_renvoyee = False
    commande_originale_obj = None
    commande_renvoi_obj = None
    etat_articles_renvoyes = None
    articles_livres = []
    articles_renvoyes = []
    
    # Vérifier si cette commande est un renvoi
    if hasattr(commande, 'commande_renvoi') and commande.commande_renvoi:
        is_commande_renvoyee = True
        commande_originale_obj = commande.commande_renvoi
        commande_renvoi_obj = commande
        
        # Récupérer l'état des articles renvoyés
        try:
            etat_articles_renvoyes = Operation.objects.filter(
                commande=commande_originale_obj,
                enum_etat__libelle='Livrée partiellement'
            ).first()
            
            if etat_articles_renvoyes:
                # Récupérer les articles livrés et renvoyés
                articles_livres = etat_articles_renvoyes.articles_livres.all() if hasattr(etat_articles_renvoyes, 'articles_livres') else []
                articles_renvoyes = etat_articles_renvoyes.articles_renvoyes.all() if hasattr(etat_articles_renvoyes, 'articles_renvoyes') else []
        except Exception as e:
            print(f"⚠️ Erreur lors de la récupération des articles renvoyés: {e}")
    
    # Préparer le contexte
    context = {
        'commande': commande,
        'paniers': paniers,
        'total_articles': total_articles,
        'operateur': operateur,
        'is_commande_renvoyee': is_commande_renvoyee,
        'articles_livres': articles_livres,
        'articles_renvoyes': articles_renvoyes,
        # Variables de debug/informations supplémentaires
        'commande_originale': commande_originale_obj,
        'commande_renvoi': commande_renvoi_obj,
        'etat_articles_renvoyes': etat_articles_renvoyes,
        # Informations de livraison
        'frais_livraison': float(frais_livraison),
        'villes': villes,
        # 'articles_renvoyes_map': articles_renvoyes_map, # Retiré car plus nécessaire
    }
    return render(request, 'Superpreparation/modifier_commande_superviseur.html', context)

@superviseur_preparation_required
def diagnostiquer_compteur(request, commande_id):
    """Diagnostiquer et corriger le compteur upsell pour une commande (vue superviseur)."""
    from commande.models import Commande
    try:
        commande = get_object_or_404(Commande, id=commande_id)

        # Tous les paniers upsell
        articles_upsell_qs = commande.paniers.filter(article__isUpsell=True)
        compteur_actuel = commande.compteur or 0

        total_quantite_upsell = articles_upsell_qs.aggregate(total=Sum('quantite'))['total'] or 0

        # Nouvelle logique: 0-1 => 0, 2+ => total - 1
        compteur_correct = total_quantite_upsell - 1 if total_quantite_upsell >= 2 else 0

        if compteur_actuel != compteur_correct:
            commande.compteur = compteur_correct
            commande.save()
            # Recalcul des totaux
            commande.recalculer_totaux_upsell()

            return JsonResponse({
                'success': True,
                'message': f'Compteur corrigé de {compteur_actuel} vers {compteur_correct}',
                'ancien_compteur': compteur_actuel,
                'nouveau_compteur': compteur_correct,
                'total_commande': float(commande.total_cmd),
                'articles_upsell': articles_upsell_qs.count(),
                'quantite_totale_upsell': total_quantite_upsell,
            })

        # Rien à corriger, retourner un diagnostic détaillé
        return JsonResponse({
            'success': True,
            'message': 'Compteur déjà correct',
            'compteur': compteur_actuel,
            'articles_upsell': articles_upsell_qs.count(),
            'quantite_totale_upsell': total_quantite_upsell,
            'total_commande': float(commande.total_cmd),
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@superviseur_preparation_required
def api_articles_disponibles_prepa(request):
    """API pour récupérer les articles disponibles pour les opérateurs de préparation"""
    from article.models import Article
    
    try:
        # Accepter PREPARATION, SUPERVISEUR_PREPARATION et ADMIN
        operateur = Operateur.objects.get(user=request.user, actif=True)
        if operateur.type_operateur not in ['PREPARATION', 'SUPERVISEUR_PREPARATION', 'ADMIN']:
            return JsonResponse({'success': False, 'message': 'Accès non autorisé'}, status=403)
    except Operateur.DoesNotExist:
        # Pas de profil: continuer (le décorateur a déjà validé l'accès via groupes)
        operateur = None
    
    try:
        search_query = request.GET.get('search', '')
        filter_type = request.GET.get('filter', 'tous')
        
        # Récupérer les articles actifs
        articles = Article.objects.filter(actif=True)
        
        # Appliquer les filtres selon le type
        if filter_type == 'disponible':
            # Filtrer les articles qui ont au moins une variante avec stock > 0
            articles = articles.filter(variantes__qte_disponible__gt=0, variantes__actif=True).distinct()
        elif filter_type == 'upsell':
            articles = articles.filter(isUpsell=True)
        elif filter_type == 'liquidation':
            articles = articles.filter(phase='LIQUIDATION')
        elif filter_type == 'test':
            articles = articles.filter(phase='EN_TEST')
        
        # Recherche textuelle
        if search_query:
            articles = articles.filter(
                Q(nom__icontains=search_query) |
                Q(reference__icontains=search_query) |
                Q(couleur__icontains=search_query) |
                Q(pointure__icontains=search_query) |
                Q(description__icontains=search_query)
            )
        
        # Filtrer les articles en promotion si nécessaire (approche plus sûre)
        if filter_type == 'promo':
            # Filtrer les articles qui ont un prix actuel inférieur au prix unitaire
            articles = articles.filter(
                prix_actuel__isnull=False,
                prix_actuel__lt=F('prix_unitaire')
            )
        
        # Limiter les résultats
        articles = articles[:50]
        
        articles_data = []
        for article in articles:
            # Mettre à jour le prix actuel si nécessaire
            if article.prix_actuel is None:
                article.prix_actuel = article.prix_unitaire
                article.save(update_fields=['prix_actuel'])
            
            # Récupérer toutes les variantes actives (inclet celles en rupture)
            variantes_actives = article.variantes.filter(actif=True)
            
            # Si pas de variantes, créer une entrée avec les propriétés de compatibilité
            if not variantes_actives.exists():
                # Propriétés de compatibilité du modèle Article
                stock = article.qte_disponible
                couleur = article.couleur
                pointure = article.pointure
                
                if stock > 0:
                    articles_data.append({
                        'id': article.id,
                        'nom': article.nom,
                        'reference': article.reference or '',
                        'couleur': couleur or '',
                        'pointure': pointure or '',
                        'description': article.description or '',
                        'prix_unitaire': float(article.prix_unitaire),
                        'prix_actuel': float(article.prix_actuel or article.prix_unitaire),
                        'qte_disponible': stock,
                        'isUpsell': bool(article.isUpsell),
                        'phase': article.phase or 'NORMAL',
                        'has_promo_active': article.has_promo_active,
                        'image_url': article.image.url if article.image else article.image_url,
                        'categorie': str(article.categorie) if article.categorie else '',
                        'genre': str(article.genre) if article.genre else '',
                        'modele': article.modele_complet(),
                        'variantes_count': 0,
                        'is_variante': False,
                        'variante_id': None,
                        # Propriétés pour compatibilité avec le template
                        'prix': float(article.prix_actuel or article.prix_unitaire),
                        'prix_original': float(article.prix_unitaire),
                        'has_reduction': article.has_promo_active,
                        'reduction_pourcentage': round(((float(article.prix_unitaire) - float(article.prix_actuel or article.prix_unitaire)) / float(article.prix_unitaire)) * 100, 0) if article.has_promo_active else 0,
                        'article_type': 'normal',
                        'type_icon': 'fas fa-box',
                        'type_color': 'text-gray-600',
                        'display_text': f"{article.nom} - {couleur or ''} - {pointure or ''} ({float(article.prix_actuel or article.prix_unitaire):.2f} DH)"
                    })
            else:
                # Préparer la liste complète des variantes pour affichage (y compris rupture)
                variantes_list = []
                for v in variantes_actives:
                    variantes_list.append({
                        'id': v.id,
                        'couleur': v.couleur.nom if v.couleur else '',
                        'pointure': v.pointure.pointure if v.pointure else '',
                        'qte_disponible': v.qte_disponible,
                        'reference_variante': v.reference_variante,
                    })

                # Créer une entrée pour chaque variante (même celles en rupture)
                for variante in variantes_actives:
                    # Déterminer le type d'article pour l'affichage
                    article_type = 'normal'
                    type_icon = 'fas fa-box'
                    type_color = 'text-gray-600'
                    
                    if article.isUpsell:
                        article_type = 'upsell'
                        type_icon = 'fas fa-arrow-up'
                        type_color = 'text-purple-600'
                    elif article.phase == 'LIQUIDATION':
                        article_type = 'liquidation'
                        type_icon = 'fas fa-money-bill-wave'
                        type_color = 'text-red-600'
                    elif article.phase == 'EN_TEST':
                        article_type = 'test'
                        type_icon = 'fas fa-flask'
                        type_color = 'text-yellow-600'
                    
                    # Vérifier si l'article est en promotion
                    if article.has_promo_active:
                        article_type = 'promo'
                        type_icon = 'fas fa-fire'
                        type_color = 'text-orange-600'
                    
                    articles_data.append({
                        'id': article.id,
                        'nom': article.nom,
                        'reference': article.reference or '',
                        'couleur': variante.couleur.nom if variante.couleur else '',
                        'pointure': variante.pointure.pointure if variante.pointure else '',
                        'description': article.description or '',
                        'prix_unitaire': float(article.prix_unitaire),
                        'prix_actuel': float(article.prix_actuel or article.prix_unitaire),
                        'qte_disponible': variante.qte_disponible,
                        'isUpsell': bool(article.isUpsell),
                        'phase': article.phase or 'NORMAL',
                        'has_promo_active': article.has_promo_active,
                        'image_url': article.image.url if article.image else article.image_url,
                        'categorie': str(article.categorie) if article.categorie else '',
                        'genre': str(article.genre) if article.genre else '',
                        'modele': article.modele_complet(),
                        'variantes_count': variantes_actives.count(),
                        'is_variante': True,
                        'variante_id': variante.id,
                        'reference_variante': variante.reference_variante,
                        'variantes_all': variantes_list,
                        # Propriétés pour compatibilité avec le template
                        'prix': float(article.prix_actuel or article.prix_unitaire),
                        'prix_original': float(article.prix_unitaire),
                        'has_reduction': article.has_promo_active,
                        'reduction_pourcentage': round(((float(article.prix_unitaire) - float(article.prix_actuel or article.prix_unitaire)) / float(article.prix_unitaire)) * 100, 0) if article.has_promo_active else 0,
                        'article_type': article_type,
                        'type_icon': type_icon,
                        'type_color': type_color,
                        'display_text': f"{article.nom} - {variante.couleur.nom if variante.couleur else ''} - {variante.pointure.pointure if variante.pointure else ''} ({float(article.prix_actuel or article.prix_unitaire):.2f} DH)"
                    })
        
        # Calculer les statistiques pour la réponse
        stats = {
            "total": len(articles_data),
            "disponible": len([a for a in articles_data if a.get('qte_disponible', 0) > 0]),
            "upsell": len([a for a in articles_data if a.get('isUpsell', False)]),
            "liquidation": len([a for a in articles_data if a.get('phase') == 'LIQUIDATION']),
            "test": len([a for a in articles_data if a.get('phase') == 'EN_TEST']),
            "promo": len([a for a in articles_data if a.get('has_promo_active', False)])
        }
        
        # Retourner le format attendu par le frontend
        return JsonResponse({
            'success': True,
            'articles': articles_data,
            'stats': stats
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()  # This will print to the server console
        return JsonResponse({
            'success': False,
            'message': 'Erreur interne du serveur lors du chargement des articles.',
            'details': str(e)
        }, status=500)

@superviseur_preparation_required
def api_panier_commande_prepa(request, commande_id):
    """API pour récupérer le panier d'une commande pour les opérateurs de préparation"""
    try:
        # Accepter PREPARATION et SUPERVISEUR_PREPARATION
        operateur = Operateur.objects.get(user=request.user, actif=True)
        if operateur.type_operateur not in ['PREPARATION', 'SUPERVISEUR_PREPARATION']:
            return JsonResponse({'success': False, 'message': 'Accès non autorisé'})
    except Operateur.DoesNotExist:
        # Pas de profil: continuer (le décorateur a déjà validé l'accès via groupes)
        operateur = None
    
    # Récupérer la commande
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'})
    
    # Pour les superviseurs, on ne vérifie pas l'affectation spécifique
    # Ils peuvent accéder à toutes les commandes en préparation
    if operateur and operateur.type_operateur == 'SUPERVISEUR_PREPARATION':
        # Vérifier seulement que la commande est en préparation (incluant les états finaux)
        etat_preparation = commande.etats.filter(
            Q(enum_etat__libelle='À imprimer') | Q(enum_etat__libelle='En préparation') | Q(enum_etat__libelle='Collectée') | Q(enum_etat__libelle='Emballée') | Q(enum_etat__libelle='Préparée'),
            date_fin__isnull=True
        ).first()
    else:
        # Pour les opérateurs normaux, vérifier l'affectation spécifique
        etat_preparation = commande.etats.filter(
            Q(enum_etat__libelle='À imprimer') | Q(enum_etat__libelle='En préparation') | Q(enum_etat__libelle='Collectée') | Q(enum_etat__libelle='Emballée') | Q(enum_etat__libelle='Préparée'),
            operateur=operateur,
            date_fin__isnull=True
        ).first()
    
    if not etat_preparation:
        return JsonResponse({'success': False, 'message': 'Commande non affectée'})
    
    # Récupérer les paniers
    paniers = commande.paniers.all().select_related('article')

    # Construire le format "articles" attendu par le front (compatibilité)
    articles = []
    total_articles_montant = 0.0
    for panier in paniers:
        prix_unitaire = float(panier.article.prix_unitaire) if panier.article and panier.article.prix_unitaire is not None else 0.0
        sous_total = float(panier.sous_total) if panier.sous_total is not None else round(prix_unitaire * (panier.quantite or 0), 2)
        total_articles_montant += sous_total
        articles.append({
            'id': panier.id,
            'article_id': getattr(panier.article, 'id', None),
            'nom': getattr(panier.article, 'nom', ''),
            'reference': getattr(panier.article, 'reference', '') or '',
            'couleur': getattr(panier.article, 'couleur', ''),
            'pointure': getattr(panier.article, 'pointure', ''),
            'quantite': panier.quantite or 0,
            'prix_unitaire': prix_unitaire,
            'sous_total': sous_total,
        })

    # Objet commande attendu par le front
    client_nom = ''
    try:
        if commande.client:
            prenom = getattr(commande.client, 'prenom', '') or ''
            nom = getattr(commande.client, 'nom', '') or ''
            client_nom = (prenom + ' ' + nom).strip()
    except Exception:
        client_nom = ''

    commande_payload = {
        'id': commande.id,
        'id_yz': getattr(commande, 'id_yz', ''),
        'client_nom': client_nom,
        'total_articles': len(articles),
        'total_montant': round(total_articles_montant, 2),
        'total_final': float(commande.total_cmd) if getattr(commande, 'total_cmd', None) is not None else round(total_articles_montant, 2),
    }

    # Réponse incluant les deux formats (nouveau et legacy) pour compatibilité
    return JsonResponse({
        'success': True,
        # Nouveau format utilisé par le front de Suivi Général
        'commande': commande_payload,
        'articles': articles,
        # Legacy (au cas où)
        'paniers': articles,
        'total_commande': commande_payload['total_final'],
        'nb_articles': len(articles),
    })

@superviseur_preparation_required
def api_finaliser_commande(request, commande_id):
    """API pour finaliser une commande depuis la liste (pour les superviseurs)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
    
    try:
        operateur_profile = request.user.profil_operateur
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            return JsonResponse({'success': False, 'message': 'Accès non autorisé'}, status=403)
    except Operateur.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Profil opérateur non trouvé'}, status=403)
    
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'}, status=404)
    
    # Vérifier que la commande est emballée (prête pour finalisation)
    etat_actuel = commande.etat_actuel
    if not etat_actuel or etat_actuel.enum_etat.libelle != 'Emballée':
        return JsonResponse({
            'success': False, 
            'message': f'La commande n\'est pas emballée (état actuel: {etat_actuel.enum_etat.libelle if etat_actuel else "Aucun"})'
        }, status=400)
    
    try:
        with transaction.atomic():
            # Marquer l'état 'Emballée' comme terminé
            etat_actuel.date_fin = timezone.now()
            etat_actuel.operateur = operateur_profile
            etat_actuel.save()
            
            # Créer le nouvel état 'Préparée' (final)
            etat_preparee, created = EnumEtatCmd.objects.get_or_create(libelle='Préparée')
            EtatCommande.objects.create(
                commande=commande,
                enum_etat=etat_preparee,
                operateur=operateur_profile
            )
            
            # Log de l'opération
            Operation.objects.create(
                commande=commande,
                type_operation='PREPARATION_TERMINEE',
                operateur=operateur_profile,
                conclusion=f"Commande préparée par {operateur_profile.nom_complet} (via API)."
            )
        
        return JsonResponse({
            'success': True,
            'message': f'La commande {commande.id_yz} a été préparée avec succès.'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la finalisation: {str(e)}'
        }, status=500)

@superviseur_preparation_required
def api_panier_commande(request, commande_id):
    """API pour récupérer le contenu du panier d'une commande (pour le modal)"""
    if request.method != 'GET':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)
    
    try:
        operateur_profile = request.user.profil_operateur
        if not (operateur_profile.is_preparation or operateur_profile.is_superviseur_preparation):
            return JsonResponse({'success': False, 'message': 'Accès non autorisé'}, status=403)
    except Operateur.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Profil opérateur non trouvé'}, status=403)
    
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'}, status=404)
    
    try:
        # Récupérer les articles du panier
        paniers = commande.paniers.all().select_related('article')
        
        if not paniers.exists():
            html_content = '''
                <div class="text-center py-8">
                    <i class="fas fa-shopping-cart text-4xl text-gray-300 mb-4"></i>
                    <p class="text-gray-500">Aucun article dans le panier</p>
                </div>
            '''
            return JsonResponse({
                'success': True,
                'html': html_content
            })
        
        # Construire le HTML du panier
        html_content = '''
            <div class="space-y-4">
                <div class="bg-gray-50 p-4 rounded-lg">
                    <h4 class="font-medium text-gray-900 mb-2">Commande #''' + str(commande.id_yz) + '''</h4>
                    <p class="text-sm text-gray-600">Client: ''' + (commande.client.prenom + ' ' + commande.client.nom if commande.client else 'Non défini') + '''</p>
                </div>
                <div class="space-y-3">
        '''
        
        total_articles = 0
        total_montant = 0
        
        for panier in paniers:
            article = panier.article
            quantite = panier.quantite or 0
            prix = getattr(article, 'prix', 0) or 0
            sous_total = quantite * prix
            
            total_articles += quantite
            total_montant += sous_total
            
            # Informations de l'article
            nom_article = getattr(article, 'nom', 'Article sans nom') or 'Article sans nom'
            reference = getattr(article, 'reference', '') or ''
            couleur = getattr(article, 'couleur', '') or ''
            pointure = getattr(article, 'pointure', '') or ''
            
            html_content += '''
                <div class="border border-gray-200 rounded-lg p-4">
                    <div class="flex justify-between items-start mb-2">
                        <div class="flex-1">
                            <h5 class="font-medium text-gray-900">''' + nom_article + '''</h5>
                            <p class="text-sm text-gray-600">Réf: ''' + reference + '''</p>
                        </div>
                        <div class="text-right">
                            <span class="text-lg font-bold text-green-600">''' + str(prix) + ''' DH</span>
                        </div>
                    </div>
                    <div class="grid grid-cols-2 gap-4 text-sm">
                        <div>
                            <span class="text-gray-500">Quantité:</span>
                            <span class="font-medium text-gray-900 ml-1">''' + str(quantite) + '''</span>
                        </div>
                        <div>
                            <span class="text-gray-500">Sous-total:</span>
                            <span class="font-medium text-green-600 ml-1">''' + str(sous_total) + ''' DH</span>
                        </div>
            '''
            
            # Ajouter couleur et pointure si disponibles
            if couleur or pointure:
                html_content += '''
                        <div class="col-span-2 flex space-x-4">
                '''
                if couleur:
                    html_content += '''
                            <div>
                                <span class="text-gray-500">Couleur:</span>
                                <span class="font-medium text-gray-900 ml-1">''' + couleur + '''</span>
                            </div>
                    '''
                if pointure:
                    html_content += '''
                            <div>
                                <span class="text-gray-500">Pointure:</span>
                                <span class="font-medium text-gray-900 ml-1">''' + pointure + '''</span>
                            </div>
                    '''
                html_content += '''
                        </div>
                '''
            
            html_content += '''
                    </div>
                </div>
            '''
        
        # Ajouter le total
        html_content += '''
                </div>
                <div class="border-t border-gray-200 pt-4">
                    <div class="flex justify-between items-center">
                        <div class="text-lg font-medium text-gray-900">
                            Total (''' + str(total_articles) + ''' article''' + ('s' if total_articles > 1 else '') + '''):
                        </div>
                        <div class="text-2xl font-bold text-green-600">
                            ''' + str(round(total_montant, 2)) + ''' DH
                        </div>
                    </div>
                </div>
            </div>
        '''
        
        return JsonResponse({
            'success': True,
            'html': html_content
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du chargement du panier: {str(e)}'
        }, status=500)

@superviseur_preparation_required
def imprimer_tickets_preparation(request):
    """
    Vue pour imprimer les tickets de préparation SANS changer l'état des commandes.
    Permet d'imprimer ou de réimprimer des tickets pour les commandes en préparation.
    """
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            return HttpResponse("Accès non autorisé.", status=403)
    except Operateur.DoesNotExist:
        return HttpResponse("Profil opérateur non trouvé.", status=403)

    commande_ids_str = request.GET.get('ids')
    if not commande_ids_str:
        return HttpResponse("Aucun ID de commande fourni.", status=400)

    try:
        commande_ids = [int(id) for id in commande_ids_str.split(',') if id.isdigit()]
    except ValueError:
        return HttpResponse("IDs de commande invalides.", status=400)
    
    # Récupérer les commandes en préparation affectées à cet opérateur
    commandes = Commande.objects.filter(
        id__in=commande_ids,
        etats__operateur=operateur_profile,
        etats__enum_etat__libelle='En préparation',
        etats__date_fin__isnull=True
    ).distinct()

    if not commandes.exists():
        messages.info(request, "L'impression des tickets est désactivée. Utilisez les outils de gestion.")
        return redirect('Superpreparation:liste_prepa')

    # Génération du code-barres pour chaque commande (sans transition d'état)
    code128 = barcode.get_barcode_class('code128')
    
    messages.info(request, "L'impression des tickets a été retirée de l'interface superviseur.")
    return redirect('Superpreparation:liste_prepa')

# === NOUVELLES FONCTIONNALITÉS : GESTION DE STOCK ===

@superviseur_preparation_required
def ajuster_stock(request, article_id):
    """Ajuster le stock d'un article - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    article = get_object_or_404(Article, pk=article_id)
    
    if request.method == 'POST':
        form = AjusterStockForm(request.POST, article=article)
        if form.is_valid():
            type_mouvement = form.cleaned_data['type_mouvement']
            quantite = form.cleaned_data['quantite']
            commentaire = form.cleaned_data['commentaire']
            variante = form.cleaned_data.get('variante')
            
            try:
                print(f"🔧 Ajustement stock - Article: {article.nom}, Type: {type_mouvement}, Quantité: {quantite}")
                if variante:
                    print(f"🔧 Variante sélectionnée: {variante.couleur.nom} - {variante.pointure.pointure} (Stock: {variante.qte_disponible})")
                print(f"🔧 Stock total avant ajustement: {article.qte_disponible}")
                
                mouvement = creer_mouvement_stock(
                    article=article,
                    quantite=quantite,
                    type_mouvement=type_mouvement,
                    operateur=operateur_profile,
                    commentaire=commentaire,
                    variante=variante
                )
                
                # Recharger l'article pour voir le stock mis à jour
                article.refresh_from_db()
                print(f"✅ Stock total après ajustement: {article.qte_disponible}")
                
                if mouvement:
                    if variante:
                        variante.refresh_from_db()
                        messages.success(request, 
                            f"Le stock de la variante '{variante.couleur.nom} - {variante.pointure.pointure}' "
                            f"a été ajusté avec succès. Nouveau stock: {variante.qte_disponible}")
                    else:
                        messages.success(request, 
                            f"Le stock de l'article '{article.nom}' a été ajusté avec succès. "
                            f"Nouveau stock total: {article.qte_disponible}")
                else:
                    messages.warning(request, "L'ajustement n'a pas pu être effectué.")
                    
                return redirect('Superpreparation:detail_article', article_id=article.id)
            except Exception as e:
                print(f"❌ Erreur dans ajuster_stock: {str(e)}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Une erreur est survenue lors de l'ajustement du stock : {e}")

    else:
            form = AjusterStockForm(article=article)
            
            # Pré-sélectionner la variante si elle est passée en paramètre
            variante_id = request.GET.get('variante')
            if variante_id:
                try:
                    variante = VarianteArticle.objects.get(id=variante_id, article=article)
                    form.fields['variante'].initial = variante
                except VarianteArticle.DoesNotExist:
                    pass

    mouvements_recents = article.mouvements.order_by('-date_mouvement')[:10]

    context = {
        'form': form,
        'article': article,
        'mouvements_recents': mouvements_recents,
        'page_title': f"Ajuster le Stock - {article.nom}",
    }
    return render(request, 'Superpreparation/stock/ajuster_stock.html', context)

@superviseur_preparation_required
def detail_article(request, article_id):
    """Afficher les détails d'un article spécifique - Service de préparation"""
    article = get_object_or_404(Article, pk=article_id)
    
    # Calculer la valeur totale du stock
    valeur_stock = article.prix_actuel * article.qte_disponible if article.prix_actuel else 0
    
    # Récupérer le dernier mouvement de stock pour cet article
    dernier_mouvement = article.mouvements.order_by('-date_mouvement').first()

    context = {
        'article': article,
        'valeur_stock': valeur_stock,
        'dernier_mouvement': dernier_mouvement,
        'page_title': f"Détail de l'article : {article.nom}",
        'page_subtitle': "Informations complètes sur l'article",
    }
    return render(request, 'Superpreparation/stock/detail_article.html', context)

@superviseur_preparation_required
def liste_articles(request):
    """Afficher la liste des articles avec filtres et statistiques - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    # Imports locaux pour les annotations
    from django.db.models import Q, F, Sum, Count, Avg
    from django.db.models.functions import Coalesce

    # Calcul des statistiques globales (avant tout filtrage)
    articles_qs = Article.objects.all().annotate(
        total_qte_disponible=Coalesce(Sum('variantes__qte_disponible'), 0)
    )
    articles_total = articles_qs.count()
    articles_actifs = articles_qs.filter(actif=True).count()
    articles_inactifs = articles_qs.filter(actif=False).count()
    articles_rupture = articles_qs.filter(total_qte_disponible__lte=0).count()
    
    # Articles créés aujourd'hui
    today = timezone.now().date()
    articles_crees_aujourd_hui = articles_qs.filter(date_creation__date=today).count()

    # Récupération des articles pour la liste, filtrée
    articles_list = Article.objects.all().annotate(
        total_qte_disponible=Coalesce(Sum('variantes__qte_disponible'), 0)
    )
    
    # Filtres de recherche améliorés
    query = request.GET.get('q', '').strip()
    categorie_filter = request.GET.get('categorie', '').strip()
    statut_filter = request.GET.get('statut', '').strip()
    stock_filter = request.GET.get('stock', '').strip()
    prix_min = request.GET.get('prix_min', '').strip()
    prix_max = request.GET.get('prix_max', '').strip()
    couleur_filter = request.GET.get('couleur', '').strip()
    phase_filter = request.GET.get('phase', '').strip()
    tri = request.GET.get('tri', 'date_creation').strip()
    
    # Recherche textuelle intelligente
    if query:
        articles_list = articles_list.filter(
            Q(nom__icontains=query) |
            Q(reference__icontains=query) |
            Q(description__icontains=query) |
            Q(categorie__nom__icontains=query) |
            Q(variantes__couleur__nom__icontains=query) |
            Q(variantes__pointure__pointure__icontains=query)
        ).distinct()
    
    # Filtre par catégorie (ID ou nom)
    if categorie_filter and categorie_filter.strip():
        if categorie_filter.isdigit():
            articles_list = articles_list.filter(categorie_id=int(categorie_filter))
        else:
            articles_list = articles_list.filter(categorie__nom__icontains=categorie_filter)
    
    # Filtre par statut
    if statut_filter:
        if statut_filter == 'actif':
            articles_list = articles_list.filter(actif=True)
        elif statut_filter == 'inactif':
            articles_list = articles_list.filter(actif=False)
    
    # Filtre par niveau de stock
    if stock_filter:
        if stock_filter == 'rupture':
            articles_list = articles_list.filter(total_qte_disponible__lte=0)
        elif stock_filter == 'faible':
            articles_list = articles_list.filter(total_qte_disponible__gt=0, total_qte_disponible__lte=10)
        elif stock_filter == 'normal':
            articles_list = articles_list.filter(total_qte_disponible__gt=10, total_qte_disponible__lte=50)
        elif stock_filter == 'eleve':
            articles_list = articles_list.filter(total_qte_disponible__gt=50)
    
    # Filtre par prix
    if prix_min:
        try:
            prix_min_val = float(prix_min.replace(',', '.'))
            articles_list = articles_list.filter(prix_unitaire__gte=prix_min_val)
        except (ValueError, TypeError):
            pass
    
    if prix_max:
        try:
            prix_max_val = float(prix_max.replace(',', '.'))
            articles_list = articles_list.filter(prix_unitaire__lte=prix_max_val)
        except (ValueError, TypeError):
            pass
    
    # Filtre par couleur (via variantes)
    if couleur_filter:
        articles_list = articles_list.filter(variantes__couleur__nom__icontains=couleur_filter).distinct()
    
    # Filtre par phase
    if phase_filter:
        articles_list = articles_list.filter(phase=phase_filter)
    
    # Tri des résultats
    if tri == 'nom':
        articles_list = articles_list.order_by('nom')
    elif tri == 'prix_asc':
        articles_list = articles_list.order_by('prix_unitaire')
    elif tri == 'prix_desc':
        articles_list = articles_list.order_by('-prix_unitaire')
    elif tri == 'stock_asc':
        articles_list = articles_list.order_by('total_qte_disponible')
    elif tri == 'stock_desc':
        articles_list = articles_list.order_by('-total_qte_disponible')
    elif tri == 'date_creation':
        articles_list = articles_list.order_by('-date_creation')
    elif tri == 'reference':
        articles_list = articles_list.order_by('reference')
    else:
        articles_list = articles_list.order_by('-date_creation')
    
    # Récupération des valeurs uniques pour les filtres (éviter comparaisons FK avec '')
    from article.models import Categorie, Couleur
    categories_uniques = list(Categorie.objects.values_list('nom', flat=True).order_by('nom'))
    couleurs_uniques = list(Couleur.objects.values_list('nom', flat=True).order_by('nom'))
    phases_uniques = list(Article.objects.values_list('phase', flat=True).distinct())

    # Pagination
    paginator = Paginator(articles_list, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'articles': page_obj,
        'categories_uniques': categories_uniques,
        'couleurs_uniques': couleurs_uniques,
        'phases_uniques': phases_uniques,
        'articles_total': articles_total,
        'articles_actifs': articles_actifs,
        'articles_inactifs': articles_inactifs,
        'articles_rupture': articles_rupture,
        'articles_crees_aujourd_hui': articles_crees_aujourd_hui,
        'page_title': "Liste des Articles",
        'page_subtitle': "Inventaire complet et gestion du stock",
        'request': request,
        'query': query,
        'current_filters': {
            'categorie': categorie_filter,
            'statut': statut_filter,
            'stock': stock_filter,
            'prix_min': prix_min,
            'prix_max': prix_max,
            'couleur': couleur_filter,
            'phase': phase_filter,
            'tri': tri,
        }
    }
    return render(request, 'Superpreparation/stock/liste_articles.html', context)

@superviseur_preparation_required
def mouvements_stock(request):
    """Vue pour afficher l'historique des mouvements de stock - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    from article.models import MouvementStock
    
    # Récupération de tous les mouvements
    mouvements_list = MouvementStock.objects.select_related('article', 'operateur').order_by('-date_mouvement')
    
    # Filtres de recherche
    article_filter = request.GET.get('article', '').strip()
    type_filter = request.GET.get('type', '').strip()
    date_filter = request.GET.get('date_range', '').strip()
    
    # Filtre par article (nom ou référence)
    if article_filter:
        mouvements_list = mouvements_list.filter(
            Q(article__nom__icontains=article_filter) |
            Q(article__reference__icontains=article_filter)
        )
    
    # Filtre par type de mouvement
    if type_filter:
        if type_filter == 'entree':
            mouvements_list = mouvements_list.filter(type_mouvement='entree')
        elif type_filter == 'sortie':
            mouvements_list = mouvements_list.filter(type_mouvement='sortie')
        elif type_filter == 'ajustement':
            mouvements_list = mouvements_list.filter(
                type_mouvement__in=['ajustement_pos', 'ajustement_neg']
            )
    
    # Filtre par date
    if date_filter:
        try:
            date_obj = datetime.strptime(date_filter, '%Y-%m-%d').date()
            mouvements_list = mouvements_list.filter(date_mouvement__date=date_obj)
        except ValueError:
            pass
    
    # Pagination
    paginator = Paginator(mouvements_list, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques rapides
    total_mouvements = mouvements_list.count()
    mouvements_aujourd_hui = MouvementStock.objects.filter(
        date_mouvement__date=timezone.now().date()
    ).count()
    
    context = {
        'mouvements': page_obj,
        'total_mouvements': total_mouvements,
        'mouvements_aujourd_hui': mouvements_aujourd_hui,
        'page_title': 'Mouvements de Stock',
        'current_filters': {
            'article': article_filter,
            'type': type_filter,
            'date_range': date_filter,
        }
    }
    return render(request, 'Superpreparation/stock/mouvements_stock.html', context)

@superviseur_preparation_required
def alertes_stock(request):
    """Vue pour afficher les alertes de stock - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    from article.models import MouvementStock
    
    # Paramètres de seuils
    SEUIL_RUPTURE = 0
    SEUIL_STOCK_FAIBLE = 10
    SEUIL_A_COMMANDER = 20
    
    # Récupération de tous les articles actifs avec annotation du stock total
    from django.db.models import Q, Sum
    from django.db.models.functions import Coalesce
    
    articles_actifs = Article.objects.filter(actif=True).annotate(
        total_qte_disponible=Coalesce(Sum('variantes__qte_disponible'), 0)
    )
    
    # Filtres par niveau d'alerte
    filtre_alerte = request.GET.get('filtre', 'tous')
    
    if filtre_alerte == 'rupture':
        articles_alerte = articles_actifs.filter(total_qte_disponible__lte=SEUIL_RUPTURE)
    elif filtre_alerte == 'faible':
        articles_alerte = articles_actifs.filter(
            total_qte_disponible__gt=SEUIL_RUPTURE,
            total_qte_disponible__lte=SEUIL_STOCK_FAIBLE
        )
    elif filtre_alerte == 'a_commander':
        articles_alerte = articles_actifs.filter(
            total_qte_disponible__gt=SEUIL_STOCK_FAIBLE,
            total_qte_disponible__lte=SEUIL_A_COMMANDER
        )
    else:
        articles_alerte = articles_actifs.filter(total_qte_disponible__lte=SEUIL_A_COMMANDER)
    
    # Tri des résultats
    tri = request.GET.get('tri', 'stock_asc')
    if tri == 'stock_asc':
        articles_alerte = articles_alerte.order_by('total_qte_disponible')
    elif tri == 'stock_desc':
        articles_alerte = articles_alerte.order_by('-total_qte_disponible')
    elif tri == 'nom':
        articles_alerte = articles_alerte.order_by('nom')
    elif tri == 'reference':
        articles_alerte = articles_alerte.order_by('reference')
    elif tri == 'categorie':
        articles_alerte = articles_alerte.order_by('categorie')
    else:
        articles_alerte = articles_alerte.order_by('total_qte_disponible')
    
    # Statistiques détaillées
    stats = {
        'total_articles': articles_actifs.count(),
        'rupture_stock': articles_actifs.filter(total_qte_disponible__lte=SEUIL_RUPTURE).count(),
        'stock_faible': articles_actifs.filter(
            total_qte_disponible__gt=SEUIL_RUPTURE,
            total_qte_disponible__lte=SEUIL_STOCK_FAIBLE
        ).count(),
        'a_commander': articles_actifs.filter(
            total_qte_disponible__gt=SEUIL_STOCK_FAIBLE,
            total_qte_disponible__lte=SEUIL_A_COMMANDER
        ).count(),
        'stock_ok': articles_actifs.filter(total_qte_disponible__gt=SEUIL_A_COMMANDER).count(),
    }
    
    # Alertes critiques
    alertes_critiques = articles_actifs.filter(total_qte_disponible__lte=SEUIL_RUPTURE).order_by('total_qte_disponible')[:5]
    
    # Analyse par catégorie
    categories_alertes = articles_actifs.values('categorie').annotate(
        total=Count('id'),
        valeur_stock=Sum('variantes__qte_disponible')
    ).exclude(categorie__isnull=True)
    
    # Calculer les statistiques par catégorie en Python
    for cat in categories_alertes:
        articles_cat = articles_actifs.filter(categorie=cat['categorie'])
        cat['rupture'] = sum(1 for article in articles_cat if (article.total_qte_disponible or 0) <= SEUIL_RUPTURE)
        cat['faible'] = sum(1 for article in articles_cat if SEUIL_RUPTURE < (article.total_qte_disponible or 0) <= SEUIL_STOCK_FAIBLE)
        cat['a_commander'] = sum(1 for article in articles_cat if SEUIL_STOCK_FAIBLE < (article.total_qte_disponible or 0) <= SEUIL_A_COMMANDER)
    
    # Trier par rupture puis faible
    categories_alertes = sorted(categories_alertes, key=lambda x: (x['rupture'], x['faible']), reverse=True)
    
    # Historique des mouvements récents
    mouvements_recents = MouvementStock.objects.filter(
        article__in=articles_alerte,
        date_mouvement__gte=timezone.now() - timedelta(days=30)
    ).select_related('article', 'operateur').order_by('-date_mouvement')[:10]
    
    # Suggestions d'actions
    suggestions = []
    
    if stats['rupture_stock'] > 0:
        suggestions.append({
            'type': 'danger',
            'titre': 'Rupture de Stock Critique',
            'message': f'{stats["rupture_stock"]} article(s) en rupture totale nécessitent un réapprovisionnement immédiat.',
            'action': 'Contacter les fournisseurs',
            'icone': 'fas fa-exclamation-triangle'
        })
    
    if stats['stock_faible'] > 0:
        suggestions.append({
            'type': 'warning',
            'titre': 'Stock Faible',
            'message': f'{stats["stock_faible"]} article(s) ont un stock faible. Planifier les commandes.',
            'action': 'Préparer les commandes',
            'icone': 'fas fa-exclamation-circle'
        })
    
    if stats['a_commander'] > 0:
        suggestions.append({
            'type': 'info',
            'titre': 'À Commander Bientôt',
            'message': f'{stats["a_commander"]} article(s) devront être commandés prochainement.',
            'action': 'Surveiller l\'évolution',
            'icone': 'fas fa-info-circle'
        })
    
    # Pagination
    paginator = Paginator(articles_alerte, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'articles': page_obj,
        'stats': stats,
        'alertes_critiques': alertes_critiques,
        'categories_alertes': categories_alertes,
        'mouvements_recents': mouvements_recents,
        'suggestions': suggestions,
        'filtre_actuel': filtre_alerte,
        'tri_actuel': tri,
        'seuils': {
            'rupture': SEUIL_RUPTURE,
            'faible': SEUIL_STOCK_FAIBLE,
            'a_commander': SEUIL_A_COMMANDER
        },
        'page_title': 'Alertes Stock',
        'page_subtitle': 'Articles nécessitant une attention immédiate'
    }
    return render(request, 'Superpreparation/stock/alertes_stock.html', context)

@superviseur_preparation_required
def statistiques_stock(request):
    """Vue pour afficher les statistiques de stock - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    from article.models import MouvementStock
    from django.db.models import Q, F, Sum, Count, Avg
    from django.db.models.functions import Coalesce
    
    # Paramètres de filtrage
    periode = int(request.GET.get('periode', 30))
    categorie_filter = request.GET.get('categorie', '')
    
    # Date de début selon la période
    date_debut = timezone.now() - timedelta(days=periode)
    
    # Articles de base
    articles_qs = Article.objects.filter(actif=True)
    
    # Filtrage par catégorie si spécifié
    if categorie_filter:
        articles_qs = articles_qs.filter(categorie=categorie_filter)
    
    # Articles avec annotation du stock total
    articles_qs = articles_qs.annotate(
        total_qte_disponible=Coalesce(Sum('variantes__qte_disponible'), 0)
    )
    
    # Valeur totale du stock
    valeur_stock = sum(
        (article.prix_unitaire or 0) * (article.total_qte_disponible or 0)
        for article in articles_qs
    )
    
    # Nombre total d'articles en stock
    articles_en_stock = articles_qs.filter(total_qte_disponible__gt=0).count()
    
    # Articles par niveau de stock
    stats_niveaux = articles_qs.aggregate(
        total_articles=Count('id')
    )
    
    # Calculer les statistiques en Python
    rupture = sum(1 for article in articles_qs if (article.total_qte_disponible or 0) == 0)
    stock_faible = sum(1 for article in articles_qs if 0 < (article.total_qte_disponible or 0) <= 10)
    stock_normal = sum(1 for article in articles_qs if 10 < (article.total_qte_disponible or 0) <= 50)
    stock_eleve = sum(1 for article in articles_qs if (article.total_qte_disponible or 0) > 50)
    
    stats_niveaux.update({
        'rupture': rupture,
        'stock_faible': stock_faible,
        'stock_normal': stock_normal,
        'stock_eleve': stock_eleve
    })
    
    # Taux de rupture
    taux_rupture = (stats_niveaux['rupture'] / stats_niveaux['total_articles'] * 100) if stats_niveaux['total_articles'] > 0 else 0
    
    # Statistiques par catégorie
    stats_categories = articles_qs.values('categorie').annotate(
        total_articles=Count('id'),
        stock_total=Coalesce(Sum('variantes__qte_disponible'), 0),
        prix_moyen=Avg('prix_unitaire')
    ).exclude(categorie__isnull=True)
    
    # Calculer la valeur totale, le stock moyen et les statistiques en Python
    for cat in stats_categories:
        articles_cat = articles_qs.filter(categorie=cat['categorie'])
        cat['valeur_totale'] = sum(
            (article.prix_unitaire or 0) * (article.total_qte_disponible or 0)
            for article in articles_cat
        )
        cat['stock_moyen'] = sum(article.total_qte_disponible or 0 for article in articles_cat) / cat['total_articles'] if cat['total_articles'] > 0 else 0
        cat['articles_rupture'] = sum(1 for article in articles_cat if (article.total_qte_disponible or 0) == 0)
        cat['articles_faible'] = sum(1 for article in articles_cat if 0 < (article.total_qte_disponible or 0) <= 10)
    
    # Trier par valeur totale
    stats_categories = sorted(stats_categories, key=lambda x: x['valeur_totale'], reverse=True)
    
    # Top articles
    top_articles_valeur = articles_qs.annotate(
        valeur_stock=F('total_qte_disponible') * F('prix_unitaire')
    ).filter(total_qte_disponible__gt=0).order_by('-valeur_stock')[:10]
    
    top_articles_quantite = articles_qs.filter(total_qte_disponible__gt=0).order_by('-total_qte_disponible')[:10]
    
    # Mouvements de stock
    mouvements_periode = MouvementStock.objects.filter(
        date_mouvement__gte=date_debut,
        article__in=articles_qs
    ).select_related('article')
    
    mouvements_sortie = mouvements_periode.filter(
        type_mouvement__in=['sortie', 'ajustement_neg']
    ).aggregate(total_sorties=Sum('quantite'))['total_sorties'] or 0
    
    rotation_stock = (mouvements_sortie / valeur_stock * 100) if valeur_stock > 0 else 0
    
    # Évolution temporelle
    evolution_donnees = []
    nb_semaines = min(periode // 7, 12)
    
    for i in range(nb_semaines):
        date_fin = timezone.now() - timedelta(days=i*7)
        valeur_semaine = sum(
            (article.prix_unitaire or 0) * (article.total_qte_disponible or 0)
            for article in articles_qs
        )
        
        evolution_donnees.append({
            'date': date_fin.strftime('%d/%m'),
            'valeur': float(valeur_semaine)
        })
    
    evolution_donnees.reverse()
    
    # Alertes
    alertes = []
    
    if stats_niveaux['rupture'] > 0:
        alertes.append({
            'type': 'danger',
            'titre': 'Articles en Rupture',
            'message': f'{stats_niveaux["rupture"]} article(s) en rupture de stock',
            'valeur': stats_niveaux['rupture']
        })
    
    if taux_rupture > 10:
        alertes.append({
            'type': 'warning',
            'titre': 'Taux de Rupture Élevé',
            'message': f'Taux de rupture de {taux_rupture:.1f}% (seuil recommandé: 5%)',
            'valeur': f'{taux_rupture:.1f}%'
        })
    
    if rotation_stock < 2:
        alertes.append({
            'type': 'info',
            'titre': 'Rotation Faible',
            'message': 'La rotation du stock est faible, optimisation possible',
            'valeur': f'{rotation_stock:.1f}'
        })
    
    # Données pour graphiques
    categories_chart_data = {
        'labels': [cat['categorie'] for cat in stats_categories],
        'values': [float(cat['valeur_totale'] or 0) for cat in stats_categories],
        'colors': ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF', '#FF9F40']
    }
    
    top_articles_chart_data = {
        'labels': [art.nom[:20] for art in top_articles_valeur[:5]],
        'values': [float((art.qte_disponible or 0) * (art.prix_unitaire or 0)) for art in top_articles_valeur[:5]]
    }
    
    categories_disponibles = Article.objects.filter(actif=True).values_list('categorie', flat=True).distinct().exclude(categorie__isnull=True).order_by('categorie')
    
    context = {
        'page_title': 'Statistiques Stock',
        'page_subtitle': 'Analyse de la performance et de la valeur de l\'inventaire',
        'valeur_stock': valeur_stock,
        'articles_en_stock': articles_en_stock,
        'rotation_stock': rotation_stock,
        'taux_rupture': taux_rupture,
        'stats_niveaux': stats_niveaux,
        'stats_categories': stats_categories,
        'top_articles_valeur': top_articles_valeur,
        'top_articles_quantite': top_articles_quantite,
        'evolution_donnees': evolution_donnees,
        'alertes': alertes,
        'categories_chart_data': categories_chart_data,
        'top_articles_chart_data': top_articles_chart_data,
        'categories_disponibles': categories_disponibles,
        'periode_actuelle': periode,
        'categorie_actuelle': categorie_filter,
    }
    return render(request, 'Superpreparation/stock/statistiques_stock.html', context)

@superviseur_preparation_required
def creer_article(request):
    """Créer un nouvel article - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    # Récupérer les données nécessaires pour le formulaire
    from article.models import Categorie, Genre, Couleur, Pointure
    
    categories = Categorie.objects.filter(actif=True).order_by('nom')
    genres = Genre.objects.filter(actif=True).order_by('nom')
    couleurs = Couleur.objects.filter(actif=True).order_by('nom')
    pointures = Pointure.objects.filter(actif=True).order_by('ordre', 'pointure')
    
    if request.method == 'POST':
        try:
            # Récupération des données du formulaire
            nom = request.POST.get('nom')
            reference = request.POST.get('reference')
            categorie_id = request.POST.get('categorie')
            genre_id = request.POST.get('genre')
            modele = request.POST.get('modele')
            phase = request.POST.get('phase')
            prix_str = request.POST.get('prix_unitaire', '').strip().replace(',', '.')
            prix_achat_str = request.POST.get('prix_achat', '').strip().replace(',', '.')
            description = request.POST.get('description')
            actif = 'actif' in request.POST
            isUpsell = 'isUpsell' in request.POST
            image = request.FILES.get('image')

            # Validation des champs obligatoires
            if not all([nom, categorie_id, genre_id, modele, prix_str]):
                messages.error(request, "Veuillez remplir tous les champs obligatoires (Nom, Catégorie, Genre, Modèle, Prix).")
                return render(request, 'Superpreparation/stock/creer_article.html', {
                    'categories': categories,
                    'genres': genres,
                    'couleurs': couleurs,
                    'pointures': pointures,
                    'article_phases': Article.PHASE_CHOICES,
                    'page_title': "Créer un Nouvel Article",
                    'page_subtitle': "Ajouter un article au catalogue"
                })

            try:
                prix_unitaire = float(prix_str)
                if prix_unitaire <= 0:
                    raise ValueError("Le prix doit être supérieur à 0")
            except ValueError:
                messages.error(request, "Le prix unitaire doit être un nombre valide supérieur à 0.")
                return render(request, 'Superpreparation/stock/creer_article.html', {
                    'categories': categories,
                    'genres': genres,
                    'couleurs': couleurs,
                    'pointures': pointures,
                    'article_phases': Article.PHASE_CHOICES,
                    'page_title': "Créer un Nouvel Article",
                    'page_subtitle': "Ajouter un article au catalogue"
                })

            # Vérifier l'unicité du modèle
            if Article.objects.filter(modele=modele).exists():
                messages.error(request, f"Un article avec le modèle {modele} existe déjà.")
                return render(request, 'Superpreparation/stock/creer_article.html', {
                    'categories': categories,
                    'genres': genres,
                    'couleurs': couleurs,
                    'pointures': pointures,
                    'article_phases': Article.PHASE_CHOICES,
                    'page_title': "Créer un Nouvel Article",
                    'page_subtitle': "Ajouter un article au catalogue"
                })

            # Créer l'article
            article = Article()
            article.nom = nom
            article.reference = reference
            article.categorie_id = categorie_id
            article.genre_id = genre_id
            article.modele = modele
            article.phase = phase or 'EN_COURS'
            article.prix_unitaire = prix_unitaire
            article.description = description
            article.actif = actif
            article.isUpsell = isUpsell
            
            # Gérer le prix d'achat
            if prix_achat_str:
                try:
                    prix_achat = float(prix_achat_str)
                    if prix_achat >= 0:
                        article.prix_achat = prix_achat
                except ValueError:
                    pass  # Ignorer les valeurs non numériques
            
            # Gérer l'image si elle est fournie
            if image:
                article.image = image
            
            # Générer automatiquement la référence si elle n'est pas fournie
            if not reference and categorie_id and genre_id and modele:
                article.save()
                article.refresh_from_db()
                reference_auto = article.generer_reference_automatique()
                if reference_auto:
                    article.reference = reference_auto
            
            article.save()
            
            # Gérer les variantes si elles sont fournies via AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Les variantes seront gérées par le JavaScript
                return JsonResponse({
                    'success': True,
                    'article_id': article.id,
                    'article_nom': article.nom,
                    'message': f"Article '{article.nom}' créé avec succès"
                })
            
            messages.success(request, f"L'article '{article.nom}' a été créé avec succès.")
            return redirect('Superpreparation:liste_articles')
            
        except Exception as e:
            messages.error(request, f"Une erreur est survenue lors de la création de l'article : {str(e)}")
            return render(request, 'Superpreparation/stock/creer_article.html', {
                'categories': categories,
                'genres': genres,
                'couleurs': couleurs,
                'pointures': pointures,
                'article_phases': Article.PHASE_CHOICES,
                'page_title': "Créer un Nouvel Article",
                'page_subtitle': "Ajouter un article au catalogue"
            })

    context = {
        'categories': categories,
        'genres': genres,
        'couleurs': couleurs,
        'pointures': pointures,
        'article_phases': Article.PHASE_CHOICES,
        'page_title': "Créer un Nouvel Article",
        'page_subtitle': "Ajouter un article au catalogue"
    }
    return render(request, 'Superpreparation/stock/creer_article.html', context)

@superviseur_preparation_required
def creer_variantes_ajax(request):
    """Créer des variantes pour un article via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'})
    
    try:
        import json
        data = json.loads(request.body)
        article_id = data.get('article_id')
        variantes_data = data.get('variantes', [])
        
        if not article_id:
            return JsonResponse({'success': False, 'error': 'ID de l\'article manquant'})
        
        article = get_object_or_404(Article, id=article_id)
        variantes_crees = 0
        erreurs = []
        
        for variante_data in variantes_data:
            try:
                couleur_id = variante_data.get('couleur_id')
                pointure_id = variante_data.get('pointure_id')
                quantite = variante_data.get('quantite', 0)
                reference = variante_data.get('reference')
                
                # Vérifier l'unicité de la combinaison
                if VarianteArticle.objects.filter(
                    article=article,
                    couleur_id=couleur_id if couleur_id else None,
                    pointure_id=pointure_id if pointure_id else None
                ).exists():
                    erreurs.append(f"Variante avec cette combinaison couleur/pointure existe déjà")
                    continue
                
                # Créer la variante
                variante = VarianteArticle()
                variante.article = article
                variante.couleur_id = couleur_id if couleur_id else None
                variante.pointure_id = pointure_id if pointure_id else None
                variante.qte_disponible = int(quantite) if quantite else 0
                variante.actif = True
                
                # Définir la référence de la variante
                if reference:
                    variante.reference_variante = reference
                else:
                    # Générer automatiquement la référence
                    variante.reference_variante = variante.generer_reference_variante_automatique()
                
                variante.save()
                variantes_crees += 1
                
            except Exception as e:
                erreurs.append(f"Erreur lors de la création de la variante: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'nombre_crees': variantes_crees,
            'erreurs': erreurs,
            'message': f'{variantes_crees} variante(s) créée(s) avec succès'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Données JSON invalides'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Erreur serveur: {str(e)}'})

@superviseur_preparation_required
def modifier_article(request, article_id):
    """Modifier un article existant - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    article = get_object_or_404(Article, pk=article_id)
    
    if request.method == 'POST':
        form = ArticleForm(request.POST, request.FILES, instance=article)
        if form.is_valid():
            # Sauvegarder l'article
            article_modifie = form.save()
            
            # Gérer l'ajustement de stock optionnel
            type_mouvement_stock = request.POST.get('type_mouvement_stock')
            quantite_ajustement = request.POST.get('quantite_ajustement')
            commentaire_ajustement = request.POST.get('commentaire_ajustement', '')
            
            if type_mouvement_stock and quantite_ajustement:
                try:
                    quantite = int(quantite_ajustement)
                    if quantite > 0:
                        print(f"🔧 Ajustement stock via modification - Article: {article_modifie.nom}, Type: {type_mouvement_stock}, Quantité: {quantite}")
                        print(f"🔧 Stock avant ajustement: {article_modifie.qte_disponible}")
                        
                        mouvement = creer_mouvement_stock(
                            article=article_modifie,
                            quantite=quantite,
                            type_mouvement=type_mouvement_stock,
                            operateur=operateur_profile,
                            commentaire=f"Ajustement via modification article. {commentaire_ajustement}".strip()
                        )
                        
                        # Recharger l'article pour voir le stock mis à jour
                        article_modifie.refresh_from_db()
                        print(f"✅ Stock après ajustement: {article_modifie.qte_disponible}")
                        
                        if mouvement:
                            messages.success(request, f"L'article '{article_modifie.nom}' a été modifié avec succès. Stock ajusté : nouveau stock = {article_modifie.qte_disponible} unités.")
                        else:
                            messages.warning(request, f"L'article '{article_modifie.nom}' a été modifié avec succès, mais l'ajustement de stock a échoué.")
                    else:
                        messages.warning(request, f"L'article '{article_modifie.nom}' a été modifié avec succès, mais la quantité d'ajustement doit être positive.")
                except (ValueError, TypeError) as e:
                    print(f"❌ Erreur lors de l'ajustement de stock: {str(e)}")
                    messages.warning(request, f"L'article '{article_modifie.nom}' a été modifié avec succès, mais l'ajustement de stock a échoué (quantité invalide).")
                except Exception as e:
                    print(f"❌ Erreur lors de l'ajustement de stock: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    messages.warning(request, f"L'article '{article_modifie.nom}' a été modifié avec succès, mais l'ajustement de stock a échoué.")
            else:
                messages.success(request, f"L'article '{article_modifie.nom}' a été modifié avec succès.")
                
            return redirect('Superpreparation:detail_article', article_id=article_modifie.id)
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    else:
        form = ArticleForm(instance=article)

    context = {
        'form': form,
        'article': article,
        'page_title': "Modifier l'Article",
        'page_subtitle': f"Mise à jour de {article.nom}"
    }
    return render(request, 'Superpreparation/stock/modifier_article.html', context)

# === NOUVELLES FONCTIONNALITÉS : RÉPARTITION AUTOMATIQUE ===

def get_operateur_display_name(operateur):
    """Fonction helper pour obtenir le nom d'affichage d'un opérateur"""
    if not operateur:
        return "Opérateur inconnu"
    
    if hasattr(operateur, 'nom_complet') and operateur.nom_complet:
        return operateur.nom_complet
    elif operateur.nom and operateur.prenom:
        return f"{operateur.prenom} {operateur.nom}"
    elif operateur.nom:
        return operateur.nom
    elif hasattr(operateur, 'user') and operateur.user:
        return operateur.user.username
    else:
        return "Opérateur inconnu"

# === VUES DE RÉPARTITION SUPPRIMÉES (DÉPLACÉES VERS ADMIN) ===
# Les vues de répartition ont été déplacées vers l'interface admin
# car ce sont les administrateurs qui s'en occupent maintenant

# === FONCTIONS UTILITAIRES DE RÉPARTITION SUPPRIMÉES (DÉPLACÉES VERS ADMIN) ===

# === VUES DE GESTION DES ENVOIS ===

@superviseur_preparation_required
def etats_livraison(request):
    """Gestion des états de livraison - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    from parametre.models import Region
    
    # Filtres
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    region_id = request.GET.get('region')
    statut = request.GET.get('statut')
    
    # Base queryset
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle__in=['En préparation', 'Prête', 'En cours de livraison', 'Livrée'],
        etats__date_fin__isnull=True
    ).select_related(
        'ville__region',
        'client'
    ).prefetch_related(
        'etats__enum_etat',
        'etats__operateur__user'
    ).distinct()
    
    # Appliquer les filtres
    if date_debut:
        commandes = commandes.filter(date_creation__gte=date_debut)
    if date_fin:
        commandes = commandes.filter(date_creation__lte=date_fin)
    if region_id and region_id.strip():
        commandes = commandes.filter(ville__region_id=region_id)
    if statut:
        commandes = commandes.filter(etats__enum_etat__libelle=statut, etats__date_fin__isnull=True)
    
    # Statistiques
    stats = {
        'total_commandes': commandes.count(),
        'en_preparation': commandes.filter(etats__enum_etat__libelle='En préparation', etats__date_fin__isnull=True).count(),
        'pretes': commandes.filter(etats__enum_etat__libelle='Prête', etats__date_fin__isnull=True).count(),
        'livrees': commandes.filter(etats__enum_etat__libelle='Livrée', etats__date_fin__isnull=True).count(),
    }
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(commandes, 50)
    page_number = request.GET.get('page')
    commandes = paginator.get_page(page_number)
    
    regions = Region.objects.all()
    
    context = {
        'commandes': commandes,
        'regions': regions,
        'stats': stats,
    }
    
    return render(request, 'Superpreparation/etats_livraison.html', context)

@superviseur_preparation_required
def export_envois(request):
    """Export des envois journaliers - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    from parametre.models import Region, Operateur
    from django.utils import timezone
    import datetime
    
    # Date par défaut : aujourd'hui
    today = timezone.now().date()
    date_envoi = request.GET.get('date_envoi', today)
    region_id = request.GET.get('region')
    livreur_id = request.GET.get('livreur')
    
    # Obtenir tous les livreurs (opérateurs de livraison)
    livreurs = Operateur.objects.filter(is_livraison=True, actif=True)
    regions = Region.objects.all()
    
    # Simuler des envois (à remplacer par votre modèle Envoi)
    envois = []
    
    # Commandes PRÉPARÉES à être envoyées
    commandes_pretes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True
    ).select_related('ville__region')
    
    if region_id and region_id.strip():
        commandes_pretes = commandes_pretes.filter(ville__region_id=region_id)
    
    # Statistiques
    stats = {
        'total_envois': len(envois),
        'total_commandes': 0,
        'commandes_pretes': commandes_pretes.count(),
        'livreurs_actifs': livreurs.filter(actif=True).count(),
    }
    
    context = {
        'envois': envois,
        'commandes_pretes': commandes_pretes,
        'livreurs': livreurs,
        'regions': regions,
        'stats': stats,
        'today': today,
    }
    
    return render(request, 'Superpreparation/export_envois.html', context)

@superviseur_preparation_required
def creer_envoi(request):
    """Créer un nouvel envoi"""
    if request.method == 'POST':
        try:
            livreur_id = request.POST.get('livreur')
            region_id = request.POST.get('region')
            notes = request.POST.get('notes', '')
            commandes_selectionnees = request.POST.get('commandes_selectionnees', '').split(',')
            
            # Ici vous devriez créer l'objet Envoi
            # envoi = Envoi.objects.create(
            #     livreur_id=livreur_id,
            #     region_id=region_id if region_id else None,
            #     notes=notes,
            #     date_creation=timezone.now()
            # )
            
            # Associer les commandes à l'envoi
            # for commande_id in commandes_selectionnees:
            #     if commande_id:
            #         commande = Commande.objects.get(id=commande_id)
            #         commande.envoi = envoi
            #         commande.save()
            
            return JsonResponse({'success': True, 'message': 'Envoi créé avec succès'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    
    return JsonResponse({'success': False, 'message': 'Méthode non autorisée'})

@superviseur_preparation_required
def details_envoi(request, envoi_id):
    """Afficher les détails d'un envoi"""
    # Ici vous devriez récupérer l'envoi par son ID
    # envoi = get_object_or_404(Envoi, id=envoi_id)
    
    # Pour l'exemple, retourner un contenu HTML simple
    html_content = f"""
    <div class="p-3">
        <h6>Envoi ENV-{envoi_id}</h6>
        <p><strong>Statut:</strong> En cours</p>
        <p><strong>Date création:</strong> {timezone.now().strftime('%d/%m/%Y %H:%M')}</p>
        <p><strong>Commandes associées:</strong> 0</p>
    </div>
    """
    
    return HttpResponse(html_content)

# === VUES D'EXPORT ET D'IMPRESSION ===

@superviseur_preparation_required
def details_region_view(request):
    """Vue détaillée pour afficher les commandes par région - Service de préparation"""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')
    
    from parametre.models import Region, Ville
    
    # Récupérer les paramètres de filtrage
    region_name = request.GET.get('region')
    ville_name = request.GET.get('ville')
    
    # Base queryset pour toutes les commandes en traitement
    commandes_reparties = Commande.objects.filter(
        etats__enum_etat__libelle__in=['Confirmée', 'À imprimer', 'Préparée', 'En cours de livraison'],
        etats__date_fin__isnull=True,
        ville__isnull=False,  # Exclure les commandes sans ville
        ville__region__isnull=False  # Exclure les commandes sans région
    ).select_related('client', 'ville', 'ville__region').prefetch_related(
        'etats__operateur', 'etats__enum_etat', 'paniers__article'
    ).distinct()
    
    # Appliquer les filtres
    if region_name:
        commandes_reparties = commandes_reparties.filter(ville__region__nom_region=region_name)
    if ville_name:
        commandes_reparties = commandes_reparties.filter(ville__nom=ville_name)
    
    # Statistiques par ville dans la région/ville filtrée
    stats_par_ville = commandes_reparties.values(
        'ville__id', 'ville__nom', 'ville__region__nom_region'
    ).annotate(
        nb_commandes=Count('id'),
        total_montant=Sum('total_cmd')
    ).order_by('ville__region__nom_region', 'ville__nom')
    
    # Statistiques des commandes PRÉPARÉES par ville dans la région/ville filtrée
    commandes_preparees = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True,
        ville__isnull=False,
        ville__region__isnull=False
    ).select_related('ville', 'ville__region')
    
    # Appliquer les mêmes filtres que pour les commandes en traitement
    if region_name:
        commandes_preparees = commandes_preparees.filter(ville__region__nom_region=region_name)
    if ville_name:
        commandes_preparees = commandes_preparees.filter(ville__nom=ville_name)
    
    stats_preparees_par_ville = commandes_preparees.values(
        'ville__nom', 'ville__region__nom_region'
    ).annotate(
        nb_commandes_preparees=Count('id')
    ).order_by('ville__region__nom_region', 'ville__nom')
    
    # Créer un dictionnaire pour un accès rapide
    preparees_par_ville = {(stat['ville__nom'], stat['ville__region__nom_region']): stat['nb_commandes_preparees'] for stat in stats_preparees_par_ville}
    
    # Calculer les totaux
    total_commandes = commandes_reparties.count()
    total_montant = commandes_reparties.aggregate(total=Sum('total_cmd'))['total'] or 0
    
    # Définir le titre selon le filtre appliqué
    if region_name:
        page_title = f"Détails - {region_name}"
        page_subtitle = f"Commandes en traitement dans la région {region_name}"
    elif ville_name:
        page_title = f"Détails - {ville_name}"
        page_subtitle = f"Commandes en traitement à {ville_name}"
    else:
        page_title = "Détails par Région"
        page_subtitle = "Répartition détaillée des commandes en traitement"
    
    context = {
        'operateur': operateur_profile,
        'commandes_reparties': commandes_reparties,
        'stats_par_ville': stats_par_ville,
        'preparees_par_ville': preparees_par_ville,
        'total_commandes': total_commandes,
        'total_montant': total_montant,
        'region_name': region_name,
        'ville_name': ville_name,
        'page_title': page_title,
        'page_subtitle': page_subtitle,
    }
    
    return render(request, 'Superpreparation/details_region.html', context)

@superviseur_preparation_required
def imprimer_commande(request, commande_id):
    """
    Imprime une commande spécifique.
    """
    commande = get_object_or_404(Commande, id=commande_id)
    # Assurez-vous que l'opérateur a le droit de voir cette commande si nécessaire
    return render(request, 'Superpreparation/impression_commande.html', {'commande': commande})

@superviseur_preparation_required 
def exporter_etats_pdf(request):
    """
    Exporte l'état actuel des livraisons en PDF.
    """
    # Votre logique d'exportation PDF ici
    return HttpResponse("Export PDF des états de livraison à implémenter.", content_type="text/plain")

@superviseur_preparation_required
def imprimer_envoi(request, envoi_id):
    """
    Imprime les détails d'un envoi.
    """
    envoi = get_object_or_404(Envoi, id=envoi_id)
    return render(request, 'Superpreparation/impression_envoi.html', {'envoi': envoi})

@superviseur_preparation_required
def exporter_envoi(request, envoi_id, format):
    """
    Exporte un envoi dans un format spécifique (CSV/PDF).
    """
    envoi = get_object_or_404(Envoi, id=envoi_id)
    if format == 'csv':
        # Logique d'export CSV
        return HttpResponse(f"Export CSV de l'envoi {envoi_id}", content_type="text/csv")
    elif format == 'pdf':
        # Logique d'export PDF
        return HttpResponse(f"Export PDF de l'envoi {envoi_id}", content_type="application/pdf")
    return HttpResponse("Format non supporté", status=400)

@superviseur_preparation_required
def exporter_envois_journaliers(request):
    """
    Exporte tous les envois du jour.
    """
    # Votre logique d'exportation ici
    return HttpResponse("Export des envois journaliers à implémenter.", content_type="text/plain")

@superviseur_preparation_required
def envois_view(request):
    """Page de gestion des envois (création/en cours)."""
    try:
        operateur_profile = request.user.profil_operateur
        if not operateur_profile.is_preparation:
            messages.error(request, "Accès non autorisé.")
            return redirect('login')
    except Operateur.DoesNotExist:
        messages.error(request, "Profil opérateur non trouvé.")
        return redirect('login')

    # Données placeholder: commandes prêtes (même logique que export_envois)
    from parametre.models import Region
    regions = Region.objects.all()
    commandes_pretes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True
    ).select_related('ville__region')
    
    # Récupérer les envois actifs (non clôturés)
    envois_actifs = Envoi.objects.exclude(
        status=False
    ).select_related('region').order_by('-date_creation')

    # Mettre à jour le compteur de commandes pour chaque envoi
    for envoi in envois_actifs:
        nb_commandes = Commande.objects.filter(
            ville__region=envoi.region,
            etats__enum_etat__libelle='Préparée',
            etats__date_fin__isnull=True
        ).count()
        
        # Mettre à jour seulement si le nombre a changé
        if envoi.nb_commandes != nb_commandes:
            envoi.nb_commandes = nb_commandes
            envoi.save(update_fields=['nb_commandes'])

    # Récupérer les IDs des régions qui ont déjà un envoi actif
    regions_avec_envoi_actif = set(envois_actifs.values_list('region_id', flat=True))

    context = {
        'regions': regions,
        'commandes_pretes': commandes_pretes,
        'envois_actifs': envois_actifs,
        'regions_avec_envoi_actif': regions_avec_envoi_actif,
        'page_title': 'Gestion des envois',
        'page_subtitle': 'Créer et suivre les envois en cours',
    }
    return render(request, 'Superpreparation/envois.html', context)

@csrf_exempt
@login_required
def historique_envois_view(request):
    """Page d'historique des envois clôturés."""
    from django.core.paginator import Paginator
    from datetime import datetime, timedelta
    
    # Récupérer les envois clôturés (status=False)
    envois_clotures = Envoi.objects.filter(
        status=False  # False = clôturé
    ).select_related('region', 'operateur_creation').order_by('-date_creation')
    
    # Mettre à jour le compteur de commandes pour chaque envoi clôturé
    for envoi in envois_clotures:
        # Pour les envois clôturés, on compte toutes les commandes de la région
        # qui étaient préparées à la date de clôture (approximatif)
        nb_commandes = Commande.objects.filter(
            ville__region=envoi.region,
            etats__enum_etat__libelle='Préparée',
            date_creation__lte=envoi.date_livraison_effective or envoi.date_creation
        ).count()
        
        # Mettre à jour seulement si le nombre a changé
        if envoi.nb_commandes != nb_commandes:
            envoi.nb_commandes = nb_commandes
            envoi.save(update_fields=['nb_commandes'])
    
    # Filtrage par période si demandé
    periode = request.GET.get('periode')
    if periode:
        today = datetime.now().date()
        if periode == '7j':
            date_debut = today - timedelta(days=7)
            envois_clotures = envois_clotures.filter(date_creation__gte=date_debut)
        elif periode == '30j':
            date_debut = today - timedelta(days=30)
            envois_clotures = envois_clotures.filter(date_creation__gte=date_debut)
        elif periode == '90j':
            date_debut = today - timedelta(days=90)
            envois_clotures = envois_clotures.filter(date_creation__gte=date_debut)
    
    # Filtrage par région si demandé
    region_filter = request.GET.get('region')
    if region_filter:
        envois_clotures = envois_clotures.filter(region__id=region_filter)
    
    # Pagination
    paginator = Paginator(envois_clotures, 12)  # 12 envois par page
    page_number = request.GET.get('page')
    envois = paginator.get_page(page_number)
    
    # Récupérer toutes les régions pour le filtre
    from parametre.models import Region
    regions = Region.objects.all().order_by('nom_region')
    
    context = {
        'envois': envois,
        'regions': regions,
        'periode_selected': periode,
        'region_selected': region_filter,
        'page_title': 'Historique des envois',
        'page_subtitle': 'Consulter les envois clôturés et exporter leurs données',
    }
    return render(request, 'Superpreparation/historique_envois.html', context)

@csrf_exempt
@login_required
def creer_envoi_region(request):
    """Créer un envoi basé sur une région (POST)."""
    # Debug: Afficher les informations de l'utilisateur
    print(f"🔍 DEBUG - Utilisateur: {request.user.username}")
    print(f"🔍 DEBUG - Authentifié: {request.user.is_authenticated}")
    print(f"🔍 DEBUG - Headers: {dict(request.headers)}")
    print(f"🔍 DEBUG - Method: {request.method}")
    
    try:
        from parametre.models import Operateur
        operateur = Operateur.objects.get(user=request.user, actif=True)
        print(f"🔍 DEBUG - Profil opérateur: {operateur.type_operateur}")
    except Operateur.DoesNotExist:
        print(f"🔍 DEBUG - Aucun profil opérateur trouvé")
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)

    try:
        from parametre.models import Region
        region_id = request.POST.get('region_id')

        if not region_id:
            return JsonResponse({'success': False, 'message': 'Région requise'}, status=400)

        region = Region.objects.get(id=region_id)

        # Vérifier qu'il n'y a pas déjà un envoi actif pour cette région
        envoi_actif_existant = Envoi.objects.filter(region=region, status=True).exists()
        if envoi_actif_existant:
            return JsonResponse({
                'success': False, 
                'message': f'Il existe déjà un envoi actif pour la région {region.nom_region}. Veuillez clôturer l\'envoi existant avant d\'en créer un nouveau.'
            }, status=400)

        # Générer un numéro d'envoi via le modèle Envoi
        from django.utils import timezone
        
        # Récupérer le profil opérateur
        try:
            from parametre.models import Operateur
            operateur_creation = Operateur.objects.get(user=request.user, actif=True)
        except Operateur.DoesNotExist:
            operateur_creation = None
            
        envoi = Envoi.objects.create(
            region=region,
            operateur_creation=operateur_creation,
            date_envoi=timezone.now().date(),
            date_livraison_prevue=timezone.now().date(),
        )

        return JsonResponse({'success': True, 'envoi_id': envoi.id, 'numero': envoi.numero_envoi})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@csrf_exempt
@login_required
def cloturer_envoi(request):
    """Clôturer un envoi (POST)."""
    print(f"🔍 DEBUG CLOTURER - Utilisateur: {request.user.username}")
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée'}, status=405)

    try:
        envoi_id = request.POST.get('envoi_id')
        if not envoi_id:
            return JsonResponse({'success': False, 'message': 'ID d\'envoi requis'}, status=400)

        envoi = Envoi.objects.get(id=envoi_id)
        
        # Vérifier que l'envoi n'est pas déjà clôturé
        if not envoi.status:  # False = déjà clôturé
            return JsonResponse({'success': False, 'message': 'Cet envoi est déjà clôturé'}, status=400)

        from django.db import transaction
        from django.utils import timezone
        from commande.models import Commande, EtatCommande, EnumEtatCmd

        with transaction.atomic():
            # Passer toutes les commandes liées à l'envoi en "En livraison"
            # On cible explicitement les commandes dont l'état courant est "Préparée"
            commandes = Commande.objects.filter(
                envoi=envoi,
                etats__enum_etat__libelle='Préparée',
                etats__date_fin__isnull=True,
            ).distinct()
            if not commandes.exists():
                # Fallback: prendre les commandes 'Préparée' de la région de l'envoi
                commandes = (
                    Commande.objects.filter(
                        ville__region=envoi.region,
                        etats__enum_etat__libelle='Préparée',
                        etats__date_fin__isnull=True,
                    )
                    .select_related('client', 'ville')
                    .distinct()
                )
            etat_enum, _ = EnumEtatCmd.objects.get_or_create(
                libelle='En livraison',
                defaults={'ordre': 17, 'couleur': '#8B5CF6'}
            )

            for commande in commandes:
                try:
                    # Lier la commande à cet envoi si ce n'est pas déjà fait
                    if commande.envoi_id != envoi.id:
                        commande.envoi = envoi
                        commande.save(update_fields=['envoi'])
                    
                    # Fermer l'état courant "Préparée" s'il existe
                    etat_actuel = commande.etats.filter(date_fin__isnull=True).order_by('-date_debut').first()
                    if etat_actuel and etat_actuel.enum_etat.libelle == 'Préparée':
                        etat_actuel.date_fin = timezone.now()
                        etat_actuel.save(update_fields=['date_fin'])

                    # Créer le nouvel état
                    EtatCommande.objects.create(
                        commande=commande,
                        enum_etat=etat_enum,
                        operateur=getattr(request.user, 'profil_operateur', None),
                        date_debut=timezone.now(),
                        commentaire=f"Envoi clôturé: {envoi.numero_envoi}"
                    )
                except Exception as state_err:
                    print(f"⚠️ Erreur transition état commande {commande.id}: {state_err}")
                    continue

            # Marquer l'envoi comme livré (clôturé)
            envoi.marquer_comme_livre(getattr(request.user, 'profil_operateur', None))

        return JsonResponse({
            'success': True, 
            'message': f'Envoi {envoi.numero_envoi} clôturé avec succès'
        })
    except Envoi.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Envoi non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)

@superviseur_preparation_required
def rafraichir_articles_commande_prepa(request, commande_id):
    """Rafraîchir la section des articles de la commande en préparation"""
    print(f"🔄 Rafraîchissement des articles pour la commande {commande_id}")
    
    try:
        operateur = Operateur.objects.get(user=request.user, actif=True)
        
        # Vérifier le type d'opérateur
        if operateur.type_operateur not in ['SUPERVISEUR_PREPARATION', 'PREPARATION', 'ADMIN']:
            return JsonResponse({'error': 'Accès non autorisé.'}, status=403)
            
    except Operateur.DoesNotExist:
        # Fallback si pas de profil: autoriser selon groupes Django
        if not request.user.groups.filter(name__in=['superviseur', 'operateur_preparation']).exists():
            return JsonResponse({'error': 'Profil d\'opérateur de préparation non trouvé.'}, status=403)
        operateur = None
    
    try:
        commande = Commande.objects.get(id=commande_id)
        
        # Pour les superviseurs, on ne vérifie pas l'affectation spécifique
        # Ils peuvent accéder à toutes les commandes en préparation
        if operateur and operateur.type_operateur == 'SUPERVISEUR_PREPARATION':
            # Vérifier seulement que la commande est en préparation
            etat_preparation = commande.etats.filter(
                enum_etat__libelle__in=['En préparation', 'À imprimer'],
                date_fin__isnull=True
            ).first()
        else:
            # Pour les opérateurs normaux, vérifier l'affectation spécifique
            etat_preparation = commande.etats.filter(
                operateur=operateur,
                enum_etat__libelle__in=['En préparation', 'À imprimer'],
                date_fin__isnull=True
            ).first()
        
        if not etat_preparation:
            return JsonResponse({'error': 'Cette commande ne vous est pas affectée.'}, status=403)
        
        # Générer le HTML directement pour éviter les erreurs de template
        paniers = commande.paniers.select_related('article').all()
        html_rows = []
        
        for panier in paniers:
            # Utiliser le filtre Django pour calculer le prix selon la logique upsell
            from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
            
            prix_calcule = get_prix_upsell_avec_compteur(panier.article, commande.compteur)
            
            # Déterminer l'affichage selon le type d'article
            if panier.article.isUpsell:
                prix_class = "text-orange-600"
                upsell_badge = f'<span class="inline-flex items-center px-2 py-1 rounded-full text-xs font-bold bg-orange-100 text-orange-700 ml-2"><i class="fas fa-arrow-up mr-1"></i>Upsell</span>'
                
                # Déterminer le niveau d'upsell affiché
                if commande.compteur >= 4 and panier.article.prix_upsell_4 is not None:
                    niveau_upsell = 4
                elif commande.compteur >= 3 and panier.article.prix_upsell_3 is not None:
                    niveau_upsell = 3
                elif commande.compteur >= 2 and panier.article.prix_upsell_2 is not None:
                    niveau_upsell = 2
                elif commande.compteur >= 1 and panier.article.prix_upsell_1 is not None:
                    niveau_upsell = 1
                else:
                    niveau_upsell = 0
                
                if niveau_upsell > 0:
                    upsell_info = f'<div class="text-xs text-orange-600 flex items-center justify-center gap-1"><i class="fas fa-arrow-up"></i>Upsell Niveau {niveau_upsell}</div>'
                else:
                    upsell_info = f'<div class="text-xs text-orange-600 flex items-center justify-center gap-1"><i class="fas fa-arrow-up"></i>Upsell (Prix normal)</div>'
            else:
                # Prix normal
                prix_class = "text-gray-700"
                upsell_badge = ""
                upsell_info = ""
            
            sous_total = prix_calcule * panier.quantite
            
            html_row = f"""
            <tr data-panier-id="{panier.id}" data-article-id="{panier.article.id}">
                <td class="px-4 py-3">
                    <div class="flex items-center">
                        <div class="flex-1">
                            <div class="font-medium text-gray-900 flex items-center">
                                {panier.article.nom}
                                {upsell_badge}
                            </div>
                            <div class="text-sm text-gray-500">
                                Réf: {panier.article.reference or 'N/A'}
                                {f" · Réf variante: {getattr(panier.variante, 'reference', '')}" if hasattr(panier, 'variante') and getattr(panier, 'variante', None) and getattr(panier.variante, 'reference', None) else ''}
                                {f" · Couleur: {getattr(getattr(panier.variante,'couleur', None),'nom','')}" if hasattr(panier, 'variante') and getattr(panier, 'variante', None) and getattr(panier.variante, 'couleur', None) else ''}
                                {f" · Pointure: {getattr(getattr(panier.variante,'pointure', None),'pointure','')}" if hasattr(panier, 'variante') and getattr(panier, 'variante', None) and getattr(panier.variante, 'pointure', None) else ''}
                            </div>
                        </div>
                    </div>
                </td>
                <td class="px-4 py-3 text-center">
                    <div class="flex items-center justify-center space-x-2">
                        <button type="button" onclick="modifierQuantite({panier.id}, -1)" 
                                class="w-8 h-8 bg-gray-200 hover:bg-gray-300 rounded text-sm transition-colors">-</button>
                        <input type="number" id="quantite-{panier.id}" value="{panier.quantite}" min="1" 
                               class="w-16 p-2 border rounded-lg text-center" 
                               onchange="modifierQuantiteDirecte({panier.id}, this.value)">
                        <button type="button" onclick="modifierQuantite({panier.id}, 1)" 
                                class="w-8 h-8 bg-gray-200 hover:bg-gray-300 rounded text-sm transition-colors">+</button>
                    </div>
                </td>
                <td class="px-4 py-3 text-center">
                    <div class="font-medium {prix_class}" id="prix-unitaire-{panier.id}">
                        {prix_calcule:.2f} DH
                    </div>
                    {upsell_info}
                </td>
                <td class="px-4 py-3 text-center">
                    <div class="font-bold text-gray-900">
                        {sous_total:.2f} DH
                    </div>
                </td>
                <td class="px-4 py-3 text-center">
                    <button type="button" onclick="supprimerArticle({panier.id})" 
                            class="px-3 py-2 bg-red-500 hover:bg-red-600 text-white text-sm rounded-lg transition-colors">
                        <i class="fas fa-trash mr-1"></i>Supprimer
                    </button>
                </td>
            </tr>
            """
            html_rows.append(html_row)
        
        html = "".join(html_rows)
        
        # Calculer les statistiques upsell
        articles_upsell = commande.paniers.filter(article__isUpsell=True)
        total_quantite_upsell = articles_upsell.aggregate(
            total=Sum('quantite')
        )['total'] or 0
        
        response_data = {
            'success': True,
            'html': html,
            'articles_count': commande.paniers.count(),
            'total_commande': float(commande.total_cmd),
            'sous_total_articles': float(commande.sous_total_articles),
            'compteur': commande.compteur,
            'articles_upsell': articles_upsell.count(),
            'quantite_totale_upsell': total_quantite_upsell
        }
        
        print(f"✅ Rafraîchissement terminé - Articles: {response_data['articles_count']}, Compteur: {response_data['compteur']}")
        print(f"🔍 Détails du compteur: type={type(commande.compteur)}, valeur={commande.compteur}")
        if commande.compteur is None:
            print("⚠️ ATTENTION: Le compteur est None dans la base de données")
        elif commande.compteur == 0:
            print("ℹ️ Le compteur est 0 (normal si pas d'articles upsell)")
        else:
            print(f"✅ Le compteur a une valeur: {commande.compteur}")
        return JsonResponse(response_data)
        
    except Commande.DoesNotExist:
        print(f"❌ Commande {commande_id} non trouvée")
        return JsonResponse({'error': 'Commande non trouvée'}, status=404)
    except Exception as e:
        print(f"❌ Erreur lors du rafraîchissement: {str(e)}")
        import traceback
        traceback.print_exc()
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@superviseur_preparation_required
def ajouter_article_commande_prepa(request, commande_id):
    """Ajouter un article à la commande en préparation"""
    print("🔄 ===== DÉBUT ajouter_article_commande_prepa =====")
    print(f"📦 Méthode HTTP: {request.method}")
    print(f"📦 Commande ID: {commande_id}")
    print(f"📦 User: {request.user}")
    print(f"📦 POST data: {dict(request.POST)}")
    print(f"📦 Headers: {dict(request.headers)}")
    
    if request.method != 'POST':
        print("❌ Méthode non autorisée")
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
    
    try:
        # Autoriser PREPARATION et SUPERVISEUR_PREPARATION
        operateur = Operateur.objects.get(user=request.user, actif=True)
        print(f"✅ Opérateur trouvé: {operateur.id} - Type: {operateur.type_operateur}")
        if operateur.type_operateur not in ['PREPARATION', 'SUPERVISEUR_PREPARATION']:
            print(f"❌ Type d'opérateur non autorisé: {operateur.type_operateur}")
            return JsonResponse({'error': "Accès réservé à l'équipe préparation"}, status=403)
    except Operateur.DoesNotExist:
        print("❌ Profil d'opérateur non trouvé")
        return JsonResponse({'error': 'Profil d\'opérateur non trouvé.'}, status=403)
    
    try:
        with transaction.atomic():
            print("🔧 Début de la transaction atomique")
            commande = Commande.objects.select_for_update().get(id=commande_id)
            print(f"✅ Commande trouvée: {commande.id} - ID YZ: {commande.id_yz}")
            
            # Vérifier que la commande est en préparation
            if operateur.type_operateur == 'SUPERVISEUR_PREPARATION':
                # Le superviseur peut agir sur toute commande en cours (sans contrainte d'affectation)
                etat_preparation = commande.etats.filter(
                    enum_etat__libelle__in=['En préparation', 'À imprimer'],
                    date_fin__isnull=True
                ).first()
            else:
                # Opérateur préparation: seulement sur les commandes qui lui sont affectées
                etat_preparation = commande.etats.filter(
                    operateur=operateur,
                    enum_etat__libelle__in=['En préparation', 'À imprimer'],
                    date_fin__isnull=True
                ).first()
            
            if not etat_preparation:
                print("❌ Commande non en préparation pour cet opérateur")
                return JsonResponse({'error': 'Cette commande n\'est pas en préparation pour vous.'}, status=403)
            
            print(f"✅ État de préparation trouvé: {etat_preparation.enum_etat.libelle}")
            
            article_id = request.POST.get('article_id')
            quantite = int(request.POST.get('quantite', 1))
            variante_id = request.POST.get('variante_id')

            print("[AJOUT VARIANTE] entrée:", {
                'commande_id': commande_id,
                'operateur': getattr(operateur, 'id', None),
                'article_id': article_id,
                'quantite': quantite,
                'variante_id': variante_id,
            })
            
            if not article_id or quantite <= 0:
                print(f"❌ Données invalides: article_id={article_id}, quantite={quantite}")
                return JsonResponse({'error': 'Données invalides'}, status=400)

            print(f"✅ Données reçues: article_id={article_id}, quantite={quantite}, variante_id={variante_id}")
            article = Article.objects.get(id=article_id)
            print(f"✅ Article trouvé: {article.id} - {article.nom}")
            variante_obj = None
            if variante_id:
                try:
                    from article.models import VarianteArticle
                    variante_obj = VarianteArticle.objects.get(id=int(variante_id), article=article)
                    print("[AJOUT VARIANTE] variante trouvée:", {
                        'id': variante_obj.id,
                        'couleur': getattr(getattr(variante_obj, 'couleur', None), 'nom', None),
                        'pointure': getattr(getattr(variante_obj, 'pointure', None), 'pointure', None),
                        'qte_disponible_avant': variante_obj.qte_disponible,
                    })
                except Exception:
                    variante_obj = None
                    print("[AJOUT VARIANTE] variante introuvable ou invalide", variante_id)
            
            # Décrémenter le stock et créer un mouvement
            print("[AJOUT VARIANTE] création mouvement stock", {
                'article': article.id,
                'quantite': quantite,
                'type': 'sortie',
                'variante': getattr(variante_obj, 'id', None),
            })
            creer_mouvement_stock(
                article=article, quantite=quantite, type_mouvement='sortie',
                commande=commande, operateur=operateur,
                commentaire=f'Ajout article pendant préparation cmd {commande.id_yz}',
                variante=variante_obj
            )
            
            # Vérifier si l'article existe déjà dans la commande
            filtre_panier = {'commande': commande, 'article': article}
            if hasattr(Panier, 'variante'):
                filtre_panier['variante'] = variante_obj
            panier_existant = Panier.objects.filter(**filtre_panier).first()
            print("[AJOUT VARIANTE] filtre panier:", {
                'commande': commande.id,
                'article': article.id,
                'variante': getattr(variante_obj, 'id', None),
                'existant': getattr(panier_existant, 'id', None)
            })
            
            if panier_existant:
                # Si l'article existe déjà, mettre à jour la quantité
                panier_existant.quantite += quantite
                panier_existant.save()
                panier = panier_existant
                print(f"[AJOUT VARIANTE] 🔄 panier existant {panier.id} mis à jour, nouvelle_quantite={panier.quantite}")
            else:
                # Si l'article n'existe pas, créer un nouveau panier
                create_kwargs = {
                    'commande': commande,
                    'article': article,
                    'quantite': quantite,
                    'sous_total': 0,
                }
                if hasattr(Panier, 'variante'):
                    create_kwargs['variante'] = variante_obj
                panier = Panier.objects.create(**create_kwargs)
                print(f"[AJOUT VARIANTE] ➕ nouveau panier créé id={panier.id}, article={article.id}, variante={getattr(variante_obj,'id',None)}, quantite={quantite}")
            
            # Recalculer le compteur après ajout (logique de confirmation)
            if article.isUpsell and hasattr(article, 'prix_upsell_1') and article.prix_upsell_1 is not None:
                # Compter la quantité totale d'articles upsell (après ajout)
                total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                    total=Sum('quantite')
                )['total'] or 0
                
                # Le compteur ne s'incrémente qu'à partir de 2 unités d'articles upsell
                # 0-1 unités upsell → compteur = 0
                # 2+ unités upsell → compteur = total_quantite_upsell - 1
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()
            else:
                # Pour les articles normaux, juste calculer le sous-total
                from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                prix_unitaire = get_prix_upsell_avec_compteur(article, commande.compteur)
                sous_total = prix_unitaire * panier.quantite
                panier.sous_total = float(sous_total)
                panier.save()
            
            # Recalculer le total
            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()
            print("[AJOUT VARIANTE] totaux mis à jour:", {
                'commande_id': commande.id,
                'total_cmd': commande.total_cmd,
                'articles_count': commande.paniers.count(),
            })
            
            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = articles_upsell.aggregate(
                total=Sum('quantite')
            )['total'] or 0
            
            # Déterminer si c'était un ajout ou une mise à jour
            message = 'Article ajouté' if not panier_existant else f'Quantité mise à jour ({panier.quantite})'
            
            # Préparer les données de l'article pour le frontend
            article_data = {
                'panier_id': panier.id,
                'article_id': article.id,
                'nom': article.nom,
                'reference': article.reference,
                'couleur_fr': (variante_obj.couleur.nom if variante_obj and variante_obj.couleur else (article.couleur or '')),
                'couleur_ar': (variante_obj.couleur.nom if variante_obj and variante_obj.couleur else (article.couleur or '')),
                'pointure': (variante_obj.pointure.pointure if variante_obj and variante_obj.pointure else (article.pointure or '')),
                'quantite': panier.quantite,
                'prix': panier.sous_total / panier.quantite,  # Prix unitaire
                'sous_total': panier.sous_total,
                'is_upsell': article.isUpsell,
                'compteur': commande.compteur,
                'phase': article.phase,
                'has_promo_active': article.has_promo_active,
                'qte_disponible': variante_obj.qte_disponible if variante_obj else article.qte_disponible,
                'description': article.description or ''
            }
            
            response_data = {
                'success': True, 
                'message': message,
                'compteur': commande.compteur,
                'total_commande': float(commande.total_cmd),
                'was_update': panier_existant is not None,
                'new_quantity': panier.quantite,
                'article_data': article_data,
                'articles_count': commande.paniers.count(),
                'sous_total_articles': float(sum(p.sous_total for p in commande.paniers.all())),
                'articles_upsell': articles_upsell.count(),
                'quantite_totale_upsell': total_quantite_upsell
            }
            
            print("✅ ===== SUCCÈS ajouter_article_commande_prepa =====")
            print(f"📦 Réponse: {response_data}")
            
            return JsonResponse(response_data)
            
    except Article.DoesNotExist:
        print("❌ Article non trouvé")
        return JsonResponse({'error': 'Article non trouvé'}, status=404)
    except Exception as e:
        print(f"❌ ===== ERREUR ajouter_article_commande_prepa =====")
        print(f"❌ Exception: {str(e)}")
        import traceback
        print(f"❌ Traceback: {traceback.format_exc()}")
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@superviseur_preparation_required
def modifier_quantite_article_prepa(request, commande_id):
    """Modifier la quantité d'un article dans la commande en préparation"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur non trouvé.'}, status=403)

    try:
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # Vérifier l'affectation
            if not commande.etats.filter(operateur=operateur, enum_etat__libelle__in=['En préparation', 'À imprimer'], date_fin__isnull=True).exists():
                return JsonResponse({'error': 'Commande non affectée.'}, status=403)
            
            panier_id = request.POST.get('panier_id')
            nouvelle_quantite = int(request.POST.get('quantite', 1))

            panier = Panier.objects.get(id=panier_id, commande=commande)
            ancienne_quantite = panier.quantite
            article = panier.article
            difference = nouvelle_quantite - ancienne_quantite

            if difference > 0:
                creer_mouvement_stock(article, difference, 'sortie', commande, operateur, f'Ajustement qté cmd {commande.id_yz}')
            elif difference < 0:
                creer_mouvement_stock(article, abs(difference), 'entree', commande, operateur, f'Ajustement qté cmd {commande.id_yz}')

            panier.quantite = nouvelle_quantite
            
            # Recalculer le compteur si c'est un article upsell
            if article.isUpsell:
                # Compter la quantité totale d'articles upsell après modification
                total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                    total=Sum('quantite')
                )['total'] or 0
                
                # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()
            else:
                # Pour les articles normaux, juste calculer le sous-total
                from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
                prix_unitaire = get_prix_upsell_avec_compteur(article, commande.compteur)
                panier.sous_total = prix_unitaire * nouvelle_quantite
            panier.save()

            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()

            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = articles_upsell.aggregate(
                total=Sum('quantite')
            )['total'] or 0

            return JsonResponse({
                'success': True, 
                'message': 'Quantité modifiée',
                'compteur': commande.compteur,
                'articles_upsell': articles_upsell.count(),
                'quantite_totale_upsell': total_quantite_upsell,
                'total_commande': float(commande.total_cmd),
                'sous_total_articles': float(commande.sous_total_articles)
            })

    except Panier.DoesNotExist:
        return JsonResponse({'error': 'Panier non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

@superviseur_preparation_required
def supprimer_article_commande_prepa(request, commande_id):
    """Supprimer un article de la commande en préparation"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur non trouvé.'}, status=403)

    try:
        with transaction.atomic():
            commande = Commande.objects.select_for_update().get(id=commande_id)
            
            # Vérifier l'affectation
            if not commande.etats.filter(operateur=operateur, enum_etat__libelle__in=['En préparation', 'À imprimer'], date_fin__isnull=True).exists():
                return JsonResponse({'error': 'Commande non affectée.'}, status=403)

            panier_id = request.POST.get('panier_id')
            panier = Panier.objects.get(id=panier_id, commande=commande)
            quantite_supprimee = panier.quantite
            article = panier.article
            
            creer_mouvement_stock(article, quantite_supprimee, 'entree', commande, operateur, f'Suppression article cmd {commande.id_yz}')
            
            # Sauvegarder l'info avant suppression
            etait_upsell = panier.article.isUpsell
            
            # Supprimer l'article
            panier.delete()

            # Recalculer le compteur après suppression (logique de confirmation)
            if etait_upsell:
                # Compter la quantité totale d'articles upsell restants (après suppression)
                total_quantite_upsell = commande.paniers.filter(article__isUpsell=True).aggregate(
                    total=Sum('quantite')
                )['total'] or 0
                
                # Appliquer la logique : compteur = max(0, total_quantite_upsell - 1)
                if total_quantite_upsell >= 2:
                    commande.compteur = total_quantite_upsell - 1
                else:
                    commande.compteur = 0
                
                commande.save()
                
                # Recalculer TOUS les articles de la commande avec le nouveau compteur
                commande.recalculer_totaux_upsell()

            commande.total_cmd = sum(p.sous_total for p in commande.paniers.all())
            commande.save()

            # Calculer les statistiques upsell
            articles_upsell = commande.paniers.filter(article__isUpsell=True)
            total_quantite_upsell = articles_upsell.aggregate(
                total=Sum('quantite')
            )['total'] or 0

            return JsonResponse({
                'success': True, 
                'message': 'Article supprimé',
                'compteur': commande.compteur,
                'articles_upsell': articles_upsell.count(),
                'quantite_totale_upsell': total_quantite_upsell,
                'total_commande': float(commande.total_cmd),
                'sous_total_articles': float(commande.sous_total_articles)
            })

    except Panier.DoesNotExist:
        return JsonResponse({'error': 'Panier non trouvé'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Erreur interne: {str(e)}'}, status=500)

# === VUES DE RÉPARTITION SUPPRIMÉES (DÉPLACÉES VERS ADMIN) ===
# Les vues de répartition ont été déplacées vers l'interface admin
# car ce sont les administrateurs qui s'en occupent maintenant

@superviseur_preparation_required
def api_panier_commande_livraison(request, commande_id):
    """API pour récupérer le panier d'une commande pour les opérateurs de livraison"""
    try:
        # Vérifier que l'utilisateur est un opérateur de livraison
        operateur = Operateur.objects.get(user=request.user, type_operateur='LOGISTIQUE')
    except Operateur.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Accès non autorisé'})
    
    # Récupérer la commande
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'})
    
    # Récupérer les paniers de la commande
    paniers = commande.paniers.all().select_related('article')
    
    paniers_data = []
    for panier in paniers:
        paniers_data.append({
            'id': panier.id,
            'article_id': panier.article.id,
            'nom': panier.article.nom,
            'reference': panier.article.reference,
            'couleur': panier.article.couleur,
            'pointure': panier.article.pointure,
            'prix_unitaire': float(panier.article.prix_unitaire),
            'quantite': panier.quantite,
            'sous_total': float(panier.sous_total),
            'qte_disponible': panier.article.qte_disponible,
        })
    
    return JsonResponse({
        'success': True,
        'paniers': paniers_data,
        'total_commande': float(commande.total_cmd)
    })

@superviseur_preparation_required
def api_articles_commande_livree_partiellement(request, commande_id):
    """API pour récupérer les détails des articles d'une commande livrée partiellement"""
    import json
    from article.models import Article
    from commande.models import Commande, EtatCommande, EnumEtatCmd, Operation
    from parametre.models import Operateur

    try:
        # Vérifier que l'utilisateur est un opérateur de préparation
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Accès non autorisé'})
    
    # Récupérer la commande
    try:
        commande = Commande.objects.get(id=commande_id)
    except Commande.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Commande non trouvée'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur lors de la récupération de la commande: {str(e)}'})
    
    # Analyser les articles pour les commandes livrées partiellement
    articles_livres = []
    articles_renvoyes = []
    
    # Récupérer tous les états de la commande
    etats_commande = commande.etats.all().select_related('enum_etat', 'operateur').order_by('date_debut')
    
    # Déterminer l'état actuel
    etat_actuel = etats_commande.filter(date_fin__isnull=True).first()
    
    # Récupérer l'état précédent pour comprendre d'où vient la commande
    etat_precedent = None
    if etat_actuel:
        # Trouver l'état précédent (le dernier état terminé avant l'état actuel)
        for etat in reversed(etats_commande):
            if etat.date_fin and etat.date_fin < etat_actuel.date_debut:
                if etat.enum_etat.libelle not in ['À imprimer', 'En préparation']:
                    etat_precedent = etat
                    break
    
    # Vérifier si c'est une commande livrée partiellement
    if etat_actuel and etat_actuel.enum_etat.libelle == 'Livrée Partiellement':
        # Les articles dans cette commande sont ceux qui ont été livrés partiellement
        for panier in commande.paniers.all():
            articles_livres.append({
                'article_id': panier.article.id,
                'nom': panier.article.nom,
                'reference': panier.article.reference,
                'couleur': panier.article.couleur,
                'pointure': panier.article.pointure,
                'quantite_livree': panier.quantite,
                'prix': float(panier.article.prix_unitaire),
                'sous_total': float(panier.sous_total)
            })
        
        # Chercher la commande de renvoi associée
        commande_renvoi = Commande.objects.filter(
            num_cmd__startswith=f"RENVOI-{commande.num_cmd}",
            client=commande.client
        ).first()
        
        if commande_renvoi:
            # Récupérer l'état des articles renvoyés depuis l'opération de livraison partielle
            etat_articles_renvoyes = {}
            operation_livraison_partielle = commande.operations.filter(
                type_operation='LIVRAISON_PARTIELLE'
            ).order_by('-date_operation').first()
            if operation_livraison_partielle:
                try:
                    details = json.loads(operation_livraison_partielle.conclusion)
                    if 'recap_articles_renvoyes' in details:
                        for item in details['recap_articles_renvoyes']:
                            etat_articles_renvoyes[item['article_id']] = item['etat']
                except Exception:
                    pass
            if commande_renvoi:
                for panier_renvoi in commande_renvoi.paniers.all():
                    etat = etat_articles_renvoyes.get(panier_renvoi.article.id, 'bon')
                    articles_renvoyes.append({
                        'article_id': panier_renvoi.article.id,
                        'nom': panier_renvoi.article.nom,
                        'reference': panier_renvoi.article.reference,
                        'couleur': panier_renvoi.article.couleur,
                        'pointure': panier_renvoi.article.pointure,
                        'quantite': panier_renvoi.quantite,
                        'prix': float(panier_renvoi.article.prix_unitaire),
                        'sous_total': float(panier_renvoi.sous_total),
                        'etat': etat
                    })
    # Vérifier si c'est une commande renvoyée après livraison partielle
    elif etat_precedent and etat_precedent.enum_etat.libelle == 'Livrée Partiellement':
        # Chercher la commande originale qui a été livrée partiellement
        commande_originale = Commande.objects.filter(
            num_cmd=commande.num_cmd.replace('RENVOI-', ''),
            client=commande.client
        ).first()
        # Récupérer l'état des articles renvoyés depuis l'opération de livraison partielle
        etat_articles_renvoyes = {}
        if commande_originale:
            operation_livraison_partielle = commande_originale.operations.filter(
                type_operation='LIVRAISON_PARTIELLE'
            ).order_by('-date_operation').first()
            if operation_livraison_partielle:
                try:
                    details = json.loads(operation_livraison_partielle.conclusion)
                    if 'recap_articles_renvoyes' in details:
                        for item in details['recap_articles_renvoyes']:
                            etat_articles_renvoyes[item['article_id']] = item['etat']
                except Exception:
                    pass
        if commande_originale:
            # Les articles dans cette commande de renvoi sont ceux qui ont été renvoyés
            for panier in paniers:
                etat = etat_articles_renvoyes.get(panier.article.id, 'bon')
                articles_renvoyes.append({
                    'article': panier.article,
                    'quantite': panier.quantite,
                    'prix': panier.article.prix_unitaire,
                    'sous_total': panier.sous_total,
                    'etat': etat
                })
    
    # Récupérer les détails de la livraison partielle
    date_livraison_partielle = None
    commentaire_livraison_partielle = None
    operateur_livraison = None
    
    if etat_actuel and etat_actuel.enum_etat.libelle == 'Livrée Partiellement':
        date_livraison_partielle = etat_actuel.date_debut
        commentaire_livraison_partielle = etat_actuel.commentaire
        operateur_livraison = etat_actuel.operateur
    elif etat_precedent and etat_precedent.enum_etat.libelle == 'Livrée Partiellement':
        date_livraison_partielle = etat_precedent.date_debut
        commentaire_livraison_partielle = etat_precedent.commentaire
        operateur_livraison = etat_precedent.operateur
    
    try:
        return JsonResponse({
            'success': True,
            'commande': {
                'id': commande.id,
                'id_yz': commande.id_yz,
                'num_cmd': commande.num_cmd,
                'total_cmd': float(commande.total_cmd),
                'date_livraison_partielle': date_livraison_partielle.isoformat() if date_livraison_partielle else None,
                'commentaire_livraison_partielle': commentaire_livraison_partielle,
                'operateur_livraison': {
                    'nom': operateur_livraison.nom_complet if operateur_livraison else None,
                    'email': operateur_livraison.mail if operateur_livraison else None
                } if operateur_livraison else None
            },
            'articles_livres': articles_livres,
            'articles_renvoyes': articles_renvoyes,
            'total_articles_livres': len(articles_livres),
            'total_articles_renvoyes': len(articles_renvoyes)
        })
    except Exception as e:
        print(f"Erreur lors de la génération de la réponse JSON: {e}")
        return JsonResponse({
            'success': False, 
            'message': f'Erreur lors de la génération de la réponse: {str(e)}'
    })





@superviseur_preparation_required
def export_region_consolidee_csv(request, region_name):
    """
    Export CSV consolidé pour une région spécifique
    """
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    # Récupérer les commandes PRÉPARÉES de la région
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True,
        ville__region__nom_region=region_name
    ).select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'paniers__article'
    ).distinct().order_by('-date_cmd')
    
    # Créer la réponse CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    filename = f"region_{region_name.lower().replace(' ', '_')}_consolidee_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Écrire l'en-tête BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-têtes
    headers = [
        'N° Commande', 'Client', 'Téléphone', 'Ville', 'Région', 
        'Articles et Quantités', 'Prix Total (DH)', 'Adresse', 'État'
    ]
    writer.writerow(headers)
    
    # Traiter chaque commande
    for commande in commandes:
        # Construire la liste des articles avec quantités
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = ", ".join(articles_list) if articles_list else "Aucun article"
        
        # État actuel de la commande
        etat_actuel = commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else "Non défini"
        
        # Écrire la ligne
        row = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}" if commande.client else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region if commande.ville and commande.ville.region else "N/A",
            articles_consolides,
            f"{commande.total_cmd:.2f}" if commande.total_cmd else "0.00",
            commande.adresse or "N/A",
            etat_actuel
        ]
        writer.writerow(row)
    
    return response


@superviseur_preparation_required
def export_region_consolidee_excel(request, region_name):
    """
    Export Excel consolidé pour une région spécifique
    """
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    # Récupérer les commandes PRÉPARÉES de la région
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True,
        ville__region__nom_region=region_name
    ).select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'paniers__article'
    ).distinct().order_by('-date_cmd')
    
    # Créer le fichier Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Région {region_name}"
    
    # En-têtes
    headers = [
        'N° Commande', 'Client', 'Téléphone', 'Ville', 'Région', 
        'Articles et Quantités', 'Prix Total (DH)', 'Adresse', 'État'
    ]
    
    # Ajouter les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Traiter chaque commande
    for row, commande in enumerate(commandes, 2):
        # Construire la liste des articles avec quantités
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = ", ".join(articles_list) if articles_list else "Aucun article"
        
        # État actuel de la commande
        etat_actuel = commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else "Non défini"
        
        # Ajouter les données
        ws.cell(row=row, column=1, value=commande.id_yz or commande.num_cmd)
        ws.cell(row=row, column=2, value=f"{commande.client.prenom} {commande.client.nom}" if commande.client else "N/A")
        ws.cell(row=row, column=3, value=commande.client.numero_tel if commande.client else "N/A")
        ws.cell(row=row, column=4, value=commande.ville.nom if commande.ville else "N/A")
        ws.cell(row=row, column=5, value=commande.ville.region.nom_region if commande.ville and commande.ville.region else "N/A")
        ws.cell(row=row, column=6, value=articles_consolides)
        ws.cell(row=row, column=7, value=float(commande.total_cmd) if commande.total_cmd else 0.00)
        ws.cell(row=row, column=8, value=commande.adresse or "N/A")
        ws.cell(row=row, column=9, value=etat_actuel)
        
        # Ajuster la hauteur de la ligne pour les articles
        if len(articles_consolides) > 100:
            ws.row_dimensions[row].height = 30
    
    # Ajuster la largeur des colonnes
    column_widths = [15, 25, 15, 15, 15, 50, 15, 40, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Ajouter une feuille de résumé
    ws_resume = wb.create_sheet("Résumé")
    
    # Statistiques de la région
    total_commandes = commandes.count()
    total_montant = sum(float(cmd.total_cmd) for cmd in commandes if cmd.total_cmd)
    
    resume_data = [
        ['Région', region_name],
        ['Nombre de commandes', total_commandes],
        ['Montant total', f"{total_montant:.2f} DH"],
        ['Date d\'export', timezone.now().strftime('%d/%m/%Y %H:%M')]
    ]
    
    for row, (label, value) in enumerate(resume_data, 1):
        ws_resume.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_resume.cell(row=row, column=2, value=value)
    
    # Ajuster la largeur des colonnes du résumé
    ws_resume.column_dimensions['A'].width = 25
    ws_resume.column_dimensions['B'].width = 20
    
    # Créer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"region_{region_name.lower().replace(' ', '_')}_consolidee_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@superviseur_preparation_required
def export_ville_consolidee_csv(request, ville_id):
    """
    Export CSV consolidé pour une ville spécifique
    """
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Accès non autorisé'}, status=403)
    
    # Récupérer la ville
    try:
        ville = Ville.objects.get(id=ville_id)
    except Ville.DoesNotExist:
        return JsonResponse({'error': 'Ville non trouvée'}, status=404)
    
    # Récupérer les commandes PRÉPARÉES de la ville
    commandes = Commande.objects.filter(
        etats__enum_etat__libelle='Préparée',
        etats__date_fin__isnull=True,
        ville=ville
    ).select_related(
        'client', 
        'ville', 
        'ville__region'
    ).prefetch_related(
        'paniers__article'
    ).distinct().order_by('-date_cmd')
    
    # Créer la réponse CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="ville_{ville.nom}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Écrire l'en-tête BOM pour Excel
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    
    # En-têtes
    writer.writerow([
        'N° Commande', 'Client', 'Téléphone', 'Ville', 'Région', 
        'Articles et Quantités', 'Prix Total (DH)', 'Adresse', 'État'
    ])
    
    # Traiter chaque commande
    for commande in commandes:
        # Construire la liste des articles avec quantités
        articles_list = []
        for panier in commande.paniers.all():
            article_info = f"{panier.article.nom}"
            if panier.article.couleur:
                article_info += f" {panier.article.couleur}"
            if panier.article.pointure:
                article_info += f" {panier.article.pointure}"
            if panier.quantite > 1:
                article_info += f" x{panier.quantite}"
            articles_list.append(article_info)
        
        # Joindre tous les articles avec des virgules
        articles_consolides = ", ".join(articles_list) if articles_list else "Aucun article"
        
        # État actuel de la commande
        etat_actuel = commande.etat_actuel.enum_etat.libelle if commande.etat_actuel else "Non défini"
        
        # Écrire la ligne
        row = [
            commande.id_yz or commande.num_cmd,
            f"{commande.client.prenom} {commande.client.nom}" if commande.client else "N/A",
            commande.client.numero_tel if commande.client else "N/A",
            commande.ville.nom if commande.ville else "N/A",
            commande.ville.region.nom_region if commande.ville and commande.ville.region else "N/A",
            articles_consolides,
            float(commande.total_cmd) if commande.total_cmd else 0.00,
            commande.adresse or "N/A",
            etat_actuel
        ]
        writer.writerow(row)
    
    return response


@superviseur_preparation_required
def export_ville_consolidee_excel(request, ville_id):
    """
    Export Excel consolidé pour une ville spécifique
    """
    try:
        ville = get_object_or_404(Ville, id=ville_id)
        
        # Récupérer toutes les commandes de cette ville
        commandes = Commande.objects.filter(
            ville=ville,
            etat_actuel__enum_etat__libelle__in=['Confirmée', 'En préparation', 'Prête', 'Livrée']
        ).select_related(
            'client', 'ville', 'ville__region', 'etat_actuel', 'etat_actuel__enum_etat'
        ).prefetch_related('paniers__article').order_by('date_cmd')
        
        if not commandes.exists():
            messages.warning(request, f"Aucune commande trouvée pour la ville {ville.nom}")
            return redirect('Superpreparation:liste_prepa')
    
        # Créer le fichier Excel
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
    
        wb = Workbook()
        ws = wb.active
        ws.title = f"Commandes {ville.nom}"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
        # En-têtes
        headers = [
            'ID YZ', 'Numéro', 'Date', 'Client', 'Téléphone', 'Adresse', 
            'Ville', 'Région', 'État', 'Total Articles', 'Frais Livraison', 
            'Total Commande', 'Compteur Upsell', 'Articles'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Données
        row = 2
        for commande in commandes:
            # Articles détaillés
            articles_detail = []
            for panier in commande.paniers.all():
                article_info = f"{panier.article.nom} (Qté: {panier.quantite}, Prix: {panier.sous_total:.2f} DH)"
                if panier.article.isUpsell:
                    article_info += " [UPSELL]"
                articles_detail.append(article_info)
            
            articles_text = "\n".join(articles_detail)
            
            ws.cell(row=row, column=1, value=commande.id_yz).border = border
            ws.cell(row=row, column=2, value=commande.num_cmd).border = border
            ws.cell(row=row, column=3, value=commande.date_cmd.strftime('%d/%m/%Y')).border = border
            ws.cell(row=row, column=4, value=f"{commande.client.nom} {commande.client.prenom}").border = border
            ws.cell(row=row, column=5, value=commande.client.numero_tel).border = border
            ws.cell(row=row, column=6, value=commande.adresse).border = border
            ws.cell(row=row, column=7, value=commande.ville.nom).border = border
            ws.cell(row=row, column=8, value=commande.ville.region.nom_region).border = border
            ws.cell(row=row, column=9, value=commande.etat_actuel.enum_etat.libelle).border = border
            ws.cell(row=row, column=10, value=float(commande.sous_total_articles)).border = border
            ws.cell(row=row, column=11, value=float(commande.ville.frais_livraison or 0)).border = border
            ws.cell(row=row, column=12, value=float(commande.total_cmd)).border = border
            ws.cell(row=row, column=13, value=commande.compteur).border = border
            ws.cell(row=row, column=14, value=articles_text).border = border
            
            # Ajuster la hauteur de ligne pour les articles
            ws.row_dimensions[row].height = max(20, len(articles_detail) * 15)
            
            row += 1
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Largeur spéciale pour la colonne articles
        ws.column_dimensions['N'].width = 40
        
        # Nom du fichier
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"commandes_consolidees_{ville.nom}_{timestamp}.xlsx"
        
        # Réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de l'export: {str(e)}")
        return redirect('Superpreparation:liste_prepa')

@superviseur_preparation_required
def api_prix_upsell_articles(request, commande_id):
    """API pour récupérer les prix upsell mis à jour des articles"""
    print(f"🔄 Récupération des prix upsell pour la commande {commande_id}")
    
    try:
        operateur = Operateur.objects.get(user=request.user, type_operateur='PREPARATION')
    except Operateur.DoesNotExist:
        return JsonResponse({'error': 'Profil d\'opérateur de préparation non trouvé.'}, status=403)
    
    try:
        commande = Commande.objects.get(id=commande_id)
        
        # Vérifier que la commande est affectée à cet opérateur
        etat_preparation = commande.etats.filter(
            operateur=operateur,
            enum_etat__libelle__in=['En préparation', 'À imprimer'],
            date_fin__isnull=True
        ).first()
        
        if not etat_preparation:
            return JsonResponse({'error': 'Cette commande ne vous est pas affectée.'}, status=403)
        
        # Récupérer tous les articles du panier avec leurs prix mis à jour
        paniers = commande.paniers.select_related('article').all()
        prix_articles = {}
        
        for panier in paniers:
            # Utiliser le filtre Django pour calculer le prix selon la logique upsell
            from commande.templatetags.commande_filters import get_prix_upsell_avec_compteur
            
            # Debug: Afficher les informations de l'article
            print(f"🔍 Article {panier.article.id} ({panier.article.nom}):")
            print(f"   - isUpsell: {panier.article.isUpsell}")
            print(f"   - prix_actuel: {panier.article.prix_actuel}")
            print(f"   - prix_unitaire: {panier.article.prix_unitaire}")
            print(f"   - prix_upsell_1: {panier.article.prix_upsell_1}")
            print(f"   - prix_upsell_2: {panier.article.prix_upsell_2}")
            print(f"   - prix_upsell_3: {panier.article.prix_upsell_3}")
            print(f"   - prix_upsell_4: {panier.article.prix_upsell_4}")
            print(f"   - compteur: {commande.compteur}")
            
            prix_calcule = get_prix_upsell_avec_compteur(panier.article, commande.compteur)
            print(f"   - prix_calcule: {prix_calcule}")
            
            # Convertir en float pour éviter les problèmes de sérialisation JSON
            prix_calcule = float(prix_calcule) if prix_calcule is not None else 0.0
            
            # Déterminer le type de prix et les informations
            if panier.article.isUpsell:
                prix_type = "upsell"
                
                # Déterminer le niveau d'upsell affiché
                if commande.compteur >= 4 and panier.article.prix_upsell_4 is not None:
                    niveau_upsell = 4
                elif commande.compteur >= 3 and panier.article.prix_upsell_3 is not None:
                    niveau_upsell = 3
                elif commande.compteur >= 2 and panier.article.prix_upsell_2 is not None:
                    niveau_upsell = 2
                elif commande.compteur >= 1 and panier.article.prix_upsell_1 is not None:
                    niveau_upsell = 1
                else:
                    niveau_upsell = 0
                
                if niveau_upsell > 0:
                    libelle = f"Upsell Niveau {niveau_upsell}"
                else:
                    libelle = "Upsell (Prix normal)"
                    
                icone = "fas fa-arrow-up"
            else:
                prix_type = "normal"
                libelle = "Prix normal"
                icone = "fas fa-tag"
                niveau_upsell = 0
            
            prix_articles[panier.id] = {
                'prix': prix_calcule,
                'type': prix_type,
                'libelle': libelle,
                'icone': icone,
                'niveau_upsell': niveau_upsell,
                'is_upsell': panier.article.isUpsell,
                'sous_total': float(prix_calcule * panier.quantite)
            }
        
        return JsonResponse({
            'success': True,
            'prix_articles': prix_articles,
            'compteur': commande.compteur,
            'message': f'Prix mis à jour pour {len(prix_articles)} articles'
        })
        
    except Commande.DoesNotExist:
        return JsonResponse({'error': 'Commande non trouvée.'}, status=404)
    except Exception as e:
        print(f"❌ Erreur lors de la récupération des prix upsell: {e}")
        return JsonResponse({'error': f'Erreur serveur: {str(e)}'}, status=500)



@superviseur_preparation_required
def commandes_confirmees(request):
    """Page des commandes confirmées"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Récupérer toutes les commandes confirmées
    commandes_confirmees = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée'
    ).select_related('client', 'ville', 'ville__region').prefetch_related('etats', 'operations').distinct()
    
    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        commandes_confirmees = commandes_confirmees.filter(
            Q(id_yz__icontains=search_query) |
            Q(num_cmd__icontains=search_query) |
            Q(client__nom__icontains=search_query) |
            Q(client__prenom__icontains=search_query) |
            Q(client__numero_tel__icontains=search_query) |
            Q(etats__operateur__nom__icontains=search_query) |
            Q(etats__operateur__prenom__icontains=search_query)
        ).distinct()
    
    # Tri par date de confirmation (plus récentes en premier)
    commandes_confirmees = commandes_confirmees.order_by('-etats__date_debut')
    
    # Créer une copie des données non paginées pour les statistiques AVANT la pagination
    commandes_non_paginees = commandes_confirmees
    
    # Paramètres de pagination flexible
    items_per_page = request.GET.get('items_per_page', 25)
    start_range = request.GET.get('start_range')
    end_range = request.GET.get('end_range')
    
    # Gestion de la pagination flexible
    if start_range and end_range:
        try:
            start_range = int(start_range)
            end_range = int(end_range)
            if start_range > 0 and end_range >= start_range:
                # Pagination par plage personnalisée
                commandes_confirmees = commandes_confirmees[start_range-1:end_range]
                paginator = Paginator(commandes_confirmees, end_range - start_range + 1)
                page_obj = paginator.get_page(1)
            else:
                # Plage invalide, utiliser la pagination normale
                items_per_page = 25
                paginator = Paginator(commandes_confirmees, items_per_page)
                page_number = request.GET.get('page', 1)
                page_obj = paginator.get_page(page_number)
        except (ValueError, TypeError):
            # Erreur de conversion, utiliser la pagination normale
            items_per_page = 25
            paginator = Paginator(commandes_confirmees, items_per_page)
            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)
    else:
        # Pagination normale
        if items_per_page == 'all':
            # Afficher toutes les commandes
            paginator = Paginator(commandes_confirmees, commandes_confirmees.count())
            page_obj = paginator.get_page(1)
        else:
            try:
                items_per_page = int(items_per_page)
                if items_per_page <= 0:
                    items_per_page = 25
            except (ValueError, TypeError):
                items_per_page = 25
            
            paginator = Paginator(commandes_confirmees, items_per_page)
            page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Statistiques
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    
    # Confirmées aujourd'hui
    confirmees_aujourd_hui = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée',
        etats__date_debut__date=today
    ).distinct().count()
    
    # Confirmées cette semaine
    confirmees_semaine = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée',
        etats__date_debut__date__gte=week_start
    ).distinct().count()
    
    # Confirmées ce mois
    confirmees_mois = Commande.objects.filter(
        etats__enum_etat__libelle='Confirmée',
        etats__date_debut__date__gte=month_start
    ).distinct().count()
    
    # Total des commandes confirmées (utiliser les données non paginées)
    total_confirmees = commandes_non_paginees.count()
    
    # Montant total des commandes confirmées (utiliser les données non paginées)
    montant_total = commandes_non_paginees.aggregate(total=Sum('total_cmd'))['total'] or 0
    
    # Ajouter l'état de confirmation à chaque commande
    for commande in page_obj:
        etats_commande = list(commande.etats.all().order_by('date_debut'))
        
        # Trouver l'état de confirmation (le premier état "Confirmée")
        etat_confirmation = None
        for etat in etats_commande:
            if etat.enum_etat.libelle == 'Confirmée':
                etat_confirmation = etat
                break
        
        commande.etat_confirmation = etat_confirmation
    
    # Récupérer les opérateurs de préparation actifs
    operateurs_preparation = Operateur.objects.filter(
        type_operateur='PREPARATION',
        actif=True
    ).order_by('nom', 'prenom')
    
    # Vérifier si c'est une requête AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Rendre les templates partiels pour AJAX
        html_table_body = render_to_string('Superpreparation/partials/_confirmees_table_body.html', {
            'page_obj': page_obj
        }, request=request)
        
        html_pagination = render_to_string('Superpreparation/partials/_confirmees_pagination.html', {
            'page_obj': page_obj,
            'search_query': search_query,
            'items_per_page': items_per_page,
            'start_range': start_range,
            'end_range': end_range
        }, request=request)
        
        html_pagination_info = render_to_string('Superpreparation/partials/_confirmees_pagination_info.html', {
            'page_obj': page_obj
        }, request=request)
        
        return JsonResponse({
            'success': True,
            'html_table_body': html_table_body,
            'html_pagination': html_pagination,
            'html_pagination_info': html_pagination_info,
            'total_count': commandes_confirmees.count()
        })
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_confirmees': total_confirmees,
        'confirmees_aujourd_hui': confirmees_aujourd_hui,
        'confirmees_semaine': confirmees_semaine,
        'confirmees_mois': confirmees_mois,
        'montant_total': montant_total,
        'operateurs_preparation': operateurs_preparation,
        'page_title': 'Suivi des Commandes Confirmées',
        'page_subtitle': 'Suivi et gestion des commandes validées prêtes pour la préparation',
        'items_per_page': items_per_page,
        'start_range': start_range,
        'end_range': end_range,
    }
    
    return render(request, 'Superpreparation/commande_confirmees.html', context)

@login_required
def get_article_variants(request, article_id):
    """
    Récupère toutes les variantes disponibles d'un article donné
    """
    try:
        from article.models import Article, VarianteArticle
        
        print(f"🔍 Recherche des variantes pour l'article ID: {article_id} (type: {type(article_id)})")
        
        # D'abord, vérifier si l'article existe (avec ou sans le filtre actif=True)
        try:
            article_test = Article.objects.get(id=article_id)
            print(f"📋 Article trouvé mais actif={article_test.actif}")
        except Article.DoesNotExist:
            print(f"❌ Aucun article trouvé avec l'ID {article_id}")
            # Lister quelques articles pour debug
            articles_existants = Article.objects.all()[:5]
            print("📚 Articles existants (premiers 5):")
            for art in articles_existants:
                print(f"   - ID: {art.id}, Nom: {art.nom}, Actif: {art.actif}")
        
        # Vérifier que l'article existe ET est actif
        article = get_object_or_404(Article, id=article_id, actif=True)
        print(f"✅ Article trouvé: {article.nom}")
        
        # Récupérer toutes les variantes actives de cet article
        variantes = VarianteArticle.objects.filter(
            article=article,
            actif=True
        ).select_related('couleur', 'pointure').order_by('couleur__nom', 'pointure__ordre')
        
        print(f"📊 Nombre de variantes trouvées: {variantes.count()}")
        
        # Construire la liste des variantes avec leurs informations
        variants_data = []
        for variante in variantes:
            print(f"🔸 Variante: {variante.id} - Couleur: {variante.couleur} - Pointure: {variante.pointure} - Stock: {variante.qte_disponible}")
            
            variant_info = {
                'id': variante.id,
                'couleur': variante.couleur.nom if variante.couleur else None,
                'pointure': variante.pointure.pointure if variante.pointure else None,
                'taille': None,  # À adapter selon votre modèle si vous avez des tailles
                'stock': variante.qte_disponible,
                'qte_disponible': variante.qte_disponible,
                'prix_unitaire': float(variante.prix_unitaire),
                'prix_actuel': float(variante.prix_actuel),
                'reference_variante': variante.reference_variante,
                'est_disponible': variante.est_disponible
            }
            variants_data.append(variant_info)
        
        response_data = {
            'success': True,
            'variants': variants_data,
            'article': {
                'id': article.id,
                'nom': article.nom,
                'reference': article.reference
            }
        }
        
        print(f"📤 Réponse envoyée: {len(variants_data)} variantes")
        return JsonResponse(response_data)
        
    except Article.DoesNotExist:
        print(f"❌ Article {article_id} non trouvé")
        return JsonResponse({
            'success': False,
            'error': 'Article non trouvé'
        }, status=404)
        
    except Exception as e:
        import traceback
        print(f"❌ Erreur dans get_article_variants: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': 'Erreur lors de la récupération des variantes'
        }, status=500)

@superviseur_preparation_required
def api_codes_barres_commandes(request):
    """
    API pour récupérer le contenu HTML des codes-barres des commandes
    """
    try:
        ids = request.GET.get('ids')
        if not ids:
            return JsonResponse({'error': 'IDs des commandes requis'}, status=400)
        
        print(f"🔍 Recherche de la commande avec id_yz: {ids}")
        
        # Récupérer la commande confirmée avec des articles
        try:
            commande = Commande.objects.filter(
                id_yz=ids,
                etats__enum_etat__libelle='Confirmée',
                paniers__isnull=False
            ).distinct().first()
            
            if not commande:
                print(f"❌ Commande {ids} non trouvée ou non confirmée ou sans articles")
                return JsonResponse({'error': f'Commande {ids} non confirmée ou sans articles'}, status=404)
            
            print(f"✅ Commande confirmée trouvée: {commande.id_yz}")
        except Exception as e:
            print(f"❌ Erreur lors de la recherche de la commande {ids}: {str(e)}")
            return JsonResponse({'error': f'Erreur lors de la recherche de la commande {ids}'}, status=500)
        
        # Générer le code-barres
        try:
            barcode_data = f"YZ{commande.id_yz}"
            print(f"📊 Génération du code-barres: {barcode_data}")
            barcode_image = barcode.Code128(barcode_data, writer=ImageWriter())
            buffer = BytesIO()
            barcode_image.write(buffer)
            barcode_base64 = base64.b64encode(buffer.getvalue()).decode()
            print(f"✅ Code-barres généré avec succès")
        except Exception as barcode_error:
            print(f"❌ Erreur lors de la génération du code-barres: {str(barcode_error)}")
            barcode_base64 = ""
        
        # Préparer les données de la commande
        commande_data = {
            'id_yz': commande.id_yz,
            'num_cmd': commande.num_cmd,
            'client_nom': f"{commande.client.nom} {commande.client.prenom}" if commande.client else "Client non défini",
            'client_telephone': commande.client.numero_tel if commande.client else "",
            'ville_client': commande.ville_init or (commande.ville.nom if commande.ville else ""),
            'ville_region': f"{commande.ville.nom} {commande.ville.region.nom_region}" if commande.ville and commande.ville.region else "",
            'adresse': commande.client.adresse if commande.client else "",
            'total': commande.total_cmd,
            'barcode_image': barcode_base64
        }
        
        # Rendre le template HTML
        html_content = render_to_string('Superpreparation/partials/_codes_barres_commande.html', {
            'commande': commande_data
        })
        
        return JsonResponse({
            'success': True,
            'html': html_content,
            'commande': commande_data
        })
        
    except Exception as e:
        import traceback
        print(f"Erreur dans api_codes_barres_commandes: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération des codes-barres: {str(e)}'
        }, status=500)

@superviseur_preparation_required
def api_ticket_commande(request):
    """
    API pour récupérer le contenu HTML du ticket de commande
    """
    try:
        ids = request.GET.get('ids')
        if not ids:
            return JsonResponse({'error': 'IDs des commandes requis'}, status=400)
        
        # Supporter plusieurs IDs séparés par des virgules
        id_list = [id.strip() for id in ids.split(',') if id.strip()]
        
        print(f"🔍 Recherche des commandes avec id_yz: {id_list} pour ticket")
        
        all_tickets_html = []
        commandes_data = []
        
        for commande_id in id_list:
            try:
                commande = Commande.objects.filter(
                    id_yz=commande_id,
                    etats__enum_etat__libelle='Confirmée',
                    paniers__isnull=False
                ).distinct().first()
                
                if not commande:
                    print(f"❌ Commande {commande_id} non trouvée ou non confirmée ou sans articles")
                    continue
                
                print(f"✅ Commande confirmée trouvée: {commande.id_yz}")
            except Exception as e:
                print(f"❌ Erreur lors de la recherche de la commande {commande_id}: {str(e)}")
                continue
            
            # Préparer la description des articles
            articles_description = ""
            total_articles = 0
            paniers = commande.paniers.all().select_related('article', 'variante')
            for panier in paniers:
                article = panier.article
                variante = panier.variante
                total_articles += panier.quantite  # Ajouter la quantité au total
                if articles_description:
                    articles_description += " + "
                articles_description += f"{article.nom} {variante.couleur if variante else ''} , P{panier.quantite}"
            
            # Préparer les données de la commande
            commande_data = {
                'id_yz': commande.id_yz,
                'num_cmd': commande.num_cmd,
                'client_nom': f"{commande.client.nom} {commande.client.prenom}" if commande.client else "Client non défini",
                'client_telephone': commande.client.numero_tel if commande.client else "",
                'ville_client': commande.ville_init or (commande.ville.nom if commande.ville else ""),
                'adresse': commande.client.adresse if commande.client else "",
                'total': commande.total_cmd,
                'date_confirmation': commande.date_creation,
                'articles_description': articles_description,
                'total_articles': total_articles
            }
            
            # Rendre le template HTML pour cette commande
            ticket_html = render_to_string('Superpreparation/partials/_ticket_commande.html', {
                'commande': commande_data
            })
            
            all_tickets_html.append(ticket_html)
            commandes_data.append(commande_data)
        
        if not all_tickets_html:
            return JsonResponse({'error': 'Aucune commande valide trouvée'}, status=404)
        
        # Combiner tous les tickets avec un conteneur de grille
        combined_html = f'''
        <div class="ticket-commande-container" style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; padding: 15px; max-width: 100%; width: 100%; -webkit-print-color-adjust: exact !important; color-adjust: exact !important; print-color-adjust: exact !important;">
            {''.join(all_tickets_html)}
        </div>
        '''
        
        return JsonResponse({
            'success': True,
            'html': combined_html,
            'commandes': commandes_data
        })
        
    except Exception as e:
        import traceback
        print(f"Erreur dans api_ticket_commande: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération du ticket: {str(e)}'
        }, status=500)

@superviseur_preparation_required
def api_etiquettes_articles(request):
    """
    API pour récupérer le contenu HTML des étiquettes des articles
    """
    try:
        ids = request.GET.get('ids')
        format_type = request.GET.get('format', 'qr')  # 'qr' ou 'barcode'
        
        if not ids:
            return JsonResponse({'error': 'IDs des commandes requis'}, status=400)
        
        print(f"🔍 Recherche de la commande avec id_yz: {ids} pour format: {format_type}")
        
        # Récupérer la commande confirmée avec des articles
        try:
            commande = Commande.objects.filter(
                id_yz=ids,
                etats__enum_etat__libelle='Confirmée',
                paniers__isnull=False
            ).distinct().first()
            
            if not commande:
                print(f"❌ Commande {ids} non trouvée ou non confirmée ou sans articles")
                return JsonResponse({'error': f'Commande {ids} non confirmée ou sans articles'}, status=404)
            
            print(f"✅ Commande confirmée trouvée: {commande.id_yz}")
        except Exception as e:
            print(f"❌ Erreur lors de la recherche de la commande {ids}: {str(e)}")
            return JsonResponse({'error': f'Erreur lors de la recherche de la commande {ids}'}, status=500)
        articles = commande.paniers.all().select_related('article', 'variante')
        
        articles_data = []
        for panier in articles:
            article = panier.article
            variante = panier.variante
            
            # Générer le QR code pour l'article
            try:
                import qrcode
                qr_data = f"ART{article.id}_{variante.id if variante else 'NO_VAR'}"
                qr = qrcode.QRCode(version=1, box_size=10, border=5)
                qr.add_data(qr_data)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="black", back_color="white")
                
                # Convertir en base64
                buffer = BytesIO()
                qr_img.save(buffer, format='PNG')
                qr_base64 = base64.b64encode(buffer.getvalue()).decode()
            except ImportError:
                # Fallback si qrcode n'est pas installé
                qr_base64 = ""
                print("Warning: qrcode library not installed, QR codes will not be generated")
            
            # Générer le code-barres pour l'article
            barcode_data = f"ART{article.reference}_{variante.reference_variante if variante else 'NO_VAR'}"
            barcode_image = barcode.Code128(barcode_data, writer=ImageWriter())
            buffer = BytesIO()
            barcode_image.write(buffer)
            barcode_base64 = base64.b64encode(buffer.getvalue()).decode()
            
            # Calculer le prix unitaire à partir du sous_total et de la quantité
            prix_unitaire = panier.sous_total / panier.quantite if panier.quantite > 0 else 0
            
            article_data = {
                'id': article.id,
                'nom': article.nom,
                'reference': article.reference,
                'prix': float(prix_unitaire),
                'quantite': panier.quantite,
                'couleur': variante.couleur.nom if variante and variante.couleur else "Standard",
                'pointure': variante.pointure.pointure if variante and variante.pointure else "Standard",
                'qr_image': qr_base64,
                'barcode_image': barcode_base64,
                'commande_id': commande.id_yz
            }
            articles_data.append(article_data)
        
        # Rendre le template HTML avec le format spécifié
        html_content = render_to_string('Superpreparation/partials/_etiquettes_articles.html', {
            'articles': articles_data,
            'commande': {
                'id_yz': commande.id_yz,
                'client_nom': f"{commande.client.nom} {commande.client.prenom}" if commande.client else "Client non défini"
            },
            'format_type': format_type
        })
        
        return JsonResponse({
            'success': True,
            'html': html_content,
            'articles': articles_data,
            'format_type': format_type
        })
        
    except Exception as e:
        import traceback
        print(f"Erreur dans api_etiquettes_articles: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération des étiquettes: {str(e)}'
        }, status=500)

@superviseur_preparation_required
def api_etiquettes_articles_multiple(request):
    """
    API pour récupérer le contenu HTML des étiquettes de tous les articles de toutes les commandes confirmées
    Supporte la sélection de commandes spécifiques ou toutes les commandes
    """
    try:
        format_type = request.GET.get('format', 'qr')  # 'qr' ou 'barcode'
        
        # Récupérer les IDs des commandes sélectionnées
        selected_ids = request.GET.get('selected_ids', '')
        
        print(f"🔍 Récupération des articles pour étiquettes multiples, format: {format_type}")
        
        # Construire le filtre de base
        base_filter = {
            'etats__enum_etat__libelle': 'Confirmée',
            'paniers__isnull': False
        }
        
        # Si des IDs sont spécifiés, filtrer par ces IDs
        if selected_ids:
            try:
                # Convertir la chaîne d'IDs en liste d'entiers
                commande_ids = [int(id.strip()) for id in selected_ids.split(',') if id.strip().isdigit()]
                if commande_ids:
                    base_filter['id__in'] = commande_ids
                    print(f"🔍 Filtrage par {len(commande_ids)} commandes sélectionnées: {commande_ids}")
                else:
                    print("⚠️ Aucun ID valide trouvé dans selected_ids")
            except (ValueError, AttributeError) as e:
                print(f"❌ Erreur lors du parsing des IDs: {e}")
                # En cas d'erreur, continuer avec toutes les commandes
        else:
            print("🔍 Aucune sélection spécifique - impression de tous les articles des commandes confirmées")
        
        # Récupérer les commandes selon le filtre
        commandes = Commande.objects.filter(**base_filter).distinct().prefetch_related('paniers__article', 'paniers__variante__couleur', 'paniers__variante__pointure')
        
        print(f"📊 Nombre de commandes confirmées avec paniers: {commandes.count()}")
        
        all_articles_data = []
        
        for commande in commandes:
            print(f"📋 Traitement de la commande: {commande.id_yz}")
            articles = commande.paniers.all()
            print(f"  📦 Nombre de paniers: {articles.count()}")
            
            if articles.count() == 0:
                print(f"  ⚠️ Aucun panier trouvé pour la commande {commande.id_yz}")
                continue
            
            for panier in articles:
                article = panier.article
                variante = panier.variante
                
                # Générer le QR code pour l'article
                try:
                    import qrcode
                    qr_data = f"ART{article.id}_{variante.id if variante else 'NO_VAR'}"
                    qr = qrcode.QRCode(version=1, box_size=10, border=5)
                    qr.add_data(qr_data)
                    qr.make(fit=True)
                    qr_img = qr.make_image(fill_color="black", back_color="white")
                    
                    # Convertir en base64
                    buffer = BytesIO()
                    qr_img.save(buffer, format='PNG')
                    qr_base64 = base64.b64encode(buffer.getvalue()).decode()
                except ImportError:
                    # Fallback si qrcode n'est pas installé
                    qr_base64 = ""
                    print("Warning: qrcode library not installed, QR codes will not be generated")
                
                # Générer le code-barres pour l'article
                barcode_data = f"ART{article.reference}_{variante.reference_variante if variante else 'NO_VAR'}"
                barcode_image = barcode.Code128(barcode_data, writer=ImageWriter())
                buffer = BytesIO()
                barcode_image.write(buffer)
                barcode_base64 = base64.b64encode(buffer.getvalue()).decode()
                
                # Calculer le prix unitaire à partir du sous_total et de la quantité
                prix_unitaire = panier.sous_total / panier.quantite if panier.quantite > 0 else 0
                
                article_data = {
                    'id': article.id,
                    'nom': article.nom,
                    'reference': article.reference,
                    'prix': float(prix_unitaire),
                    'quantite': panier.quantite,
                    'couleur': variante.couleur.nom if variante and variante.couleur else "Standard",
                    'pointure': variante.pointure.pointure if variante and variante.pointure else "Standard",
                    'qr_image': qr_base64,
                    'barcode_image': barcode_base64,
                    'commande_id': commande.id_yz
                }
                all_articles_data.append(article_data)
        
        print(f"📦 Nombre total d'articles: {len(all_articles_data)}")
        
        if not all_articles_data:
            return JsonResponse({'error': 'Aucun article trouvé dans les commandes confirmées'}, status=404)
        
        # Rendre le template HTML avec le format spécifié
        html_content = render_to_string('Superpreparation/partials/_etiquettes_articles.html', {
            'articles': all_articles_data,
            'commande': {
                'id_yz': 'MULTIPLE',
                'client_nom': 'Toutes les commandes confirmées'
            },
            'format_type': format_type
        })
        
        return JsonResponse({
            'success': True,
            'html': html_content,
            'articles': all_articles_data,
            'format_type': format_type,
            'total_articles': len(all_articles_data)
        })
        
    except Exception as e:
        import traceback
        print(f"Erreur dans api_etiquettes_articles_multiple: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération des étiquettes multiples: {str(e)}'
        }, status=500)

@superviseur_preparation_required
def api_ticket_commande_multiple(request):
    """
    API pour récupérer le contenu HTML des tickets de commande multiples
    Supporte la sélection de commandes spécifiques ou toutes les commandes
    """
    print("🚀 === API TICKET COMMANDE MULTIPLE DÉMARRÉE ===")
    print(f"📡 Méthode HTTP: {request.method}")
    print(f"📡 URL: {request.path}")
    print(f"📡 Paramètres GET: {dict(request.GET)}")
    
    try:
        # Vérifier si c'est pour l'impression directe
        direct_print = request.GET.get('direct_print', 'false').lower() == 'true'
        print(f"🖨️ Impression directe: {direct_print}")
        
        # Récupérer les IDs des commandes sélectionnées
        selected_ids = request.GET.get('selected_ids', '')
        print(f"📋 Paramètre selected_ids reçu: '{selected_ids}' (type: {type(selected_ids)})")
        
        # Construire le filtre de base
        base_filter = {
            'etats__enum_etat__libelle': 'Confirmée',
            'paniers__isnull': False
        }
        print(f"🔧 Filtre de base: {base_filter}")
        
        # Si des IDs sont spécifiés, filtrer par ces IDs
        if selected_ids:
            print(f"🔍 Traitement de la sélection: '{selected_ids}'")
            try:
                # Convertir la chaîne d'IDs en liste d'entiers
                commande_ids = [int(id.strip()) for id in selected_ids.split(',') if id.strip().isdigit()]
                print(f"🔍 IDs parsés: {commande_ids}")
                
                if commande_ids:
                    base_filter['id__in'] = commande_ids
                    print(f"✅ Filtrage par {len(commande_ids)} commandes sélectionnées: {commande_ids}")
                else:
                    print("⚠️ Aucun ID valide trouvé dans selected_ids")
            except (ValueError, AttributeError) as e:
                print(f"❌ Erreur lors du parsing des IDs: {e}")
                print(f"❌ Traceback: {traceback.format_exc()}")
                # En cas d'erreur, continuer avec toutes les commandes
        else:
            print("🔍 Aucune sélection spécifique - impression de toutes les commandes confirmées")
        
        print(f"🔍 Filtre final appliqué: {base_filter}")
        
        # Récupérer les commandes selon le filtre
        print("🔍 Exécution de la requête de base de données...")
        commandes = Commande.objects.filter(**base_filter).distinct().select_related('client', 'ville').prefetch_related(
            'paniers__article', 
            'paniers__variante__couleur', 
            'paniers__variante__pointure'
        )
        
        print(f"🔍 Récupération des commandes confirmées avec paniers pour tickets multiples")
        
        # Log pour debug - voir combien de commandes confirmées
        print(f"📊 Nombre de commandes confirmées avec paniers: {commandes.count()}")
        
        # Log détaillé des commandes trouvées
        print("🔍 Détail des commandes trouvées:")
        for cmd in commandes:
            print(f"  📋 Commande: ID={cmd.id}, ID_YZ={cmd.id_yz}, Client={cmd.client.nom if cmd.client else 'N/A'}")
        
        all_tickets_html = []
        commandes_data = []
        
        for commande in commandes:
            print(f"✅ Traitement de la commande: {commande.id_yz}")
            
            # Préparer la description des articles avec les vraies variantes
            articles_description = ""
            total_articles = 0
            # Utiliser les paniers déjà préchargés
            paniers = commande.paniers.all()
            print(f"📦 Nombre de paniers pour {commande.id_yz}: {paniers.count()}")
            
            if paniers.count() == 0:
                print(f"  ⚠️ Aucun panier trouvé pour la commande {commande.id_yz}")
                articles_description = "Aucun article"
                total_articles = 0
            else:
                for panier in paniers:
                    article = panier.article
                    variante = panier.variante
                    total_articles += panier.quantite  # Ajouter la quantité au total
                    
                    print(f"  📋 Article: {article.nom}, Variante: {variante}, Quantité: {panier.quantite}")
                    
                if articles_description:
                    articles_description += " + "
                
                # Construire la description détaillée
                if variante and variante.couleur and variante.pointure:
                    articles_description += f"{article.nom} {variante.couleur.nom} P{variante.pointure.pointure} , Q{panier.quantite}"
                elif variante and variante.couleur:
                    articles_description += f"{article.nom} {variante.couleur.nom} , Q{panier.quantite}"
                elif variante and variante.pointure:
                    articles_description += f"{article.nom} P{variante.pointure.pointure} , Q{panier.quantite}"
                else:
                    articles_description += f"{article.nom} , Q{panier.quantite}"
            
            print(f"  📝 Description finale: '{articles_description}'")
            print(f"  🔢 Total articles: {total_articles}")
            
            # Préparer les données de la commande
            commande_data = {
                'id_yz': commande.id_yz,
                'num_cmd': commande.num_cmd,
                'client_nom': f"{commande.client.nom} {commande.client.prenom}" if commande.client else "Client non défini",
                'client_telephone': commande.client.numero_tel if commande.client else "",
                'ville_client': commande.ville_init or (commande.ville.nom if commande.ville else ""),
                'adresse': commande.client.adresse if commande.client else "",
                'total': commande.total_cmd,
                'date_confirmation': commande.date_creation,
                'articles_description': articles_description,
                'total_articles': total_articles
            }
            
            # Rendre le template HTML pour cette commande
            ticket_html = render_to_string('Superpreparation/partials/_ticket_commande.html', {
                'commande': commande_data
            })
            
            # Si c'est pour l'impression directe, supprimer les checkboxes mais garder les badges
            if direct_print:
                import re
                # Supprimer les checkboxes du HTML
                ticket_html = re.sub(r'<input[^>]*class="[^"]*ticket-checkbox[^"]*"[^>]*>', '', ticket_html)
                # S'assurer que les badges d'articles sont visibles
                ticket_html = ticket_html.replace('class="articles-count-badge"', 'class="articles-count-badge" style="display: inline-block !important;"')
            
            all_tickets_html.append(ticket_html)
            commandes_data.append(commande_data)
        
        if not all_tickets_html:
            return JsonResponse({'error': 'Aucune commande confirmée trouvée'}, status=404)
        
        print(f"✅ {len(all_tickets_html)} tickets générés avec succès")
        
        # Combiner tous les tickets avec un conteneur de grille
        combined_html = f'''
        <div class="ticket-commande-container" style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; padding: 15px; max-width: 100%; width: 100%; -webkit-print-color-adjust: exact !important; color-adjust: exact !important; print-color-adjust: exact !important;">
            {''.join(all_tickets_html)}
        </div>
        '''
        
        return JsonResponse({
            'success': True,
            'html': combined_html,
            'commandes': commandes_data
        })
        
    except Exception as e:
        import traceback
        print(f"Erreur dans api_ticket_commande_multiple: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération des tickets multiples: {str(e)}'
        }, status=500)


@superviseur_preparation_required
@transaction.atomic
def api_finaliser_preparation(request, commande_id):
    """
    API pour finaliser la préparation d'une commande (changer l'état vers "Préparée")
    Seul le superviseur peut finaliser la préparation
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode non autorisée'}, status=405)
    
    try:
        # Récupérer la commande
        commande = get_object_or_404(Commande, id=commande_id)
        
        # Vérifier que l'opérateur est bien un superviseur
        operateur = request.user.profil_operateur
        if not operateur or operateur.type_operateur != 'SUPERVISEUR_PREPARATION':
            return JsonResponse({
                'success': False, 
                'error': 'Seuls les superviseurs de préparation peuvent finaliser la préparation'
            }, status=403)
        
        # Vérifier que la commande est dans l'état "Emballée"
        etat_actuel = commande.etat_actuel
        if not etat_actuel or etat_actuel.enum_etat.libelle != 'Emballée':
            return JsonResponse({
                'success': False, 
                'error': 'La commande doit être dans l\'état "Emballée" pour être finalisée'
            }, status=400)
        
        # Récupérer l'état "Préparée"
        try:
            etat_preparee = EnumEtatCmd.objects.get(libelle='Préparée')
        except EnumEtatCmd.DoesNotExist:
            return JsonResponse({
                'success': False, 
                'error': 'L\'état "Préparée" n\'existe pas dans le système'
            }, status=500)
        
        # Fermer l'état actuel (Emballée)
        etat_actuel.date_fin = timezone.now()
        etat_actuel.save()
        
        # Créer le nouvel état (Préparée)
        nouvel_etat = EtatCommande.objects.create(
            commande=commande,
            enum_etat=etat_preparee,
            operateur=operateur,
            date_debut=timezone.now(),
            date_fin=None  # État actif
        )
        
        # Créer une opération pour tracer l'action
        Operation.objects.create(
            commande=commande,
            type_operation='FINALISATION_PREPARATION',
            operateur=operateur,
            date_operation=timezone.now(),
            description=f'Préparation finalisée par le superviseur {operateur.nom} {operateur.prenom}'
        )
        
        print(f"✅ Préparation finalisée pour la commande {commande.id_yz} par {operateur.nom}")
        
        return JsonResponse({
            'success': True,
            'message': f'La préparation de la commande {commande.id_yz} a été finalisée avec succès',
            'nouvel_etat': 'Préparée',
            'operateur': f'{operateur.nom} {operateur.prenom}',
            'date_finalisation': timezone.now().strftime('%d/%m/%Y %H:%M')
        })
        
    except Commande.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Commande non trouvée'
        }, status=404)
    except Exception as e:
        print(f"Erreur lors de la finalisation de la préparation: {str(e)}")
        return JsonResponse({
            'success': False, 
            'error': f'Erreur lors de la finalisation: {str(e)}'
        }, status=500)

@csrf_exempt
@login_required
def export_commandes_envoi_excel(request, envoi_id):
    """Exporter les commandes préparées d'un envoi en fichier Excel"""
    from django.http import HttpResponse
    from django.utils import timezone
    
    try:
        print(f"🔍 DEBUG - Export Excel pour envoi ID: {envoi_id}")
        
        # Récupérer l'envoi
        envoi = Envoi.objects.select_related('region').get(id=envoi_id)
        print(f"🔍 DEBUG - Envoi trouvé: {envoi.numero_envoi}, Région: {envoi.region.nom_region}")
        
        # Importer openpyxl après avoir vérifié l'envoi
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            print("🔍 DEBUG - Imports openpyxl réussis")
        except ImportError as import_error:
            print(f"🚫 ERREUR - Import openpyxl échoué: {import_error}")
            return JsonResponse({
                'success': False,
                'error': f'Erreur d\'import openpyxl: {import_error}'
            }, status=500)
        
        # Récupérer les commandes associées à cet envoi. Si aucune n'est encore liée,
        # fallback: prendre les commandes "Préparée" de la région de l'envoi.
        commandes = (
            Commande.objects.filter(envoi=envoi)
            .select_related('client', 'ville', 'ville__region')
            .prefetch_related(
                'paniers__article',
                'paniers__variante',
                'paniers__variante__couleur',
                'paniers__variante__pointure',
            )
            .distinct()
        )

        if not commandes.exists():
            print("🔎 Aucun lien direct commande→envoi. Fallback sur commandes 'Préparée' de la région.")
            commandes = (
                Commande.objects.filter(
                    ville__region=envoi.region,
                    etats__enum_etat__libelle='Préparée',
                    etats__date_fin__isnull=True,
                )
                .select_related('client', 'ville', 'ville__region')
                .prefetch_related(
                    'paniers__article',
                    'paniers__variante',
                    'paniers__variante__couleur',
                    'paniers__variante__pointure',
                )
                .distinct()
            )
        
        print(f"🔍 DEBUG - Nombre de commandes trouvées: {commandes.count()}")
        
        # Test: Créer un workbook simple d'abord
        try:
            wb = openpyxl.Workbook()
            print("🔍 DEBUG - Workbook créé avec succès")
            ws = wb.active
            ws.title = f"Envoi_{envoi.numero_envoi}"[:31]  # Limiter à 31 caractères
            print(f"🔍 DEBUG - Worksheet configuré: {ws.title}")
        except Exception as wb_error:
            print(f"🚫 ERREUR - Création workbook échouée: {wb_error}")
            raise wb_error
        
        # TEST SIMPLE: Ajoutons juste quelques données de base
        try:
            # En-tête simple
            ws['A1'] = f"EXPORT COMMANDES - ENVOI {envoi.numero_envoi}"
            ws['A2'] = f"Région: {envoi.region.nom_region}"
            ws['A3'] = f"Date: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
            
            # En-têtes des colonnes
            headers = ['N° Commande', 'Client', 'Téléphone', 'Adresse', 'Ville', 'Région', 'Total', 'Panier']
            for col, header in enumerate(headers, 1):
                ws.cell(row=5, column=col, value=header)
            
            print(f"🔍 DEBUG - En-têtes ajoutés")
            
            # Données simplifiées des commandes
            row = 6
            for commande in commandes:
                try:
                    # Données de base
                    client_nom = (
                        f"{(commande.client.prenom or '').strip()} {(commande.client.nom or '').strip()}".strip()
                        if commande.client else "N/A"
                    )
                    adresse_livraison = commande.adresse or (commande.client.adresse if getattr(commande, 'client', None) else '') or ""
                    ville_nom = commande.ville.nom if commande.ville else "N/A"
                    region_nom = commande.ville.region.nom_region if getattr(commande, 'ville', None) and commande.ville.region else (
                        envoi.region.nom_region if envoi and envoi.region else "N/A"
                    )

                    # Détails du panier (items)
                    items_texts = []
                    for p in commande.paniers.all():
                        try:
                            article_nom = p.article.nom if p.article else "Article N/A"
                            variante_text = ""
                            if p.variante:
                                couleur = getattr(p.variante.couleur, 'nom', None)
                                pointure = getattr(p.variante.pointure, 'pointure', None)
                                details = [v for v in [couleur, pointure] if v]
                                if details:
                                    variante_text = f" ({' - '.join(details)})"
                            items_texts.append(f"{p.quantite}x {article_nom}{variante_text}")
                        except Exception as p_err:
                            print(f"⚠️ ERREUR panier commande {commande.id}: {p_err}")
                            continue
                    panier_str = "\n".join(items_texts) if items_texts else ""

                    # Ecrire la ligne
                    ws.cell(row=row, column=1, value=commande.num_cmd)
                    ws.cell(row=row, column=2, value=client_nom)
                    ws.cell(row=row, column=3, value=getattr(commande.client, 'numero_tel', '') if commande.client else '')
                    ws.cell(row=row, column=4, value=adresse_livraison)
                    ws.cell(row=row, column=5, value=ville_nom)
                    ws.cell(row=row, column=6, value=region_nom)
                    # Total commande (montant total des articles)
                    ws.cell(row=row, column=7, value=float(getattr(commande, 'total_cmd', 0) or 0))
                    panier_cell = ws.cell(row=row, column=8, value=panier_str)
                    panier_cell.alignment = Alignment(wrap_text=True, vertical='top')

                    row += 1

                except Exception as row_error:
                    print(f"🚫 ERREUR - Ligne commande {commande.id}: {row_error}")
                    continue
            
            print(f"🔍 DEBUG - {row-6} lignes de données ajoutées")
            
        except Exception as data_error:
            print(f"🚫 ERREUR - Ajout des données: {data_error}")
            raise data_error
        
        # Préparer la réponse HTTP
        try:
            print("🔍 DEBUG - Préparation de la réponse HTTP")
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Envoi_{envoi.numero_envoi}_Commandes.xlsx"'
            print(f"🔍 DEBUG - Headers HTTP configurés")
            
            # Sauvegarder le workbook dans la réponse
            print("🔍 DEBUG - Sauvegarde du workbook...")
            wb.save(response)
            print("🔍 DEBUG - Workbook sauvegardé avec succès")
            
            return response
            
        except Exception as save_error:
            print(f"🚫 ERREUR - Sauvegarde échouée: {save_error}")
            raise save_error
        
    except Envoi.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Envoi non trouvé'}, status=404)
    except Exception as e:
        import traceback
        print(f"Erreur dans export_commandes_envoi_excel: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors de la génération du fichier Excel: {str(e)}'
        }, status=500)


@superviseur_preparation_required
def api_commande_info(request, commande_id):
    """API pour récupérer les informations détaillées d'une commande pour la modale"""
    try:
        # Récupérer la commande
        commande = get_object_or_404(Commande, id=commande_id)
        
        # Récupérer les informations détaillées de manière sécurisée
        try:
            etat_actuel = commande.etat_actuel
        except:
            etat_actuel = None
            
        try:
            etat_precedent = commande.etat_precedent
        except:
            etat_precedent = None
            
        try:
            etat_confirmation = commande.etat_confirmation
        except:
            etat_confirmation = None
            
        try:
            operations = commande.operations.all().order_by('-date_operation')
        except:
            operations = []
        
        # Récupérer les informations détaillées
        context = {
            'commande': commande,
            'etat_actuel': etat_actuel,
            'etat_precedent': etat_precedent,
            'etat_confirmation': etat_confirmation,
            'operations': operations,
        }
        
        # Rendre le template HTML pour les informations détaillées
        html_content = render_to_string('Superpreparation/partials/_commande_info_modal.html', context)
        
        return JsonResponse({
            'success': True,
            'html_content': html_content
        })
        
    except Exception as e:
        import traceback
        print(f"Erreur dans api_commande_info: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'Erreur lors du chargement des informations: {str(e)}'
        }, status=500)
